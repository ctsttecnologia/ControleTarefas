
# usuario/services/excel_export.py
"""
Service dedicado à geração do relatório Excel de usuários.
Isola a lógica pesada de formatação OpenPyXL fora da view.
"""
import io

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from usuario.models import Filial


# Paleta de cores
_AZUL_ESCURO = "1F3864"
_CINZA_CLARO = "F2F2F2"
_VERDE = "548235"
_VERMELHO = "C00000"
_BORDA_COR = "B4C6E7"

_COLUNAS = [
    {"header": "Nº", "width": 5, "field": "numero"},
    {"header": "Nome Completo", "width": 28, "field": "nome"},
    {"header": "Usuário", "width": 16, "field": "username"},
    {"header": "E-mail", "width": 30, "field": "email"},
    {"header": "Grupos", "width": 22, "field": "grupos"},
    {"header": "Filial Ativa", "width": 18, "field": "filial_ativa"},
    {"header": "Filiais Permitidas", "width": 26, "field": "filiais_permitidas"},
    {"header": "Status", "width": 10, "field": "status"},
    {"header": "Staff", "width": 8, "field": "is_staff"},
    {"header": "Último Acesso", "width": 18, "field": "last_login"},
]


def gerar_excel_usuarios(queryset, request_user, active_filial_id=None, search=""):
    """
    Gera um arquivo Excel (BytesIO) com relatório de usuários.
    Retorna (buffer, filename).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuários Cadastrados"

    _configurar_pagina(ws)
    _configurar_cabecalho_impressao(ws)

    # Estilos
    estilos = _criar_estilos()

    # Larguras
    total_colunas = len(_COLUNAS)
    for idx, col in enumerate(_COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col["width"]

    # Título + subtítulo + cabeçalho
    agora = timezone.localtime(timezone.now())
    _escrever_titulo(ws, total_colunas, estilos)
    _escrever_subtitulo(ws, total_colunas, estilos, request_user,
                        agora, active_filial_id, search, queryset.count())
    ws.row_dimensions[3].height = 6

    row_header = 4
    _escrever_cabecalho_tabela(ws, row_header, estilos)
    ws.freeze_panes = ws.cell(row=row_header + 1, column=1).coordinate

    # Dados
    row_start = row_header + 1
    for idx, usuario in enumerate(queryset, start=1):
        _escrever_linha_usuario(ws, row_start + idx - 1, idx, usuario, estilos)

    # Resumo
    total = queryset.count()
    ativos = queryset.filter(is_active=True).count()
    row_resumo = row_start + total + 2
    _escrever_resumo(ws, row_resumo, total_colunas, total, ativos, estilos)

    # Área de impressão
    ultima_col = get_column_letter(total_colunas)
    ws.print_area = f"A1:{ultima_col}{row_resumo}"
    ws.print_title_rows = f'1:{row_header}'

    # Gera buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"usuarios_cadastrados_{agora.strftime('%Y%m%d_%H%M')}.xlsx"
    return buffer, filename


# ─────────────────────────────────────────────────────────────
# Helpers privados
# ─────────────────────────────────────────────────────────────

def _configurar_pagina(ws):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.6, bottom=0.6,
        header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True


def _configurar_cabecalho_impressao(ws):
    ws.oddHeader.center.text = "&B&14Relatório de Usuários Cadastrados"
    ws.oddFooter.left.text = "Emitido em: &D às &T"
    ws.oddFooter.center.text = "Página &P de &N"
    ws.oddFooter.right.text = "Confidencial"


def _criar_estilos():
    return {
        'borda_fina': Border(
            left=Side(style='thin', color=_BORDA_COR),
            right=Side(style='thin', color=_BORDA_COR),
            top=Side(style='thin', color=_BORDA_COR),
            bottom=Side(style='thin', color=_BORDA_COR),
        ),
        'borda_cabecalho': Border(
            left=Side(style='thin', color="FFFFFF"),
            right=Side(style='thin', color="FFFFFF"),
            top=Side(style='medium', color=_AZUL_ESCURO),
            bottom=Side(style='medium', color=_AZUL_ESCURO),
        ),
        'font_titulo': Font(name='Calibri', bold=True, size=16, color="FFFFFF"),
        'font_subtitulo': Font(name='Calibri', size=10, color="666666", italic=True),
        'font_cabecalho': Font(name='Calibri', bold=True, size=10, color="FFFFFF"),
        'font_dados': Font(name='Calibri', size=9, color="333333"),
        'font_ativo': Font(name='Calibri', size=9, bold=True, color=_VERDE),
        'font_inativo': Font(name='Calibri', size=9, bold=True, color=_VERMELHO),
        'font_rodape': Font(name='Calibri', size=8, color="999999", italic=True),
        'fill_titulo': PatternFill(start_color=_AZUL_ESCURO, end_color=_AZUL_ESCURO, fill_type='solid'),
        'fill_cabecalho': PatternFill(start_color=_AZUL_ESCURO, end_color=_AZUL_ESCURO, fill_type='solid'),
        'fill_zebra': PatternFill(start_color=_CINZA_CLARO, end_color=_CINZA_CLARO, fill_type='solid'),
        'fill_branco': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid'),
        'align_centro': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'align_esquerda': Alignment(horizontal='left', vertical='center', wrap_text=True),
    }


def _escrever_titulo(ws, total_colunas, estilos):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_colunas)
    cel = ws.cell(row=1, column=1)
    cel.value = "📋 RELATÓRIO DE USUÁRIOS CADASTRADOS"
    cel.font = estilos['font_titulo']
    cel.fill = estilos['fill_titulo']
    cel.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    for col_idx in range(2, total_colunas + 1):
        ws.cell(row=1, column=col_idx).fill = estilos['fill_titulo']


def _escrever_subtitulo(ws, total_colunas, estilos, user, agora,
                        active_filial_id, search, total_registros):
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_colunas)

    filial_filtro = "Todas"
    if active_filial_id:
        filial = Filial.objects.filter(pk=active_filial_id).only('nome').first()
        filial_filtro = filial.nome if filial else "N/D"

    texto = (
        f"Emitido por: {user.get_full_name() or user.username}  |  "
        f"Data: {agora.strftime('%d/%m/%Y %H:%M')}  |  "
        f"Filial filtro: {filial_filtro}  |  "
        f"Total de registros: {total_registros}"
    )
    if search:
        texto += f'  |  Busca: "{search}"'

    cel = ws.cell(row=2, column=1)
    cel.value = texto
    cel.font = estilos['font_subtitulo']
    cel.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20


def _escrever_cabecalho_tabela(ws, row, estilos):
    ws.row_dimensions[row].height = 26
    for col_idx, col_def in enumerate(_COLUNAS, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = col_def["header"]
        cell.font = estilos['font_cabecalho']
        cell.fill = estilos['fill_cabecalho']
        cell.alignment = estilos['align_centro']
        cell.border = estilos['borda_cabecalho']


def _escrever_linha_usuario(ws, row_num, idx, usuario, estilos):
    is_zebra = idx % 2 == 0
    fill_row = estilos['fill_zebra'] if is_zebra else estilos['fill_branco']

    # Preparar dados
    nome = usuario.get_full_name() or usuario.username
    grupos = ", ".join(usuario.groups.values_list('name', flat=True)) or "—"
    filial_ativa = str(usuario.filial_ativa) if usuario.filial_ativa else "—"
    filiais_perm = ", ".join(
        usuario.filiais_permitidas.values_list('nome', flat=True)
    ) or "—"
    status = "Ativo" if usuario.is_active else "Inativo"
    staff = "Sim" if usuario.is_staff else "Não"

    if usuario.last_login:
        last_login = timezone.localtime(usuario.last_login).strftime('%d/%m/%Y %H:%M')
    else:
        last_login = "Nunca"

    dados = [idx, nome, usuario.username, usuario.email, grupos,
             filial_ativa, filiais_perm, status, staff, last_login]

    ws.row_dimensions[row_num].height = 22

    for col_idx, valor in enumerate(dados, start=1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = valor
        cell.font = estilos['font_dados']
        cell.fill = fill_row
        cell.border = estilos['borda_fina']

        campo = _COLUNAS[col_idx - 1]["field"]
        if campo in ("numero", "status", "is_staff", "last_login"):
            cell.alignment = estilos['align_centro']
        else:
            cell.alignment = estilos['align_esquerda']

        if campo == "status":
            cell.font = estilos['font_ativo'] if valor == "Ativo" else estilos['font_inativo']


def _escrever_resumo(ws, row, total_colunas, total, ativos, estilos):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_colunas)
    cel = ws.cell(row=row, column=1)
    cel.value = (
        f"Total: {total} usuário(s)  |  "
        f"Ativos: {ativos}  |  "
        f"Inativos: {total - ativos}"
    )
    cel.font = estilos['font_rodape']
    cel.alignment = Alignment(horizontal='right', vertical='center')

