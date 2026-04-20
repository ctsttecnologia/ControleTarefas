# usuario/views.py

import datetime
from django.db.models import Q
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.views import (
    LoginView, LogoutView,
    PasswordChangeView,
    PasswordResetView, PasswordResetDoneView,
    PasswordResetConfirmView, PasswordResetCompleteView
)
from django.contrib.auth.forms import SetPasswordForm
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, FormView
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import ProtectedError

from usuario.models import Usuario, Group, Filial, GroupCardPermissions
from usuario.forms import (
    CustomUserCreationForm, CustomUserChangeForm,
    GrupoForm, CustomPasswordChangeForm, FilialForm
)
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from django.utils import timezone

# =============================================================================
# MIXINS DE PERMISSÃO
# =============================================================================

class StaffRequiredMixin(UserPassesTestMixin):
    """Restringe acesso a usuários staff."""
    def test_func(self):
        return self.request.user.is_staff


class SuperuserRequiredMixin(UserPassesTestMixin):
    """Restringe acesso a superusuários."""
    def test_func(self):
        return self.request.user.is_superuser


class AdminOrManagerMixin(UserPassesTestMixin):
    """Restringe acesso a superusuários, gerentes ou administradores."""
    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.is_gerente or user.is_administrador


# =============================================================================
# AUTENTICAÇÃO (Login / Logout / Seleção de Filial)
# =============================================================================

class CustomLoginView(LoginView):
    """Login com gerenciamento automático de filial ativa na sessão."""
    template_name = 'usuario/login.html'
    redirect_authenticated_user = True

    def form_valid(self, form):
        response = super().form_valid(form)
        filial_ativa = self._determinar_filial_ativa()

        if filial_ativa:
            self._registrar_filial_na_sessao(filial_ativa)
        else:
            return self._handle_usuario_sem_filial()

        return response

    def _determinar_filial_ativa(self):
        user = self.request.user
        if user.filial_ativa:
            return user.filial_ativa
        if user.filiais_permitidas.exists():
            return user.filiais_permitidas.only('id', 'nome').first()
        return None

    def _registrar_filial_na_sessao(self, filial):
        self.request.session['active_filial_id'] = filial.id
        messages.info(self.request, f"Você está conectado na filial: {filial.nome}")

    def _handle_usuario_sem_filial(self):
        messages.error(
            self.request,
            "Você não está associado a nenhuma filial. Contate o administrador."
        )
        logout(self.request)
        return redirect('usuario:login')


class CustomLogoutView(LogoutView):
    """Logout com limpeza da sessão de filial."""
    next_page = reverse_lazy('usuario:login')

    def dispatch(self, request, *args, **kwargs):
        request.session.pop('active_filial_id', None)
        return super().dispatch(request, *args, **kwargs)


class SelecionarFilialView(LoginRequiredMixin, View):
    """Permite ao usuário trocar a filial ativa."""

    def post(self, request, *args, **kwargs):
        filial_id_str = request.POST.get('filial_id')

        # Superusuário limpando filtro ("Todas as Filiais")
        if filial_id_str == '0' and request.user.is_superuser:
            request.session.pop('active_filial_id', None)
            request.user.filial_ativa = None
            request.user.save(update_fields=['filial_ativa'])
            messages.success(request, "Exibindo dados de todas as filiais.")

        elif filial_id_str:
            try:
                filial_id = int(filial_id_str)
                filial_selecionada = request.user.filiais_permitidas.get(pk=filial_id)

                request.session['active_filial_id'] = filial_selecionada.id
                request.user.filial_ativa = filial_selecionada
                request.user.save(update_fields=['filial_ativa'])
                messages.success(request, f"Filial alterada para: {filial_selecionada.nome}")

            except (Filial.DoesNotExist, ValueError):
                messages.error(request, "Seleção inválida ou sem permissão para esta filial.")

        return redirect(request.POST.get('next', reverse('usuario:profile')))


# =============================================================================
# PERFIL E SENHA DO USUÁRIO LOGADO
# =============================================================================

class ProfileView(LoginRequiredMixin, DetailView):
    """Página inicial do usuário com cards de acesso rápido."""
    model = Usuario
    template_name = 'usuario/profile.html'

    def get_object(self, queryset=None):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        all_cards = self._get_all_cards()
        allowed_card_ids = self._get_allowed_card_ids(user)

        visible_cards = []
        for card in all_cards:
            # Verifica se o card é visível
            if not (user.is_superuser
                    or user.has_perm(card['permission'])
                    or card['id'] in allowed_card_ids):
                continue

            # Filtra os links que o usuário pode ver
            filtered_links = [
                link for link in card['links']
                if user.is_superuser
                or user.has_perm(link.get('permission', ''))
                or card['id'] in allowed_card_ids          # ← LINHA ADICIONADA
            ]

            # Só exibe o card se tiver pelo menos 1 link visível
            if filtered_links:
                visible_cards.append({
                    **card,
                    'links': filtered_links,
                })

        context['allowed_cards'] = visible_cards
        return context



    def _get_allowed_card_ids(self, user):
        """Retorna IDs de cards permitidos pelos grupos do usuário."""
        if user.is_superuser:
            return [c['id'] for c in self._get_all_cards()]

        ids = []
        perms = GroupCardPermissions.objects.filter(
            group__in=user.groups.all()
        ).values_list('cards_visiveis', flat=True)

        for cards_list in perms:
            if cards_list:
                ids.extend(cards_list)
        return ids


    def _get_all_cards(self):
        """Definição centralizada de todos os cards do sistema."""
        return [
            {
                'id': 'clientes',
                'title': 'Clientes',
                'permission': 'cliente.view_cliente',
                'icon': 'images/cliente.gif',
                'links': [ 
                    {'url': 'cliente:cliente_create', 'text': 'Cadastrar', 'permission': 'cliente.add_cliente', },
                    
                    {'url': 'cliente:lista_clientes', 'text': 'Lista de Clientes', 'permission': 'cliente.view_cliente'},
                ]
            },
            {
                'id': 'dp',
                'title': 'Departamento Pessoal',
                'permission': 'departamento_pessoal.view_painel_dp',
                'icon': 'images/dp.gif',
                'links': [
                    {'url': 'departamento_pessoal:painel_dp', 'text': 'Painel DP', 
                        'permission': 'departamento_pessoal.view_funcionario'},
                    {'url': 'treinamentos:dashboard', 'text': 'Treinamentos'},
                ]
            },
            {
                'id': 'sst',
                'title': 'Segurança do Trabalho',
                'permission': 'seguranca_trabalho.view_fichaepi',
                'icon': 'images/tst.gif',
                'links': [
                    {'url': 'seguranca_trabalho:dashboard', 'text': 'Painel SST', 'permission': 'seguranca_trabalho.view_fichaepi'},
                    {'url': 'seguranca_trabalho:ficha_list', 'text': 'Fichas de EPI', 'permission': 'seguranca_trabalho.view_fichaepi'},
                    {'url': 'gestao_riscos:lista_riscos', 'text': 'Gestão de Riscos', 'permission': 'gestao_riscos.view_risco'},
                ]
            },
            {
                'id': 'endereco',
                'title': 'Logradouro',
                'permission': 'logradouro.view_logradouro',
                'icon': 'images/cadastro.gif',
                'links': [
                    {'url': 'logradouro:cadastrar_logradouro', 'text': 'Cadastrar', 'permission': 'logradouro.add_logradouro'},
                    {'url': 'logradouro:listar_logradouros', 'text': 'Lista de Logradouros', 'permission': 'logradouro.view_logradouro'},
                ]
            },
            # ══════════════════════════════════════════════════════════
            # ANTIGO "ga" DIVIDIDO EM CARDS INDIVIDUAIS
            # ══════════════════════════════════════════════════════════
            {
                'id': 'ata_reuniao',
                'title': 'Atas de Reunião',
                'permission': 'ata_reuniao.view_atareuniao',
                'icon': 'images/reuniao.png',
                'links': [
                    {'url': 'ata_reuniao:ata_reuniao_dashboard', 'text': 'Painel de Atas', 'permission': 'ata_reuniao.view_atareuniao'},
                    {'url': 'ata_reuniao:ata_reuniao_list', 'text': 'Lista de Atas', 'permission': 'ata_reuniao.view_atareuniao'},
                ]
            },
            {
                'id': 'suprimentos',
                'title': 'Suprimentos',
                'permission': 'suprimentos.view_pedido',
                'icon': 'images/suprimentos.gif', 
                'links': [
                    {'url': 'suprimentos:dashboard', 'text': 'Suprimentos', 'permission': 'suprimentos.view_pedido'},
                ]
            },
            {
                'id': 'documentos',
                'title': 'Documentos',
                'permission': 'documentos.view_documento',
                'icon': 'images/documentos.gif',  
                'links': [
                    {'url': 'documentos:lista', 'text': 'Gestão de Documentos', 'permission': 'documentos.view_documento'},
                ]
            },
            {
                'id': 'telefones',
                'title': 'Controle de Telefones',
                'permission': 'controle_de_telefone.view_linhatelefonica',
                'icon': 'images/telefones.gif',  
                'links': [
                    {'url': 'controle_de_telefone:dashboard', 'text': 'Gestão de Telefones', 'permission': 'controle_de_telefone.view_linhatelefonica'},
                ]
            },
            {
                'id': 'estoque',
                'title': 'Estoque',
                'permission': 'seguranca_trabalho.view_equipamento',
                'icon': 'images/estoque.gif',  
                'links': [
                    {'url': 'seguranca_trabalho:equipamento_list', 'text': 'Equipamentos e Material', 'permission': 'seguranca_trabalho.view_equipamento'},
                ]
            },
            # ══════════════════════════════════════════════════════════
            # OUTROS CARDS (mantém igual)
            # ══════════════════════════════════════════════════════════
            {
                'id': 'veiculos',
                'title': 'Veículos',
                'permission': 'automovel.view_carro',
                'icon': 'images/carro.gif',
                'links': [
                    {'url': 'automovel:carro_list', 'text': 'Frota', 'permission': 'automovel.view_carro'},
                    {'url': 'automovel:agendamento_list', 'text': 'Agendamentos', 'permission': 'automovel.view_agendamento'},
                    {'url': 'automovel:dashboard', 'text': 'Relatórios', 'permission': 'automovel.view_dashboard'},
                ]
            },
            {
                'id': 'operacao',
                'title': 'Operação',
                'permission': 'ferramentas.view_ferramentas',
                'icon': 'images/serviço.gif',
                'links': [
                    {'url': 'tarefas:dashboard', 'text': 'Tarefas', 'permission': 'tarefas.view_dashboard'},
                    {'url': 'ferramentas:dashboard', 'text': 'Ferramentas', 'permission': 'ferramentas.view_dashboard'},
                ]
            },
            {
                'id': 'main_dashboard',
                'title': 'Dashboard Integrado',
                'permission': 'GARANT_ALL',
                'icon': 'images/favicon.ico',
                'links': [
                    {'url': 'dashboard:dashboard_geral', 'text': 'Visão Geral', 'permission': 'dashboard.view_dashboard_geral'},
                ]
            },
        ]



class CustomPasswordChangeView(LoginRequiredMixin, PasswordChangeView):
    """Alteração de senha do próprio usuário logado."""
    form_class = CustomPasswordChangeForm
    template_name = 'usuario/alterar_senha.html'
    success_url = reverse_lazy('usuario:profile')

    def form_valid(self, form):
        messages.success(self.request, 'Sua senha foi alterada com sucesso!')
        return super().form_valid(form)


# =============================================================================
# CRUD DE USUÁRIOS (Staff/Admin)
# =============================================================================

class UserListView(LoginRequiredMixin, StaffRequiredMixin, ListView):
    """Lista de usuários com busca e filtro por filial ativa."""
    model = Usuario
    template_name = 'usuario/lista_usuarios.html'
    context_object_name = 'usuarios'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('filial_ativa').order_by('first_name')
        active_filial_id = self.request.session.get('active_filial_id')

        if not self.request.user.is_superuser and active_filial_id:
            qs = qs.filter(filiais_permitidas__id=active_filial_id).distinct()

        search = self.request.GET.get('q', '').strip()
        if search:
            qs = qs.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )
        return qs


class UserCreateView(LoginRequiredMixin, AdminOrManagerMixin, CreateView):
    """Criação de novo usuário."""
    model = Usuario
    form_class = CustomUserCreationForm
    template_name = 'usuario/form_usuario.html'
    success_url = reverse_lazy('usuario:lista_usuarios')

    def form_valid(self, form):
        response = super().form_valid(form)
        user = self.object

        # Define a primeira filial permitida como ativa
        primeira_filial = user.filiais_permitidas.first()
        if primeira_filial:
            user.filial_ativa = primeira_filial
            user.save(update_fields=['filial_ativa'])

        messages.success(self.request, f"Usuário '{user.username}' criado com sucesso.")
        return response


class UserUpdateView(LoginRequiredMixin, AdminOrManagerMixin, UpdateView):
    """Edição de usuário existente."""
    model = Usuario
    form_class = CustomUserChangeForm
    template_name = 'usuario/form_usuario.html'
    success_url = reverse_lazy('usuario:lista_usuarios')

    def get_queryset(self):
        return super().get_queryset().select_related('filial_ativa')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        if self.object:
            kwargs['filiais_permitidas_qs'] = self.object.filiais_permitidas.all()
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)

        # Garante que filial_ativa esteja entre as permitidas
        user = self.object
        if user.filial_ativa and user.filial_ativa not in user.filiais_permitidas.all():
            user.filial_ativa = user.filiais_permitidas.first()
            user.save(update_fields=['filial_ativa'])

        messages.success(self.request, f"Usuário '{form.instance.username}' atualizado com sucesso.")
        return response



class UserToggleActiveView(LoginRequiredMixin, AdminOrManagerMixin, View):
    """Ativar/desativar um usuário."""

    def post(self, request, *args, **kwargs):
        user_to_toggle = get_object_or_404(Usuario, pk=self.kwargs.get('pk'))

        if user_to_toggle == request.user:
            messages.error(request, "Você não pode desativar a si mesmo.")
            return redirect('usuario:lista_usuarios')

        # Staff não pode desativar superusuários
        if user_to_toggle.is_superuser and not request.user.is_superuser:
            messages.error(request, "Apenas superusuários podem desativar outros superusuários.")
            return redirect('usuario:lista_usuarios')

        user_to_toggle.is_active = not user_to_toggle.is_active
        user_to_toggle.save(update_fields=['is_active'])

        status = "ativado" if user_to_toggle.is_active else "desativado"
        messages.success(request, f"Usuário '{user_to_toggle.username}' {status} com sucesso.")
        return redirect('usuario:lista_usuarios')


class UserSetPasswordView(LoginRequiredMixin, SuperuserRequiredMixin, FormView):
    """Permite superusuário definir nova senha para outro usuário."""
    form_class = SetPasswordForm
    template_name = 'usuario/definir_senha_form.html'
    success_url = reverse_lazy('usuario:lista_usuarios')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = get_object_or_404(Usuario, pk=self.kwargs['pk'])
        return kwargs

    def form_valid(self, form):
        form.save()
        messages.success(self.request, f"Senha de '{form.user.username}' redefinida com sucesso.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario_alvo'] = get_object_or_404(Usuario, pk=self.kwargs['pk'])
        return context


# =============================================================================
# CRUD DE GRUPOS (Superusuário)
# =============================================================================

class GroupListView(LoginRequiredMixin, SuperuserRequiredMixin, ListView):
    model = Group
    template_name = 'usuario/grupo_lista.html'
    context_object_name = 'grupos'


class GroupCreateView(LoginRequiredMixin, SuperuserRequiredMixin, CreateView):
    model = Group
    form_class = GrupoForm
    template_name = 'usuario/grupo_form.html'
    success_url = reverse_lazy('usuario:grupo_lista')

    def form_valid(self, form):
        messages.success(self.request, f"Grupo '{form.instance.name}' criado com sucesso.")
        return super().form_valid(form)


class GroupUpdateView(LoginRequiredMixin, SuperuserRequiredMixin, UpdateView):
    model = Group
    form_class = GrupoForm
    template_name = 'usuario/grupo_form.html'
    success_url = reverse_lazy('usuario:grupo_lista')

    def form_valid(self, form):
        messages.success(self.request, f"Grupo '{form.instance.name}' atualizado com sucesso.")
        return super().form_valid(form)


class GroupDeleteView(LoginRequiredMixin, SuperuserRequiredMixin, DeleteView):
    model = Group
    template_name = 'usuario/grupo_confirmar_exclusao.html'
    success_url = reverse_lazy('usuario:grupo_lista')

    def form_valid(self, form):
        messages.success(self.request, f"Grupo excluído com sucesso.")
        return super().form_valid(form)


class GerenciarGruposUsuarioView(LoginRequiredMixin, SuperuserRequiredMixin, View):
    """Adicionar/remover grupos de um usuário específico."""
    template_name = 'usuario/gerenciar_grupos_usuario.html'

    def get(self, request, *args, **kwargs):
        usuario = get_object_or_404(Usuario, pk=self.kwargs.get('pk'))
        grupos_usuario = usuario.groups.all()
        context = {
            'usuario_alvo': usuario,
            'grupos_usuario': grupos_usuario,
            'grupos_disponiveis': Group.objects.exclude(pk__in=grupos_usuario),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        usuario = get_object_or_404(Usuario, pk=self.kwargs.get('pk'))
        grupo_id = request.POST.get('grupo')
        acao = request.POST.get('acao')

        if grupo_id and acao:
            grupo = get_object_or_404(Group, pk=grupo_id)
            if acao == 'adicionar':
                usuario.groups.add(grupo)
                messages.success(request, f"Grupo '{grupo.name}' adicionado a '{usuario.username}'.")
            elif acao == 'remover':
                usuario.groups.remove(grupo)
                messages.success(request, f"Grupo '{grupo.name}' removido de '{usuario.username}'.")

        return redirect('usuario:gerenciar_grupos_usuario', pk=usuario.pk)


# =============================================================================
# CRUD DE FILIAIS (Superusuário)
# =============================================================================

class FilialListView(LoginRequiredMixin, SuperuserRequiredMixin, ListView):
    model = Filial
    template_name = 'usuario/filial_list.html'
    context_object_name = 'filiais'


class FilialCreateView(LoginRequiredMixin, SuperuserRequiredMixin, CreateView):
    model = Filial
    form_class = FilialForm
    template_name = 'usuario/filial_form.html'
    success_url = reverse_lazy('usuario:filial_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = 'Adicionar Nova Filial'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Filial cadastrada com sucesso!')
        return super().form_valid(form)


class FilialUpdateView(LoginRequiredMixin, SuperuserRequiredMixin, UpdateView):
    model = Filial
    form_class = FilialForm
    template_name = 'usuario/filial_form.html'
    success_url = reverse_lazy('usuario:filial_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f'Editar Filial: {self.object.nome}'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Filial atualizada com sucesso!')
        return super().form_valid(form)


class FilialDeleteView(LoginRequiredMixin, SuperuserRequiredMixin, DeleteView):
    model = Filial
    template_name = 'usuario/filial_confirm_delete.html'
    success_url = reverse_lazy('usuario:filial_list')

    def post(self, request, *args, **kwargs):
        try:
            response = super().post(request, *args, **kwargs)
            messages.success(request, 'Filial excluída com sucesso!')
            return response
        except ProtectedError:
            messages.error(
                request,
                'Esta filial não pode ser excluída pois existem registros associados a ela.'
            )
            return redirect('usuario:filial_list')


# =============================================================================
# GERENCIAMENTO DE CARDS POR GRUPO
# =============================================================================

class ManageCardPermissionsView(LoginRequiredMixin, SuperuserRequiredMixin, View):
    """Configura quais cards cada grupo pode ver."""
    template_name = 'usuario/gerenciar_cards.html'

    ALL_CARDS = [
        {'id': 'clientes', 'title': 'Clientes'},
        {'id': 'dp', 'title': 'Departamento Pessoal'},
        {'id': 'sst', 'title': 'Segurança do Trabalho'},
        {'id': 'endereco', 'title': 'Logradouro'},
        {'id': 'veiculos', 'title': 'Veículos'},
        {'id': 'operacao', 'title': 'Operação'},
        {'id': 'main_dashboard', 'title': 'Dashboard Integrado'},
        {'id': 'ata_reuniao', 'title': 'Atas de Reunião'},
        {'id': 'suprimentos', 'title': 'Suprimentos'},
        {'id': 'documentos', 'title': 'Documentos'},
        {'id': 'telefones', 'title': 'Controle de Telefones'},
        {'id': 'estoque', 'title': 'Estoque'},
    ]

    def get(self, request, *args, **kwargs):
        context = {
            'grupos': Group.objects.all().order_by('name'),
            'todos_os_cards': self.ALL_CARDS,
            'permissoes_por_grupo': {
                p.group_id: p.cards_visiveis
                for p in GroupCardPermissions.objects.all()
            },
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        card_ids = [c['id'] for c in self.ALL_CARDS]

        for group in Group.objects.all():
            cards_selecionados = [
                cid for cid in card_ids
                if f'group_{group.id}_{cid}' in request.POST
            ]
            GroupCardPermissions.objects.update_or_create(
                group=group,
                defaults={'cards_visiveis': cards_selecionados}
            )

        messages.success(request, "Permissões de cards atualizadas com sucesso!")
        return redirect('usuario:gerenciar_cards')


# =============================================================================
# RECUPERAÇÃO DE SENHA
# =============================================================================

class CustomPasswordResetView(PasswordResetView):
    template_name = 'usuario/password_reset/form.html'
    email_template_name = 'usuario/password_reset/email.html'
    subject_template_name = 'usuario/password_reset/subject.txt'
    success_url = reverse_lazy('usuario:password_reset_done')


class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'usuario/password_reset/done.html'


class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'usuario/password_reset/confirm.html'
    success_url = reverse_lazy('usuario:password_reset_complete')


class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'usuario/password_reset/complete.html'

# =============================================================================
# EXPORTAÇÃO EXCEL — PLANILHA DE USUÁRIOS CADASTRADOS
# =============================================================================

class ExportarUsuariosExcelView(LoginRequiredMixin, StaffRequiredMixin, View):
    """
    Exporta a lista de usuários cadastrados em formato Excel (.xlsx),
    formatado e otimizado para impressão em papel A4 paisagem.
    """

    def get(self, request, *args, **kwargs):
        # ─── Buscar dados (respeitando filtros de filial e busca) ───
        qs = Usuario.objects.select_related('filial_ativa').prefetch_related(
            'filiais_permitidas', 'groups'
        ).order_by('first_name', 'last_name')

        active_filial_id = request.session.get('active_filial_id')
        if not request.user.is_superuser and active_filial_id:
            qs = qs.filter(filiais_permitidas__id=active_filial_id).distinct()

        search = request.GET.get('q', '').strip()
        if search:
            qs = qs.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )

        # ─── Criar Workbook ───
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuários Cadastrados"

        # ─── Configuração de página para impressão (A4 Paisagem) ───
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # quantas páginas forem necessárias
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(
            left=0.4, right=0.4, top=0.6, bottom=0.6,
            header=0.3, footer=0.3
        )
        ws.print_options.horizontalCentered = True

        # Cabeçalho e rodapé de impressão
        ws.oddHeader.center.text = "&B&14Relatório de Usuários Cadastrados"
        ws.oddFooter.left.text = "Emitido em: &D às &T"
        ws.oddFooter.center.text = "Página &P de &N"
        ws.oddFooter.right.text = "Confidencial"

        # ─── Definição de estilos ───
        AZUL_ESCURO = "1F3864"
        AZUL_CLARO = "D6E4F0"
        CINZA_CLARO = "F2F2F2"
        VERDE = "548235"
        VERMELHO = "C00000"
        BORDA_COR = "B4C6E7"

        borda_fina = Border(
            left=Side(style='thin', color=BORDA_COR),
            right=Side(style='thin', color=BORDA_COR),
            top=Side(style='thin', color=BORDA_COR),
            bottom=Side(style='thin', color=BORDA_COR),
        )

        borda_cabecalho = Border(
            left=Side(style='thin', color="FFFFFF"),
            right=Side(style='thin', color="FFFFFF"),
            top=Side(style='medium', color=AZUL_ESCURO),
            bottom=Side(style='medium', color=AZUL_ESCURO),
        )

        font_titulo = Font(name='Calibri', bold=True, size=16, color="FFFFFF")
        font_subtitulo = Font(name='Calibri', size=10, color="666666", italic=True)
        font_cabecalho = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
        font_dados = Font(name='Calibri', size=9, color="333333")
        font_status_ativo = Font(name='Calibri', size=9, bold=True, color=VERDE)
        font_status_inativo = Font(name='Calibri', size=9, bold=True, color=VERMELHO)
        font_rodape = Font(name='Calibri', size=8, color="999999", italic=True)

        fill_titulo = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type='solid')
        fill_cabecalho = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type='solid')
        fill_zebra = PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type='solid')
        fill_branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

        align_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_esquerda = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # ─── Colunas e larguras ───
        colunas = [
            {"header": "Nº",              "width": 5,   "field": "numero"},
            {"header": "Nome Completo",    "width": 28,  "field": "nome"},
            {"header": "Usuário",          "width": 16,  "field": "username"},
            {"header": "E-mail",           "width": 30,  "field": "email"},
            {"header": "Grupos",           "width": 22,  "field": "grupos"},
            {"header": "Filial Ativa",     "width": 18,  "field": "filial_ativa"},
            {"header": "Filiais Permitidas", "width": 26, "field": "filiais_permitidas"},
            {"header": "Status",           "width": 10,  "field": "status"},
            {"header": "Staff",            "width": 8,   "field": "is_staff"},
            {"header": "Último Acesso",    "width": 18,  "field": "last_login"},
        ]

        total_colunas = len(colunas)

        # Definir larguras das colunas
        for idx, col in enumerate(colunas, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = col["width"]

        # ─── LINHA 1: Título (mesclada) ───
        row_titulo = 1
        ws.merge_cells(
            start_row=row_titulo, start_column=1,
            end_row=row_titulo, end_column=total_colunas
        )
        celula_titulo = ws.cell(row=row_titulo, column=1)
        celula_titulo.value = "📋 RELATÓRIO DE USUÁRIOS CADASTRADOS"
        celula_titulo.font = font_titulo
        celula_titulo.fill = fill_titulo
        celula_titulo.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_titulo].height = 36

        # Preencher fundo das células mescladas
        for col_idx in range(2, total_colunas + 1):
            c = ws.cell(row=row_titulo, column=col_idx)
            c.fill = fill_titulo

        # ─── LINHA 2: Subtítulo / Informações do relatório ───
        row_sub = 2
        ws.merge_cells(
            start_row=row_sub, start_column=1,
            end_row=row_sub, end_column=total_colunas
        )
        agora = timezone.localtime(timezone.now())
        filial_filtro = "Todas"
        if active_filial_id:
            try:
                filial_filtro = Filial.objects.get(pk=active_filial_id).nome
            except Filial.DoesNotExist:
                filial_filtro = "N/D"

        info_texto = (
            f"Emitido por: {request.user.get_full_name() or request.user.username}  |  "
            f"Data: {agora.strftime('%d/%m/%Y %H:%M')}  |  "
            f"Filial filtro: {filial_filtro}  |  "
            f"Total de registros: {qs.count()}"
        )
        if search:
            info_texto += f"  |  Busca: \"{search}\""

        celula_sub = ws.cell(row=row_sub, column=1)
        celula_sub.value = info_texto
        celula_sub.font = font_subtitulo
        celula_sub.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_sub].height = 20

        # ─── LINHA 3: Espaçamento ───
        ws.row_dimensions[3].height = 6

        # ─── LINHA 4: Cabeçalho da tabela ───
        row_header = 4
        ws.row_dimensions[row_header].height = 26

        for col_idx, col_def in enumerate(colunas, start=1):
            cell = ws.cell(row=row_header, column=col_idx)
            cell.value = col_def["header"]
            cell.font = font_cabecalho
            cell.fill = fill_cabecalho
            cell.alignment = align_centro
            cell.border = borda_cabecalho

        # ─── Congelar painel (fixar cabeçalho ao rolar) ───
        ws.freeze_panes = ws.cell(row=row_header + 1, column=1).coordinate

        # ─── LINHAS DE DADOS ───
        row_start = row_header + 1

        for idx, usuario in enumerate(qs, start=1):
            row_num = row_start + idx - 1
            is_zebra = idx % 2 == 0
            fill_row = fill_zebra if is_zebra else fill_branco

            # Preparar dados
            nome_completo = usuario.get_full_name() or usuario.username
            grupos = ", ".join(
                usuario.groups.values_list('name', flat=True)
            ) or "—"
            filial_ativa_nome = str(usuario.filial_ativa) if usuario.filial_ativa else "—"
            filiais_perm = ", ".join(
                usuario.filiais_permitidas.values_list('nome', flat=True)
            ) or "—"
            status_texto = "Ativo" if usuario.is_active else "Inativo"
            staff_texto = "Sim" if usuario.is_staff else "Não"

            if usuario.last_login:
                last_login_local = timezone.localtime(usuario.last_login)
                last_login_texto = last_login_local.strftime('%d/%m/%Y %H:%M')
            else:
                last_login_texto = "Nunca"

            dados_linha = [
                idx,                          # Nº
                nome_completo,                # Nome Completo
                usuario.username,             # Usuário
                usuario.email,                # E-mail
                grupos,                       # Grupos
                filial_ativa_nome,            # Filial Ativa
                filiais_perm,                 # Filiais Permitidas
                status_texto,                 # Status
                staff_texto,                  # Staff
                last_login_texto,             # Último Acesso
            ]

            ws.row_dimensions[row_num].height = 22

            for col_idx, valor in enumerate(dados_linha, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = valor
                cell.font = font_dados
                cell.fill = fill_row
                cell.border = borda_fina

                # Alinhamento por coluna
                campo = colunas[col_idx - 1]["field"]
                if campo in ("numero", "status", "is_staff", "last_login"):
                    cell.alignment = align_centro
                else:
                    cell.alignment = align_esquerda

                # Estilo especial para Status
                if campo == "status":
                    if valor == "Ativo":
                        cell.font = font_status_ativo
                    else:
                        cell.font = font_status_inativo

        # ─── LINHA DE RODAPÉ / RESUMO ───
        total_usuarios = qs.count()
        total_ativos = qs.filter(is_active=True).count()
        total_inativos = total_usuarios - total_ativos

        row_rodape = row_start + total_usuarios + 1
        ws.row_dimensions[row_rodape].height = 6  # Espaçamento

        row_resumo = row_rodape + 1
        ws.merge_cells(
            start_row=row_resumo, start_column=1,
            end_row=row_resumo, end_column=total_colunas
        )
        celula_resumo = ws.cell(row=row_resumo, column=1)
        celula_resumo.value = (
            f"Total: {total_usuarios} usuário(s)  |  "
            f"Ativos: {total_ativos}  |  "
            f"Inativos: {total_inativos}"
        )
        celula_resumo.font = font_rodape
        celula_resumo.alignment = Alignment(horizontal='right', vertical='center')

        # ─── Área de impressão ───
        ultima_col_letter = get_column_letter(total_colunas)
        ws.print_area = f"A1:{ultima_col_letter}{row_resumo}"

        # Repetir cabeçalho em cada página impressa
        ws.print_title_rows = f'{row_titulo}:{row_header}'

        # ─── Gerar resposta HTTP ───
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        data_arquivo = agora.strftime('%Y%m%d_%H%M')
        filename = f"usuarios_cadastrados_{data_arquivo}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response