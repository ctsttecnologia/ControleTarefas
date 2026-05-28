

# suprimentos/services.py
"""
Serviços de negócio reutilizáveis do app Suprimentos.
Extraídos das views para facilitar testes e manutenção.
"""
import uuid
import logging
from decimal import Decimal, InvalidOperation
from io import BytesIO
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from django.db import transaction
from django.utils.translation import gettext_lazy as _

from seguranca_trabalho.models import Equipamento
from ferramentas.models import Ferramenta as FerramentaModel
from suprimentos.models import (
    Material, CategoriaMaterial, TipoMaterial, UnidadeMedida,
)
from tributacao.models import NCM, GrupoTributario

logger = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES DE IMPORTAÇÃO
# ═════════════════════════════════════════════════════════════════════

COLUNAS = [
    'descricao', 'classificacao', 'tipo', 'marca',
    'unidade', 'valor_unitario',
    'ncm_codigo', 'grupo_tributario_codigo',
    'ativo',
]

CABECALHOS_AMIGAVEIS = {
    'descricao': 'Descrição *',
    'classificacao': 'Classificação *',
    'tipo': 'Tipo *',
    'marca': 'Marca',
    'unidade': 'Unidade',
    'valor_unitario': 'Valor Unitário (R$)',
    'ncm_codigo': 'NCM (código)',
    'grupo_tributario_codigo': 'Grupo Tributário (código)',
    'ativo': 'Ativo (S/N)',
}

COLUNAS_OBRIGATORIAS = {'descricao', 'classificacao', 'tipo'}


# ═════════════════════════════════════════════════════════════════════
# DATACLASSES DE RESULTADO
# ═════════════════════════════════════════════════════════════════════

@dataclass
class LinhaResultado:
    linha: int
    dados: dict
    erros: list = field(default_factory=list)
    material: Optional[Material] = None

    @property
    def valida(self):
        return not self.erros


@dataclass
class ResultadoImportacao:
    linhas: list = field(default_factory=list)
    total_criados: int = 0
    total_erros: int = 0

    @property
    def tem_erros(self):
        return self.total_erros > 0


# ═════════════════════════════════════════════════════════════════════
# SERVICE DE IMPORTAÇÃO EM MASSA
# ═════════════════════════════════════════════════════════════════════

class MaterialImportService:
    """Processa upload em lote de Materiais (Excel/CSV)."""

    def __init__(self, arquivo, filial, *, commit=True):
        self.arquivo = arquivo
        self.filial = filial
        self.commit = commit
        self._cache_ncm = {}
        self._cache_grupo = {}

    # ── Cache de FKs ─────────────────────────────────────────────────
    def _get_ncm(self, codigo):
        if not codigo:
            return None
        codigo = str(codigo).strip()
        if codigo not in self._cache_ncm:
            self._cache_ncm[codigo] = NCM.objects.filter(
                codigo=codigo, ativo=True
            ).first()
        return self._cache_ncm[codigo]

    def _get_grupo(self, codigo):
        if not codigo:
            return None
        codigo = str(codigo).strip()
        if codigo not in self._cache_grupo:
            self._cache_grupo[codigo] = GrupoTributario.objects.filter(
                codigo=codigo, ativo=True
            ).first()
        return self._cache_grupo[codigo]

    # ── Leitura do arquivo ───────────────────────────────────────────
    def _ler_arquivo(self):
        nome = self.arquivo.name.lower()
        if nome.endswith(('.xlsx', '.xlsm')):
            return self._ler_excel()
        elif nome.endswith('.csv'):
            return self._ler_csv()
        raise ValueError("Formato não suportado. Use .xlsx ou .csv")

    def _ler_excel(self):
        wb = openpyxl.load_workbook(BytesIO(self.arquivo.read()), data_only=True)
        ws = wb.active
        linhas = []
        # Pula linha 1 (cabeçalho) e linha 2 (legenda). Dados a partir da linha 3.
        for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not any(c not in (None, '') for c in row):
                continue
            dados = {col: row[i] if i < len(row) else None
                     for i, col in enumerate(COLUNAS)}
            linhas.append((idx, dados))
        return linhas

    def _ler_csv(self):
        import csv
        from io import StringIO
        conteudo = self.arquivo.read().decode('utf-8-sig')
        reader = csv.DictReader(StringIO(conteudo), delimiter=';')
        linhas = []
        for idx, row in enumerate(reader, start=2):
            dados = {}
            for col in COLUNAS:
                valor = row.get(col) or row.get(CABECALHOS_AMIGAVEIS[col])
                dados[col] = valor
            if not any(v not in (None, '') for v in dados.values()):
                continue
            linhas.append((idx, dados))
        return linhas

    # ── Validação ────────────────────────────────────────────────────
    def _validar_linha(self, dados):
        erros = []
        limpo = {}

        desc = (str(dados.get('descricao') or '')).strip()
        if not desc:
            erros.append("Descrição é obrigatória")
        limpo['descricao'] = desc[:500]

        classif = (str(dados.get('classificacao') or '')).strip().upper()
        if classif not in dict(CategoriaMaterial.choices):
            erros.append(
                f"Classificação inválida '{classif}'. "
                f"Use: {', '.join(dict(CategoriaMaterial.choices).keys())}"
            )
        limpo['classificacao'] = classif

        tipo = (str(dados.get('tipo') or '')).strip().upper()
        if tipo not in dict(TipoMaterial.choices):
            erros.append(
                f"Tipo inválido '{tipo}'. "
                f"Use um dos: {', '.join(list(dict(TipoMaterial.choices).keys())[:5])}..."
            )
        limpo['tipo'] = tipo

        limpo['marca'] = (str(dados.get('marca') or '')).strip()[:100]

        unidade = (str(dados.get('unidade') or 'PC')).strip().upper()
        if unidade not in dict(UnidadeMedida.choices):
            erros.append(f"Unidade inválida '{unidade}'")
            unidade = 'PC'
        limpo['unidade'] = unidade

        valor_bruto = dados.get('valor_unitario')
        try:
            if valor_bruto in (None, ''):
                valor = Decimal('0.00')
            else:
                s = str(valor_bruto)
                if ',' in s:
                    s = s.replace('.', '').replace(',', '.')
                valor = Decimal(s)
            if valor < 0:
                erros.append("Valor unitário não pode ser negativo")
        except (InvalidOperation, ValueError):
            erros.append(f"Valor unitário inválido: '{valor_bruto}'")
            valor = Decimal('0.00')
        limpo['valor_unitario'] = valor

        ncm_codigo = dados.get('ncm_codigo')
        if ncm_codigo:
            ncm = self._get_ncm(ncm_codigo)
            if not ncm:
                erros.append(f"NCM '{ncm_codigo}' não encontrado/ativo")
            limpo['ncm'] = ncm
        else:
            limpo['ncm'] = None

        gt_codigo = dados.get('grupo_tributario_codigo')
        if gt_codigo:
            gt = self._get_grupo(gt_codigo)
            if not gt:
                erros.append(f"Grupo Tributário '{gt_codigo}' não encontrado/ativo")
            limpo['grupo_tributario'] = gt
        else:
            limpo['grupo_tributario'] = None

        ativo_bruto = str(dados.get('ativo') or 'S').strip().upper()
        limpo['ativo'] = ativo_bruto in ('S', 'SIM', '1', 'TRUE', 'V', 'VERDADEIRO')

        return limpo, erros

    # ── Execução ─────────────────────────────────────────────────────
    def executar(self):
        resultado = ResultadoImportacao()
        linhas_arquivo = self._ler_arquivo()

        for num_linha, dados in linhas_arquivo:
            limpo, erros = self._validar_linha(dados)
            resultado.linhas.append(LinhaResultado(
                linha=num_linha, dados=limpo, erros=erros,
            ))

        resultado.total_erros = sum(1 for l in resultado.linhas if not l.valida)

        if not self.commit or resultado.tem_erros:
            return resultado

        with transaction.atomic():
            for linha_res in resultado.linhas:
                d = linha_res.dados
                material = Material(
                    descricao=d['descricao'],
                    classificacao=d['classificacao'],
                    tipo=d['tipo'],
                    marca=d['marca'],
                    unidade=d['unidade'],
                    valor_unitario=d['valor_unitario'],
                    ncm=d['ncm'],
                    grupo_tributario=d['grupo_tributario'],
                    ativo=d['ativo'],
                    filial=self.filial,
                )
                material.save()
                linha_res.material = material
                resultado.total_criados += 1

        return resultado


# ═════════════════════════════════════════════════════════════════════
# ESTILOS DO TEMPLATE EXCEL
# ═════════════════════════════════════════════════════════════════════

def _estilos():
    return {
        'header_obrig': PatternFill("solid", fgColor="C0392B"),
        'header_opc':   PatternFill("solid", fgColor="2E5984"),
        'header_aux':   PatternFill("solid", fgColor="27AE60"),
        'font_white':   Font(bold=True, color="FFFFFF", size=11),
        'font_bold':    Font(bold=True, size=11),
        'center':       Alignment(horizontal="center", vertical="center", wrap_text=True),
        'left':         Alignment(horizontal="left", vertical="center", wrap_text=True),
        'border':       Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        ),
        'zebra':        PatternFill("solid", fgColor="F4F6F7"),
    }


# ═════════════════════════════════════════════════════════════════════
# GERAÇÃO DO TEMPLATE EXCEL
# ═════════════════════════════════════════════════════════════════════

def gerar_template_excel():
    """Gera template .xlsx com 3 abas: Materiais, Referências e Instruções."""
    wb = openpyxl.Workbook()
    st = _estilos()

    # ── ABA 1: MATERIAIS ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Materiais"

    cabecalhos = [
        ('descricao', 'Descrição', True),
        ('classificacao', 'Classificação', True),
        ('tipo', 'Tipo', True),
        ('marca', 'Marca', False),
        ('unidade', 'Unidade', False),
        ('valor_unitario', 'Valor Unitário (R$)', False),
        ('ncm_codigo', 'NCM (código)', False),
        ('grupo_tributario_codigo', 'Grupo Tributário (código)', False),
        ('ativo', 'Ativo (S/N)', False),
    ]

    for idx, (campo, label, obrig) in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = f"{label} *" if obrig else label
        cell.font = st['font_white']
        cell.alignment = st['center']
        cell.fill = st['header_obrig'] if obrig else st['header_opc']
        cell.border = st['border']

    # Linha 2: legenda
    legenda = ws.cell(
        row=2, column=1,
        value="🔴 Vermelho = Obrigatório   🔵 Azul = Opcional"
    )
    legenda.font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cabecalhos))

    # Exemplos
    exemplos = [
        ['Fita Isolante 3M Scotch 33+', 'CONSUMO', 'ELETRICA',
         '3M', 'PC', '12.50', '8546.90.00', 'GT-CONSUMO', 'S'],
        ['Capacete de Segurança Branco', 'EPI', 'EPI',
         '3M', 'PC', '45.00', '6506.10.00', 'GT-EPI', 'S'],
        ['Chave de Fenda Phillips 1/4"', 'FERRAMENTA', 'CIVIL',
         'GEDORE', 'PC', '28.90', '', '', 'S'],
        ['Detergente Neutro 5L', 'CONSUMO', 'LIMPEZA',
         'YPÊ', 'GALAO', '24.90', '', '', 'S'],
    ]
    for i, linha in enumerate(exemplos, start=3):
        for j, val in enumerate(linha, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = st['border']
            cell.alignment = st['left']
            if i % 2 == 0:
                cell.fill = st['zebra']

    # Dropdowns
    dv_classif = DataValidation(
        type="list",
        formula1=f'"{",".join(dict(CategoriaMaterial.choices).keys())}"',
        allow_blank=False,
    )
    dv_classif.error = "Use um dos valores: " + ", ".join(dict(CategoriaMaterial.choices).keys())
    dv_classif.errorTitle = "Classificação inválida"
    dv_classif.add('B3:B1000')
    ws.add_data_validation(dv_classif)

    dv_tipo = DataValidation(
        type="list",
        formula1=f'"{",".join(dict(TipoMaterial.choices).keys())}"',
        allow_blank=False,
    )
    dv_tipo.error = "Use um dos valores válidos de Tipo"
    dv_tipo.errorTitle = "Tipo inválido"
    dv_tipo.add('C3:C1000')
    ws.add_data_validation(dv_tipo)

    dv_unid = DataValidation(
        type="list",
        formula1=f'"{",".join(dict(UnidadeMedida.choices).keys())}"',
        allow_blank=True,
    )
    dv_unid.error = "Use um dos valores válidos de Unidade"
    dv_unid.errorTitle = "Unidade inválida"
    dv_unid.add('E3:E1000')
    ws.add_data_validation(dv_unid)

    dv_ativo = DataValidation(type="list", formula1='"S,N"', allow_blank=True)
    dv_ativo.add('I3:I1000')
    ws.add_data_validation(dv_ativo)

    larguras = [40, 18, 22, 20, 12, 18, 18, 25, 12]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A3'

    # ── ABA 2: REFERÊNCIAS ───────────────────────────────────────────
    ws_ref = wb.create_sheet("Referências")

    def render_tabela(ws, start_col, titulo, choices):
        end_col = start_col + 1
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=end_col)
        c = ws.cell(row=1, column=start_col, value=titulo)
        c.font = st['font_white']
        c.fill = st['header_aux']
        c.alignment = st['center']
        c.border = st['border']

        h1 = ws.cell(row=2, column=start_col, value="Código (usar no Excel)")
        h2 = ws.cell(row=2, column=end_col, value="Descrição")
        for h in (h1, h2):
            h.font = st['font_bold']
            h.fill = PatternFill("solid", fgColor="D5F5E3")
            h.alignment = st['center']
            h.border = st['border']

        for i, (codigo, descricao) in enumerate(choices, start=3):
            cc = ws.cell(row=i, column=start_col, value=codigo)
            cd = ws.cell(row=i, column=end_col, value=descricao)
            cc.font = Font(bold=True, color="1B5E20")
            cc.alignment = st['center']
            cd.alignment = st['left']
            cc.border = st['border']
            cd.border = st['border']
            if i % 2 == 0:
                cc.fill = st['zebra']
                cd.fill = st['zebra']

        ws.column_dimensions[get_column_letter(start_col)].width = 22
        ws.column_dimensions[get_column_letter(end_col)].width = 28

    render_tabela(ws_ref, 1, "📦 CLASSIFICAÇÃO", list(CategoriaMaterial.choices))
    render_tabela(ws_ref, 4, "🏷️ TIPO",         list(TipoMaterial.choices))
    render_tabela(ws_ref, 7, "📏 UNIDADE",      list(UnidadeMedida.choices))

    ws_ref.column_dimensions['C'].width = 3
    ws_ref.column_dimensions['F'].width = 3
    ws_ref.row_dimensions[1].height = 28
    ws_ref.row_dimensions[2].height = 22
    ws_ref.freeze_panes = 'A3'

    # ── ABA 3: INSTRUÇÕES ────────────────────────────────────────────
    ws_inst = wb.create_sheet("Instruções")
    ws_inst.column_dimensions['A'].width = 28
    ws_inst.column_dimensions['B'].width = 15
    ws_inst.column_dimensions['C'].width = 65

    for j, val in enumerate(["Campo", "Obrigatório", "Observações"], start=1):
        c = ws_inst.cell(row=1, column=j, value=val)
        c.font = st['font_white']
        c.fill = st['header_opc']
        c.alignment = st['center']
        c.border = st['border']

    instrucoes = [
        ("Descrição", "✅ SIM", "Texto livre até 500 caracteres."),
        ("Classificação", "✅ SIM", "Use o código da aba 'Referências'. Ex: EPI, CONSUMO, FERRAMENTA"),
        ("Tipo", "✅ SIM", "Use o código da aba 'Referências'. Ex: ELETRICA, LIMPEZA, EPI"),
        ("Marca", "—",     "Texto livre até 100 caracteres."),
        ("Unidade", "—",   "Use o código da aba 'Referências'. Padrão: PC (Peça)."),
        ("Valor Unitário", "—", "Decimal. Aceita '12.50' ou '12,50'. Padrão: 0,00."),
        ("NCM (código)", "—", "Código NCM já cadastrado e ativo no sistema."),
        ("Grupo Tributário", "—", "Código do grupo já cadastrado e ativo."),
        ("Ativo", "—", "'S' ou 'N'. Padrão: S (ativo)."),
    ]
    for i, (campo, obrig, obs) in enumerate(instrucoes, start=2):
        cc = ws_inst.cell(row=i, column=1, value=campo)
        co = ws_inst.cell(row=i, column=2, value=obrig)
        ct = ws_inst.cell(row=i, column=3, value=obs)
        cc.font = st['font_bold']
        co.alignment = st['center']
        ct.alignment = st['left']
        if "SIM" in obrig:
            co.font = Font(bold=True, color="C0392B")
        for c in (cc, co, ct):
            c.border = st['border']
        if i % 2 == 0:
            for c in (cc, co, ct):
                c.fill = st['zebra']

    ws_inst.row_dimensions[1].height = 25

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ═════════════════════════════════════════════════════════════════════
# HELPERS DE CRIAÇÃO AUTOMÁTICA (EPI E FERRAMENTA)
# ═════════════════════════════════════════════════════════════════════

def criar_equipamento_epi_from_form(material, form, filial):
    """Cria um Equipamento EPI no módulo SST vinculado ao Material."""
    equipamento = Equipamento.objects.create(
        nome=material.descricao,
        modelo=form.cleaned_data.get('epi_modelo', '') or '',
        fabricante=form.cleaned_data['epi_fabricante'],
        certificado_aprovacao=form.cleaned_data.get('epi_ca', ''),
        vida_util_dias=form.cleaned_data['epi_vida_util_dias'],
        filial=filial,
    )
    material.equipamento_epi = equipamento
    material.save(update_fields=['equipamento_epi'])
    logger.info(
        f"Equipamento EPI criado: {equipamento.nome} "
        f"(CA: {equipamento.certificado_aprovacao}) para filial {filial}"
    )
    return equipamento


def criar_ferramenta_from_form(material, form, filial):
    """Cria uma Ferramenta no módulo Ferramentas vinculada ao Material."""
    codigo = form.cleaned_data.get('ferr_codigo', '').strip()
    if not codigo:
        codigo = f"FERR-{uuid.uuid4().hex[:8].upper()}"

    ferramenta = FerramentaModel.objects.create(
        nome=material.descricao,
        codigo_identificacao=codigo,
        patrimonio=form.cleaned_data.get('ferr_patrimonio') or None,
        fabricante_marca=material.marca or None,
        localizacao_padrao=form.cleaned_data['ferr_localizacao'],
        data_aquisicao=form.cleaned_data['ferr_data_aquisicao'],
        quantidade=form.cleaned_data.get('ferr_quantidade') or 0,
        fornecedor=form.cleaned_data.get('ferr_fornecedor'),
        filial=filial,
        status=FerramentaModel.Status.DISPONIVEL,
    )
    material.ferramenta_ref = ferramenta
    material.save(update_fields=['ferramenta_ref'])
    logger.info(
        f"Ferramenta criada: {ferramenta.nome} ({codigo}) para filial {filial}"
    )
    return ferramenta


