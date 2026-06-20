
# suprimentos/tests/test_services_import.py
"""
Testes do MaterialImportService — foco na validação de linhas,
na execução transacional da importação em massa, leitura de arquivos
e geração do template Excel.
"""
import csv
from io import BytesIO, StringIO
from decimal import Decimal
from unittest.mock import Mock
import openpyxl
from django.test import TestCase
from usuario.models import Filial
from suprimentos.models import (
    Material, CategoriaMaterial, TipoMaterial, UnidadeMedida,
)
from suprimentos.services import (
    MaterialImportService,
    gerar_template_excel,
    COLUNAS,
    CABECALHOS_AMIGAVEIS,
)
from suprimentos.models import UnidadeMedida
from tributacao.models import NCM, GrupoTributario, CFOP





# Pega a PRIMEIRA unidade válida do enum, seja qual for
UNIDADE_VALIDA = list(dict(UnidadeMedida.choices).keys())[0] 

# ═════════════════════════════════════════════════════════════════════
# HELPERS DE ARQUIVO EM MEMÓRIA
# ═════════════════════════════════════════════════════════════════════

class _FakeUpload:
    """Simula um arquivo de upload (tem .name e .read())."""
    def __init__(self, conteudo: bytes, name: str):
        self._buffer = BytesIO(conteudo)
        self.name = name

    def read(self):
        return self._buffer.read()


def _build_xlsx(linhas_dados):
    """
    Monta um .xlsx no layout esperado pelo service:
    linha 1 = cabeçalho, linha 2 = legenda, dados a partir da linha 3.
    `linhas_dados` é uma lista de listas (9 colunas em COLUNAS).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Linha 1: cabeçalho
    for j, col in enumerate(COLUNAS, start=1):
        ws.cell(row=1, column=j, value=CABECALHOS_AMIGAVEIS[col])
    # Linha 2: legenda (ignorada pelo leitor)
    ws.cell(row=2, column=1, value="legenda qualquer")
    # Linha 3+: dados
    for i, linha in enumerate(linhas_dados, start=3):
        for j, val in enumerate(linha, start=1):
            ws.cell(row=i, column=j, value=val)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_csv(linhas_dados):
    """Monta um CSV (delimitador ';') usando os nomes técnicos de COLUNAS."""
    out = StringIO()
    writer = csv.writer(out, delimiter=';')
    writer.writerow(COLUNAS)
    for linha in linhas_dados:
        writer.writerow(linha)
    return out.getvalue().encode('utf-8-sig')


# ═════════════════════════════════════════════════════════════════════
# VALIDAÇÃO DE LINHA (regra de negócio pura — sem tocar no banco)
# ═════════════════════════════════════════════════════════════════════

class ValidarLinhaTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.filial = Filial.objects.create(nome="Filial Teste")

    def _service(self, commit=False):
        return MaterialImportService(arquivo=None, filial=self.filial, commit=commit)

    def _linha_valida(self):
        return {
            'descricao': 'Fita Isolante',
            'classificacao': CategoriaMaterial.CONSUMO,
            'tipo': TipoMaterial.LIMPEZA,
            'marca': '3M',
            'unidade': UNIDADE_VALIDA,   # ← era 'PC'
            'valor_unitario': '12.50',
            'ncm_codigo': '',
            'grupo_tributario_codigo': '',
            'ativo': 'S',
        }

    # ── Caminho feliz ────────────────────────────────────────────────
    def test_linha_valida_sem_erros(self):
        limpo, erros = self._service()._validar_linha(self._linha_valida())
        self.assertEqual(erros, [])
        self.assertEqual(limpo['descricao'], 'Fita Isolante')
        self.assertEqual(limpo['valor_unitario'], Decimal('12.50'))
        self.assertTrue(limpo['ativo'])
        self.assertIsNone(limpo['ncm'])
        self.assertIsNone(limpo['grupo_tributario'])

    # ── Descrição ────────────────────────────────────────────────────
    def test_descricao_vazia_gera_erro(self):
        dados = self._linha_valida()
        dados['descricao'] = '   '
        _, erros = self._service()._validar_linha(dados)
        self.assertTrue(any('Descrição' in e for e in erros))

    def test_descricao_none_gera_erro(self):
        dados = self._linha_valida()
        dados['descricao'] = None
        _, erros = self._service()._validar_linha(dados)
        self.assertTrue(any('Descrição' in e for e in erros))

    def test_descricao_truncada_em_500(self):
        dados = self._linha_valida()
        dados['descricao'] = 'X' * 600
        limpo, _ = self._service()._validar_linha(dados)
        self.assertEqual(len(limpo['descricao']), 500)

    # ── Classificação / Tipo ─────────────────────────────────────────
    def test_classificacao_invalida_gera_erro(self):
        dados = self._linha_valida()
        dados['classificacao'] = 'XPTO'
        _, erros = self._service()._validar_linha(dados)
        self.assertTrue(any('Classificação' in e for e in erros))

    def test_classificacao_normaliza_para_maiuscula(self):
        dados = self._linha_valida()
        dados['classificacao'] = str(CategoriaMaterial.CONSUMO).lower()
        limpo, erros = self._service()._validar_linha(dados)
        self.assertEqual(erros, [])
        self.assertEqual(limpo['classificacao'], CategoriaMaterial.CONSUMO)

    def test_tipo_invalido_gera_erro(self):
        dados = self._linha_valida()
        dados['tipo'] = 'INEXISTENTE'
        _, erros = self._service()._validar_linha(dados)
        self.assertTrue(any('Tipo' in e for e in erros))

    # ── Valor unitário (ponto mais sensível) ─────────────────────────
    def test_valor_com_virgula_brasileira(self):
        dados = self._linha_valida()
        dados['valor_unitario'] = '1.234,56'
        limpo, erros = self._service()._validar_linha(dados)
        self.assertEqual(erros, [])
        self.assertEqual(limpo['valor_unitario'], Decimal('1234.56'))

    def test_valor_com_ponto_decimal(self):
        dados = self._linha_valida()
        dados['valor_unitario'] = '99.90'
        limpo, _ = self._service()._validar_linha(dados)
        self.assertEqual(limpo['valor_unitario'], Decimal('99.90'))

    def test_valor_vazio_assume_zero(self):
        dados = self._linha_valida()
        dados['valor_unitario'] = ''
        limpo, _ = self._service()._validar_linha(dados)
        self.assertEqual(limpo['valor_unitario'], Decimal('0.00'))

    def test_valor_none_assume_zero(self):
        dados = self._linha_valida()
        dados['valor_unitario'] = None
        limpo, _ = self._service()._validar_linha(dados)
        self.assertEqual(limpo['valor_unitario'], Decimal('0.00'))

    def test_valor_negativo_gera_erro(self):
        dados = self._linha_valida()
        dados['valor_unitario'] = '-10.00'
        _, erros = self._service()._validar_linha(dados)
        self.assertTrue(any('negativo' in e.lower() for e in erros))

    def test_valor_invalido_gera_erro(self):
        dados = self._linha_valida()
        dados['valor_unitario'] = 'abc'
        limpo, erros = self._service()._validar_linha(dados)
        self.assertTrue(any('inválido' in e.lower() for e in erros))
        # mesmo com erro, deve cair no fallback 0.00
        self.assertEqual(limpo['valor_unitario'], Decimal('0.00'))

    # ── Unidade ──────────────────────────────────────────────────────
    def test_unidade_vazia_assume_pc(self):
        dados = self._linha_valida()
        dados['unidade'] = ''
        limpo, erros = self._service()._validar_linha(dados)
        self.assertEqual(limpo['unidade'], 'PÇ')
        self.assertEqual(erros, [])

    def test_unidade_invalida_gera_erro_e_fallback_pc(self):
        dados = self._linha_valida()
        dados['unidade'] = 'ZZZ'
        limpo, erros = self._service()._validar_linha(dados)
        self.assertTrue(any('Unidade' in e for e in erros))
        self.assertEqual(limpo['unidade'], 'PÇ')

    # ── Campo ativo (parsing de booleano) ────────────────────────────
    def test_ativo_aceita_variacoes_verdadeiras(self):
        svc = self._service()
        for val in ('S', 'SIM', '1', 'TRUE', 'V', 'VERDADEIRO', 's', 'sim'):
            dados = self._linha_valida()
            dados['ativo'] = val
            limpo, _ = svc._validar_linha(dados)
            self.assertTrue(limpo['ativo'], f"'{val}' deveria ser True")

    def test_ativo_falso(self):
        dados = self._linha_valida()
        dados['ativo'] = 'N'
        limpo, _ = self._service()._validar_linha(dados)
        self.assertFalse(limpo['ativo'])

    def test_ativo_vazio_assume_verdadeiro(self):
        dados = self._linha_valida()
        dados['ativo'] = ''
        limpo, _ = self._service()._validar_linha(dados)
        self.assertTrue(limpo['ativo'])

    # ── Múltiplos erros acumulados ───────────────────────────────────
    def test_multiplos_erros_na_mesma_linha(self):
        dados = self._linha_valida()
        dados['descricao'] = ''
        dados['classificacao'] = 'XXX'
        dados['tipo'] = 'YYY'
        _, erros = self._service()._validar_linha(dados)
        self.assertGreaterEqual(len(erros), 3)

    def test_unidade_minuscula_normaliza(self):
        dados = self._linha_valida()
        dados['unidade'] = 'pç'
        limpo, erros = self._service()._validar_linha(dados)
        self.assertEqual(limpo['unidade'], 'PÇ')
        self.assertEqual(erros, [])


class GetFKsCacheTests(TestCase):
    """Cobre _get_ncm e _get_grupo: lookup, ativo, strip e cache (linhas 101-118)."""

    def setUp(self):
        self.filial = Filial.objects.create(nome='Matriz')

    def _service(self):
        return MaterialImportService(arquivo=Mock(), filial=self.filial)

    def _criar_grupo(self, codigo):
        cfop = CFOP.objects.create(codigo=f'5102-{codigo}', descricao='Venda')
        return GrupoTributario.objects.create(
            nome=f'Grupo {codigo}',
            codigo=codigo,
            cfop=cfop,
            filial=self.filial,
        )

    # ── NCM ──────────────────────────────────────────────────────────
    def test_get_ncm_vazio_retorna_none(self):
        self.assertIsNone(self._service()._get_ncm(''))
        self.assertIsNone(self._service()._get_ncm(None))

    def test_get_ncm_existente_retorna_objeto(self):
        ncm = NCM.objects.create(codigo='8546.90.00', descricao='Fita isolante')
        self.assertEqual(self._service()._get_ncm('8546.90.00'), ncm)

    def test_get_ncm_inexistente_retorna_none(self):
        self.assertIsNone(self._service()._get_ncm('0000.00.00'))

    def test_get_ncm_inativo_retorna_none(self):
        NCM.objects.create(codigo='1111.11.11', descricao='Inativo', ativo=False)
        self.assertIsNone(self._service()._get_ncm('1111.11.11'))

    def test_get_ncm_usa_cache(self):
        NCM.objects.create(codigo='2222.22.22', descricao='Cache')
        svc = self._service()
        svc._get_ncm('2222.22.22')
        with self.assertNumQueries(0):
            svc._get_ncm('2222.22.22')

    def test_get_ncm_strip_espacos(self):
        ncm = NCM.objects.create(codigo='3333.33.33', descricao='Strip')
        self.assertEqual(self._service()._get_ncm('  3333.33.33  '), ncm)

    # ── Grupo Tributário ─────────────────────────────────────────────
    def test_get_grupo_vazio_retorna_none(self):
        self.assertIsNone(self._service()._get_grupo(''))
        self.assertIsNone(self._service()._get_grupo(None))

    def test_get_grupo_existente_retorna_objeto(self):
        gt = self._criar_grupo('GT-CONSUMO')
        self.assertEqual(self._service()._get_grupo('GT-CONSUMO'), gt)

    def test_get_grupo_inexistente_retorna_none(self):
        self.assertIsNone(self._service()._get_grupo('GT-XXX'))

    def test_get_grupo_usa_cache(self):
        self._criar_grupo('GT-CACHE')
        svc = self._service()
        svc._get_grupo('GT-CACHE')
        with self.assertNumQueries(0):
            svc._get_grupo('GT-CACHE')

    def test_get_grupo_strip_espacos(self):
        gt = self._criar_grupo('GT-EPI')
        self.assertEqual(self._service()._get_grupo('  GT-EPI  '), gt)

# ═════════════════════════════════════════════════════════════════════
# EXECUÇÃO DA IMPORTAÇÃO (com transação atômica)
# ═════════════════════════════════════════════════════════════════════

class ExecutarImportacaoTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.filial = Filial.objects.create(nome="Filial Import")

    def _linha_ok(self, descricao):
        return [
            descricao, CategoriaMaterial.CONSUMO, TipoMaterial.LIMPEZA,
            '3M', UNIDADE_VALIDA, '10.00', '', '', 'S',   # ← era 'PC'
        ]

    def test_importacao_xlsx_cria_materiais(self):
        conteudo = _build_xlsx([
            self._linha_ok('Material 1'),
            self._linha_ok('Material 2'),
        ])
        upload = _FakeUpload(conteudo, 'materiais.xlsx')
        svc = MaterialImportService(upload, self.filial, commit=True)

        resultado = svc.executar()

        self.assertFalse(resultado.tem_erros)
        self.assertEqual(resultado.total_criados, 2)
        self.assertEqual(Material.objects.filter(filial=self.filial).count(), 2)

    def test_commit_false_nao_persiste(self):
        conteudo = _build_xlsx([self._linha_ok('Material X')])
        upload = _FakeUpload(conteudo, 'materiais.xlsx')
        svc = MaterialImportService(upload, self.filial, commit=False)

        resultado = svc.executar()

        self.assertFalse(resultado.tem_erros)
        self.assertEqual(resultado.total_criados, 0)
        self.assertEqual(Material.objects.count(), 0)

    def test_erro_em_uma_linha_faz_rollback_total(self):
        """Se QUALQUER linha tem erro, nenhuma deve ser criada (atomicidade)."""
        linha_ruim = self._linha_ok('Material Ruim')
        linha_ruim[1] = 'CLASSIFICACAO_INVALIDA'  # quebra a validação
        conteudo = _build_xlsx([
            self._linha_ok('Material Bom'),
            linha_ruim,
        ])
        upload = _FakeUpload(conteudo, 'materiais.xlsx')
        svc = MaterialImportService(upload, self.filial, commit=True)

        resultado = svc.executar()

        self.assertTrue(resultado.tem_erros)
        self.assertEqual(resultado.total_erros, 1)
        self.assertEqual(resultado.total_criados, 0)
        self.assertEqual(Material.objects.count(), 0)

    def test_linhas_vazias_sao_ignoradas(self):
        conteudo = _build_xlsx([
            self._linha_ok('Material Único'),
            [None, None, None, None, None, None, None, None, None],
        ])
        upload = _FakeUpload(conteudo, 'materiais.xlsx')
        svc = MaterialImportService(upload, self.filial, commit=True)

        resultado = svc.executar()

        self.assertEqual(len(resultado.linhas), 1)
        self.assertEqual(resultado.total_criados, 1)

    def test_material_criado_com_valores_corretos(self):
        conteudo = _build_xlsx([
            ['Cola Forte', CategoriaMaterial.CONSUMO, TipoMaterial.LIMPEZA,
            'Loctite', UNIDADE_VALIDA, '15,90', '', '', 'S'],   # ← era 'PC'
        ])

        upload = _FakeUpload(conteudo, 'materiais.xlsx')
        MaterialImportService(upload, self.filial, commit=True).executar()

        mat = Material.objects.get(descricao='Cola Forte')
        self.assertEqual(mat.classificacao, CategoriaMaterial.CONSUMO)
        self.assertEqual(mat.tipo, TipoMaterial.LIMPEZA)
        self.assertEqual(mat.marca, 'Loctite')
        self.assertEqual(mat.valor_unitario, Decimal('15.90'))
        self.assertTrue(mat.ativo)
        self.assertEqual(mat.filial, self.filial)


# ═════════════════════════════════════════════════════════════════════
# LEITURA DE ARQUIVOS
# ═════════════════════════════════════════════════════════════════════

class LeituraArquivoTests(TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.filial = Filial.objects.create(nome="Filial Leitura")

    def test_formato_nao_suportado_levanta_valueerror(self):
        upload = _FakeUpload(b'qualquer coisa', 'dados.txt')
        svc = MaterialImportService(upload, self.filial)
        with self.assertRaises(ValueError):
            svc._ler_arquivo()

    def test_leitura_csv(self):
        conteudo = _build_csv([
            ['Material CSV', CategoriaMaterial.CONSUMO, TipoMaterial.LIMPEZA,
             '3M', UNIDADE_VALIDA, '10.00', '', '', 'S'],   # ← era 'PC'
        ])
        upload = _FakeUpload(conteudo, 'materiais.csv')
        svc = MaterialImportService(upload, self.filial, commit=True)

        resultado = svc.executar()

        self.assertFalse(resultado.tem_erros)
        self.assertEqual(resultado.total_criados, 1)
        self.assertTrue(Material.objects.filter(descricao='Material CSV').exists())

    def test_leitura_xlsx_pula_cabecalho_e_legenda(self):
        conteudo = _build_xlsx([
            ['Primeiro Material', CategoriaMaterial.CONSUMO,
             TipoMaterial.LIMPEZA, '', UNIDADE_VALIDA, '0', '', '', 'S'],  # ← era 'PC'
        ])
        upload = _FakeUpload(conteudo, 'materiais.xlsx')
        svc = MaterialImportService(upload, self.filial)

        linhas = svc._ler_arquivo()

        # Só uma linha de dado, começando na linha 3
        self.assertEqual(len(linhas), 1)
        idx, dados = linhas[0]
        self.assertEqual(idx, 3)
        self.assertEqual(dados['descricao'], 'Primeiro Material')


# ═════════════════════════════════════════════════════════════════════
# TEMPLATE EXCEL (smoke tests)
# ═════════════════════════════════════════════════════════════════════

class TemplateExcelTests(TestCase):

    def test_gera_xlsx_valido(self):
        conteudo = gerar_template_excel()
        self.assertIsInstance(conteudo, bytes)
        self.assertGreater(len(conteudo), 0)

    def test_template_tem_tres_abas(self):
        wb = openpyxl.load_workbook(BytesIO(gerar_template_excel()))
        self.assertEqual(
            wb.sheetnames, ["Materiais", "Referências", "Instruções"]
        )

    def test_template_marca_obrigatorios_com_asterisco(self):
        wb = openpyxl.load_workbook(BytesIO(gerar_template_excel()))
        ws = wb["Materiais"]
        self.assertIn("*", ws.cell(row=1, column=1).value)  # Descrição
        self.assertIn("*", ws.cell(row=1, column=2).value)  # Classificação
        self.assertIn("*", ws.cell(row=1, column=3).value)  # Tipo



# .\run_tests.ps1 suprimentos.tests.test_services_import
# pytest suprimentos/tests/test_services_import.py --cov=suprimentos.services -v
