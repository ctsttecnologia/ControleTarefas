
# suprimentos/views.py
"""
Views do app Suprimentos — implementam o fluxo completo:
PEDIDO → APROVAR → SOLICITAÇÃO/COTAÇÃO (NxN) → APROVAR COTAÇÃO →
MONTAR PEDIDO DE COMPRA → ACOMPANHAR ENTREGA → FINALIZADO
"""
import json
from multiprocessing import context
from django.core.serializers.json import DjangoJSONEncoder
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, Sum, Q, DecimalField, Value
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.generic import (
    CreateView, DetailView, ListView, UpdateView, DeleteView, View, TemplateView,
)
import openpyxl

from suprimentos.tests.conftest import material, solicitacao
from collections import OrderedDict
from .forms import (
    AnexoPedidoForm, CotacaoCabecalhoForm, CotacaoItemValorFormSet, PedidoForm, ItemPedidoFormSet,  AprovarPedidoForm,
    EntregaPedidoCompraForm, ParceiroForm, MaterialForm, ContratoForm, VerbaContratoForm, MaterialForm
   
)
from .models import (
    AnexoPedido, CategoriaMaterial, HistoricoSolicitacao, Parceiro, Material, Contrato,
    Pedido, HistoricoPedido, SolicitacaoCompra, ItemSolicitacao, Cotacao,
    PedidoCompra, ItemPedidoCompra, EntregaAnexo, Material,
    TipoMaterial, UnidadeMedida, VerbaContrato
)
from collections import defaultdict
from django.db import IntegrityError
import logging
from datetime import timedelta
from django.db.models.functions import Coalesce, TruncMonth
from .utils import _registrar_historico
from core.mixins import (
    AppPermissionMixin,
    ViewFilialScopedMixin,
    FilialCreateMixin,
    RequireActiveFilialMixin,
)
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import io
from core.utils import get_filial_ativa


logger = logging.getLogger(__name__)


# Constantes de módulo (não recriam a cada request)
CAMPOS_VERBA = {
    CategoriaMaterial.EPI:        ("verba_epi", "saldo_epi"),
    CategoriaMaterial.CONSUMO:    ("verba_consumo", "saldo_consumo"),
    CategoriaMaterial.FERRAMENTA: ("verba_ferramenta", "saldo_ferramenta"),
}
SALDO_POR_CLASSIF = {cat: campos[1] for cat, campos in CAMPOS_VERBA.items()}


# método auxiliar

class MaterialPrecosMixin:
    """Disponibiliza um JSON {material_id: valor} para o template."""
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        precos = {
            str(pk): valor
            for pk, valor in Material.objects.values_list("pk", "valor_unitario")
        }
        ctx["material_precos_json"] = json.dumps(precos, cls=DjangoJSONEncoder)
        return ctx


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────
def _obter_verba_do_mes(sol):
    """
    Resolve a verba do mês de referência da solicitação (defensivo).
    Retorna (verba_ou_None, erro_ou_None).
    """
    ref = sol.data_necessaria or timezone.now().date()
    try:
        verba = sol.contrato.verba_do_mes(ref.year, ref.month)
    except Exception:
        logger.exception("Falha ao obter verba do mês (contrato=%s)", sol.contrato_id)
        return None, "Não foi possível consultar a verba do contrato."
    return verba, None


def _registrar_historico(func, **kwargs):
    """Wrapper seguro para registro de histórico — nunca quebra o fluxo."""
    try:
        func(**kwargs)
    except Exception:
        logger.exception("Falha ao registrar histórico (%s)", func.__qualname__)


# ═════════════════════════════════════════════════════════════
# DASHBOARD
# ═════════════════════════════════════════════════════════════
class SuprimentosDashboard(LoginRequiredMixin, TemplateView):
    template_name = "suprimentos/dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        hoje = timezone.now()
        ano, mes = hoje.year, hoje.month
        dec = DecimalField(max_digits=14, decimal_places=2)
        zero = Value(Decimal("0.00"), output_field=dec)

        SC = SolicitacaoCompra.StatusChoices
        PC = PedidoCompra.StatusPC

        # ── Contagem de Solicitações em uma única query ───────
        # Agrupa por status e transforma em dict {status: total}
        sc_counts = dict(
            SolicitacaoCompra.objects
            .values_list("status")
            .annotate(total=Count("id"))
        )

        def sc(*status):
            """Soma as contagens de um ou mais status de solicitação."""
            return sum(sc_counts.get(s, 0) for s in status)

        # ── PCs abertos (uma query reaproveitada) ─────────────
        pcs_abertos = PedidoCompra.objects.filter(
            status__in=[
                PC.ENVIADO_FORNECEDOR,
                PC.EMITIDO,
                PC.ENTREGA_PARCIAL,
            ]
        )

        # ── KPIs de contagem ──────────────────────────────────
        ctx["pedidos_pendentes"] = Pedido.objects.filter(
            status=Pedido.StatusChoices.PENDENTE
        ).count()
        ctx["solicitacoes_cotacao"] = sc(SC.FAZER_COTACAO)
        ctx["solicitacoes_aprovacao"] = sc(SC.EM_APROVACAO)
        ctx["pcs_pendentes_entrega"] = pcs_abertos.count()

        # ── PCs atrasados (data prevista vencida) ─────────────
        ctx["pcs_atrasados"] = pcs_abertos.filter(
            data_entrega_prevista__lt=hoje.date()
        ).count()

        # ── Valor comprado no mês (PCs emitidos) ──────────────
        ctx["valor_comprado_mes"] = PedidoCompra.objects.filter(
            data_emissao__year=ano, data_emissao__month=mes,
        ).exclude(status=PC.CANCELADO).aggregate(
            t=Coalesce(Sum("valor_total"), zero)
        )["t"]

        # ── Compras por classificação (mês) ───────────────────
        compras_classif = (
            ItemPedidoCompra.objects
            .filter(
                pedido_compra__data_emissao__year=ano,
                pedido_compra__data_emissao__month=mes,
            )
            .exclude(pedido_compra__status=PC.CANCELADO)
            .values("material__classificacao")
            .annotate(total=Coalesce(Sum("valor_total"), zero))
            .order_by("-total")
        )
        labels_cat = dict(CategoriaMaterial.choices)
        ctx["compras_classif"] = [
            {
                "label": labels_cat.get(
                    r["material__classificacao"], r["material__classificacao"]
                ),
                "total": r["total"],
            }
            for r in compras_classif
        ]

        # ── Evolução de compras (últimos 6 meses) ─────────────
        seis_meses = hoje - timedelta(days=180)
        evolucao = (
            PedidoCompra.objects
            .filter(data_emissao__gte=seis_meses.date())
            .exclude(status=PC.CANCELADO)
            .annotate(mes_ref=TruncMonth("data_emissao"))
            .values("mes_ref")
            .annotate(total=Coalesce(Sum("valor_total"), zero))
            .order_by("mes_ref")
        )
        ctx["evolucao_labels"] = [
            e["mes_ref"].strftime("%m/%Y") for e in evolucao if e["mes_ref"]
        ]
        ctx["evolucao_valores"] = [
            float(e["total"]) for e in evolucao if e["mes_ref"]
        ]

        # ── Top 5 fornecedores (mês) ──────────────────────────
        ctx["top_fornecedores"] = (
            PedidoCompra.objects
            .filter(data_emissao__year=ano, data_emissao__month=mes)
            .exclude(status=PC.CANCELADO)
            .values("fornecedor__nome_fantasia")
            .annotate(
                total=Coalesce(Sum("valor_total"), zero),
                qtd=Count("id"),
            )
            .order_by("-total")[:5]
        )

        # ── Funil do fluxo (quantos em cada etapa) ────────────
        ctx["funil"] = {
            "pedidos_aprovados": Pedido.objects.filter(
                status=Pedido.StatusChoices.APROVADO
            ).count(),
            "em_cotacao":   sc(SC.FAZER_COTACAO, SC.COTACAO_ENVIADA),
            "em_aprovacao": sc(SC.EM_APROVACAO, SC.APROVADO),
            "pc_gerado":    sc(SC.ENVIAR_PEDIDO, SC.PEDIDO_GERADO),
            "em_entrega":   sc(SC.EM_ENTREGA),    # ⬅️ agora populado de verdade
            "finalizado":   sc(SC.FINALIZADO),    # ⬅️ sem o remendo de + PedidoCompra.ENTREGUE
        }

        # ── Últimos pedidos ───────────────────────────────────
        ctx["ultimos_pedidos"] = (
            Pedido.objects.select_related("contrato", "solicitante")
            .order_by("-data_pedido")[:8]
        )

        # ── Entregas em atraso (detalhe) ──────────────────────
        ctx["entregas_atrasadas"] = (
            pcs_abertos.filter(data_entrega_prevista__lt=hoje.date())
            .select_related("fornecedor")
            .order_by("data_entrega_prevista")[:5]
        )
        return ctx



# ═════════════════════════════════════════════════════════════
# 1. PEDIDO — CRUD + Submissão
# ═════════════════════════════════════════════════════════════
class PedidoListView(LoginRequiredMixin, ListView):
    model = Pedido
    template_name = "suprimentos/pedido_list.html"
    context_object_name = "pedidos"
    paginate_by = 20

    def get_queryset(self):
        qs = Pedido.objects.select_related("contrato", "solicitante")
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)
        return qs.visiveis_para(self.request.user) if hasattr(qs, "visiveis_para") else qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["status_choices"] = Pedido.StatusChoices.choices
        return ctx


class PedidoDetailView(LoginRequiredMixin, DetailView):
    model = Pedido
    template_name = "suprimentos/pedido_detail.html"
    context_object_name = "pedido"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        pedido = self.object

        itens = list(pedido.itens.select_related("material"))
        ctx["anexos"] = pedido.anexos.all()
        ctx["historico"] = pedido.historico.all()

        ok, erros = pedido.verificar_verba()
        ctx["verba_ok"], ctx["verba_erros"] = ok, erros

        # ── Localiza a verba do contrato no mês/ano do pedido ──
        verba = (
            VerbaContrato.objects
            .filter(
                contrato=pedido.contrato,
                ano=pedido.data_pedido.year,
                mes=pedido.data_pedido.month,
            )
            .first()
        )
        ctx["verba_mes"] = verba

        # Mapa: classificação -> saldo disponível
        if verba:
            saldo_por_cat = {
                CategoriaMaterial.EPI: verba.saldo_epi,
                CategoriaMaterial.CONSUMO: verba.saldo_consumo,
                CategoriaMaterial.FERRAMENTA: verba.saldo_ferramenta,
            }
        else:
            saldo_por_cat = {}

        # Anexa o saldo da categoria a cada item
        for item in itens:
            item.verba_saldo = saldo_por_cat.get(
                item.material.classificacao, None
            )

        ctx["itens"] = itens
        return ctx


class PedidoCreateView(LoginRequiredMixin, MaterialPrecosMixin, CreateView):
    model = Pedido
    form_class = PedidoForm
    template_name = "suprimentos/pedido_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx["formset"] = ItemPedidoFormSet(self.request.POST)
        else:
            ctx["formset"] = ItemPedidoFormSet()
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx["formset"]
        form.instance.solicitante = self.request.user
        if not form.instance.filial_id:
            form.instance.filial = getattr(self.request.user, "filial", None)
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            HistoricoPedido.registrar(
                pedido=self.object,
                descricao="Pedido criado (rascunho).",
                responsavel=self.request.user,
                status_novo=self.object.status,
            )
            messages.success(self.request, f"Pedido {self.object.numero} criado.")
            return redirect(self.object.get_absolute_url())
        return self.render_to_response(self.get_context_data(form=form))


class PedidoUpdateView(LoginRequiredMixin, MaterialPrecosMixin, UpdateView):
    model = Pedido
    form_class = PedidoForm
    template_name = "suprimentos/pedido_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx["formset"] = ItemPedidoFormSet(self.request.POST, instance=self.object)
        else:
            ctx["formset"] = ItemPedidoFormSet(instance=self.object)
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx["formset"]
        if formset.is_valid():
            self.object = form.save()
            formset.save()
            HistoricoPedido.registrar(
                pedido=self.object,
                descricao="Pedido editado.",
                responsavel=self.request.user,
            )
            messages.success(self.request, "Pedido atualizado.")
            return redirect(self.object.get_absolute_url())
        return self.render_to_response(self.get_context_data(form=form))


class AnexoPedidoUploadView(LoginRequiredMixin, View):
    """Upload de um ou vários anexos para um Pedido."""

    def post(self, request, pk):
        pedido = get_object_or_404(Pedido, pk=pk)
        form = AnexoPedidoForm(request.POST, request.FILES)

        if form.is_valid():
            arquivos = request.FILES.getlist("arquivos")
            observacao = form.cleaned_data.get("observacao", "")

            criados = 0
            for arquivo in arquivos:
                AnexoPedido.objects.create(
                    pedido=pedido,
                    arquivo=arquivo,
                    observacao=observacao,
                    enviado_por=request.user,
                )
                criados += 1

            messages.success(
                request,
                f"{criados} anexo(s) enviado(s) com sucesso!"
            )
        else:
            messages.error(request, "Erro ao enviar anexos. Verifique os arquivos.")

        return redirect("suprimentos:pedido_detalhe", pk=pedido.pk)


@login_required
def pedido_submeter(request, pk):
    """RASCUNHO/REVISAO → PENDENTE (envia para aprovação)."""
    pedido = get_object_or_404(Pedido, pk=pk)
    if pedido.status not in (Pedido.StatusChoices.RASCUNHO, Pedido.StatusChoices.REVISAO):
        messages.error(request, "Pedido não pode ser submetido neste status.")
        return redirect(pedido.get_absolute_url())
    if not pedido.itens.exists():
        messages.error(request, "Adicione ao menos um item antes de submeter.")
        return redirect(pedido.get_absolute_url())
    anterior = pedido.status
    pedido.status = Pedido.StatusChoices.PENDENTE
    pedido.save(update_fields=["status", "atualizado_em"])
    HistoricoPedido.registrar(
        pedido=pedido, descricao="Pedido submetido para aprovação.",
        responsavel=request.user,
        status_anterior=anterior, status_novo=pedido.status,
    )
    messages.success(request, "Pedido enviado para aprovação.")
    return redirect(pedido.get_absolute_url())


# ═════════════════════════════════════════════════════════════
# 2. APROVAR PEDIDO (Gerente)
# ═════════════════════════════════════════════════════════════
@login_required
@permission_required("suprimentos.pode_aprovar_pedido", raise_exception=True)
def pedido_aprovar(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    if pedido.status != Pedido.StatusChoices.PENDENTE:
        messages.error(request, "Apenas pedidos PENDENTES podem ser avaliados.")
        return redirect(pedido.get_absolute_url())

    if request.method == "POST":
        form = AprovarPedidoForm(request.POST)
        if form.is_valid():
            decisao = form.cleaned_data["decisao"]
            motivo = form.cleaned_data["motivo"]

            # 🔒 Bloqueia o registro para evitar dupla submissão (clique duplo)
            with transaction.atomic():
                pedido = Pedido.objects.select_for_update().get(pk=pedido.pk)

                # Revalida o status dentro do lock
                if pedido.status != Pedido.StatusChoices.PENDENTE:
                    messages.warning(
                        request, "Este pedido já foi avaliado por outra ação."
                    )
                    return redirect(pedido.get_absolute_url())

                anterior = pedido.status
                pedido.aprovador = request.user

                if decisao == "APROVAR":
                    pedido.status = Pedido.StatusChoices.APROVADO
                    pedido.data_aprovacao = timezone.now()
                    pedido.save()
                    HistoricoPedido.registrar(
                        pedido=pedido, descricao="Pedido aprovado.",
                        responsavel=request.user,
                        status_anterior=anterior, status_novo=pedido.status,
                    )
                    # Gera (ou reaproveita) a solicitação de compra
                    try:
                        sol = pedido.gerar_solicitacao_compra(request.user)
                        messages.success(
                            request,
                            f"Pedido aprovado. Solicitação {sol.numero} gerada.",
                        )
                        return redirect(sol.get_absolute_url())
                    except ValidationError as e:
                        messages.warning(
                            request,
                            f"Pedido aprovado, mas: {'; '.join(e.messages)}",
                        )
                        return redirect(pedido.get_absolute_url())

                elif decisao == "REVISAR":
                    pedido.status = Pedido.StatusChoices.REVISAO
                    pedido.motivo_revisao = motivo
                    pedido.save()
                    HistoricoPedido.registrar(
                        pedido=pedido,
                        descricao=f"Devolvido para revisão: {motivo}",
                        responsavel=request.user,
                        status_anterior=anterior, status_novo=pedido.status,
                    )
                    messages.info(request, "Pedido devolvido para revisão.")

                else:  # REPROVAR
                    pedido.status = Pedido.StatusChoices.REPROVADO
                    pedido.motivo_reprovacao = motivo
                    pedido.save()
                    HistoricoPedido.registrar(
                        pedido=pedido, descricao=f"Reprovado: {motivo}",
                        responsavel=request.user,
                        status_anterior=anterior, status_novo=pedido.status,
                    )
                    messages.warning(request, "Pedido reprovado.")

            return redirect(pedido.get_absolute_url())
    else:
        form = AprovarPedidoForm()

    return render(request, "suprimentos/pedido_aprovar.html", {
        "pedido": pedido, "form": form,
        "itens": pedido.itens.select_related("material"),
    })


# ═════════════════════════════════════════════════════════════
# 3. SOLICITAÇÃO — Listagem / Detalhe / Cotação (NxN)
# ═════════════════════════════════════════════════════════════
class SolicitacaoListView(LoginRequiredMixin, ListView):
    model = SolicitacaoCompra
    template_name = "suprimentos/solicitacao_list.html"
    context_object_name = "solicitacoes"
    paginate_by = 20

    def get_queryset(self):
        qs = SolicitacaoCompra.objects.select_related("contrato", "solicitante", "comprador")
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["status_choices"] = SolicitacaoCompra.StatusChoices.choices
        return ctx


class SolicitacaoDetailView(LoginRequiredMixin, DetailView):
    model = SolicitacaoCompra
    template_name = "suprimentos/solicitacao_detail.html"
    context_object_name = "solicitacao"

    # ──────────────────────────────────────────────────────────────
    # Ordem oficial do fluxo da solicitação (CANCELADO fica fora)
    # ──────────────────────────────────────────────────────────────
    @property
    def fluxo(self):
        S = SolicitacaoCompra.StatusChoices
        return [
            S.FAZER_COTACAO,
            S.COTACAO_ENVIADA,
            S.EM_APROVACAO,
            S.APROVADO,
            S.ENVIAR_PEDIDO,
            S.PEDIDO_GERADO,
            S.EM_ENTREGA,
            S.FINALIZADO,
        ]

    # ──────────────────────────────────────────────────────────────
    def get_queryset(self):
        """Otimiza o carregamento do objeto principal e relações diretas."""
        return (
            super()
            .get_queryset()
            .select_related(
                "pedido",
                "pedido__contrato",
                "solicitante",
                "comprador",
                "aprovador_cotacao",
            )
        )

    # ──────────────────────────────────────────────────────────────
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        sol = self.object
        user = self.request.user
        S = SolicitacaoCompra.StatusChoices
        status = sol.status

        # ───────── Anexos gerais (esconde confidenciais sem permissão) ─────────
        anexos_qs = sol.anexos.select_related("enviado_por")
        if not user.has_perm("suprimentos.view_anexosolicitacao_confidencial"):
            anexos_qs = anexos_qs.filter(confidencial=False)

        # ───────── Anexos de cotação (dedup por fornecedor + arquivo) ─────────
        cotacoes_anexo = (
            Cotacao.objects
            .filter(
                item_solicitacao__solicitacao=sol,
                anexo_cotacao__isnull=False,
            )
            .exclude(anexo_cotacao="")
            .select_related("fornecedor", "item_solicitacao__material")
        )
        anexos_cotacao = OrderedDict()
        for cot in cotacoes_anexo:
            chave = (cot.fornecedor_id, cot.anexo_cotacao.name)
            anexos_cotacao.setdefault(chave, cot)

        # ───────── Posição atual no fluxo ─────────
        fluxo = self.fluxo
        try:
            pos = fluxo.index(status)
        except ValueError:          # CANCELADO ou status fora do fluxo
            pos = -1

        def passou(etapa) -> bool:
            """True se o fluxo já atingiu (ou ultrapassou) a etapa informada."""
            return pos >= fluxo.index(etapa)

        # ───────── Querysets / dados base ─────────
        itens = (
            sol.itens
            .select_related("material")
            .prefetch_related("cotacoes__fornecedor")
        )
        pedidos_compra = sol.pedidos_compra.select_related("fornecedor")
        tem_cotacoes = itens.filter(cotacoes__isnull=False).exists()

        # ───────── Permissões ─────────
        pode_cotar_perm     = user.has_perm("suprimentos.pode_cotar")
        pode_aprovar_perm   = user.has_perm("suprimentos.pode_aprovar_cotacao")
        pode_montar_pc_perm = user.has_perm("suprimentos.pode_montar_pc")

        # ───────── Estados pontuais ─────────
        em_cotacao   = status in (S.FAZER_COTACAO, S.COTACAO_ENVIADA)
        em_aprovacao = status == S.EM_APROVACAO
        pronto_pc    = status in (S.APROVADO, S.ENVIAR_PEDIDO)  # pode montar PC

        # ───────── Monta contexto (um único update) ─────────
        ctx.update(
            {
                # Dados
                "itens": itens,
                "anexos": anexos_qs,                       # ← respeita confidenciais
                "anexos_cotacao": list(anexos_cotacao.values()),
                "pedidos_compra": pedidos_compra,
                "historico": sol.historico.all(),
                "comparativo": self._montar_comparativo(itens),
                "tem_cotacoes": tem_cotacoes,
                "todos_cotados": sol.todos_itens_cotados,

                # Stepper — "done" = já passou da etapa; "active" = está nela
                "step_cotacao_done":   passou(S.EM_APROVACAO),
                "step_cotacao_active": em_cotacao,
                "step_aprov_done":     passou(S.APROVADO),
                "step_aprov_active":   em_aprovacao,
                "step_pc_done":        passou(S.PEDIDO_GERADO),
                "step_pc_active":      pronto_pc,

                # Navegação lateral
                "nav_cotacao":   status != S.FAZER_COTACAO,
                "nav_aprovacao": passou(S.EM_APROVACAO),
                "pc_gerado":     passou(S.PEDIDO_GERADO),

                # Botões de ação (estado + permissão)
                "pode_cotar": em_cotacao and pode_cotar_perm,
                "pode_enviar_aprovacao": (
                    status == S.COTACAO_ENVIADA
                    and tem_cotacoes
                    and pode_cotar_perm
                ),
                "pode_aprovar_cotacao": em_aprovacao and pode_aprovar_perm,
                "pode_montar_pc": pronto_pc and pode_montar_pc_perm,
            }
        )
        return ctx

    # ──────────────────────────────────────────────────────────────
    @staticmethod
    def _montar_comparativo(itens):
        """Total que cada fornecedor representaria se fornecesse tudo o que cotou."""
        agregado = defaultdict(lambda: {"total": Decimal("0.00"), "qtd_itens": 0})

        for item in itens:
            for cotacao in item.cotacoes.all():
                nome = cotacao.fornecedor.nome_fantasia
                agregado[nome]["total"] += cotacao.valor_total
                agregado[nome]["qtd_itens"] += 1

        return sorted(
            ({"fornecedor": nome, **dados} for nome, dados in agregado.items()),
            key=lambda x: x["total"],
        )
    

@login_required
@permission_required("suprimentos.pode_cotar", raise_exception=True)
def cotacao_adicionar(request, solicitacao_pk):
    """
    Lança a cotação de UM fornecedor para os itens da solicitação.
    Gera 1 Cotacao por ItemSolicitacao que tiver valor preenchido.
    """
    arquivo_anexo = cab_form.cleaned_data.get("anexo_cotacao")
    solicitacao = get_object_or_404(SolicitacaoCompra, pk=solicitacao_pk)
    itens = list(
        solicitacao.itens.select_related("material")
        .exclude(status=ItemSolicitacao.StatusItem.CANCELADO)
    )

    # initial do formset: uma linha por item da solicitação
    initial_itens = [{"item_id": item.pk} for item in itens]

    if request.method == "POST":
        cab_form = CotacaoCabecalhoForm(request.POST, request.FILES)
        formset = CotacaoItemValorFormSet(request.POST, initial=initial_itens)

        if cab_form.is_valid() and formset.is_valid():
            fornecedor = cab_form.cleaned_data["fornecedor"]

            # Coleta apenas os itens com valor > 0 preenchido
            linhas_validas = [
                f.cleaned_data for f in formset
                if f.cleaned_data.get("valor_unitario")
            ]

            if not linhas_validas:
                messages.error(request, "Informe ao menos um valor por item.")
            else:
                # Mapa rápido id -> ItemSolicitacao
                mapa_itens = {item.pk: item for item in itens}
                try:
                    with transaction.atomic():
                        criadas = 0
                        for linha in linhas_validas:
                            item = mapa_itens.get(linha["item_id"])
                            if not item:
                                continue

                            # Verifica duplicidade ANTES (respeita o UniqueConstraint)
                            ja_existe = Cotacao.objects.filter(
                                item_solicitacao=item,
                                fornecedor=fornecedor,
                            ).exists()
                            if ja_existe:
                                raise IntegrityError(
                                    f"item:{item.material.descricao}"
                                )

                            Cotacao.objects.create(
                                item_solicitacao=item,
                                fornecedor=fornecedor,
                                valor_unitario=linha["valor_unitario"],
                                prazo_entrega_dias=cab_form.cleaned_data.get("prazo_entrega_dias"),
                                condicoes_pagamento=cab_form.cleaned_data.get("condicoes_pagamento", ""),
                                validade_cotacao=cab_form.cleaned_data.get("validade_cotacao"),
                                observacoes=cab_form.cleaned_data.get("observacoes", ""),
                                anexo_cotacao=arquivo_anexo if criadas == 0 else None,
                                criado_por=request.user,
                            )
                            # Atualiza status do item
                            if item.status == ItemSolicitacao.StatusItem.PENDENTE_COTACAO:
                                item.status = ItemSolicitacao.StatusItem.COTADO
                                item.save(update_fields=["status"])
                            criadas += 1

                        # Atualiza status da solicitação se já está cotando
                        if solicitacao.status == SolicitacaoCompra.StatusChoices.FAZER_COTACAO:
                            solicitacao.status = SolicitacaoCompra.StatusChoices.COTACAO_ENVIADA
                            solicitacao.save(update_fields=["status", "atualizado_em"])

                    messages.success(
                        request,
                        f"Cotação de {fornecedor.nome_fantasia} lançada para {criadas} item(ns)!"
                    )
                    return redirect("suprimentos:solicitacao_detalhe", pk=solicitacao.pk)

                except IntegrityError as e:
                    messages.error(
                        request,
                        f"O fornecedor '{fornecedor.nome_fantasia}' já cotou um ou mais "
                        f"destes itens. Edite a cotação existente, se necessário."
                    )
    else:
        cab_form = CotacaoCabecalhoForm()
        formset = CotacaoItemValorFormSet(initial=initial_itens)

    # Junta cada subform com seu item para o template (mesma ordem!)
    linhas = list(zip(formset.forms, itens))

    return render(request, "suprimentos/cotacao_form.html", {
        "cab_form": cab_form,
        "formset": formset,
        "linhas": linhas,
        "solicitacao": solicitacao,
    })


@login_required
@permission_required("suprimentos.pode_cotar", raise_exception=True)
def cotacao_excluir(request, pk):
    cot = get_object_or_404(Cotacao, pk=pk)
    sol_pk = cot.item_solicitacao.solicitacao.pk
    cot.delete()
    messages.info(request, "Cotação removida.")
    return redirect("suprimentos:solicitacao_detalhe", pk=sol_pk)


@login_required
@permission_required("suprimentos.pode_cotar", raise_exception=True)
def solicitacao_enviar_aprovacao(request, pk):
    """FAZER_COTACAO → EM_APROVACAO (todos os itens cotados)."""
    sol = get_object_or_404(SolicitacaoCompra, pk=pk)
    if not sol.todos_itens_cotados:
        messages.error(request, "Todos os itens precisam de ao menos 1 cotação.")
        return redirect(sol.get_absolute_url())
    sol.comprador = request.user
    sol.data_cotacao = timezone.now().date()
    sol.status = SolicitacaoCompra.StatusChoices.EM_APROVACAO
    sol.save()
    messages.success(request, "Cotações enviadas para aprovação.")
    return redirect(sol.get_absolute_url())


# ─────────────────────────────────────────────────────────────
# HELPERS — cotacao_aprovar
# ─────────────────────────────────────────────────────────────
def _coletar_escolhas(request, itens):
    """
    Lê as escolhas do POST e soma por classificação.
    Retorna: (escolhas, total_por_classif, erro_str_ou_None)
    """
    escolhas = []  # [(item, cotacao_ou_None)]
    total_por_classif = defaultdict(lambda: Decimal("0.00"))

    for item in itens:
        cotacao_id = request.POST.get(f"cotacao_{item.pk}", "").strip()

        if cotacao_id:
            # valida que o ID é numérico antes de consultar
            if not cotacao_id.isdigit():
                return [], total_por_classif, (
                    f"Cotação inválida para o item “{item.material.descricao}”."
                )
            cotacao = item.cotacoes.filter(pk=cotacao_id).first()
            if not cotacao:
                return [], total_por_classif, (
                    f"Cotação inválida para o item “{item.material.descricao}”."
                )
            escolhas.append((item, cotacao))
            total_por_classif[item.material.classificacao] += cotacao.valor_total
        else:
            escolhas.append((item, None))  # "Nenhum" → não comprar

    return escolhas, total_por_classif, None


def _checar_estouros(total_por_classif, verba, labels):
    """Retorna lista de mensagens de estouro por classificação."""
    estouros = []
    for classif, total in total_por_classif.items():
        campo_saldo = SALDO_POR_CLASSIF.get(classif)
        if not campo_saldo:
            continue
        saldo = getattr(verba, campo_saldo, None) or Decimal("0.00")
        if total > saldo:
            label = labels.get(classif, classif)
            estouros.append(
                f"{label}: escolhido R$ {total:.2f} > saldo R$ {saldo:.2f}"
            )
    return estouros


def _devolver_para_revisao(request, sol, msg_estouro, S):
    """Devolve o pedido vinculado para REVISÃO de forma atômica."""
    with transaction.atomic():
        pedido = getattr(sol, "pedido", None)
        if pedido is not None:
            status_anterior = pedido.status
            pedido.status = Pedido.StatusChoices.REVISAO
            pedido.motivo_revisao = msg_estouro
            pedido.save(update_fields=[
                "status", "motivo_revisao", "atualizado_em",
            ])
            _registrar_historico(
                HistoricoPedido.registrar,
                pedido=pedido,
                descricao=f"Devolvido para revisão (verba): {msg_estouro}",
                responsavel=request.user,
                status_anterior=status_anterior,
                status_novo=Pedido.StatusChoices.REVISAO,
            )

        _registrar_historico(
            HistoricoSolicitacao.registrar,
            solicitacao=sol,
            descricao=f"Cotação não aprovada — {msg_estouro}",
            responsavel=request.user,
            status_anterior=sol.status,
            status_novo=sol.status,
        )


def _montar_comparativo(itens):
    """Agrega cotações por fornecedor para o comparativo (GET)."""
    agg = {}
    for item in itens:
        for cot in item.cotacoes.all():
            f = cot.fornecedor
            entry = agg.setdefault(f.id, {
                "fornecedor": f,
                "total": Decimal("0.00"),
                "qtd_itens": 0,
                "prazos": [],
                "pagamentos": set(),
            })
            entry["total"] += cot.valor_total
            entry["qtd_itens"] += 1
            if cot.prazo_entrega_dias is not None:
                entry["prazos"].append(cot.prazo_entrega_dias)
            if cot.condicoes_pagamento:
                entry["pagamentos"].add(cot.condicoes_pagamento.strip())

    comparativo = []
    for entry in agg.values():
        prazos = entry.pop("prazos")
        pagamentos = entry.pop("pagamentos")
        entry["prazo"] = max(prazos) if prazos else None
        entry["pagamentos_lista"] = " | ".join(sorted(pagamentos)) if pagamentos else ""
        if len(pagamentos) == 1:
            entry["pagamento"] = next(iter(pagamentos))
        elif len(pagamentos) > 1:
            entry["pagamento"] = "Variadas"
        else:
            entry["pagamento"] = None
        comparativo.append(entry)

    comparativo.sort(key=lambda d: d["total"])
    return comparativo


def _montar_painel_verba(sol, labels):
    """Monta o painel de verba e o dict de saldos para o JS (GET)."""
    verba, _ = _obter_verba_do_mes(sol)

    painel_verba = []
    saldos_json = {}
    for classif, (campo_verba, campo_saldo) in CAMPOS_VERBA.items():
        v_verba = (getattr(verba, campo_verba, None) or Decimal("0.00")) if verba else Decimal("0.00")
        v_saldo = (getattr(verba, campo_saldo, None) or Decimal("0.00")) if verba else Decimal("0.00")
        painel_verba.append({
            "classif": classif,
            "label": labels.get(classif, classif),
            "verba": v_verba,
            "saldo": v_saldo,
            "estimado": Decimal("0.00"),
        })
        saldos_json[classif] = float(v_saldo)

    return painel_verba, saldos_json


# ─────────────────────────────────────────────────────────────
# 4. APROVAR COTAÇÃO (escolha por item) + CHECAGEM DE VERBA
# ─────────────────────────────────────────────────────────────
@login_required
@permission_required("suprimentos.pode_aprovar_cotacao", raise_exception=True)
def cotacao_aprovar(request, pk):
    """
    Gerente escolhe a cotação vencedora de cada item.

    Reforço de servidor:
        - Só processa se a solicitação estiver EM_APROVACAO (idempotência).
        - select_for_update() evita aprovação concorrente (duplo clique / abas).
        - Toda gravação é atômica; verba insuficiente devolve para REVISÃO.

    Regra de verba (por classificação EPI/CONSUMO/FERRAMENTA):
        - Verba suficiente   → APROVA a solicitação.
        - Verba insuficiente → devolve o Pedido para REVISÃO.
    """
    sol = get_object_or_404(
        SolicitacaoCompra.objects.select_related("contrato"),
        pk=pk,
    )

    S = SolicitacaoCompra.StatusChoices

    # ═════════════════════════════════════════════════════════
    # 🔒 GUARDA DE STATUS (server-side, antes de qualquer coisa)
    # ═════════════════════════════════════════════════════════
    if sol.status != S.EM_APROVACAO:
        messages.warning(
            request,
            f"Esta solicitação não está disponível para aprovação "
            f"(status atual: {sol.get_status_display()}).",
        )
        return redirect(sol.get_absolute_url())

    itens = (
        sol.itens
        .select_related("material")
        .prefetch_related("cotacoes__fornecedor")
    )
    labels = dict(CategoriaMaterial.choices)

    # ═════════════════════════════════════════════════════════
    # POST — coleta escolhas → valida verba → grava (atômico)
    # ═════════════════════════════════════════════════════════
    if request.method == "POST":
        escolhas, total_por_classif, erro_post = _coletar_escolhas(request, itens)
        if erro_post:
            messages.error(request, erro_post)
            return redirect(request.path)

        # Exige ao menos uma cotação escolhida
        if not any(cot for _, cot in escolhas):
            messages.warning(request, "Escolha ao menos uma cotação para aprovar.")
            return redirect(request.path)

        # ── Verba do mês (defensivo) ─────────────────────────
        verba, erro = _obter_verba_do_mes(sol)
        if erro:
            messages.error(request, erro)
            return redirect(request.path)
        if verba is None:
            messages.error(
                request,
                "Verba do contrato não cadastrada para o período da solicitação.",
            )
            return redirect(request.path)

        # ── Checagem de saldo por classificação ──────────────
        estouros = _checar_estouros(total_por_classif, verba, labels)

        # ── VERBA INSUFICIENTE → devolve para REVISÃO ────────
        if estouros:
            msg_estouro = "Verba insuficiente — " + " | ".join(estouros)
            _devolver_para_revisao(request, sol, msg_estouro, S)
            messages.error(
                request,
                f"❌ {msg_estouro}. Pedido devolvido para revisão do solicitante.",
            )
            return redirect(sol.get_absolute_url())

        # ── VERBA OK → grava tudo de forma atômica e APROVA ──
        try:
            with transaction.atomic():
                # 🔒 Lock + revalidação do status sob trava (anti-concorrência)
                sol_locked = (
                    SolicitacaoCompra.objects
                    .select_for_update()
                    .get(pk=sol.pk)
                )
                if sol_locked.status != S.EM_APROVACAO:
                    messages.warning(
                        request,
                        "Esta solicitação já foi processada por outro usuário.",
                    )
                    return redirect(sol_locked.get_absolute_url())

                for item, cot in escolhas:
                    if cot:
                        item.cotacao_escolhida = cot
                        item.status = ItemSolicitacao.StatusItem.APROVADO
                    else:
                        item.cotacao_escolhida = None
                        item.status = ItemSolicitacao.StatusItem.REPROVADO  # "não comprar"
                    item.save(update_fields=[
                        "cotacao_escolhida", "status", "atualizado_em",
                    ])

                sol_locked.status = S.APROVADO
                sol_locked.aprovador_cotacao = request.user
                sol_locked.data_validacao_cotacao = timezone.now().date()
                sol_locked.save(update_fields=[
                    "status", "aprovador_cotacao",
                    "data_validacao_cotacao", "atualizado_em",
                ])

                _registrar_historico(
                    HistoricoSolicitacao.registrar,
                    solicitacao=sol_locked,
                    descricao="Cotações aprovadas dentro da verba.",
                    responsavel=request.user,
                    status_anterior=S.EM_APROVACAO,
                    status_novo=S.APROVADO,
                )
        except SolicitacaoCompra.DoesNotExist:
            messages.error(request, "Solicitação não encontrada.")
            return redirect("suprimentos:lista_solicitacoes")

        messages.success(
            request,
            "✅ Cotações aprovadas dentro da verba. Monte os Pedidos de Compra.",
        )
        return redirect("suprimentos:montar_pedido_compra", pk=sol.pk)

    # ═════════════════════════════════════════════════════════
    # GET — comparativo por fornecedor + painel de verba + saldos
    # ═════════════════════════════════════════════════════════
    comparativo = _montar_comparativo(itens)
    painel_verba, saldos_json = _montar_painel_verba(sol, labels)

    return render(request, "suprimentos/cotacao_aprovar.html", {
        "solicitacao": sol,
        "itens": itens,
        "comparativo": comparativo,
        "painel_verba": painel_verba,
        "saldos_json": saldos_json,
    })


# ═════════════════════════════════════════════════════════════
# 5. MONTAR PEDIDO DE COMPRA (NxN — agrupa por fornecedor)
# ═════════════════════════════════════════════════════════════
@login_required
@permission_required("suprimentos.pode_montar_pc", raise_exception=True)
@transaction.atomic
def montar_pedido_compra(request, pk):
    """Agrupa itens aprovados por fornecedor e gera 1 PedidoCompra por fornecedor."""
    sol = get_object_or_404(SolicitacaoCompra, pk=pk)

    # Alias local para legibilidade
    PC_GERADO = SolicitacaoCompra.StatusChoices.PEDIDO_GERADO

    if request.method == "POST":
        sol = SolicitacaoCompra.objects.select_for_update().get(pk=sol.pk)

        # ── 1) Idempotência ──────────────────────────────────────
        if sol.status == PC_GERADO:
            messages.info(request, "Os Pedidos de Compra já foram gerados.")
            return redirect(sol.get_absolute_url())

        status_anterior = sol.status


        # 🔁 Reavalia DENTRO do lock (consistência garantida)
        itens_aprovados = sol.itens.filter(
            status=ItemSolicitacao.StatusItem.APROVADO,
            cotacao_escolhida__isnull=False,
        ).select_related("cotacao_escolhida__fornecedor", "material")

        if not itens_aprovados.exists():
            messages.error(request, "Nenhum item aprovado com cotação escolhida.")
            return redirect(sol.get_absolute_url())

        # ✅ Respeita os fornecedores marcados nos checkboxes
        fornecedores_sel = set(request.POST.getlist("fornecedores"))
        observacoes = request.POST.get("observacoes", "").strip()

        # Agrupa por fornecedor (apenas os selecionados)
        grupos = {}
        for item in itens_aprovados:
            forn = item.cotacao_escolhida.fornecedor
            if fornecedores_sel and str(forn.pk) not in fornecedores_sel:
                continue
            grupos.setdefault(forn.pk, {"fornecedor": forn, "itens": []})
            grupos[forn.pk]["itens"].append(item)

        if not grupos:
            messages.error(request, "Selecione ao menos um fornecedor para gerar o PC.")
            return redirect(request.path)

        pcs_criados = []
        for dados in grupos.values():
            pc = PedidoCompra.objects.create(
                solicitacao=sol,
                fornecedor=dados["fornecedor"],
                filial=sol.filial,
                status=PedidoCompra.StatusPC.EMITIDO,
                data_emissao=timezone.now().date(),
                criado_por=request.user,
                observacoes=observacoes,
            )
            for item in dados["itens"]:
                cot = item.cotacao_escolhida
                ItemPedidoCompra.objects.create(
                    pedido_compra=pc,
                    cotacao=cot,
                    item_solicitacao=item,
                    material=item.material,
                    quantidade=item.quantidade,
                    valor_unitario=cot.valor_unitario,
                )
            pc.recalcular_total()
            pcs_criados.append(pc)

        # ── Atualiza status da solicitação ───────────────────────
        sol.status = PC_GERADO  # ← era SOLICITACAO_GERADA
        sol.data_criacao_pedido = timezone.now().date()
        sol.save(update_fields=["status", "data_criacao_pedido", "atualizado_em"])

         # ── Histórico ────────────────────────────────────────────
        try:
            HistoricoSolicitacao.registrar(
                solicitacao=sol,
                descricao=(
                    f"{len(pcs_criados)} Pedido(s) de Compra emitido(s): "
                    + ", ".join(pc.numero for pc in pcs_criados)
                ),
                responsavel=request.user,
                status_anterior=status_anterior,
                status_novo=PC_GERADO,  # ← era SOLICITACAO_GERADA
            )
        except Exception:
            pass

        messages.success(request, f"{len(pcs_criados)} Pedido(s) de Compra emitido(s).")

        if len(pcs_criados) == 1:
            return redirect(pcs_criados[0].get_absolute_url())
        return redirect(sol.get_absolute_url())

    # ── Preview (GET) ─────────────────────────────────────────
    itens_aprovados = sol.itens.filter(
        status=ItemSolicitacao.StatusItem.APROVADO,
        cotacao_escolhida__isnull=False,
    ).select_related("cotacao_escolhida__fornecedor", "material")

    if not itens_aprovados.exists():
        messages.error(request, "Nenhum item aprovado com cotação escolhida.")
        return redirect(sol.get_absolute_url())

    # Agrupa por fornecedor + consolida prazo/pagamento divergentes
    #   prazo:     maior (pior caso de entrega)
    #   pagamento: condição única, "Variadas" ou None
    grupos = {}
    for item in itens_aprovados:
        cot = item.cotacao_escolhida
        forn = cot.fornecedor

        g = grupos.setdefault(forn.pk, {
            "fornecedor": forn,
            "itens": [],
            "total": Decimal("0.00"),
            "prazos": [],
            "pagamentos": set(),
        })

        item.valor_cotado = cot.valor_unitario
        item.total_cotado = cot.valor_total
        g["itens"].append(item)
        g["total"] += cot.valor_total

        if cot.prazo_entrega_dias is not None:
            g["prazos"].append(cot.prazo_entrega_dias)
        if cot.condicoes_pagamento:
            g["pagamentos"].add(cot.condicoes_pagamento.strip())

    # Pós-processa: consolida prazo e pagamento por fornecedor
    grupos_fornecedor = []
    for g in grupos.values():
        prazos = g.pop("prazos")
        pagamentos = g.pop("pagamentos")

        # Prazo → maior (pior caso); None se não houver
        g["prazo"] = max(prazos) if prazos else None

        # Pagamento → único, "Variadas" ou None
        if len(pagamentos) == 1:
            g["pagamento"] = next(iter(pagamentos))
        elif len(pagamentos) > 1:
            g["pagamento"] = "Variadas"
        else:
            g["pagamento"] = None

        # Para tooltip quando houver divergência
        g["pagamentos_lista"] = " | ".join(sorted(pagamentos)) if pagamentos else ""

        grupos_fornecedor.append(g)

    grupos_fornecedor = sorted(grupos_fornecedor, key=lambda d: d["total"])

    return render(request, "suprimentos/montar_pc.html", {
        "solicitacao": sol,
        "grupos_fornecedor": grupos_fornecedor,
        "itens": list(itens_aprovados),
        "verba_disponivel": sol.valor_estimado,
    })

# ═════════════════════════════════════════════════════════════
# 6. PEDIDO DE COMPRA — Detalhe / Acompanhar Entrega / Finalizar
# ═════════════════════════════════════════════════════════════
class PedidoCompraDetailView(LoginRequiredMixin, DetailView):
    model = PedidoCompra
    template_name = "suprimentos/pedido_compra_detail.html"
    context_object_name = "pc"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["itens"] = self.object.itens.select_related("material")
        ctx["anexos_entrega"] = self.object.anexos_entrega.all()  # ← novo
        return ctx

def _render_entrega(request, pc, itens, form):
    context = {
        'pc': pc,
        'itens': itens,
        'form': form,
    }
    return render(request, 'suprimentos/pc_entrega.html', context)

@login_required
@permission_required("suprimentos.pode_receber_pedido_compra", raise_exception=True)
@transaction.atomic
def pc_finalizar(request, pk):
    """RECEBIDO — completa o saldo restante e, se todos da SOL, conclui."""
    pc = get_object_or_404(PedidoCompra, pk=pk)
    pc.status = PedidoCompra.StatusPC.RECEBIDO
    pc.recebido_por = request.user
    if not pc.data_entrega_efetiva:
        pc.data_entrega_efetiva = timezone.now().date()
    pc.save()

    # Completa apenas o que faltar (não sobrescreve entregas parciais)
    itens = list(pc.itens.all())
    for item in itens:
        if (item.quantidade_recebida or 0) < item.quantidade:
            item.quantidade_recebida = item.quantidade
    if itens:
        ItemPedidoCompra.objects.bulk_update(itens, ["quantidade_recebida"])

    sol = pc.solicitacao
    sol.sincronizar_status_entrega(responsavel=request.user)
    if sol.status == SolicitacaoCompra.StatusChoices.FINALIZADO:
        messages.success(request, "PC recebido. Solicitação CONCLUÍDA!")
    else:
        messages.success(request, "Pedido de Compra recebido.")
    return redirect(pc.get_absolute_url())

@login_required
@permission_required("suprimentos.pode_emitir_pedido_compra", raise_exception=True)
def pc_enviar_fornecedor(request, pk):
    pc = get_object_or_404(PedidoCompra, pk=pk)
    pc.status = PedidoCompra.StatusPC.ENVIADO_FORNECEDOR
    pc.data_envio = timezone.now().date()
    pc.save(update_fields=["status", "data_envio", "atualizado_em"])

    # ⬇️ TRANSIÇÃO QUE FALTAVA: a solicitação entra em fase de entrega
    sol = pc.solicitacao
    S = SolicitacaoCompra.StatusChoices
    if sol.status in (S.PEDIDO_GERADO,):
        sol.status = S.EM_ENTREGA
        sol.save(update_fields=["status", "atualizado_em"])
        _registrar_historico(
            HistoricoSolicitacao.registrar,
            solicitacao=sol,
            descricao=f"PC {pc.numero} enviado ao fornecedor — solicitação em entrega.",
            responsavel=request.user,
            status_anterior=S.PEDIDO_GERADO,
            status_novo=S.EM_ENTREGA,
        )

    messages.success(request, "Pedido enviado ao fornecedor.")
    return redirect(pc.get_absolute_url())



@login_required
@permission_required("suprimentos.pode_receber_pedido_compra", raise_exception=True)
@transaction.atomic
def pc_acompanhar_entrega(request, pk):
    """
    Registra recebimento de um Pedido de Compra.

    Reforço de servidor:
        - Bloqueia GET/POST se o PC estiver ENTREGUE / RECEBIDO / CANCELADO.
        - select_for_update() evita recebimento concorrente.
    """
    pc = get_object_or_404(
        PedidoCompra.objects.prefetch_related("itens__material"), pk=pk
    )

    S = PedidoCompra.StatusPC

    # Status que impedem novo recebimento
    STATUS_FINALIZADOS = (S.ENTREGUE, S.RECEBIDO, S.CANCELADO)

    # ═════════════════════════════════════════════════════════
    # 🔒 GUARDA DE STATUS
    # ═════════════════════════════════════════════════════════
    if pc.status in STATUS_FINALIZADOS:
        messages.warning(
            request,
            f"Este pedido não está disponível para entrega "
            f"(status atual: {pc.get_status_display()}).",
        )
        return redirect("suprimentos:pc_detalhe", pk=pc.pk)

    itens = list(pc.itens.select_related("material"))

    if request.method == "POST":
        form = EntregaPedidoCompraForm(request.POST, request.FILES, instance=pc)

        if form.is_valid():
            # 🔒 Lock + revalidação sob trava
            pc_locked = (
                PedidoCompra.objects
                .select_for_update()
                .get(pk=pc.pk)
            )
            if pc_locked.status in STATUS_FINALIZADOS:
                messages.warning(
                    request,
                    "Este pedido já foi finalizado por outro usuário.",
                )
                return redirect("suprimentos:pc_detalhe", pk=pc.pk)

            pc = form.save()

            # 1) Processa recebimento item a item
            houve_recebimento = False
            for item in itens:
                raw = request.POST.get(f"recebido_{item.pk}", "").strip()
                if not raw:
                    continue
                try:
                    qtd = Decimal(raw.replace(",", "."))
                except (InvalidOperation, AttributeError):
                    messages.warning(
                        request,
                        f"Valor inválido para o item {item.material.descricao}.",
                    )
                    continue

                if qtd <= 0:
                    continue
                if qtd > item.saldo:
                    messages.error(
                        request,
                        f"Quantidade recebida ({qtd}) maior que o saldo "
                        f"({item.saldo}) para {item.material.descricao}.",
                    )
                    transaction.set_rollback(True)
                    return redirect(
                        "suprimentos:pc_acompanhar_entrega", pk=pc.pk
                    )  # ✅ redireciona, sai do atomic

                item.quantidade_recebida = (item.quantidade_recebida or Decimal("0")) + qtd
                item.save(update_fields=["quantidade_recebida"])
                houve_recebimento = True

            # 2) Salva anexos
            for arquivo in form.cleaned_data.get("anexos", []):
                EntregaAnexo.objects.create(
                    pedido_compra=pc,
                    arquivo=arquivo,
                    nota_fiscal=pc.numero_nota_fiscal or "",
                    enviado_por=request.user,
                )

            # 3) Registra quem recebeu + atualiza status
            if houve_recebimento and not pc.recebido_por:
                pc.recebido_por = request.user
                pc.save(update_fields=["recebido_por", "atualizado_em"])

            pc.atualizar_status_entrega()
            pc.solicitacao.sincronizar_status_entrega(responsavel=request.user)

            if houve_recebimento:
                messages.success(request, "Recebimento registrado com sucesso!")
            else:
                messages.info(request, "Dados de entrega atualizados.")
            return redirect("suprimentos:pc_detalhe", pk=pc.pk)
    else:
        form = EntregaPedidoCompraForm(instance=pc)

    return _render_entrega(request, pc, itens, form)


@login_required
@permission_required("suprimentos.pode_visualizar_pedido_compra", raise_exception=True)
def pc_detalhe(request, pk):
    """Detalhe do Pedido de Compra com todas as entregas/anexos e progresso."""
    pc = get_object_or_404(
        PedidoCompra.objects.select_related("fornecedor", "filial"),
        pk=pk,
    )
    itens = list(pc.itens.select_related("material"))

    # Anexos de entrega (ordenados do mais recente p/ o mais antigo)
    anexos = (
        pc.anexos_entrega
        .select_related("enviado_por")
        .order_by("enviado_em")  # ajuste se o campo de data tiver outro nome
    )

    # Totais agregados
    total_pedido = sum((i.quantidade or Decimal("0") for i in itens), Decimal("0"))
    total_recebido = sum((i.quantidade_recebida or Decimal("0") for i in itens), Decimal("0"))
    total_saldo = total_pedido - total_recebido

    progresso_pct = Decimal("0")
    if total_pedido > 0:
        progresso_pct = (total_recebido / total_pedido * 100).quantize(Decimal("0.1"))

    valor_total = sum(
        ((i.quantidade or Decimal("0")) * (i.valor_unitario or Decimal("0")) for i in itens),
        Decimal("0"),
    )

    # Agrupa anexos por nota fiscal (cada "recebimento" = uma NF)
    entregas = {}
    for anexo in anexos:
        chave = anexo.nota_fiscal or "Sem NF"
        entregas.setdefault(chave, []).append(anexo)

    context = {
        "pc": pc,
        "itens": itens,
        "anexos": anexos,
        "entregas": entregas,
        "total_pedido": total_pedido,
        "total_recebido": total_recebido,
        "total_saldo": total_saldo,
        "progresso_pct": progresso_pct,
        "valor_total": valor_total,
    }
    return render(request, "suprimentos/pc_detalhe.html", context)


class PedidoCompraListView(LoginRequiredMixin, ListView):
    model = PedidoCompra
    template_name = "suprimentos/pc_list.html"
    context_object_name = "pedidos_compra"
    paginate_by = 20

    def get_queryset(self):
        qs = (
            PedidoCompra.objects
            .select_related("fornecedor", "filial", "solicitacao")
            .prefetch_related("itens")
            .order_by("-criado_em")
        )

        # ── Filtros ──────────────────────────────────────────
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        busca = self.request.GET.get("q")
        if busca:
            qs = qs.filter(
                Q(numero__icontains=busca)
                | Q(numero_pedido__icontains=busca)
                | Q(fornecedor__nome_fantasia__icontains=busca)
                | Q(numero_nota_fiscal__icontains=busca)
            )

        atrasados = self.request.GET.get("atrasados")
        if atrasados == "1":
            qs = qs.filter(
                data_entrega_prevista__lt=timezone.now().date(),
            ).exclude(
                status__in=[
                    PedidoCompra.StatusPC.ENTREGUE,
                    PedidoCompra.StatusPC.RECEBIDO,
                    PedidoCompra.StatusPC.CANCELADO,
                ]
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        hoje = timezone.now().date()
        S = PedidoCompra.StatusPC

        base = PedidoCompra.objects.all()
        
        ctx["status_choices"] = S.choices
        ctx["status_atual"] = self.request.GET.get("status", "")
        ctx["busca_atual"] = self.request.GET.get("q", "")
        ctx["atrasados_atual"] = self.request.GET.get("atrasados", "")

        # ── KPIs do topo ─────────────────────────────────────
        ctx["total_pcs"] = base.count()
        ctx["pcs_abertos"] = base.filter(
            status__in=[S.EMITIDO, S.ENVIADO_FORNECEDOR, S.ENTREGA_PARCIAL]
        ).count()
        ctx["pcs_recebidos"] = base.filter(
            status__in=[S.ENTREGUE, S.RECEBIDO]
        ).count()
        ctx["pcs_atrasados"] = base.filter(
            data_entrega_prevista__lt=hoje
        ).exclude(
            status__in=[S.ENTREGUE, S.RECEBIDO, S.CANCELADO]
        ).count()
        return ctx


# ═════════════════════════════════════════════════════════════
# CADASTROS AUXILIARES (Parceiro / Material / Contrato)
# ═════════════════════════════════════════════════════════════
# ═════════════════════════════════════════════════════════════
# PARCEIRO — CRUD
# ═════════════════════════════════════════════════════════════
class ParceiroListView(LoginRequiredMixin, AppPermissionMixin,
                       ViewFilialScopedMixin, ListView):
    model = Parceiro
    template_name = "suprimentos/parceiro_list.html"
    context_object_name = "parceiros"
    paginate_by = 25
    app_label_required = "suprimentos"

    def get_queryset(self):
        qs = super().get_queryset()  # já vem filtrado pela filial ativa
        qs = qs.select_related("endereco", "filial")
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(
                Q(nome_fantasia__icontains=q)
                | Q(razao_social__icontains=q)
                | Q(cnpj__icontains=q)
            )
        tipo = self.request.GET.get("tipo")
        if tipo == "fornecedor":
            qs = qs.filter(eh_fornecedor=True)
        elif tipo == "fabricante":
            qs = qs.filter(eh_fabricante=True)
        return qs
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["filial_ativa"] = self.get_filial_ativa()  # método do ViewFilialScopedMixin
        return ctx

class ParceiroDetailView(LoginRequiredMixin, AppPermissionMixin,
                         ViewFilialScopedMixin, DetailView):
    model = Parceiro
    template_name = "suprimentos/parceiro_detail.html"
    context_object_name = "parceiro"
    app_label_required = "suprimentos"

    def get_queryset(self):
        return super().get_queryset().select_related("endereco", "filial")


class ParceiroCreateView(LoginRequiredMixin, AppPermissionMixin,
                         RequireActiveFilialMixin, FilialCreateMixin, CreateView):
    model = Parceiro
    form_class = ParceiroForm
    template_name = "suprimentos/parceiro_form.html"
    permission_required = "suprimentos.add_parceiro"
    success_url = reverse_lazy("suprimentos:parceiro_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = "Novo Parceiro"
        return ctx
    # ✅ A mensagem de sucesso é tratada pelo FilialCreateMixin


class ParceiroUpdateView(LoginRequiredMixin, AppPermissionMixin,
                         ViewFilialScopedMixin, UpdateView):
    model = Parceiro
    form_class = ParceiroForm
    template_name = "suprimentos/parceiro_form.html"
    permission_required = "suprimentos.change_parceiro"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = f"Editar Parceiro: {self.object.nome_fantasia}"
        return ctx

    def get_success_url(self):
        return reverse_lazy("suprimentos:parceiro_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Parceiro atualizado com sucesso!")
        return super().form_valid(form)


# ═════════════════════════════════════════════════════════════
# PARCEIRO — IMPORTAÇÃO EM MASSA (.xlsx)
# ═════════════════════════════════════════════════════════════
class ParceiroUploadMassaView(LoginRequiredMixin, AppPermissionMixin,
                              RequireActiveFilialMixin, View):
    template_name = "suprimentos/parceiro_upload_massa.html"
    permission_required = "suprimentos.add_parceiro"

    COLUNAS = [
        "razao_social", "nome_fantasia", "cnpj", "inscricao_estadual",
        "contato", "telefone", "celular", "email", "site",
        "observacoes", "eh_fabricante", "eh_fornecedor", "ativo",
    ]
    BOOLEANOS = {"eh_fabricante", "eh_fornecedor", "ativo"}

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        arquivo = request.FILES.get("planilha")

        if not arquivo:
            messages.error(request, "Selecione um arquivo .xlsx para importar.")
            return redirect("suprimentos:parceiro_upload_massa")

        if not arquivo.name.lower().endswith(".xlsx"):
            messages.error(request, "Formato inválido. Envie um arquivo Excel (.xlsx).")
            return redirect("suprimentos:parceiro_upload_massa")

        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Não foi possível ler a planilha. Verifique o arquivo.")
            return redirect("suprimentos:parceiro_upload_massa")

        # ✅ Filial ativa garantida pelo RequireActiveFilialMixin
        filial = request.filial_ativa

        criados, atualizados, erros = 0, 0, []
        linhas = list(ws.iter_rows(min_row=2, values_only=True))

        try:
            with transaction.atomic():
                for idx, linha in enumerate(linhas, start=2):
                    if linha is None or all(c is None or str(c).strip() == "" for c in linha):
                        continue

                    dados = self._mapear_linha(linha)

                    if not dados.get("nome_fantasia"):
                        erros.append(f"Linha {idx}: 'nome_fantasia' é obrigatório.")
                        continue

                    cnpj = dados.get("cnpj") or None
                    defaults = {k: v for k, v in dados.items() if k != "cnpj"}
                    defaults["filial"] = filial  # ✅ sempre na filial ativa

                    try:
                        if cnpj:
                            # update_or_create escopado por filial + cnpj
                            obj, created = Parceiro.objects.update_or_create(
                                cnpj=cnpj, filial=filial, defaults=defaults
                            )
                        else:
                            Parceiro.objects.create(cnpj=None, **defaults)
                            created = True
                        criados += 1 if created else 0
                        atualizados += 0 if created else 1
                    except Exception as e:
                        erros.append(f"Linha {idx}: {e}")
        except Exception as e:
            messages.error(request, f"Importação cancelada por erro geral: {e}")
            return redirect("suprimentos:parceiro_upload_massa")

        if criados:
            messages.success(request, f"{criados} parceiro(s) criado(s).")
        if atualizados:
            messages.info(request, f"{atualizados} parceiro(s) atualizado(s).")
        if erros:
            preview = " | ".join(erros[:10])
            messages.warning(
                request,
                f"{len(erros)} linha(s) com problema: {preview}"
                + (" ..." if len(erros) > 10 else "")
            )
        if not criados and not atualizados and not erros:
            messages.info(request, "Nenhum dado importável encontrado na planilha.")

        return redirect("suprimentos:parceiro_list")

    # ── Helpers ──────────────────────────────────────────────
    def _mapear_linha(self, linha):
        dados = {}
        for i, coluna in enumerate(self.COLUNAS):
            valor = linha[i] if i < len(linha) else None
            if coluna in self.BOOLEANOS:
                dados[coluna] = self._to_bool(valor)
            else:
                dados[coluna] = (str(valor).strip() if valor is not None else "")
        return dados

    @staticmethod
    def _to_bool(valor):
        if isinstance(valor, bool):
            return valor
        if valor is None:
            return False
        return str(valor).strip().lower() in {"1", "sim", "true", "verdadeiro", "x", "s"}

class ParceiroModeloDownloadView(LoginRequiredMixin, AppPermissionMixin,
                                 RequireActiveFilialMixin, View):
    """
    Gera um modelo .xlsx formatado para importação em massa de Parceiros:
      - Cabeçalhos estilizados com comentários (dicas)
      - Dropdowns (Sim/Não) para os campos booleanos
      - Linha de exemplo
      - Aba de instruções detalhadas
    """
    permission_required = "suprimentos.add_parceiro"

    COLUNAS = [
        ("razao_social",       "Razão social completa do parceiro."),
        ("nome_fantasia",      "OBRIGATÓRIO. Nome fantasia / nome usual."),
        ("cnpj",               "CNPJ. Se já existir na filial, atualiza o registro."),
        ("inscricao_estadual", "Inscrição estadual (deixe em branco se isento)."),
        ("contato",            "Nome da pessoa de contato."),
        ("telefone",           "Telefone fixo. Ex: (11) 3000-0000"),
        ("celular",            "Celular. Ex: (11) 99999-0000"),
        ("email",              "E-mail de contato válido."),
        ("site",               "Site institucional (opcional)."),
        ("observacoes",        "Observações livres (opcional)."),
        ("eh_fabricante",      "Selecione Sim ou Não no menu suspenso."),
        ("eh_fornecedor",      "Selecione Sim ou Não no menu suspenso."),
        ("ativo",              "Selecione Sim ou Não no menu suspenso."),
    ]

    EXEMPLO = [
        "ACME Comércio de Materiais LTDA", "ACME Materiais", "12.345.678/0001-90",
        "123.456.789.000", "João da Silva", "(11) 3000-0000", "(11) 99999-0000",
        "contato@acme.com.br", "https://www.acme.com.br", "Fornecedor homologado",
        "Sim", "Sim", "Sim",
    ]

    BOOLEANOS = {"eh_fabricante", "eh_fornecedor", "ativo"}

    def get(self, request):
        filial = request.filial_ativa
        wb = Workbook()

        # ── Estilos ──────────────────────────────────────────
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="0D6EFD")
        obrig_fill = PatternFill("solid", fgColor="DC3545")  # vermelho p/ obrigatório
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ═════════════════════════════════════════════════════
        # ABA 1 — Parceiros
        # ═════════════════════════════════════════════════════
        ws = wb.active
        ws.title = "Parceiros"

        for col_idx, (nome, dica) in enumerate(self.COLUNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nome)
            cell.font = header_font
            cell.fill = obrig_fill if nome == "nome_fantasia" else header_fill
            cell.alignment = center
            cell.border = border
            cell.comment = Comment(dica, "Sistema")
            largura = max(len(nome), len(str(self.EXEMPLO[col_idx - 1]))) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(largura, 42)

        # Linha de exemplo (cinza claro/itálico)
        exemplo_font = Font(italic=True, color="888888")
        for col_idx, valor in enumerate(self.EXEMPLO, start=1):
            cell = ws.cell(row=2, column=col_idx, value=valor)
            cell.font = exemplo_font
            cell.border = border

        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A3"  # congela cabeçalho + linha de exemplo

        # ── Dropdowns Sim/Não para campos booleanos ──────────
        dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
        dv.error = "Selecione Sim ou Não."
        dv.errorTitle = "Valor inválido"
        dv.prompt = "Escolha Sim ou Não"
        dv.promptTitle = "Campo Sim/Não"
        ws.add_data_validation(dv)

        for col_idx, (nome, _) in enumerate(self.COLUNAS, start=1):
            if nome in self.BOOLEANOS:
                letra = get_column_letter(col_idx)
                dv.add(f"{letra}3:{letra}1000")  # da linha 3 em diante

        # ═════════════════════════════════════════════════════
        # ABA 2 — Instruções
        # ═════════════════════════════════════════════════════
        ws_info = wb.create_sheet("Instruções")
        ws_info.column_dimensions["A"].width = 24
        ws_info.column_dimensions["B"].width = 70

        titulo = ws_info.cell(row=1, column=1, value="COMO PREENCHER")
        titulo.font = Font(bold=True, size=14, color="0D6EFD")
        ws_info.merge_cells("A1:B1")

        ws_info.cell(row=3, column=1, value="Campo").font = header_font
        ws_info.cell(row=3, column=1).fill = header_fill
        ws_info.cell(row=3, column=2, value="Descrição").font = header_font
        ws_info.cell(row=3, column=2).fill = header_fill

        for i, (nome, dica) in enumerate(self.COLUNAS, start=4):
            ws_info.cell(row=i, column=1, value=nome).font = Font(bold=True)
            ws_info.cell(row=i, column=2, value=dica)

        notas_inicio = len(self.COLUNAS) + 6
        notas = [
            "OBSERVAÇÕES GERAIS:",
            "• A linha 2 é um EXEMPLO — apague-a antes de importar seus dados reais.",
            "• Não altere os nomes dos cabeçalhos (linha 1).",
            f"• Os parceiros serão importados na filial ativa: {filial.nome if filial else '—'}.",
            "• Se o CNPJ já existir nesta filial, o cadastro será atualizado.",
            "• Erros em uma linha não impedem o processamento das demais.",
        ]
        for j, texto in enumerate(notas, start=notas_inicio):
            cell = ws_info.cell(row=j, column=1, value=texto)
            ws_info.merge_cells(start_row=j, start_column=1, end_row=j, end_column=2)
            if j == notas_inicio:
                cell.font = Font(bold=True, color="DC3545")

        # ═════════════════════════════════════════════════════
        # ABA 3 — Referência (dados cadastrados na filial)
        # ═════════════════════════════════════════════════════
        ws_ref = wb.create_sheet("Referência")
        ws_ref.column_dimensions["A"].width = 50
        ws_ref.cell(row=1, column=1, value="Parceiros já cadastrados nesta filial").font = header_font
        ws_ref.cell(row=1, column=1).fill = header_fill

        if filial:
            existentes = (
                Parceiro.objects.filter(filial=filial)
                .order_by("nome_fantasia")
                .values_list("nome_fantasia", "cnpj")
            )
            ws_ref.cell(row=2, column=1, value="Nome Fantasia").font = Font(bold=True)
            ws_ref.cell(row=2, column=2, value="CNPJ").font = Font(bold=True)
            ws_ref.column_dimensions["B"].width = 24
            for r, (nome_f, cnpj) in enumerate(existentes, start=3):
                ws_ref.cell(row=r, column=1, value=nome_f)
                ws_ref.cell(row=r, column=2, value=cnpj or "—")

        # ── Resposta HTTP ────────────────────────────────────
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="modelo_importacao_parceiros.xlsx"'
        )
        wb.save(response)
        return response
  

# Contrato

class ContratoListView(LoginRequiredMixin, ListView):
    model = Contrato
    template_name = "suprimentos/contrato_list.html"
    context_object_name = "contratos"
    paginate_by = 25

    def get_queryset(self):
        qs = super().get_queryset().select_related("filial")
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(Q(cm__icontains=q) | Q(cliente__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = "Lista de Contratos"
        return ctx


class ContratoDetailView(LoginRequiredMixin, DetailView):
    model = Contrato
    template_name = "suprimentos/contrato_detail.html"
    context_object_name = "contrato"

    def get_queryset(self):
        return super().get_queryset().select_related("filial").prefetch_related("verbas")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = f"Contrato {self.object.cm}"
        ctx["verba_mes"] = self.object.verba_do_mes()
        return ctx


class ContratoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Contrato
    form_class = ContratoForm
    template_name = "suprimentos/contrato_form.html"
    permission_required = "suprimentos.add_contrato"
    success_message = "Contrato %(cm)s criado com sucesso!"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = "Novo Contrato"
        return ctx


class ContratoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Contrato
    form_class = ContratoForm
    template_name = "suprimentos/contrato_form.html"
    permission_required = "suprimentos.change_contrato"
    success_message = "Contrato %(cm)s atualizado com sucesso!"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = f"Editar Contrato {self.object.cm}"
        return ctx


class ContratoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Contrato
    template_name = "suprimentos/contrato_confirm_delete.html"
    permission_required = "suprimentos.delete_contrato"
    success_url = reverse_lazy("suprimentos:contrato_lista")
    success_message = "Contrato excluído com sucesso!"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = f"Excluir Contrato {self.object.cm}"
        return ctx


# ═════════════════════════════════════════════════════════════
# 7. PEDIDO DE COMPRA — Impressão (A4 / PDF)
# ═════════════════════════════════════════════════════════════
class PedidoCompraImprimirView(LoginRequiredMixin, DetailView):
    model = PedidoCompra
    template_name = "suprimentos/pc_imprimir.html"
    context_object_name = "pc"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["itens"] = self.object.itens.select_related("material", "cotacao")
        ctx["valor_total"] = self.object.valor_total  # ← campo real do model
        ctx["emitido_em"] = timezone.now()
        return ctx


# ─────────────────────────────────────────────────────────────────────────────
# 8. MATERIAL
# ─────────────────────────────────────────────────────────────────────────────

class MaterialListView(
    LoginRequiredMixin,
    AppPermissionMixin,
    RequireActiveFilialMixin,
    ViewFilialScopedMixin,
    ListView,
):
    app_label_required = "suprimentos"
    model = Material
    template_name = "suprimentos/material_list.html"
    context_object_name = "materiais"
    paginate_by = 25
    filial_field = "filial"

    def get_queryset(self):
        # ViewFilialScopedMixin já filtra por filial ativa.
        # Mas Material permite filial nula (catálogo global) → ampliamos o escopo.
        filial = get_filial_ativa(self.request.user, self.request)
        qs = Material.objects.select_related(
            "ncm", "grupo_tributario", "filial",
            "equipamento_epi", "ferramenta_ref",
        )
        if filial is not None:
            qs = qs.filter(Q(filial=filial) | Q(filial__isnull=True))
        else:
            qs = qs.none()

        # Filtros de busca
        busca = self.request.GET.get("q", "").strip()
        if busca:
            qs = qs.filter(
                Q(descricao__icontains=busca) |
                Q(codigo__icontains=busca) |
                Q(marca__icontains=busca)
            )
        classificacao = self.request.GET.get("classificacao", "").strip()
        if classificacao:
            qs = qs.filter(classificacao=classificacao)
        tipo = self.request.GET.get("tipo", "").strip()
        if tipo:
            qs = qs.filter(tipo=tipo)
        ativo = self.request.GET.get("ativo", "").strip()
        if ativo in ("0", "1"):
            qs = qs.filter(ativo=(ativo == "1"))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["classificacoes"] = CategoriaMaterial.choices
        ctx["tipos"] = TipoMaterial.choices
        ctx["filtros"] = {
            "q": self.request.GET.get("q", ""),
            "classificacao": self.request.GET.get("classificacao", ""),
            "tipo": self.request.GET.get("tipo", ""),
            "ativo": self.request.GET.get("ativo", ""),
        }
        ctx["total"] = self.get_queryset().count()
        return ctx


class MaterialDetailView(
    LoginRequiredMixin,
    AppPermissionMixin,
    RequireActiveFilialMixin,
    DetailView,
):
    app_label_required = "suprimentos"
    model = Material
    template_name = "suprimentos/material_detail.html"
    context_object_name = "material"

    def get_queryset(self):
        filial = get_filial_ativa(self.request.user, self.request)
        qs = Material.objects.select_related("ncm", "grupo_tributario", "filial")
        if filial is not None:
            return qs.filter(Q(filial=filial) | Q(filial__isnull=True))
        return qs.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["info_tributaria"] = self.object.info_tributaria_unitaria
        return ctx


class MaterialCreateView(
    LoginRequiredMixin,
    AppPermissionMixin,
    RequireActiveFilialMixin,
    FilialCreateMixin,
    CreateView,
):
    app_label_required = "suprimentos"
    model = Material
    form_class = MaterialForm
    template_name = "suprimentos/material_form.html"
    success_url = reverse_lazy("suprimentos:material_list")
    success_message = "Material cadastrado com sucesso!"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["filial"] = get_filial_ativa(self.request.user, self.request)
        return kwargs

    def form_valid(self, form):
        # FilialCreateMixin atribui form.instance.filial_id automaticamente
        response = super().form_valid(form)
        if self.success_message:
            messages.success(self.request, self.success_message)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo"] = "Novo Material"
        ctx["material_unidades_json"] = {
            str(m.pk): m.unidade_medida for m in material
        }
        return ctx


class MaterialUpdateView(
    LoginRequiredMixin,
    AppPermissionMixin,
    RequireActiveFilialMixin,
    ViewFilialScopedMixin,
    UpdateView,
):
    app_label_required = "suprimentos"
    model = Material
    form_class = MaterialForm
    template_name = "suprimentos/material_form.html"
    success_url = reverse_lazy("suprimentos:material_list")
    filial_field = "filial"

    def get_queryset(self):
        # Permite editar materiais da filial ativa OU globais
        filial = get_filial_ativa(self.request.user, self.request)
        qs = Material.objects.all()
        if filial is not None:
            return qs.filter(Q(filial=filial) | Q(filial__isnull=True))
        return qs.none()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["filial"] = get_filial_ativa(self.request.user, self.request)
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, "Material atualizado com sucesso!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo"] = "Editar Material"
        return ctx


class MaterialDeleteView(
    LoginRequiredMixin,
    AppPermissionMixin,
    RequireActiveFilialMixin,
    DeleteView,
):
    app_label_required = "suprimentos"
    model = Material
    template_name = "suprimentos/material_confirm_delete.html"
    success_url = reverse_lazy("suprimentos:material_list")

    def get_queryset(self):
        filial = get_filial_ativa(self.request.user, self.request)
        qs = Material.objects.all()
        if filial is not None:
            return qs.filter(Q(filial=filial) | Q(filial__isnull=True))
        return qs.none()

    def form_valid(self, form):
        messages.success(self.request, "Material excluído com sucesso!")
        return super().form_valid(form)


# ─────────────────────────────────────────────────────────────────────────────
# LANÇAMENTO EM MASSA — MATERIAL
# ─────────────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLUNAS = [
    ("descricao",       "Descrição*",            45, "Descrição do material (obrigatório)"),
    ("classificacao",   "Classificação*",        20, "EPI / CONSUMO / FERRAMENTA"),
    ("tipo",            "Tipo*",                 25, "Tipo do material"),
    ("marca",           "Marca",                 20, "Marca (opcional)"),
    ("unidade",         "Unidade*",              15, "PÇ, UN, KG, etc."),
    ("valor_unitario",  "Valor Unitário (R$)",   18, "Ex.: 12.50"),
    ("ncm",             "NCM",                   18, "Código NCM (opcional)"),
    ("ativo",           "Ativo",                 12, "Sim / Não"),
]


class MaterialUploadMassaView(
    LoginRequiredMixin,
    AppPermissionMixin,
    RequireActiveFilialMixin,
    View,
):
    """Tela e processamento de importação em massa de materiais."""
    app_label_required = "suprimentos"
    template_name = "suprimentos/material_upload_massa.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        arquivo = request.FILES.get("planilha")
        if not arquivo:
            messages.error(request, "Nenhum arquivo enviado.")
            return redirect("suprimentos:material_upload_massa")
        if not arquivo.name.lower().endswith(".xlsx"):
            messages.error(request, "Envie um arquivo no formato .xlsx.")
            return redirect("suprimentos:material_upload_massa")

        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb["Materiais"] if "Materiais" in wb.sheetnames else wb.active
        except Exception as exc:
            messages.error(request, f"Erro ao abrir a planilha: {exc}")
            return redirect("suprimentos:material_upload_massa")

        # Filial ativa via mixin global (RequireActiveFilialMixin já garante != None)
        filial = get_filial_ativa(request.user, request)
        criados, atualizados, erros = 0, 0, []

        map_class = self._build_choice_map(CategoriaMaterial.choices)
        map_tipo = self._build_choice_map(TipoMaterial.choices)
        map_unidade = self._build_choice_map(UnidadeMedida.choices)

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            row = list(row) + [None] * (len(COLUNAS) - len(row))

            (descricao, classificacao, tipo, marca, unidade,
             valor_unitario, ncm_codigo, ativo) = row[:8]

            try:
                descricao = (str(descricao).strip() if descricao else "")
                if not descricao:
                    raise ValueError("Descrição é obrigatória.")

                classificacao_val = map_class.get(
                    str(classificacao).strip().upper() if classificacao else ""
                )
                if not classificacao_val:
                    raise ValueError(f"Classificação inválida: '{classificacao}'.")

                tipo_val = map_tipo.get(
                    str(tipo).strip().upper() if tipo else ""
                )
                if not tipo_val:
                    raise ValueError(f"Tipo inválido: '{tipo}'.")

                unidade_val = map_unidade.get(
                    str(unidade).strip().upper() if unidade else ""
                ) or UnidadeMedida.PÇ

                valor = Decimal("0.00")
                if valor_unitario not in (None, ""):
                    try:
                        valor = Decimal(
                            str(valor_unitario).replace(".", "").replace(",", ".")
                            if "," in str(valor_unitario)
                            else str(valor_unitario)
                        )
                    except (InvalidOperation, ValueError):
                        raise ValueError(f"Valor unitário inválido: '{valor_unitario}'.")

                ncm_obj = None
                if ncm_codigo:
                    from tributacao.models import NCM
                    ncm_obj = NCM.objects.filter(
                        codigo=str(ncm_codigo).strip()
                    ).first()

                ativo_bool = True
                if ativo is not None:
                    ativo_bool = str(ativo).strip().lower() in (
                        "sim", "s", "1", "true", "verdadeiro"
                    )

                material, created = Material.objects.update_or_create(
                    descricao=descricao,
                    marca=(str(marca).strip() if marca else ""),
                    filial=filial,
                    defaults={
                        "classificacao": classificacao_val,
                        "tipo": tipo_val,
                        "unidade": unidade_val,
                        "valor_unitario": valor,
                        "ncm": ncm_obj,
                        "ativo": ativo_bool,
                    },
                )
                if created:
                    criados += 1
                else:
                    atualizados += 1

            except Exception as exc:
                erros.append(f"Linha {idx}: {exc}")

        if criados:
            messages.success(request, f"{criados} material(is) criado(s).")
        if atualizados:
            messages.info(request, f"{atualizados} material(is) atualizado(s).")
        if erros:
            preview = "; ".join(erros[:10])
            extra = f" (+{len(erros) - 10} erros)" if len(erros) > 10 else ""
            messages.warning(request, f"Erros encontrados: {preview}{extra}")
        if not criados and not atualizados and not erros:
            messages.warning(request, "Nenhuma linha válida encontrada na planilha.")

        return redirect("suprimentos:material_list")

    @staticmethod
    def _build_choice_map(choices):
        m = {}
        for value, label in choices:
            m[str(value).strip().upper()] = value
            m[str(label).strip().upper()] = value
        return m


class MaterialModeloDownloadView(
    LoginRequiredMixin,
    AppPermissionMixin,
    RequireActiveFilialMixin,
    View,
):
    """Gera e baixa a planilha modelo .xlsx para importação de materiais."""
    app_label_required = "suprimentos"

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Materiais"

        for col_idx, (_, header, width, dica) in enumerate(COLUNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            cell.comment = openpyxl.comments.Comment(dica, "Sistema")
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"

        exemplo = [
            "Luva de Segurança Nitrílica", "EPI", "Proteção das Mãos",
            "Volk", "PÇ", "12.50", "40159000", "Sim",
        ]
        for col_idx, val in enumerate(exemplo, start=1):
            c = ws.cell(row=2, column=col_idx, value=val)
            c.font = Font(italic=True, color="9CA3AF")
            c.border = BORDER

        self._add_dropdown(ws, "B", [v for v, _ in CategoriaMaterial.choices])
        self._add_dropdown(ws, "C", [v for v, _ in TipoMaterial.choices])
        self._add_dropdown(ws, "E", [v for v, _ in UnidadeMedida.choices])
        self._add_dropdown(ws, "H", ["Sim", "Não"])

        ws_inst = wb.create_sheet("Instruções")
        instrucoes = [
            "INSTRUÇÕES PARA IMPORTAÇÃO DE MATERIAIS",
            "",
            "1. Preencha uma linha por material.",
            "2. Campos com * são obrigatórios.",
            "3. Apague a linha de exemplo antes de importar.",
            "4. Classificação aceita: " +
            ", ".join(v for v, _ in CategoriaMaterial.choices),
            "5. Tipo aceita: " + ", ".join(v for v, _ in TipoMaterial.choices),
            "6. Unidade aceita: " + ", ".join(v for v, _ in UnidadeMedida.choices),
            "7. Valor unitário: use ponto ou vírgula (ex.: 12.50 ou 12,50).",
            "8. NCM: informe o código existente no cadastro de tributação.",
            "9. Materiais com mesma Descrição + Marca serão ATUALIZADOS.",
            "10. Erros em uma linha não impedem o processamento das demais.",
        ]
        for i, txt in enumerate(instrucoes, start=1):
            c = ws_inst.cell(row=i, column=1, value=txt)
            if i == 1:
                c.font = Font(bold=True, size=13, color="2563EB")
        ws_inst.column_dimensions["A"].width = 90

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        resp = HttpResponse(
            buffer.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        resp["Content-Disposition"] = (
            'attachment; filename="modelo_importacao_materiais.xlsx"'
        )
        return resp

    @staticmethod
    def _add_dropdown(ws, coluna, valores):
        formula = '"{}"'.format(",".join(valores))
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Selecione um valor válido da lista."
        dv.errorTitle = "Valor inválido"
        ws.add_data_validation(dv)
        dv.add(f"{coluna}2:{coluna}1000")


class VerbaContratoListView(LoginRequiredMixin,
    AppPermissionMixin,
    RequireActiveFilialMixin,
    ViewFilialScopedMixin,
    ListView,):
    app_label_required = "suprimentos"
    model = VerbaContrato
    template_name = "suprimentos/verba_list.html"
    context_object_name = "verbas"
    paginate_by = 25

    def get_queryset(self):
        qs = (
            VerbaContrato.objects
            .select_related("contrato", "contrato__filial")
            .order_by("-ano", "-mes", "contrato__cm")
        )
        contrato_id = self.request.GET.get("contrato")
        if contrato_id:
            qs = qs.filter(contrato_id=contrato_id)
        ano = self.request.GET.get("ano")
        if ano:
            qs = qs.filter(ano=ano)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = "Verbas Mensais"
        ctx["contratos"] = Contrato.objects.filter(ativo=True).order_by("cm")
        ctx["contrato_atual"] = self.request.GET.get("contrato", "")
        ctx["ano_atual"] = self.request.GET.get("ano", "")
        return ctx


class VerbaContratoDetailView(LoginRequiredMixin, DetailView):
    model = VerbaContrato
    template_name = "suprimentos/verba_detail.html"
    context_object_name = "verba"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        v = self.object
        ctx["titulo_pagina"] = f"Verba {v.mes:02d}/{v.ano} — {v.contrato.cm}"

        def pct(compra, verba):
            """Percentual consumido da verba (0-100+)."""
            if verba and verba > 0:
                return round((compra / verba) * 100, 1)
            return 0

        ctx["linhas"] = [
            {
                "label": "EPI",
                "verba": v.verba_epi,
                "compra": v.compra_epi,
                "saldo": v.saldo_epi,
                "pct": pct(v.compra_epi, v.verba_epi),
            },
            {
                "label": "Consumo",
                "verba": v.verba_consumo,
                "compra": v.compra_consumo,
                "saldo": v.saldo_consumo,
                "pct": pct(v.compra_consumo, v.verba_consumo),
            },
            {
                "label": "Ferramenta",
                "verba": v.verba_ferramenta,
                "compra": v.compra_ferramenta,
                "saldo": v.saldo_ferramenta,
                "pct": pct(v.compra_ferramenta, v.verba_ferramenta),
            },
        ]
        ctx["pct_total"] = pct(v.compra_total, v.verba_total)

        # ── Verba do mês anterior ──────────────────────────────
        if v.mes == 1:
            ano_ant, mes_ant = v.ano - 1, 12
        else:
            ano_ant, mes_ant = v.ano, v.mes - 1

        anterior = (
            VerbaContrato.objects
            .filter(contrato=v.contrato, ano=ano_ant, mes=mes_ant)
            .first()
        )
        ctx["verba_anterior"] = anterior
        ctx["mes_anterior"] = f"{mes_ant:02d}/{ano_ant}"
        if anterior:
            ctx["linhas_anterior"] = [
                {"label": "EPI", "verba": anterior.verba_epi, "compra": anterior.compra_epi, "saldo": anterior.saldo_epi},
                {"label": "Consumo", "verba": anterior.verba_consumo, "compra": anterior.compra_consumo, "saldo": anterior.saldo_consumo},
                {"label": "Ferramenta", "verba": anterior.verba_ferramenta, "compra": anterior.compra_ferramenta, "saldo": anterior.saldo_ferramenta},
            ]
        return ctx


class VerbaContratoCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = VerbaContrato
    form_class = VerbaContratoForm
    template_name = "suprimentos/verba_form.html"
    permission_required = "suprimentos.add_verbacontrato"
    success_url = reverse_lazy("suprimentos:verba_list")

    def get_initial(self):
        initial = super().get_initial()
        contrato_id = self.request.GET.get("contrato")
        if contrato_id:
            initial["contrato"] = contrato_id
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = "Nova Verba Mensal"
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Verba mensal cadastrada com sucesso!")
        return super().form_valid(form)


class VerbaContratoUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = VerbaContrato
    form_class = VerbaContratoForm
    template_name = "suprimentos/verba_form.html"
    permission_required = "suprimentos.change_verbacontrato"
    success_url = reverse_lazy("suprimentos:verba_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = f"Editar Verba {self.object.mes:02d}/{self.object.ano}"
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Verba mensal atualizada com sucesso!")
        return super().form_valid(form)


class VerbaContratoDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = VerbaContrato
    template_name = "suprimentos/verba_confirm_delete.html"
    permission_required = "suprimentos.delete_verbacontrato"
    success_url = reverse_lazy("suprimentos:verba_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["titulo_pagina"] = "Excluir Verba"
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Verba excluída com sucesso!")
        return super().form_valid(form)
