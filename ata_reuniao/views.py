# ata_reuniao/views.py

import io
import json
from typing import Any
from django.shortcuts import get_object_or_404, redirect, render
import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Count, F, Q
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.template.loader import get_template
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import (
    CreateView, DeleteView, ListView, TemplateView, UpdateView, View, DetailView,
)
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
import pandas as pd
from io import BytesIO

from cliente.models import Cliente
from departamento_pessoal.models import Funcionario
from core.mixins import FuncionarioRequiredMixin, ViewFilialScopedMixin, AppPermissionMixin
from core.decorators import app_permission_required

from .forms import AtaReuniaoForm, HistoricoAtaForm, ComentarioForm, UploadAtaReuniaoForm
from .models import AtaReuniao, HistoricoAta, Filial, Comentario

from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib.messages import get_messages
from django.core.exceptions import ObjectDoesNotExist
from ata_reuniao import models


User = get_user_model()


# ═══════════════════════════════════════════════════════════════════════════════
# MIXIN CENTRAL DE FILIAL
# ═══════════════════════════════════════════════════════════════════════════════

class FilialAtivaMixin:
    """
    Mixin central para obter a filial ativa do usuário.
    Usa a chave 'active_filial_id' da sessão (definida pelo seletor do header).

    NOTA: Este mixin é independente do ViewFilialScopedMixin do core.
    - ViewFilialScopedMixin → age no get_queryset() via for_request()
    - FilialAtivaMixin → fornece métodos utilitários para filial ativa
    Ambos coexistem sem conflito.
    """

    def get_filial_ativa(self):
        filial_id = self.request.session.get('active_filial_id')

        if filial_id:
            try:
                return Filial.objects.get(pk=filial_id)
            except Filial.DoesNotExist:
                pass

        user = self.request.user
        filial_ativa = getattr(user, 'filial_ativa', None)
        if filial_ativa:
            return filial_ativa

        try:
            funcionario = Funcionario.objects.select_related('filial').get(usuario=user)
            return funcionario.filial
        except Funcionario.DoesNotExist:
            pass

        return None

    def get_filial_ativa_id(self):
        filial = self.get_filial_ativa()
        return filial.id if filial else None

    def filter_queryset_by_filial(self, queryset):
        filial = self.get_filial_ativa()
        if filial:
            return queryset.filter(filial=filial)
        elif not self.request.user.is_superuser:
            return queryset.none()
        return queryset

    def filter_related_by_filial(self, queryset, filial_field='filial'):
        filial = self.get_filial_ativa()
        if filial:
            return queryset.filter(**{filial_field: filial})
        return queryset


# ═══════════════════════════════════════════════════════════════════════════════
# MIXIN DE VISIBILIDADE — CONTROLE CENTRAL
# ═══════════════════════════════════════════════════════════════════════════════

class AtaVisibilityMixin(FilialAtivaMixin):
    """
    Controla a visibilidade das atas de reunião conforme o perfil do usuário.

    Hierarquia de camadas (não conflita com core/mixins.py):
    ┌──────────────────────────────────────────────────────────────────────┐
    │ Camada 1: ViewFilialScopedMixin (core)                              │
    │   → Filtra por filial no get_queryset() via for_request()           │
    │                                                                      │
    │ Camada 2: AtaVisibilityMixin (este mixin)                           │
    │   → Filtra por perfil do usuário via apply_visibility()              │
    │   → Chamado EXPLICITAMENTE após o queryset já filtrado por filial    │
    │                                                                      │
    │ Camada 3: Filtros GET (status, contrato, coordenador)               │
    │   → Aplicados no AtaQuerysetMixin                                    │
    └──────────────────────────────────────────────────────────────────────┘

    Regras de visibilidade:
    ┌─────────────────────────────────┬────────────────────────────────────────┐
    │ Perfil                          │ Visibilidade                           │
    ├─────────────────────────────────┼────────────────────────────────────────┤
    │ Superuser                       │ Todas as atas (sem filtro)             │
    │ Permissão view_all_ata_reuniao  │ Todas as atas da filial ativa         │
    │ Usuário comum                   │ Apenas atas onde é coordenador ou      │
    │                                 │ responsável (via Funcionario)          │
    └─────────────────────────────────┴────────────────────────────────────────┘
    """

    def apply_visibility(self, queryset):
        """Aplica filtro de visibilidade ao queryset já filtrado por filial."""
        user = self.request.user

        # Superuser → tudo
        if user.is_superuser:
            return queryset

        # Permissão global da filial → tudo da filial
        if user.has_perm('ata_reuniao.view_all_ata_reuniao'):
            return queryset

        # Usuário comum → apenas atas relacionadas ao seu Funcionario
        funcionario = getattr(user, 'funcionario', None)

        if funcionario:
            return queryset.filter(
                Q(coordenador=funcionario)
                | Q(responsavel=funcionario)
            ).distinct()

        # Sem vínculo de funcionário → nenhuma ata
        return queryset.none()


# ═══════════════════════════════════════════════════════════════════════════════
# MIXINS BASE
# ═══════════════════════════════════════════════════════════════════════════════

class AtaReuniaoBaseMixin(FuncionarioRequiredMixin,
    LoginRequiredMixin, AppPermissionMixin,
    AtaVisibilityMixin, ViewFilialScopedMixin,
):
    """
    Mixin base para as views de Ata de Reunião.

    Ordem de herança (MRO):
    1. LoginRequiredMixin     → garante autenticação
    2. AppPermissionMixin     → verifica permissão do app (core)
    3. AtaVisibilityMixin     → fornece apply_visibility() + FilialAtivaMixin
    4. ViewFilialScopedMixin  → filtra get_queryset() por filial via for_request()

    NOTA: ViewFilialScopedMixin.get_queryset() aplica filtro de filial.
    AtaVisibilityMixin.apply_visibility() é chamado separadamente onde necessário.
    Não há conflito porque operam em momentos diferentes.
    """
    model = AtaReuniao
    form_class = AtaReuniaoForm
    success_url = reverse_lazy('ata_reuniao:ata_reuniao_list')
    app_label_required = 'ata_reuniao'
    modulo_nome = 'Ata de Reunião'


class AtaQuerysetMixin(AtaVisibilityMixin):
    """
    Mixin para filtrar AtaReuniao por filial, visibilidade e parâmetros GET.

    Pipeline de filtragem:
    1. for_request() ou filter_queryset_by_filial() → escopo da filial
    2. apply_visibility()                            → escopo do perfil
    3. filtros GET (contrato, status, coordenador)   → refinamento
    """

    def get_ata_queryset(self, request, model_class):
        # Camada 1: Filial scope
        if hasattr(model_class.objects, 'for_request'):
            queryset = model_class.objects.for_request(request)
        else:
            queryset = model_class.objects.all()
            queryset = self.filter_queryset_by_filial(queryset)

        # Camada 2: Visibilidade por perfil
        queryset = self.apply_visibility(queryset)

        # Camada 3: Filtros GET
        contrato_id = request.GET.get('contrato')
        status = request.GET.get('status')
        coordenador_id = request.GET.get('coordenador')

        if contrato_id:
            queryset = queryset.filter(contrato_id=contrato_id)
        if status:
            queryset = queryset.filter(status=status)
        if coordenador_id:
            queryset = queryset.filter(coordenador__usuario__id=coordenador_id)

        return queryset.order_by('prazo', '-criado_em')


# ═══════════════════════════════════════════════════════════════════════════════
# MIXIN PARA FILTROS COMUNS (coordenadores, clientes, filial no contexto)
# ═══════════════════════════════════════════════════════════════════════════════

class AtaFilterContextMixin(FilialAtivaMixin):
    """
    Adiciona coordenadores, clientes e filial_ativa ao contexto.
    Usado pelo Dashboard, Kanban e ListView para evitar repetição.
    """

    def get_filter_context(self):
        filial_ativa = self.get_filial_ativa()

        coordenadores_qs = Funcionario.objects.filter(status='ATIVO')
        coordenadores_qs = self.filter_related_by_filial(coordenadores_qs)

        clientes_qs = Cliente.objects.filter(estatus=True)
        clientes_qs = self.filter_related_by_filial(clientes_qs)

        return {
            'filial_ativa': filial_ativa,
            'is_superuser': self.request.user.is_superuser,
            'can_view_all': (
                self.request.user.is_superuser
                or self.request.user.has_perm('ata_reuniao.view_all_ata_reuniao')
            ),
            'coordenadores': coordenadores_qs.select_related('usuario').order_by('nome_completo'),
            'clientes': clientes_qs.order_by('nome')[:100],
        }


# ═══════════════════════════════════════════════════════════════════════════════
# VIEWS DE CRUD
# ═══════════════════════════════════════════════════════════════════════════════

class AtaReuniaoListView(AtaReuniaoBaseMixin, AtaQuerysetMixin, AtaFilterContextMixin, ListView):
    template_name = 'ata_reuniao/ata_reuniao_list.html'
    context_object_name = 'atas'
    paginate_by = 25

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            try:
                _ = request.user.funcionario
            except ObjectDoesNotExist:
                return render(request, 'ata_reuniao/acesso_negado.html', {
                    'titulo': 'Acesso Restrito',
                    'mensagem': (
                        'Sua conta de usuário não está vinculada a um registro de funcionário, '
                        'por isso você não pode acessar o módulo de Atas de Reunião.'
                    )
                }, status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return self.get_ata_queryset(self.request, model_class=self.model)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Filtros e filial via mixin
        context.update(self.get_filter_context())

        # Renomeia 'clientes' para 'contratos' para manter compatibilidade com template
        context['contratos'] = context.pop('clientes')

        context['user_data'] = {
            'username': self.request.user.username,
            'filial': str(context['filial_ativa']) if context['filial_ativa'] else 'Todas as filiais',
        }
        context['current_contrato'] = self.request.GET.get('contrato', '')
        context['current_status'] = self.request.GET.get('status', '')
        context['current_coordenador'] = self.request.GET.get('coordenador', '')

        return context


class AtaReuniaoCreateView(AtaReuniaoBaseMixin, SuccessMessageMixin, CreateView):
    template_name = 'ata_reuniao/ata_form.html'
    success_message = "✅ Ata de reunião criada com sucesso!"

    def form_valid(self, form):
        filial_ativa = self.get_filial_ativa()
        if filial_ativa:
            form.instance.filial = filial_ativa
        elif hasattr(self.request.user, 'funcionario'):
            form.instance.filial = self.request.user.funcionario.filial
        return super().form_valid(form)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filial_ativa'] = self.get_filial_ativa()
        return context


class AtaReuniaoUpdateView(AtaReuniaoBaseMixin, SuccessMessageMixin, UpdateView):
    template_name = 'ata_reuniao/ata_form.html'
    success_message = "🔄 Ata de reunião atualizada com sucesso!"

    def get_queryset(self):
        """Garante que o usuário só edite atas que pode ver."""
        qs = super().get_queryset()
        return self.apply_visibility(qs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        comentario_texto = form.cleaned_data.get('comentario')
        if comentario_texto:
            HistoricoAta.objects.create(
                ata=self.object,
                usuario=self.request.user,
                comentario=comentario_texto,
                filial=self.object.filial,
            )
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filial_ativa'] = self.get_filial_ativa()
        return context


class AtaReuniaoDetailView(AtaReuniaoBaseMixin, DetailView):
    template_name = 'ata_reuniao/ata_reuniao_detail.html'
    context_object_name = 'ata'

    def get_queryset(self):
        """Garante que o usuário só veja detalhes de atas que pode ver."""
        qs = super().get_queryset().prefetch_related('historico__usuario')
        return self.apply_visibility(qs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filial_ativa'] = self.get_filial_ativa()
        return context


class AtaReuniaoAddCommentView(AtaReuniaoBaseMixin, View):
    """Adiciona comentário a uma ata existente."""

    def post(self, request, pk, *args, **kwargs):
        # Busca apenas atas que o usuário pode ver
        base_qs = AtaReuniao.objects.all()
        base_qs = self.filter_queryset_by_filial(base_qs)
        base_qs = self.apply_visibility(base_qs)

        ata = get_object_or_404(base_qs, pk=pk)

        form = ComentarioForm(request.POST)
        if form.is_valid():
            HistoricoAta.objects.create(
                ata=ata,
                usuario=request.user,
                comentario=form.cleaned_data['comentario'],
                filial=ata.filial,
            )
            messages.success(request, 'Comentário adicionado com sucesso!')
        else:
            messages.error(request, 'Erro ao adicionar o comentário.')

        return redirect('ata_reuniao:ata_reuniao_detail', pk=pk)


class AtaReuniaoDeleteView(AtaReuniaoBaseMixin, SuccessMessageMixin, DeleteView):

    """
    View para exclusão de Ata de Reunião.
    
    NOTA: NÃO usar SuccessMessageMixin com DeleteView no Django 5.x.
    O mixin falha silenciosamente porque DeleteView.form_valid()
    não expõe form.cleaned_data da mesma forma que CreateView/UpdateView.
    A mensagem de sucesso é adicionada manualmente no post().
    """
    template_name = 'ata_reuniao/ata_confirm_delete.html'
    success_url = reverse_lazy('ata_reuniao:ata_reuniao_list')

    def get_queryset(self):
        """Garante que o usuário só exclua atas que pode ver."""
        qs = super().get_queryset()
        return self.apply_visibility(qs)

    def post(self, request, *args, **kwargs):
        """
        Sobrescreve o POST para controlar o fluxo completo:
        1. Busca o objeto
        2. Tenta deletar
        3. Exibe mensagem de sucesso ou erro
        """
        self.object = self.get_object()
        
        try:
            self.object.delete()
            messages.success(
                request,
                f'🗑️ Ata "{self.object}" excluída com sucesso!'
            )
            return redirect(self.get_success_url())
        
        except models.ProtectedError:
            messages.error(
                request,
                "❌ Não foi possível excluir esta ata. "
                "Existem registros vinculados que impedem a exclusão."
            )
            return redirect(
                'ata_reuniao:ata_reuniao_detail', pk=self.object.pk
            )
        except Exception as e:
            messages.error(
                request,
                f"❌ Erro inesperado ao excluir: {e}"
            )
            return redirect(
                'ata_reuniao:ata_reuniao_detail', pk=self.object.pk
            )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filial_ativa'] = self.get_filial_ativa()
        return context


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD VIEW
# ═══════════════════════════════════════════════════════════════════════════════

class AtaReuniaoDashboardView(
    AtaReuniaoBaseMixin,
    AtaQuerysetMixin, AtaFilterContextMixin, TemplateView,
):
    template_name = 'ata_reuniao/ata_reuniao_dashboard.html'
    app_label_required = 'ata_reuniao'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Filtros compartilhados via mixin
        context.update(self.get_filter_context())

        # Queryset já com visibilidade aplicada
        base_queryset = self.get_ata_queryset(self.request, model_class=AtaReuniao)

        # ── KPIs ──
        total_concluido = base_queryset.filter(status=AtaReuniao.Status.CONCLUIDO).count()
        total_cancelado = base_queryset.filter(status=AtaReuniao.Status.CANCELADO).count()

        concluido_no_prazo = 0
        concluido_com_atraso = 0

        atas_concluidas = base_queryset.filter(
            status=AtaReuniao.Status.CONCLUIDO
        ).only('prazo', 'atualizado_em')

        for ata in atas_concluidas:
            if ata.prazo and ata.atualizado_em:
                if ata.prazo >= ata.atualizado_em.date():
                    concluido_no_prazo += 1
                else:
                    concluido_com_atraso += 1
            else:
                concluido_no_prazo += 1

        total_concluido_efetivo = concluido_no_prazo + concluido_com_atraso
        percentual_no_prazo = (
            (concluido_no_prazo / total_concluido_efetivo * 100)
            if total_concluido_efetivo > 0 else 0
        )
        percentual_com_atraso = (
            (concluido_com_atraso / total_concluido_efetivo * 100)
            if total_concluido_efetivo > 0 else 0
        )

        context.update({
            'total_concluido': total_concluido,
            'concluido_no_prazo': concluido_no_prazo,
            'concluido_com_atraso': concluido_com_atraso,
            'total_cancelado': total_cancelado,
            'percentual_no_prazo': percentual_no_prazo,
            'percentual_com_atraso': percentual_com_atraso,
        })

        # ── Gráfico: Pendências por Responsável ──
        pendencias = (
            base_queryset
            .filter(status__in=[AtaReuniao.Status.PENDENTE, AtaReuniao.Status.ANDAMENTO])
            .values('responsavel__nome_completo')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        context['pendencias_labels'] = [
            item['responsavel__nome_completo'] or 'Sem responsável'
            for item in pendencias
        ]
        context['pendencias_data'] = [item['count'] for item in pendencias]

        # ── Gráfico: Qualidade das Conclusões ──
        context['qualidade_labels'] = ['No Prazo', 'Com Atraso']
        context['qualidade_data'] = [concluido_no_prazo, concluido_com_atraso]

        # ── Kanban (embutido no Dashboard) ──
        context['kanban_status_choices'] = AtaReuniao.Status.choices

        kanban_items = {}
        for status_value, status_label in AtaReuniao.Status.choices:
            kanban_items[status_label] = list(
                base_queryset.filter(status=status_value)
                .select_related(
                    'responsavel', 'responsavel__usuario',
                    'contrato', 'coordenador', 'coordenador__usuario',
                )
                .order_by('prazo')
            )
        context['kanban_items'] = kanban_items
        context['titulo_pagina'] = "Dashboard de Atas"

        return context


# ═══════════════════════════════════════════════════════════════════════════════
# KANBAN VIEW (Página dedicada)
# ═══════════════════════════════════════════════════════════════════════════════

class AtaReuniaoKanbanView(
    AtaReuniaoBaseMixin,
    AtaQuerysetMixin, AtaFilterContextMixin, TemplateView,
):
    template_name = 'ata_reuniao/ata_reuniao_kanban.html'
    app_label_required = 'ata_reuniao'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Filtros compartilhados
        context.update(self.get_filter_context())

        # Queryset já com visibilidade aplicada
        base_queryset = self.get_ata_queryset(self.request, model_class=AtaReuniao)

        kanban_items = {}
        for status_value, status_label in AtaReuniao.Status.choices:
            kanban_items[status_label] = list(
                base_queryset.filter(status=status_value)
                .select_related(
                    'responsavel', 'responsavel__usuario',
                    'contrato', 'coordenador',
                )
                .order_by('prazo')
            )

        # Responsáveis (extra, usado no Kanban dedicado)
        responsaveis_qs = Funcionario.objects.filter(status='ATIVO')
        responsaveis_qs = self.filter_related_by_filial(responsaveis_qs)
        context['responsaveis'] = responsaveis_qs.select_related('usuario').order_by('nome_completo')[:50]

        context.update({
            'titulo_pagina': "Kanban de Atas",
            'titulo_secao': "Quadro Kanban",
            'kanban_status_choices': AtaReuniao.Status.choices,
            'kanban_items': kanban_items,
            'show_filters': True,
            'kanban_id': 'kanban-page',
        })

        return context


# ═══════════════════════════════════════════════════════════════════════════════
# API - ATUALIZAÇÃO DE STATUS (Unificada)
# ═══════════════════════════════════════════════════════════════════════════════

@method_decorator(csrf_exempt, name='dispatch')
class AtaUpdateStatusAPIView(AtaReuniaoBaseMixin, View):
    """
    Endpoint unificado para atualizar status de ata via drag & drop (Kanban).
    Aceita POST com JSON: {"new_status": "concluido"}
    """
    app_label_required = 'ata_reuniao'

    def post(self, request, pk):
        try:
            data = json.loads(request.body)
            new_status = data.get('new_status')

            if not new_status:
                return JsonResponse({
                    'status': 'error', 'message': 'Status não informado'
                }, status=400)

            # Validação do status
            valid_statuses = [choice[0] for choice in AtaReuniao.Status.choices]
            if new_status not in valid_statuses:
                return JsonResponse({
                    'status': 'error', 'message': f'Status inválido: {new_status}'
                }, status=400)

            # Busca com filtro de filial + visibilidade
            base_qs = AtaReuniao.objects.all()
            base_qs = self.filter_queryset_by_filial(base_qs)
            base_qs = self.apply_visibility(base_qs)

            try:
                ata = base_qs.get(pk=pk)
            except AtaReuniao.DoesNotExist:
                return JsonResponse({
                    'status': 'error', 'message': 'Ata não encontrada ou acesso negado'
                }, status=404)

            old_status = ata.status
            ata.status = new_status
            ata.save(update_fields=['status', 'atualizado_em'])

            return JsonResponse({
                'status': 'success',
                'message': f'Status alterado de "{old_status}" para "{new_status}"',
                'data': {
                    'id': ata.pk,
                    'old_status': old_status,
                    'new_status': new_status,
                    'updated_at': ata.atualizado_em.isoformat() if ata.atualizado_em else None,
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error', 'message': 'JSON inválido'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'status': 'error', 'message': str(e)
            }, status=500)


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

class AtaReuniaoPDFExportView(AtaReuniaoBaseMixin, AtaQuerysetMixin, View):
    """Exporta PDF respeitando visibilidade."""
    app_label_required = 'ata_reuniao'
    

    def get(self, request, *args, **kwargs):
        # Queryset já com visibilidade aplicadata_reuniao'
        atas = self.get_ata_queryset(request, model_class=AtaReuniao)
        filial_ativa = self.get_filial_ativa()

        periodo_relatorio = "Período: Todas as datas"
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')

        if data_inicio and data_fim:
            atas = atas.filter(criado_em__range=[data_inicio, data_fim])
            data_inicio_fmt = timezone.datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
            data_fim_fmt = timezone.datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')
            periodo_relatorio = f"Período: {data_inicio_fmt} a {data_fim_fmt}"

        template = get_template('ata_reuniao/ata_pdf.html')
        context = {
            'atas': atas,
            'data_exportacao': timezone.now(),
            'usuario_exportacao': request.user.get_full_name(),
            'periodo_relatorio': periodo_relatorio,
            'kanban_status_choices': AtaReuniao.Status.choices,
            'filial_ativa': filial_ativa,
        }

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_atas.pdf"'

        pisa_status = pisa.CreatePDF(template.render(context), dest=response)
        if pisa_status.err:
            messages.error(request, "Ocorreu um erro ao gerar o PDF.")
            return HttpResponse('Erro ao gerar PDF', status=500)

        return response


class AtaReuniaoExcelExportView(AtaReuniaoBaseMixin, AtaQuerysetMixin, View):
    """Exporta Excel respeitando visibilidade."""
    app_label_required = 'ata_reuniao'

    def get(self, request, *args, **kwargs):
        # Queryset já com visibilidade aplicada
        atas = self.get_ata_queryset(request, model_class=AtaReuniao)

        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Relatório de Atas"

        headers = [
            'ID', 'Título', 'Contrato', 'Coordenador', 'Responsável', 'Natureza',
            'Ação', 'Entrada', 'Prazo', 'Status', 'Filial', 'Última Atualização', 'Comentário',
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for ata in atas:
            sheet.append([
                ata.id,
                ata.titulo,
                ata.contrato.nome if ata.contrato else '',
                ata.coordenador.nome_completo if ata.coordenador else '',
                ata.responsavel.nome_completo if ata.responsavel else '',
                ata.get_natureza_display(),
                ata.acao,
                ata.entrada.strftime('%d/%m/%Y') if ata.entrada else '',
                ata.prazo.strftime('%d/%m/%Y') if ata.prazo else '',
                ata.get_status_display(),
                ata.filial.nome if ata.filial else '',
                ata.atualizado_em.strftime('%d/%m/%Y %H:%M') if ata.atualizado_em else '',
                ata.historico.last().comentario if ata.historico.exists() else '',
            ])

        for column_cells in sheet.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        workbook.save(buffer)
        buffer.seek(0)

        filename = f'relatorio_atas_{timezone.now().strftime("%Y-%m-%d")}.xlsx'
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ═══════════════════════════════════════════════════════════════════════════════
# UPLOAD / DOWNLOAD
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@app_permission_required('ata_reuniao')
def download_ata_reuniao_template(request):
    """Gera planilha Excel modelo para importação em massa."""
    output = BytesIO()
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Dados das Atas"

    headers = [
        "Contrato (ID)", "Coordenador (ID)", "Responsável (ID)", "Natureza",
        "Título", "Ação/Proposta", "Data de Entrada (DD/MM/AAAA)",
        "Prazo Final (DD/MM/AAAA)", "Status",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        worksheet.column_dimensions[get_column_letter(col_num)].width = 25

    # Aba de instruções
    inst_ws = workbook.create_sheet(title="Instruções e Opções")
    inst_ws.column_dimensions['A'].width = 30
    inst_ws.column_dimensions['B'].width = 50

    instructions_data = [
        ("INSTRUÇÕES GERAIS", ""),
        ("Contrato (ID)", "Insira o número ID do cliente/contrato. Ex: 15"),
        ("Coordenador (ID)", "Insira o número ID do funcionário coordenador. Ex: 7"),
        ("Responsável (ID)", "Insira o número ID do funcionário responsável. Ex: 12"),
        ("Datas", "Use o formato DD/MM/AAAA. Ex: 25/12/2025"),
        ("", ""),
        ("OPÇÕES VÁLIDAS", "Não altere os valores desta coluna."),
        ("Natureza", ", ".join([c[0] for c in AtaReuniao.Natureza.choices])),
        ("Status", ", ".join([c[0] for c in AtaReuniao.Status.choices])),
    ]

    for row, data in enumerate(instructions_data, 1):
        inst_ws.cell(row=row, column=1, value=data[0]).font = Font(bold=True)
        inst_ws.cell(row=row, column=2, value=data[1])

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="modelo_importacao_atas.xlsx"'
    return response


class UploadAtaReuniaoView(AtaReuniaoBaseMixin, View):
    template_name = 'ata_reuniao/ata_reuniao_upload.html'
    form_class = UploadAtaReuniaoForm
    success_url = reverse_lazy('ata_reuniao:ata_reuniao_list')
    app_label_required = 'ata_reuniao'

    def get(self, request, *args, **kwargs):
        form = self.form_class()

        storage = get_messages(request)
        has_upload_errors = any('warning' in msg.tags for msg in storage)

        if not has_upload_errors and 'upload_error_details' in request.session:
            del request.session['upload_error_details']

        return render(request, self.template_name, {
            'form': form,
            'has_upload_errors': has_upload_errors,
            'filial_ativa': self.get_filial_ativa(),
        })

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if not form.is_valid():
            if 'upload_error_details' in request.session:
                del request.session['upload_error_details']
            return render(request, self.template_name, {'form': form})

        file = request.FILES['file']
        filial_ativa = self.get_filial_ativa()

        if not filial_ativa and not request.user.is_superuser:
            messages.error(request, "Nenhuma filial selecionada. Selecione uma filial no menu superior.")
            return redirect(request.path)

        try:
            df = pd.read_excel(file, dtype=str).fillna('')
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect(request.path)

        if df.empty:
            messages.warning(request, "A planilha está vazia ou não contém dados para importar.")
            return redirect(request.path)

        headers_map = {
            "Contrato (ID)": "contrato_id",
            "Coordenador (ID)": "coordenador_id",
            "Responsável (ID)": "responsavel_id",
            "Natureza": "natureza",
            "Título": "titulo",
            "Ação/Proposta": "acao",
            "Data de Entrada (DD/MM/AAAA)": "entrada",
            "Prazo Final (DD/MM/AAAA)": "prazo",
            "Status": "status",
        }
        df.rename(columns=headers_map, inplace=True)

        success_count = 0
        errors = []
        error_details_for_session = []

        valid_natureza = AtaReuniao.Natureza.values
        valid_status = AtaReuniao.Status.values

        for index, row in df.iterrows():
            linha = index + 2
            original_row_data = row.to_dict()

            try:
                # Contrato
                try:
                    contrato_qs = Cliente.objects.filter(pk=int(row['contrato_id']))
                    if filial_ativa:
                        contrato_qs = contrato_qs.filter(filial=filial_ativa)
                    contrato = contrato_qs.get()
                except (Cliente.DoesNotExist, ValueError):
                    raise ValueError(f"Contrato com ID '{row['contrato_id']}' não encontrado na filial atual.")

                # Coordenador
                try:
                    coord_qs = Funcionario.objects.filter(pk=int(row['coordenador_id']))
                    if filial_ativa:
                        coord_qs = coord_qs.filter(filial=filial_ativa)
                    coordenador = coord_qs.get()
                except (Funcionario.DoesNotExist, ValueError):
                    raise ValueError(f"Coordenador com ID '{row['coordenador_id']}' não encontrado na filial atual.")

                # Responsável
                try:
                    resp_qs = Funcionario.objects.filter(pk=int(row['responsavel_id']))
                    if filial_ativa:
                        resp_qs = resp_qs.filter(filial=filial_ativa)
                    responsavel = resp_qs.get()
                except (Funcionario.DoesNotExist, ValueError):
                    raise ValueError(f"Responsável com ID '{row['responsavel_id']}' não encontrado na filial atual.")

                # Validações
                if row['natureza'] not in valid_natureza:
                    raise ValueError(f"Natureza '{row['natureza']}' inválida.")
                if row['status'] not in valid_status:
                    raise ValueError(f"Status '{row['status']}' inválido.")

                # Datas
                entrada = None
                if row['entrada']:
                    try:
                        entrada = pd.to_datetime(row['entrada'], dayfirst=True).date()
                    except Exception:
                        raise ValueError(f"Data de entrada '{row['entrada']}' inválida.")

                prazo = None
                if row['prazo']:
                    try:
                        prazo = pd.to_datetime(row['prazo'], dayfirst=True).date()
                    except Exception:
                        raise ValueError(f"Prazo '{row['prazo']}' inválido.")

                AtaReuniao.objects.create(
                    contrato=contrato,
                    coordenador=coordenador,
                    responsavel=responsavel,
                    natureza=row['natureza'],
                    titulo=row['titulo'],
                    acao=row['acao'],
                    entrada=entrada,
                    prazo=prazo,
                    status=row['status'],
                    filial=filial_ativa,
                )
                success_count += 1

            except Exception as e:
                errors.append(f"Linha {linha}: {e}")
                original_row_data['motivo_do_erro'] = str(e)
                error_details_for_session.append(original_row_data)

        if error_details_for_session:
            request.session['upload_error_details'] = error_details_for_session

        if success_count > 0:
            messages.success(request, f"{success_count} ata(s) de reunião importada(s) com sucesso!")

        if errors:
            messages.warning(
                request,
                f"<strong>{len(errors)} linha(s) não puderam ser importadas.</strong><br>"
                f"Baixe o relatório de erros para corrigir e tente novamente.",
                extra_tags='safe',
            )

        return redirect(request.path)


@login_required
@app_permission_required('ata_reuniao')
def download_error_report(request):
    """Gera arquivo Excel com linhas que falharam durante upload."""
    error_details = request.session.get('upload_error_details')

    if not error_details:
        messages.error(request, "Nenhum relatório de erros encontrado na sessão.")
        return redirect('ata_reuniao:ata_reuniao_upload')

    df_errors = pd.DataFrame(error_details)

    original_columns_order = [
        "contrato_id", "coordenador_id", "responsavel_id", "natureza", "titulo",
        "acao", "entrada", "prazo", "status",
    ]
    df_errors = df_errors[original_columns_order + ['motivo_do_erro']]

    friendly_headers = {
        "contrato_id": "Contrato (ID)",
        "coordenador_id": "Coordenador (ID)",
        "responsavel_id": "Responsável (ID)",
        "natureza": "Natureza",
        "titulo": "Título",
        "acao": "Ação/Proposta",
        "entrada": "Data de Entrada (DD/MM/AAAA)",
        "prazo": "Prazo Final (DD/MM/AAAA)",
        "status": "Status",
        "motivo_do_erro": "Motivo do Erro",
    }
    df_errors.rename(columns=friendly_headers, inplace=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_errors.to_excel(writer, index=False, sheet_name='Erros_Para_Correcao')

        worksheet = writer.sheets['Erros_Para_Correcao']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

        for col_num in range(1, len(df_errors.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            worksheet.column_dimensions[get_column_letter(col_num)].width = 25

        # Coluna de erro mais larga
        worksheet.column_dimensions[get_column_letter(len(df_errors.columns))].width = 50

    output.seek(0)
    del request.session['upload_error_details']

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio_erros_importacao.xlsx"'
    return response
