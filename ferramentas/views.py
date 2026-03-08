# ferramentas/views.py

import logging
from io import BytesIO
import base64
import json
import subprocess
import sys
import tempfile
import zipfile
from datetime import timedelta, datetime
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import (
    CreateView, DetailView, FormView,
    ListView, TemplateView, UpdateView
)

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.mixins import SSTPermissionMixin, ViewFilialScopedMixin, AtividadeLogMixin
from usuario.models import Filial

from .forms import (
    DevolucaoForm, FerramentaForm, MovimentacaoForm,
    UploadFileForm, MalaFerramentasForm, TermoResponsabilidadeForm
)
from .models import (
    Atividade, Ferramenta, MalaFerramentas,
    Movimentacao, TermoDeResponsabilidade, ItemTermo
)

logger = logging.getLogger(__name__)


# =============================================================================
# MIXINS ESPECÍFICOS DO APP
# =============================================================================

class FilialAtribuicaoMixin:
    """
    Mixin para atribuir automaticamente a filial ao criar objetos.
    Prioriza: sessão (superuser) > filial_ativa do usuário.
    """
    def _atribuir_filial(self, instance):
        user = self.request.user
        session_filial = self.request.session.get('active_filial_id')

        if user.is_superuser and session_filial:
            instance.filial_id = session_filial
        elif user.filial_ativa:
            instance.filial = user.filial_ativa
        else:
            raise ValueError("Usuário sem filial ativa. Impossível criar registro.")


class ItemRetrievalMixin:
    """
    Mixin para buscar ferramenta ou mala de forma segura (filtrada por filial).
    """
    def _get_item_seguro(self, model_class, pk):
        qs = model_class.objects.for_request(self.request)
        return get_object_or_404(qs, pk=pk)


# =============================================================================
# DASHBOARD
# =============================================================================

class DashboardView(LoginRequiredMixin, ViewFilialScopedMixin, TemplateView):
    """Dashboard com estatísticas otimizadas (mínimo de queries)."""
    template_name = 'ferramentas/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request

        # QuerySets base filtrados por filial
        ferramentas_qs = Ferramenta.objects.for_request(request).ativas()
        malas_qs = MalaFerramentas.objects.for_request(request)
        movimentacoes_qs = Movimentacao.objects.for_request(request)
        atividades_qs = Atividade.objects.for_request(request)

        # Stats em uma única query cada
        f_stats = self._ferramenta_stats(ferramentas_qs)
        m_stats = self._mala_stats(malas_qs)

        # Movimentações ativas (em uso)
        movs_ativas = movimentacoes_qs.filter(
            data_devolucao__isnull=True
        ).select_related('ferramenta', 'mala', 'retirado_por')

        # Movimentações atrasadas
        movs_atrasadas = movs_ativas.filter(
            data_devolucao_prevista__lt=timezone.now()
        )

        context.update({
            'titulo_pagina': "Dashboard de Operações",
            'status_ferramentas': f_stats,
            'status_malas': m_stats,
            'status_total': self._totais(f_stats, m_stats, movimentacoes_qs, atividades_qs),

            # Gráficos
            'ferramentas_chart_data': {
                'labels': ['Disponível', 'Em Uso', 'Em Manutenção'],
                'data': [
                    f_stats.get('disponivel_count', 0),
                    f_stats.get('em_uso_count', 0),
                    f_stats.get('em_manutencao_count', 0),
                ],
                'title': 'Distribuição de Ferramentas'
            },
            'malas_chart_data': {
                'labels': ['Disponível', 'Em Uso'],
                'data': [
                    m_stats.get('disponivel_count', 0),
                    m_stats.get('em_uso_count', 0),
                ],
                'title': 'Distribuição de Malas/Kits'
            },

            # Listas
            'ferramentas_em_uso': movs_ativas.filter(ferramenta__isnull=False),
            'malas_em_uso': movs_ativas.filter(mala__isnull=False),
            'movimentacoes_atrasadas': movs_atrasadas,
            'total_atrasadas': movs_atrasadas.count(),
            'ultimas_atividades': atividades_qs.order_by('-timestamp')[:10],
        })
        return context

    def _ferramenta_stats(self, qs):
        em_uso_q = Q(status=Ferramenta.Status.EM_USO) | Q(mala__status=MalaFerramentas.Status.EM_USO)
        em_manutencao_q = Q(status=Ferramenta.Status.EM_MANUTENCAO)
        disponivel_q = ~em_uso_q & ~em_manutencao_q

        return qs.aggregate(
            total_count=Count('id'),
            em_uso_count=Count('id', filter=em_uso_q),
            em_manutencao_count=Count('id', filter=em_manutencao_q),
            disponivel_count=Count('id', filter=disponivel_q),
            total_sum=Sum('quantidade'),
            em_uso_sum=Sum('quantidade', filter=em_uso_q),
            disponivel_sum=Sum('quantidade', filter=disponivel_q),
        )

    def _mala_stats(self, qs):
        return qs.aggregate(
            total_count=Count('id'),
            disponivel_count=Count('id', filter=Q(status=MalaFerramentas.Status.DISPONIVEL)),
            em_uso_count=Count('id', filter=Q(status=MalaFerramentas.Status.EM_USO)),
            total_sum=Sum('quantidade'),
            disponivel_sum=Sum('quantidade', filter=Q(status=MalaFerramentas.Status.DISPONIVEL)),
        )

    def _totais(self, f, m, movs_qs, ativ_qs):
        return {
            'total': (f.get('total_count') or 0) + (m.get('total_count') or 0),
            'disponivel': (f.get('disponivel_count') or 0) + (m.get('disponivel_count') or 0),
            'em_uso': (f.get('em_uso_count') or 0) + (m.get('em_uso_count') or 0),
            'em_manutencao': f.get('em_manutencao_count') or 0,
            'total_itens_ferramentas': f.get('total_sum') or 0,
            'total_itens_malas': m.get('total_sum') or 0,
            'total_itens_geral': (f.get('total_sum') or 0) + (m.get('total_sum') or 0),
            'quantidade_ferramentas_disponiveis': f.get('disponivel_count') or 0,
            'quantidade_malas_disponiveis': m.get('disponivel_count') or 0,
            'total_malas': m.get('total_count') or 0,
            'total_ferramentas': f.get('total_count') or 0,
            'total_movimentacoes': movs_qs.count(),
            'total_atividades': ativ_qs.count(),
        }


# =============================================================================
# FERRAMENTAS — CRUD
# =============================================================================

class FerramentaListView(LoginRequiredMixin, ViewFilialScopedMixin, ListView):
    model = Ferramenta
    template_name = 'ferramentas/ferramenta_list.html'
    context_object_name = 'ferramentas'
    paginate_by = 30

    def get_queryset(self):
        qs = super().get_queryset().select_related('mala', 'filial')
        search = self.request.GET.get('q', '').strip()
        status_filter = self.request.GET.get('status', '')

        if search:
            qs = qs.filter(
                Q(nome__icontains=search) |
                Q(codigo_identificacao__icontains=search) |
                Q(patrimonio__icontains=search)
            )

        if status_filter:
            if status_filter == Ferramenta.Status.EM_USO:
                qs = qs.filter(Q(status=status_filter) | Q(mala__status=MalaFerramentas.Status.EM_USO))
            elif status_filter == Ferramenta.Status.DISPONIVEL:
                qs = qs.filter(status=status_filter).filter(
                    Q(mala__isnull=True) | Q(mala__status=MalaFerramentas.Status.DISPONIVEL)
                )
            else:
                qs = qs.filter(status=status_filter)
        else:
            qs = qs.ativas()

        return qs.distinct().order_by('nome')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = Ferramenta.Status.choices
        context['status_atual'] = self.request.GET.get('status', '')
        context['query'] = self.request.GET.get('q', '')
        return context


class FerramentaDetailView(LoginRequiredMixin, ViewFilialScopedMixin, DetailView):
    model = Ferramenta
    template_name = 'ferramentas/ferramenta_detail.html'
    context_object_name = 'ferramenta'

    def get_queryset(self):
        return super().get_queryset().select_related('mala', 'filial', 'fornecedor').prefetch_related(
            'movimentacoes__retirado_por',
            'movimentacoes__recebido_por',
            'atividades__usuario'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ferramenta = self.object

        context['status_efetivo'] = ferramenta.status_efetivo
        context['status_efetivo_display'] = ferramenta.get_status_efetivo_display
        context['movimentacoes'] = ferramenta.movimentacoes.all()
        context['atividades'] = ferramenta.atividades.all()[:20]
        context['movimentacao_ativa'] = next(
            (m for m in context['movimentacoes'] if m.esta_ativa), None
        )

        # Gráfico de uso (últimos 6 meses)
        six_months_ago = timezone.now() - timedelta(days=180)
        usage_data = (
            Movimentacao.objects.filter(
                ferramenta=ferramenta, data_retirada__gte=six_months_ago
            )
            .annotate(month=TruncMonth('data_retirada'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )
        context['chart_labels'] = json.dumps([d['month'].strftime('%b/%Y') for d in usage_data])
        context['chart_data'] = json.dumps([d['count'] for d in usage_data])
        context['titulo_pagina'] = f"Painel de Controle: {ferramenta.nome}"
        return context


class FerramentaCreateView(LoginRequiredMixin, ViewFilialScopedMixin, FilialAtribuicaoMixin, AtividadeLogMixin, CreateView):
    model = Ferramenta
    form_class = FerramentaForm
    template_name = 'ferramentas/ferramenta_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request  # ← NOVO
        return kwargs

    def form_valid(self, form):
        self._atribuir_filial(form.instance)
        response = super().form_valid(form)
        self._log_atividade(
            ferramenta=self.object,
            tipo=Atividade.TipoAtividade.CRIACAO,
            descricao=f"Ferramenta '{self.object.nome}' registrada."
        )
        messages.success(self.request, f"Ferramenta '{self.object.nome}' adicionada com sucesso.")
        return response

    def get_success_url(self):
        return self.object.get_absolute_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Adicionar Nova Ferramenta"
        context['is_create'] = True
        return context


class FerramentaUpdateView(LoginRequiredMixin, ViewFilialScopedMixin, AtividadeLogMixin, UpdateView):
    model = Ferramenta
    form_class = FerramentaForm
    template_name = 'ferramentas/ferramenta_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request  # ← NOVO
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        self._log_atividade(
            ferramenta=self.object,
            tipo=Atividade.TipoAtividade.ALTERACAO,
            descricao="Dados da ferramenta foram atualizados."
        )
        messages.success(self.request, f"'{self.object.nome}' atualizada com sucesso.")
        return response

    def get_success_url(self):
        return self.object.get_absolute_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Editar: {self.object.nome}"
        context['is_create'] = False
        return context

class FerramentaUpdateView(LoginRequiredMixin, ViewFilialScopedMixin, AtividadeLogMixin, UpdateView):
    model = Ferramenta
    form_class = FerramentaForm
    template_name = 'ferramentas/ferramenta_form.html'

    def form_valid(self, form):
        response = super().form_valid(form)
        self._log_atividade(
            ferramenta=self.object,
            tipo=Atividade.TipoAtividade.ALTERACAO,
            descricao="Dados da ferramenta foram atualizados."
        )
        messages.success(self.request, f"'{self.object.nome}' atualizada com sucesso.")
        return response

    def get_success_url(self):
        return self.object.get_absolute_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Editar: {self.object.nome}"
        context['is_create'] = False
        return context


# =============================================================================
# MALAS DE FERRAMENTAS — CRUD
# =============================================================================

class MalaListView(LoginRequiredMixin, ViewFilialScopedMixin, ListView):
    model = MalaFerramentas
    template_name = 'ferramentas/mala_list.html'
    context_object_name = 'malas'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset()
        query = self.request.GET.get('q', '').strip()
        if query:
            qs = qs.filter(
                Q(nome__icontains=query) |
                Q(codigo_identificacao__icontains=query)
            )
        return qs.annotate(item_count=Count('itens')).prefetch_related('itens').order_by('nome')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        return context


class MalaDetailView(LoginRequiredMixin, ViewFilialScopedMixin, DetailView):
    model = MalaFerramentas
    template_name = 'ferramentas/mala_detail.html'
    context_object_name = 'mala'

    def get_queryset(self):
        return super().get_queryset().prefetch_related(
            'movimentacoes__retirado_por',
            'atividades__usuario',
            'itens'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mala = self.object
        context['movimentacoes'] = mala.movimentacoes.all()
        context['atividades'] = mala.atividades.all()[:20]
        context['movimentacao_ativa'] = next(
            (m for m in context['movimentacoes'] if m.esta_ativa), None
        )
        context['titulo_pagina'] = f"Painel de Controle: {mala.nome}"
        return context


class MalaCreateView(LoginRequiredMixin, ViewFilialScopedMixin, FilialAtribuicaoMixin, AtividadeLogMixin, CreateView):
    model = MalaFerramentas
    form_class = MalaFerramentasForm
    template_name = 'ferramentas/mala_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request  # ← NOVO
        return kwargs

    def form_valid(self, form):
        self._atribuir_filial(form.instance)
        response = super().form_valid(form)
        self._log_atividade(
            mala=self.object,
            tipo=Atividade.TipoAtividade.CRIACAO,
            descricao=f"Mala '{self.object.nome}' registrada."
        )
        messages.success(self.request, f"Mala '{self.object.nome}' criada com sucesso.")
        return response

    def get_success_url(self):
        return self.object.get_absolute_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = "Nova Mala de Ferramentas"
        return context

class MalaUpdateView(LoginRequiredMixin, ViewFilialScopedMixin, AtividadeLogMixin, UpdateView):
    model = MalaFerramentas
    form_class = MalaFerramentasForm
    template_name = 'ferramentas/mala_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request  # ← NOVO
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        self._log_atividade(
            mala=self.object,
            tipo=Atividade.TipoAtividade.ALTERACAO,
            descricao=f"Dados da mala '{self.object.nome}' atualizados."
        )
        messages.success(self.request, f"Mala '{self.object.nome}' atualizada com sucesso.")
        return response

    def get_success_url(self):
        return self.object.get_absolute_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f"Editar: {self.object.nome}"
        return context


# =============================================================================
# AÇÕES DE FERRAMENTA (Manutenção, Inativação)
# =============================================================================

class AcaoFerramentaBaseView(LoginRequiredMixin, ItemRetrievalMixin, AtividadeLogMixin, View):
    """Base para ações POST sobre ferramentas."""

    def get_ferramenta(self):
        return self._get_item_seguro(Ferramenta, self.kwargs['pk'])


class IniciarManutencaoView(AcaoFerramentaBaseView):
    def post(self, request, *args, **kwargs):
        ferramenta = self.get_ferramenta()
        if ferramenta.status != Ferramenta.Status.DISPONIVEL:
            messages.error(request, "A ferramenta precisa estar 'Disponível' para entrar em manutenção.")
            return redirect(ferramenta.get_absolute_url())

        ferramenta.status = Ferramenta.Status.EM_MANUTENCAO
        ferramenta.save(update_fields=['status'])
        self._log_atividade(ferramenta, Atividade.TipoAtividade.MANUTENCAO_INICIO, "Manutenção iniciada.")
        messages.success(request, f"'{ferramenta.nome}' colocada em manutenção.")
        return redirect(ferramenta.get_absolute_url())


class FinalizarManutencaoView(AcaoFerramentaBaseView):
    def post(self, request, *args, **kwargs):
        ferramenta = self.get_ferramenta()
        if ferramenta.status != Ferramenta.Status.EM_MANUTENCAO:
            messages.error(request, "A ferramenta precisa estar 'Em Manutenção'.")
            return redirect(ferramenta.get_absolute_url())

        ferramenta.status = Ferramenta.Status.DISPONIVEL
        ferramenta.save(update_fields=['status'])
        self._log_atividade(ferramenta, Atividade.TipoAtividade.MANUTENCAO_FIM, "Manutenção finalizada.")
        messages.success(request, f"Manutenção de '{ferramenta.nome}' finalizada.")
        return redirect(ferramenta.get_absolute_url())


class InativarFerramentaView(AcaoFerramentaBaseView):
    def post(self, request, *args, **kwargs):
        ferramenta = self.get_ferramenta()
        if ferramenta.status != Ferramenta.Status.DISPONIVEL:
            messages.error(request, "A ferramenta precisa estar 'Disponível' para ser inativada.")
            return redirect(ferramenta.get_absolute_url())

        ferramenta.status = Ferramenta.Status.DESCARTADA
        ferramenta.data_descarte = timezone.now().date()
        ferramenta.save(update_fields=['status', 'data_descarte'])
        self._log_atividade(
            ferramenta=ferramenta,
            tipo=Atividade.TipoAtividade.DESCARTE,
            descricao="Ferramenta marcada como descartada/inativa."
        )
        messages.success(request, f"'{ferramenta.nome}' inativada com sucesso.")
        return redirect('ferramentas:ferramenta_list')


# =============================================================================
# MOVIMENTAÇÃO (Retirada / Devolução)
# =============================================================================

class MovimentacaoCreateView(LoginRequiredMixin, ItemRetrievalMixin, AtividadeLogMixin, CreateView):
    """Retirada de ferramenta ou mala."""
    model = Movimentacao
    form_class = MovimentacaoForm
    template_name = 'ferramentas/retirada_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.ferramenta = None
        self.mala = None
        if 'ferramenta_pk' in self.kwargs:
            self.ferramenta = self._get_item_seguro(Ferramenta, self.kwargs['ferramenta_pk'])
        elif 'mala_pk' in self.kwargs:
            self.mala = self._get_item_seguro(MalaFerramentas, self.kwargs['mala_pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['ferramenta'] = self.ferramenta
        kwargs['mala'] = self.mala
        kwargs['request'] = self.request  # ← NOVO
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = self.ferramenta or self.mala
        context['item'] = item
        context['titulo_pagina'] = f"Checklist de Retirada: {item.nome}"
        return context

    @transaction.atomic
    def form_valid(self, form):
        item = self.ferramenta or self.mala
        if item.status != 'disponivel':
            messages.error(self.request, f"'{item.nome}' não está disponível para retirada.")
            return redirect(item.get_absolute_url())

        movimentacao = form.save(commit=False)
        movimentacao.filial = self.request.user.filial_ativa

        # Processa assinatura
        assinatura_data = form.cleaned_data.get('assinatura_base64')
        if assinatura_data and ';base64,' in assinatura_data:
            fmt, imgstr = assinatura_data.split(';base64,')
            ext = fmt.split('/')[-1]
            fname = f'sig_ret_{item.pk}_{timezone.now().timestamp()}.{ext}'
            movimentacao.assinatura_retirada = ContentFile(
                base64.b64decode(imgstr), name=fname
            )

        movimentacao.save()

        item.status = 'em_uso'
        item.save(update_fields=['status'])

        self._log_atividade(
            tipo=Atividade.TipoAtividade.RETIRADA,
            descricao=f"Retirada por {movimentacao.retirado_por.get_username()}.",
            ferramenta=self.ferramenta,
            mala=self.mala
        )
        messages.success(self.request, f"'{item.nome}' retirada com sucesso.")
        return redirect(item.get_absolute_url())


class DevolucaoUpdateView(LoginRequiredMixin, ViewFilialScopedMixin, AtividadeLogMixin, UpdateView):
    """Devolução de ferramenta individual."""
    model = Movimentacao
    form_class = DevolucaoForm
    template_name = 'ferramentas/devolucao_form.html'
    context_object_name = 'movimentacao'

    def get_queryset(self):
        return super().get_queryset().filter(
            ferramenta__isnull=False,
            data_devolucao__isnull=True
        )

    @transaction.atomic
    def form_valid(self, form):
        movimentacao = form.save(commit=False)
        movimentacao.data_devolucao = timezone.now()
        movimentacao.recebido_por = self.request.user
        movimentacao.save()

        ferramenta = movimentacao.ferramenta
        ferramenta.status = Ferramenta.Status.DISPONIVEL
        ferramenta.save(update_fields=['status'])

        self._log_atividade(
            ferramenta=ferramenta,
            tipo=Atividade.TipoAtividade.DEVOLUCAO,
            descricao=f"Devolvida. Responsável: {movimentacao.retirado_por.get_username()}."
        )
        messages.success(self.request, f"'{ferramenta.nome}' devolvida com sucesso.")
        return redirect(ferramenta.get_absolute_url())

    def get_success_url(self):
        return self.object.ferramenta.get_absolute_url()


class MalaDevolucaoUpdateView(LoginRequiredMixin, ViewFilialScopedMixin, AtividadeLogMixin, UpdateView):
    """Devolução de mala de ferramentas."""
    model = Movimentacao
    form_class = DevolucaoForm
    template_name = 'ferramentas/mala_devolucao_form.html'
    context_object_name = 'movimentacao'

    def get_queryset(self):
        return super().get_queryset().filter(
            mala__isnull=False,
            data_devolucao__isnull=True
        )

    @transaction.atomic
    def form_valid(self, form):
        movimentacao = form.save(commit=False)
        movimentacao.data_devolucao = timezone.now()
        movimentacao.recebido_por = self.request.user
        movimentacao.save()

        mala = movimentacao.mala
        mala.status = MalaFerramentas.Status.DISPONIVEL
        mala.save(update_fields=['status'])

        # Itens da mala que não estão em manutenção voltam a ficar disponíveis
        mala.itens.exclude(
            status=Ferramenta.Status.EM_MANUTENCAO
        ).update(status=Ferramenta.Status.DISPONIVEL)

        self._log_atividade(
            mala=mala,
            tipo=Atividade.TipoAtividade.DEVOLUCAO,
            descricao=f"Mala devolvida. Responsável: {movimentacao.retirado_por.get_username()}."
        )
        messages.success(self.request, f"Mala '{mala.nome}' devolvida com sucesso.")
        return redirect(mala.get_absolute_url())

    def get_success_url(self):
        return self.object.mala.get_absolute_url()


# =============================================================================
# IMPORTAÇÃO E UTILITÁRIOS
# =============================================================================

class DownloadTemplateView(LoginRequiredMixin, View):
    """Gera planilha modelo para importação."""

    HEADERS = [
        "Nome da Ferramenta*", "Código de Identificação*",
        "Data de Aquisição (dd/mm/aaaa)*", "Localização Padrão*",
        "Nº de Patrimônio", "Fabricante", "Modelo", "Série",
        "Tamanho da Polegada", "Numero Laudo Técnico",
        "Mala", "Filial", "Quantidade", "Observações",
    ]

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modelo de Importação"

        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='004C99', end_color='004C99', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, title in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_num)].width = 30

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="modelo_importacao_ferramentas.xlsx"'
        wb.save(response)
        return response


class ImportarFerramentasView(LoginRequiredMixin, FormView):
    """Importa ferramentas via planilha Excel."""
    template_name = 'ferramentas/importar_ferramentas.html'
    form_class = UploadFileForm
    success_url = reverse_lazy('ferramentas:ferramenta_list')

    def form_valid(self, form):
        file = form.cleaned_data['file']
        active_filial_id = self.request.session.get('active_filial_id')

        if not active_filial_id:
            messages.error(self.request, "Selecione uma filial ativa antes de importar.")
            return self.form_invalid(form)

        try:
            result = self._processar_planilha(file, active_filial_id)
        except Exception as e:
            messages.error(self.request, f"Erro inesperado ao processar o arquivo: {e}")
            return self.form_invalid(form)

        if result['erros']:
            for erro in result['erros']:
                messages.error(self.request, erro)
            return self.form_invalid(form)

        with transaction.atomic():
            Ferramenta.objects.bulk_create(result['ferramentas'])

        messages.success(self.request, f"{len(result['ferramentas'])} ferramentas importadas com sucesso!")
        return super().form_valid(form)

    def _processar_planilha(self, file, active_filial_id):
        """Processa planilha e retorna dict com ferramentas e erros."""
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        ferramentas, erros = [], []
        codigos_processados = set()

        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=14, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            (
                nome, codigo, data_str, localizacao, patrimonio,
                fabricante, modelo, serie, tamanho, laudo,
                mala_nome, filial_nome, quantidade, observacoes
            ) = row

            # Validação de campos obrigatórios
            if not all([nome, codigo, data_str, localizacao]):
                erros.append(f"Linha {i}: Dados obrigatórios faltando.")
                continue

            codigo = str(codigo).strip()
            if codigo in codigos_processados:
                erros.append(f"Linha {i}: Código '{codigo}' duplicado na planilha.")
                continue

            # Parse da data
            try:
                if isinstance(data_str, datetime):
                    data_aquisicao = data_str.date()
                else:
                    data_aquisicao = datetime.strptime(str(data_str).split(" ")[0], '%d/%m/%Y').date()
            except (ValueError, TypeError):
                erros.append(f"Linha {i}: Data inválida '{data_str}'. Use dd/mm/aaaa.")
                continue

            # Verifica duplicidade no banco
            if Ferramenta.objects.filter(codigo_identificacao=codigo).exists():
                erros.append(f"Linha {i}: Código '{codigo}' já existe no sistema.")
                continue

            # Resolve filial
            filial_obj = self._resolver_filial(filial_nome, active_filial_id, i, erros)
            if filial_obj is None and filial_nome:
                continue

            # Resolve mala
            mala_obj = self._resolver_mala(mala_nome, filial_obj, i, erros)
            if mala_obj is None and mala_nome:
                continue

            ferramentas.append(Ferramenta(
                nome=nome,
                codigo_identificacao=codigo.upper(),
                data_aquisicao=data_aquisicao,
                localizacao_padrao=localizacao,
                patrimonio=patrimonio or None,
                fabricante_marca=fabricante or None,
                modelo=modelo or None,
                serie=serie or None,
                tamanho_polegadas=tamanho or None,
                numero_laudo_tecnico=laudo or None,
                quantidade=quantidade or 0,
                mala=mala_obj,
                filial=filial_obj or Filial.objects.get(pk=active_filial_id),
                observacoes=observacoes or None,
            ))
            codigos_processados.add(codigo)

        return {'ferramentas': ferramentas, 'erros': erros}

    def _resolver_filial(self, nome, default_id, linha, erros):
        if nome:
            try:
                return Filial.objects.get(nome__iexact=str(nome).strip())
            except Filial.DoesNotExist:
                erros.append(f"Linha {linha}: Filial '{nome}' não encontrada.")
                return None
        return Filial.objects.get(pk=default_id)

    def _resolver_mala(self, nome, filial, linha, erros):
        if nome and filial:
            try:
                return MalaFerramentas.objects.get(nome__iexact=str(nome).strip(), filial=filial)
            except MalaFerramentas.DoesNotExist:
                erros.append(f"Linha {linha}: Mala '{nome}' não encontrada na filial '{filial.nome}'.")
                return None
        return None


class ImprimirQRCodesView(LoginRequiredMixin, ViewFilialScopedMixin, ListView):
    model = Ferramenta
    template_name = 'ferramentas/imprimir_qrcodes.html'
    context_object_name = 'ferramentas'

    def get_queryset(self):
        return super().get_queryset().ativas().com_qr_code().order_by('nome')


class ResultadoScanView(LoginRequiredMixin, ViewFilialScopedMixin, DetailView):
    model = Ferramenta
    template_name = 'ferramentas/resultado_scan.html'
    context_object_name = 'ferramenta'
    slug_field = 'codigo_identificacao'
    slug_url_kwarg = 'codigo_identificacao'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['movimentacao_ativa'] = self.object.movimentacoes.filter(
            data_devolucao__isnull=True
        ).select_related('retirado_por').first()
        return context


class GerarQRCodesView(LoginRequiredMixin, SSTPermissionMixin, View):
    """Aciona geração de QR Codes em background."""
    permission_required = 'ferramentas.change_ferramenta'

    def post(self, request, *args, **kwargs):
        subprocess.Popen([
            sys.executable,
            str(settings.BASE_DIR / "manage.py"),
            "generate_qrcodes",
        ])
        messages.success(request, "Geração de QR Codes iniciada em segundo plano.")
        return redirect('ferramentas:ferramenta_list')


# =============================================================================
# TERMOS DE RESPONSABILIDADE
# =============================================================================

class TermoListView(LoginRequiredMixin, ViewFilialScopedMixin, ListView):
    model = TermoDeResponsabilidade
    template_name = 'ferramentas/termo_list.html'
    context_object_name = 'termos'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('responsavel', 'filial', 'movimentado_por')

        # Filtros
        status = self.request.GET.get('status', '')
        search = self.request.GET.get('q', '').strip()

        if status:
            qs = qs.filter(status=status)
        if search:
            qs = qs.filter(
                Q(contrato__icontains=search) |
                Q(responsavel__nome_completo__icontains=search) |
                Q(pk__icontains=search)
            )

        return qs.order_by('-data_emissao')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = TermoDeResponsabilidade.StatusTermo.choices
        context['status_atual'] = self.request.GET.get('status', '')
        context['query'] = self.request.GET.get('q', '')
        return context


class TermoDetailView(LoginRequiredMixin, ViewFilialScopedMixin, DetailView):
    model = TermoDeResponsabilidade
    template_name = 'ferramentas/termo_detail.html'
    context_object_name = 'termo'

    def get_queryset(self):
        return super().get_queryset().prefetch_related(
            'itens__ferramenta', 'itens__mala',
            'movimentacoes_geradas__retirado_por'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        termo = self.object
        context['pode_reverter'] = termo.pode_reverter
        context['titulo_pagina'] = f"Termo de Responsabilidade #{termo.pk}"
        return context


class CriarTermoResponsabilidadeView(LoginRequiredMixin, ViewFilialScopedMixin, FormView):
    template_name = 'ferramentas/termo_responsabilidade_form.html'
    form_class = TermoResponsabilidadeForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request 
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ferramentas'] = Ferramenta.objects.for_request(self.request).disponiveis()
        context['malas'] = MalaFerramentas.objects.for_request(self.request).filter(
            status=MalaFerramentas.Status.DISPONIVEL
        )
        return context

    @transaction.atomic
    def form_valid(self, form):
        # Validação da assinatura
        assinatura_base64 = self.request.POST.get('assinatura_base64')
        if not assinatura_base64:
            form.add_error(None, "É obrigatório que o responsável assine o termo.")
            return self.form_invalid(form)

        # Validação dos itens
        try:
            itens_json = json.loads(self.request.POST.get('itens_termo_json', '[]'))
            if not itens_json:
                messages.error(self.request, "Nenhum item selecionado para o termo.")
                return self.form_invalid(form)
        except json.JSONDecodeError:
            messages.error(self.request, "Erro ao processar itens. Tente novamente.")
            return self.form_invalid(form)

        # Cria o termo
        termo = form.save(commit=False)
        termo.movimentado_por = self.request.user
        termo.assinatura_data = assinatura_base64
        termo.data_recebimento = timezone.now()
        termo.filial = self.request.user.filial_ativa
        termo.save()

        data_devolucao = timezone.now() + timedelta(days=7)

        # Processa cada item
        for item_data in itens_json:
            item_pk = item_data.get('pk')
            if not item_pk:
                continue

            ferramenta_obj, mala_obj, item_obj = None, None, None

            if termo.tipo_uso == TermoDeResponsabilidade.TipoUso.FERRAMENTAL:
                ferramenta_obj = get_object_or_404(Ferramenta, pk=item_pk)
                item_obj = ferramenta_obj
            elif termo.tipo_uso == TermoDeResponsabilidade.TipoUso.MALA:
                mala_obj = get_object_or_404(MalaFerramentas, pk=item_pk)
                item_obj = mala_obj

            if not item_obj or item_obj.status != 'disponivel':
                messages.error(self.request, f"'{item_obj}' não está mais disponível.")
                raise ValueError("Item indisponível")

            # Cria ItemTermo
            ItemTermo.objects.create(
                termo=termo,
                ferramenta=ferramenta_obj,
                mala=mala_obj,
                quantidade=item_data['quantidade'],
                unidade=item_data['unidade'],
                item=item_data['item']
            )

            # Cria Movimentação
            Movimentacao.objects.create(
                termo_responsabilidade=termo,
                ferramenta=ferramenta_obj,
                mala=mala_obj,
                retirado_por=termo.movimentado_por,
                data_devolucao_prevista=data_devolucao,
                condicoes_retirada=f"Retirada via Termo #{termo.pk}",
                filial=termo.filial,
            )

            # Atualiza status
            item_obj.status = 'em_uso'
            item_obj.save(update_fields=['status'])

            # Log de atividade
            Atividade.objects.create(
                ferramenta=ferramenta_obj,
                mala=mala_obj,
                tipo_atividade=Atividade.TipoAtividade.RETIRADA,
                descricao=f"Retirada por {termo.responsavel} via Termo #{termo.pk}.",
                usuario=self.request.user,
                filial=termo.filial,
            )

        messages.success(self.request, f"Termo #{termo.pk} criado com sucesso.")
        self.termo_criado = termo
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('ferramentas:termo_detail', kwargs={'pk': self.termo_criado.pk})


class ReverterTermoView(LoginRequiredMixin, ViewFilialScopedMixin, View):
    """Reverte/estorna um Termo de Responsabilidade."""

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        termo = get_object_or_404(
            TermoDeResponsabilidade.objects.for_request(request),
            pk=self.kwargs['pk']
        )

        if not termo.pode_reverter:
            messages.warning(request, "Este termo não pode ser revertido.")
            return redirect('ferramentas:termo_detail', pk=termo.pk)

        movs_ativas = termo.movimentacoes_geradas.filter(data_devolucao__isnull=True)

        for mov in movs_ativas:
            item = mov.ferramenta or mov.mala
            if item and item.status != Ferramenta.Status.EM_MANUTENCAO:
                item.status = 'disponivel'
                item.save(update_fields=['status'])

            mov.data_devolucao = timezone.now()
            mov.recebido_por = request.user
            mov.condicoes_devolucao = f"Estorno automático do Termo #{termo.pk}."
            mov.save()

        # Atualiza status do termo
        termo.status = TermoDeResponsabilidade.StatusTermo.ESTORNADO
        termo.save(update_fields=['status'])

        Atividade.objects.create(
            descricao=f"Termo #{termo.pk} estornado por {request.user.get_username()}.",
            tipo_atividade=Atividade.TipoAtividade.DEVOLUCAO,
            usuario=request.user,
            filial=termo.filial,
        )

        messages.success(request, f"Termo #{termo.pk} estornado. Itens devolvidos.")
        return redirect('ferramentas:termo_detail', pk=termo.pk)


class DownloadTermoPDFView(LoginRequiredMixin, ViewFilialScopedMixin, View):
    """
    Gera o PDF do Termo de Responsabilidade.
    Usa WeasyPrint como biblioteca Python (sem subprocess).
    Fallback para xhtml2pdf se WeasyPrint não estiver disponível.
    """

    def get(self, request, *args, **kwargs):
        termo = get_object_or_404(
            TermoDeResponsabilidade.objects.for_request(request),
            pk=self.kwargs['pk']
        )

        html_string = render_to_string(
            'ferramentas/termo_pdf_template.html',
            {'termo': termo},
            request=request,
        )

        pdf_bytes = self._gerar_pdf(html_string, request)

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="termo_{termo.pk}.pdf"'
        return response

    def _gerar_pdf(self, html_string, request):
        """Tenta WeasyPrint (lib), depois xhtml2pdf como fallback."""

        # ── Tentativa 1: WeasyPrint como biblioteca ──
        try:
            from weasyprint import HTML
            return HTML(
                string=html_string,
                base_url=request.build_absolute_uri('/')
            ).write_pdf()
        except (ImportError, OSError) as e:
            logger.warning("WeasyPrint indisponível (%s). Tentando xhtml2pdf...", e)

        # ── Tentativa 2: xhtml2pdf (sem dependência de GTK) ──
        try:
            from xhtml2pdf import pisa

            buffer = BytesIO()
            pisa_status = pisa.CreatePDF(html_string, dest=buffer)

            if pisa_status.err:
                raise RuntimeError(
                    f"xhtml2pdf retornou {pisa_status.err} erro(s) na geração."
                )

            return buffer.getvalue()
        except ImportError:
            raise RuntimeError(
                "Nenhum gerador de PDF disponível. "
                "Instale com: pip install xhtml2pdf"
            )
        
class DownloadTermosLoteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        termos_ids = request.POST.getlist('termo_ids')
        if not termos_ids:
            messages.warning(request, "Nenhum termo selecionado.")
            return redirect('ferramentas:termoderesponsabilidade_list')

        qs = TermoDeResponsabilidade.objects.for_request(request).filter(pk__in=termos_ids)

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for termo in qs:
                html = render_to_string('ferramentas/termo_pdf_template.html', {'termo': termo})
                with tempfile.NamedTemporaryFile(delete=True, suffix='.html') as tmp_h:
                    tmp_h.write(html.encode('UTF-8'))
                    tmp_h.flush()
                    with tempfile.NamedTemporaryFile(delete=True, suffix='.pdf') as tmp_p:
                        subprocess.run(['weasyprint', tmp_h.name, tmp_p.name], check=True)
                        tmp_p.seek(0)
                        zf.writestr(f'termo_{termo.pk}.pdf', tmp_p.read())

        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="termos_responsabilidade.zip"'
        return response

    
