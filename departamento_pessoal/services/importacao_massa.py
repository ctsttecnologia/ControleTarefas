
"""
Service para importação em massa de Funcionários via planilha Excel.

Responsabilidades:
  - Gerar planilha modelo (.xlsx) com instruções e validações
  - Processar upload: validar, criar/buscar Departamento, Cargo, Cliente
    e criar Funcionário + Documentos (CPF, RG, CTPS, PIS)
  - Retornar relatório detalhado (sucessos, erros por linha)
"""

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cliente.models import Cliente
from departamento_pessoal.models import (
    Cargo,
    Departamento,
    Documento,
    Funcionario,
)


# ============================================================================
# CONSTANTES — COLUNAS DA PLANILHA
# ============================================================================

STATUS_CHOICES_PLAN = ["ATIVO", "INATIVO", "FERIAS", "AFASTADO"]
SEXO_CHOICES_PLAN = ["M", "F", "O"]

COLUNAS_FUNCIONARIO = [
    # (header, campo, largura, obrigatório, dica)
    ("Nome Completo*", "nome_completo", 35, True, "Nome completo do funcionário"),
    ("Matrícula*", "matricula", 15, True, "Código único de matrícula"),
    ("Data Admissão*", "data_admissao", 16, True, "Formato: DD/MM/AAAA"),
    ("Data Demissão", "data_demissao", 16, False, "Formato: DD/MM/AAAA (em branco se ativo)"),
    ("Data Nascimento", "data_nascimento", 16, False, "Formato: DD/MM/AAAA"),
    ("Sexo", "sexo", 10, False, "M = Masculino / F = Feminino / O = Outro"),
    ("E-mail Pessoal", "email_pessoal", 30, False, "email@exemplo.com.br"),
    ("Telefone", "telefone", 20, False, "(00) 00000-0000"),
    ("Salário*", "salario", 14, True, "Valor numérico. Ex: 3500.00"),
    ("Status*", "status", 12, True, "ATIVO, INATIVO, FERIAS ou AFASTADO"),
]

COLUNAS_VINCULO = [
    ("Departamento*", "departamento", 25, True, "Nome exato do departamento"),
    ("Cargo*", "cargo", 25, True, "Nome exato do cargo"),
    ("Cliente (Nome Fantasia)", "cliente", 30, False, "Nome Fantasia do cliente/contrato"),
]

COLUNAS_DOCUMENTOS = [
    ("CPF", "cpf", 18, False, "000.000.000-00 (11 dígitos)"),
    ("RG Número", "rg_numero", 18, False, "Número do RG"),
    ("RG Órgão Expedidor", "rg_orgao", 18, False, "Ex: SSP-SP"),
    ("RG UF", "rg_uf", 8, False, "Sigla UF (ex: SP)"),
    ("RG Data Emissão", "rg_data_emissao", 16, False, "DD/MM/AAAA"),
    ("RG Nome da Mãe", "rg_nome_mae", 30, False, "Nome completo da mãe"),
    ("RG Nome do Pai", "rg_nome_pai", 30, False, "Nome completo do pai"),
    ("RG Naturalidade", "rg_naturalidade", 20, False, "Ex: São Paulo/SP"),
    ("CTPS Número", "ctps_numero", 18, False, "Número da CTPS"),
    ("CTPS Série", "ctps_serie", 12, False, "Série da CTPS"),
    ("CTPS UF", "ctps_uf", 8, False, "Sigla UF (ex: SP)"),
    ("PIS/PASEP", "pis", 18, False, "Número do PIS/PASEP/NIT"),
]

TODAS_COLUNAS = COLUNAS_FUNCIONARIO + COLUNAS_VINCULO + COLUNAS_DOCUMENTOS

UF_SIGLAS = [
    "AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO", "MA",
    "MG", "MS", "MT", "PA", "PB", "PE", "PI", "PR", "RJ", "RN",
    "RO", "RR", "RS", "SC", "SE", "SP", "TO",
]


# ============================================================================
# ESTILOS
# ============================================================================

HEADER_FILL = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
REQUIRED_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
INSTRUCAO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INSTRUCAO_FONT = Font(name="Calibri", size=10, italic=True, color="7F6000")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B4332")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2D6A4F")
TEXT_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="center")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ============================================================================
# GERAÇÃO DA PLANILHA MODELO
# ============================================================================


def gerar_planilha_modelo(filial):
    """
    Gera planilha Excel modelo para importação em massa de funcionários.

    Args:
        filial: Instância de Filial para listar departamentos, cargos e clientes.

    Returns:
        BytesIO: Buffer com o arquivo .xlsx pronto para download.
    """
    wb = Workbook()

    # ----- ABA DE INSTRUÇÕES -----
    ws_instrucoes = wb.active
    ws_instrucoes.title = "Instruções"
    ws_instrucoes.sheet_properties.tabColor = "FFC000"
    _criar_aba_instrucoes(ws_instrucoes)

    # ----- ABA DE DADOS -----
    ws_dados = wb.create_sheet(title="Dados Funcionários")
    ws_dados.sheet_properties.tabColor = "1B4332"
    _criar_aba_dados(ws_dados)

    # ----- ABA DE REFERÊNCIA -----
    ws_ref = wb.create_sheet(title="Referência")
    ws_ref.sheet_properties.tabColor = "70AD47"
    _criar_aba_referencia(ws_ref, filial)

    # Proteger abas de instruções e referência
    ws_instrucoes.protection.sheet = True
    ws_instrucoes.protection.password = "modelo"
    ws_ref.protection.sheet = True
    ws_ref.protection.password = "modelo"

    # Aba de dados como ativa ao abrir
    wb.active = ws_dados

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _criar_aba_instrucoes(ws):
    """Cria a aba com instruções detalhadas de preenchimento."""
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 85

    # Título
    cell = ws["B2"]
    cell.value = "INSTRUÇÕES PARA IMPORTAÇÃO EM MASSA DE FUNCIONÁRIOS"
    cell.font = TITLE_FONT

    instrucoes = [
        ("", ""),
        ("REGRAS GERAIS", None),
        ("", "1. Preencha os dados na aba 'Dados Funcionários'. NÃO altere os cabeçalhos."),
        ("", "2. Campos marcados com * (asterisco) são OBRIGATÓRIOS."),
        ("", "3. Cada linha representa UM funcionário com seus documentos."),
        ("", "4. A Matrícula deve ser única — duplicatas serão rejeitadas."),
        ("", "5. Não deixe linhas em branco entre os registros."),
        ("", ""),
        ("FORMATO DOS CAMPOS", None),
        ("", "• Datas: DD/MM/AAAA (ex: 15/03/2024)"),
        ("", "• CPF: 000.000.000-00 (com pontuação) — 11 dígitos"),
        ("", "• Telefone: (00) 00000-0000 ou (00) 0000-0000"),
        ("", "• Salário: Valor numérico com ponto. Ex: 3500.00"),
        ("", "• Sexo: M (Masculino), F (Feminino) ou O (Outro)"),
        ("", "• Status: ATIVO, INATIVO, FERIAS ou AFASTADO"),
        ("", ""),
        ("VÍNCULOS (Departamento, Cargo, Cliente)", None),
        ("", "• Departamento e Cargo devem existir previamente no sistema."),
        ("", "• Consulte a aba 'Referência' para ver os nomes cadastrados."),
        ("", "• O nome deve ser EXATAMENTE igual ao cadastrado."),
        ("", "• Cliente é opcional — informe o Nome Fantasia exato se aplicável."),
        ("", ""),
        ("DOCUMENTOS", None),
        ("", "• CPF, RG, CTPS e PIS são opcionais na importação."),
        ("", "• Se informados, serão criados como documentos vinculados."),
        ("", "• RG: pode informar órgão expedidor, UF, data emissão, filiação e naturalidade."),
        ("", "• CTPS: pode informar número, série e UF."),
        ("", "• PIS/PASEP: apenas o número."),
        ("", ""),
        ("OBSERVAÇÕES IMPORTANTES", None),
        ("", "• Se a matrícula já existir no sistema, a linha será REJEITADA."),
        ("", "• Se o e-mail já existir vinculado a outro funcionário, será REJEITADO."),
        ("", "• Erros em uma linha NÃO impedem o processamento das demais."),
        ("", "• Após o upload, você receberá um relatório com sucessos e erros."),
    ]

    row = 3
    for col_a, col_b in instrucoes:
        row += 1
        if col_b is None:
            ws.cell(row=row, column=2, value=col_a).font = SUBTITLE_FONT
        else:
            ws.cell(row=row, column=2, value=col_b).font = TEXT_FONT

    # Tabelas de campos
    row += 2
    ws.cell(row=row, column=2, value="CAMPOS DO FUNCIONÁRIO (Colunas A–J)").font = SUBTITLE_FONT
    row += 1
    _inserir_tabela_campos(ws, row, COLUNAS_FUNCIONARIO)

    row += len(COLUNAS_FUNCIONARIO) + 3
    ws.cell(row=row, column=2, value="CAMPOS DE VÍNCULO (Colunas K–M)").font = SUBTITLE_FONT
    row += 1
    _inserir_tabela_campos(ws, row, COLUNAS_VINCULO)

    row += len(COLUNAS_VINCULO) + 3
    ws.cell(row=row, column=2, value="CAMPOS DE DOCUMENTOS (Colunas N–Y)").font = SUBTITLE_FONT
    row += 1
    _inserir_tabela_campos(ws, row, COLUNAS_DOCUMENTOS)


def _inserir_tabela_campos(ws, start_row, colunas):
    """Insere tabela descritiva dos campos na aba de instruções."""
    headers = ["Campo", "Obrigatório", "Descrição/Formato"]
    for j, h in enumerate(headers, start=2):
        cell = ws.cell(row=start_row, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT

    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 50

    for i, (header, _, _, obrig, dica) in enumerate(colunas, start=1):
        r = start_row + i
        ws.cell(row=r, column=2, value=header.replace("*", "")).font = TEXT_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

        obrig_cell = ws.cell(row=r, column=3, value="SIM" if obrig else "NÃO")
        obrig_cell.font = Font(
            name="Calibri", size=10, bold=obrig,
            color="C00000" if obrig else "548235",
        )
        obrig_cell.alignment = CENTER_ALIGNMENT
        obrig_cell.border = THIN_BORDER

        ws.cell(row=r, column=4, value=dica).font = TEXT_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER


def _criar_aba_dados(ws):
    """Cria a aba principal de preenchimento de dados."""

    # Campos que devem ser formatados como texto para preservar zeros
    campos_texto = {
        "matricula", "cpf", "rg_numero", "ctps_numero",
        "ctps_serie", "pis", "telefone",
    }

    indices_texto = set()
    for i, (_, campo, _, _, _) in enumerate(TODAS_COLUNAS, start=1):
        if campo in campos_texto:
            indices_texto.add(i)

    for i, (header, _, _, obrig, dica) in enumerate(TODAS_COLUNAS, start=1):
        col_letter = get_column_letter(i)

        # Dica na linha 1
        hint_cell = ws.cell(row=1, column=i, value=f"💡 {dica}")
        hint_cell.font = INSTRUCAO_FONT
        hint_cell.fill = INSTRUCAO_FILL
        hint_cell.alignment = WRAP_ALIGNMENT
        hint_cell.protection = Protection(locked=True)

        # Header na linha 2
        header_cell = ws.cell(row=2, column=i, value=header)
        header_cell.font = HEADER_FONT
        header_cell.fill = HEADER_FILL
        header_cell.alignment = CENTER_ALIGNMENT
        header_cell.border = THIN_BORDER
        header_cell.protection = Protection(locked=True)

        # Largura
        ws.column_dimensions[col_letter].width = TODAS_COLUNAS[i - 1][2]

    # Formatação das células de dados (linhas 3 a 502)
    for row in range(3, 503):
        for i, (_, _, _, obrig, _) in enumerate(TODAS_COLUNAS, start=1):
            cell = ws.cell(row=row, column=i)
            cell.fill = REQUIRED_FILL if obrig else OPTIONAL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            cell.protection = Protection(locked=False)

            # Forçar formato TEXTO
            if i in indices_texto:
                cell.number_format = FORMAT_TEXT

    # --- DATA VALIDATIONS ---

    # Sexo (M/F/O)
    idx_sexo = [c[1] for c in COLUNAS_FUNCIONARIO].index("sexo") + 1
    col_sexo = get_column_letter(idx_sexo)
    dv_sexo = DataValidation(
        type="list",
        formula1='"M,F,O"',
        allow_blank=True,
    )
    dv_sexo.error = "Selecione M, F ou O"
    dv_sexo.errorTitle = "Sexo inválido"
    dv_sexo.prompt = "M = Masculino / F = Feminino / O = Outro"
    dv_sexo.promptTitle = "Sexo"
    dv_sexo.showInputMessage = True
    dv_sexo.showErrorMessage = True
    ws.add_data_validation(dv_sexo)
    dv_sexo.add(f"{col_sexo}3:{col_sexo}502")

    # Status
    idx_status = [c[1] for c in COLUNAS_FUNCIONARIO].index("status") + 1
    col_status = get_column_letter(idx_status)
    dv_status = DataValidation(
        type="list",
        formula1=f'"{ ",".join(STATUS_CHOICES_PLAN) }"',
        allow_blank=False,
    )
    dv_status.error = "Selecione um status válido"
    dv_status.errorTitle = "Status inválido"
    dv_status.prompt = "ATIVO, INATIVO, FERIAS ou AFASTADO"
    dv_status.promptTitle = "Status"
    dv_status.showInputMessage = True
    dv_status.showErrorMessage = True
    ws.add_data_validation(dv_status)
    dv_status.add(f"{col_status}3:{col_status}502")

    # RG UF
    offset_docs = len(COLUNAS_FUNCIONARIO) + len(COLUNAS_VINCULO)
    idx_rg_uf = offset_docs + [c[1] for c in COLUNAS_DOCUMENTOS].index("rg_uf") + 1
    col_rg_uf = get_column_letter(idx_rg_uf)
    dv_rg_uf = DataValidation(
        type="list",
        formula1=f'"{ ",".join(UF_SIGLAS) }"',
        allow_blank=True,
    )
    dv_rg_uf.error = "Selecione uma UF válida"
    dv_rg_uf.showInputMessage = True
    dv_rg_uf.showErrorMessage = True
    ws.add_data_validation(dv_rg_uf)
    dv_rg_uf.add(f"{col_rg_uf}3:{col_rg_uf}502")

    # CTPS UF
    idx_ctps_uf = offset_docs + [c[1] for c in COLUNAS_DOCUMENTOS].index("ctps_uf") + 1
    col_ctps_uf = get_column_letter(idx_ctps_uf)
    dv_ctps_uf = DataValidation(
        type="list",
        formula1=f'"{ ",".join(UF_SIGLAS) }"',
        allow_blank=True,
    )
    dv_ctps_uf.error = "Selecione uma UF válida"
    dv_ctps_uf.showInputMessage = True
    dv_ctps_uf.showErrorMessage = True
    ws.add_data_validation(dv_ctps_uf)
    dv_ctps_uf.add(f"{col_ctps_uf}3:{col_ctps_uf}502")

    # Congelar painéis
    ws.freeze_panes = "A3"

    # Proteção (linhas 1-2 bloqueadas, dados editáveis)
    ws.protection.sheet = True
    ws.protection.password = "modelo"
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.enable()

    # Alturas
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 25


def _criar_aba_referencia(ws, filial):
    """Cria aba com listas de referência (Departamentos, Cargos, Clientes)."""
    col = 1

    # ── DEPARTAMENTOS ──
    ws.cell(row=1, column=col, value="DEPARTAMENTOS CADASTRADOS").font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)

    ws.cell(row=2, column=col, value="Nome").font = HEADER_FONT
    ws.cell(row=2, column=col).fill = HEADER_FILL
    ws.cell(row=2, column=col + 1, value="Centro Custo").font = HEADER_FONT
    ws.cell(row=2, column=col + 1).fill = HEADER_FILL

    ws.column_dimensions[get_column_letter(col)].width = 35
    ws.column_dimensions[get_column_letter(col + 1)].width = 15

    departamentos = Departamento.objects.model.objects.filter(
        filial=filial, ativo=True
    ).order_by("nome")
    for i, dep in enumerate(departamentos, start=3):
        ws.cell(row=i, column=col, value=dep.nome).font = TEXT_FONT
        ws.cell(row=i, column=col).border = THIN_BORDER
        ws.cell(row=i, column=col + 1, value=dep.centro_custo or "").font = TEXT_FONT
        ws.cell(row=i, column=col + 1).border = THIN_BORDER

    # ── CARGOS ──
    col = 4
    ws.cell(row=1, column=col, value="CARGOS CADASTRADOS").font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)

    ws.cell(row=2, column=col, value="Nome").font = HEADER_FONT
    ws.cell(row=2, column=col).fill = HEADER_FILL
    ws.cell(row=2, column=col + 1, value="CBO").font = HEADER_FONT
    ws.cell(row=2, column=col + 1).fill = HEADER_FILL

    ws.column_dimensions[get_column_letter(col)].width = 35
    ws.column_dimensions[get_column_letter(col + 1)].width = 12

    cargos = Cargo.objects.model.objects.filter(
        filial=filial, ativo=True
    ).order_by("nome")
    for i, cargo in enumerate(cargos, start=3):
        ws.cell(row=i, column=col, value=cargo.nome).font = TEXT_FONT
        ws.cell(row=i, column=col).border = THIN_BORDER
        ws.cell(row=i, column=col + 1, value=cargo.cbo or "").font = TEXT_FONT
        ws.cell(row=i, column=col + 1).border = THIN_BORDER

    # ── CLIENTES ──
    col = 7
    ws.cell(row=1, column=col, value="CLIENTES / CONTRATOS").font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)

    ws.cell(row=2, column=col, value="Nome Fantasia").font = HEADER_FONT
    ws.cell(row=2, column=col).fill = HEADER_FILL
    ws.cell(row=2, column=col + 1, value="Contrato").font = HEADER_FONT
    ws.cell(row=2, column=col + 1).fill = HEADER_FILL

    ws.column_dimensions[get_column_letter(col)].width = 40
    ws.column_dimensions[get_column_letter(col + 1)].width = 12

    clientes = Cliente.objects.model.objects.filter(
        filial=filial, estatus=True
    ).order_by("nome")
    for i, cli in enumerate(clientes, start=3):
        ws.cell(row=i, column=col, value=cli.nome).font = TEXT_FONT
        ws.cell(row=i, column=col).border = THIN_BORDER
        ws.cell(row=i, column=col + 1, value=cli.contrato or "").font = TEXT_FONT
        ws.cell(row=i, column=col + 1).border = THIN_BORDER

    # ── STATUS ──
    col = 10
    ws.cell(row=1, column=col, value="STATUS VÁLIDOS").font = SUBTITLE_FONT
    ws.cell(row=2, column=col, value="Valor").font = HEADER_FONT
    ws.cell(row=2, column=col).fill = HEADER_FILL
    ws.column_dimensions[get_column_letter(col)].width = 14

    for i, s in enumerate(STATUS_CHOICES_PLAN, start=3):
        ws.cell(row=i, column=col, value=s).font = TEXT_FONT
        ws.cell(row=i, column=col).border = THIN_BORDER

    # ── SEXO ──
    col = 12
    ws.cell(row=1, column=col, value="SEXO").font = SUBTITLE_FONT
    ws.cell(row=2, column=col, value="Sigla").font = HEADER_FONT
    ws.cell(row=2, column=col).fill = HEADER_FILL
    ws.cell(row=2, column=col + 1, value="Descrição").font = HEADER_FONT
    ws.cell(row=2, column=col + 1).fill = HEADER_FILL
    ws.column_dimensions[get_column_letter(col)].width = 8
    ws.column_dimensions[get_column_letter(col + 1)].width = 14

    sexos = [("M", "Masculino"), ("F", "Feminino"), ("O", "Outro")]
    for i, (sigla, desc) in enumerate(sexos, start=3):
        ws.cell(row=i, column=col, value=sigla).font = TEXT_FONT
        ws.cell(row=i, column=col).border = THIN_BORDER
        ws.cell(row=i, column=col + 1, value=desc).font = TEXT_FONT
        ws.cell(row=i, column=col + 1).border = THIN_BORDER


# ============================================================================
# PROCESSAMENTO DA PLANILHA ENVIADA
# ============================================================================


def processar_planilha(arquivo, filial):
    """
    Processa planilha Excel enviada e cria Funcionários + Documentos.

    Args:
        arquivo: InMemoryUploadedFile (arquivo .xlsx)
        filial: Instância de Filial do usuário logado

    Returns:
        dict com total, sucessos, erros, detalhes_sucesso, detalhes_erro
    """
    resultado = {
        "total": 0,
        "sucessos": 0,
        "erros": 0,
        "detalhes_sucesso": [],
        "detalhes_erro": [],
    }

    try:
        wb = load_workbook(arquivo, data_only=True)
    except Exception:
        resultado["detalhes_erro"].append(
            {"linha": 0, "erros": ["Arquivo inválido. Envie um arquivo .xlsx válido."]}
        )
        return resultado

    # Procura aba "Dados Funcionários"
    if "Dados Funcionários" in wb.sheetnames:
        ws = wb["Dados Funcionários"]
    else:
        ws = wb.worksheets[0]

    start_row = 3  # Linha 1=dica, 2=header, 3+=dados
    total_colunas = len(TODAS_COLUNAS)

    for row_idx in range(start_row, ws.max_row + 1):
        valores = []
        for col_idx in range(1, total_colunas + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            valores.append(cell_value)

        # Pular linhas totalmente vazias
        if all(v is None or str(v).strip() == "" for v in valores):
            continue

        resultado["total"] += 1
        erros_linha = []

        # Mapear valores para dicionários
        dados_func = {}
        for i, (_, campo, _, _, _) in enumerate(COLUNAS_FUNCIONARIO):
            dados_func[campo] = valores[i]

        dados_vinculo = {}
        offset_v = len(COLUNAS_FUNCIONARIO)
        for i, (_, campo, _, _, _) in enumerate(COLUNAS_VINCULO):
            dados_vinculo[campo] = valores[offset_v + i]

        dados_docs = {}
        offset_d = offset_v + len(COLUNAS_VINCULO)
        for i, (_, campo, _, _, _) in enumerate(COLUNAS_DOCUMENTOS):
            dados_docs[campo] = valores[offset_d + i]

        # Validações
        erros_linha.extend(_validar_funcionario(dados_func))
        erros_linha.extend(_validar_vinculo(dados_vinculo, filial))
        erros_linha.extend(_validar_documentos(dados_docs))

        if erros_linha:
            resultado["erros"] += 1
            resultado["detalhes_erro"].append(
                {"linha": row_idx, "erros": erros_linha}
            )
            continue

        # Criar registros
        try:
            with transaction.atomic():
                funcionario = _criar_funcionario(dados_func, dados_vinculo, filial)
                _criar_documentos(dados_docs, funcionario, filial)
                resultado["sucessos"] += 1
                resultado["detalhes_sucesso"].append(
                    f"Linha {row_idx}: {funcionario.nome_completo} "
                    f"(Mat: {funcionario.matricula}) — importado com sucesso."
                )
        except Exception as e:
            resultado["erros"] += 1
            resultado["detalhes_erro"].append(
                {"linha": row_idx, "erros": [f"Erro ao salvar: {str(e)}"]}
            )

    return resultado


# ============================================================================
# VALIDAÇÕES
# ============================================================================


def _validar_funcionario(dados):
    """Valida campos do funcionário e retorna lista de erros."""
    erros = []

    # Nome
    if not _tem_valor(dados.get("nome_completo")):
        erros.append("Nome Completo é obrigatório.")

    # Matrícula
    matricula = dados.get("matricula")
    if not _tem_valor(matricula):
        erros.append("Matrícula é obrigatória.")
    else:
        mat_str = str(matricula).strip()
        if Funcionario.objects.model.objects.filter(matricula=mat_str).exists():
            erros.append(f"Matrícula '{mat_str}' já existe no sistema.")

    # Data admissão
    if not dados.get("data_admissao"):
        erros.append("Data de Admissão é obrigatória.")
    else:
        if _parse_data(dados["data_admissao"]) is None:
            erros.append(f"Data de Admissão '{dados['data_admissao']}' inválida. Use DD/MM/AAAA.")

    # Data demissão (opcional)
    data_dem = dados.get("data_demissao")
    if data_dem and _tem_valor(data_dem):
        if _parse_data(data_dem) is None:
            erros.append(f"Data de Demissão '{data_dem}' inválida. Use DD/MM/AAAA.")

    # Data nascimento (opcional)
    data_nasc = dados.get("data_nascimento")
    if data_nasc and _tem_valor(data_nasc):
        if _parse_data(data_nasc) is None:
            erros.append(f"Data de Nascimento '{data_nasc}' inválida. Use DD/MM/AAAA.")

    # Sexo
    sexo = dados.get("sexo")
    if sexo and _tem_valor(sexo):
        if str(sexo).strip().upper() not in SEXO_CHOICES_PLAN:
            erros.append(f"Sexo '{sexo}' inválido. Use M, F ou O.")

    # E-mail
    email = dados.get("email_pessoal")
    if email and _tem_valor(email):
        email_str = str(email).strip()
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email_str):
            erros.append(f"E-mail inválido: '{email_str}'.")
        elif Funcionario.objects.model.objects.filter(email_pessoal=email_str).exists():
            erros.append(f"E-mail '{email_str}' já está vinculado a outro funcionário.")

    # Salário
    salario = dados.get("salario")
    if not _tem_valor(salario):
        erros.append("Salário é obrigatório.")
    else:
        try:
            val = Decimal(str(salario).strip().replace(",", "."))
            if val < 0:
                erros.append("Salário não pode ser negativo.")
        except (InvalidOperation, ValueError):
            erros.append(f"Salário '{salario}' inválido. Use formato numérico (ex: 3500.00).")

    # Status
    status = dados.get("status")
    if not _tem_valor(status):
        erros.append("Status é obrigatório.")
    elif str(status).strip().upper() not in STATUS_CHOICES_PLAN:
        erros.append(f"Status '{status}' inválido. Use: {', '.join(STATUS_CHOICES_PLAN)}.")

    return erros


def _validar_vinculo(dados, filial):
    """Valida campos de vínculo (departamento, cargo, cliente)."""
    erros = []

    # Departamento
    dep_nome = dados.get("departamento")
    if not _tem_valor(dep_nome):
        erros.append("Departamento é obrigatório.")
    else:
        dep_str = str(dep_nome).strip()
        if not Departamento.objects.model.objects.filter(
            nome__iexact=dep_str, filial=filial, ativo=True
        ).exists():
            erros.append(
                f"Departamento '{dep_str}' não encontrado. "
                f"Verifique na aba 'Referência' ou cadastre-o antes."
            )

    # Cargo
    cargo_nome = dados.get("cargo")
    if not _tem_valor(cargo_nome):
        erros.append("Cargo é obrigatório.")
    else:
        cargo_str = str(cargo_nome).strip()
        if not Cargo.objects.model.objects.filter(
            nome__iexact=cargo_str, filial=filial, ativo=True
        ).exists():
            erros.append(
                f"Cargo '{cargo_str}' não encontrado. "
                f"Verifique na aba 'Referência' ou cadastre-o antes."
            )

    # Cliente (opcional)
    cliente_nome = dados.get("cliente")
    if cliente_nome and _tem_valor(cliente_nome):
        cli_str = str(cliente_nome).strip()
        if not Cliente.objects.model.objects.filter(
            nome__iexact=cli_str, filial=filial
        ).exists():
            erros.append(
                f"Cliente '{cli_str}' não encontrado. "
                f"Verifique na aba 'Referência' ou cadastre-o antes."
            )

    return erros


def _validar_documentos(dados):
    """Valida campos de documentos."""
    erros = []

    # CPF
    cpf = dados.get("cpf")
    if cpf and _tem_valor(cpf):
        cpf_digitos = re.sub(r"\D", "", str(cpf).strip())
        if len(cpf_digitos) != 11:
            erros.append(f"CPF '{cpf}' inválido. Deve ter 11 dígitos.")

    # RG UF
    rg_uf = dados.get("rg_uf")
    if rg_uf and _tem_valor(rg_uf):
        if str(rg_uf).strip().upper() not in UF_SIGLAS:
            erros.append(f"RG UF '{rg_uf}' inválida.")

    # RG Data Emissão
    rg_dt = dados.get("rg_data_emissao")
    if rg_dt and _tem_valor(rg_dt):
        if _parse_data(rg_dt) is None:
            erros.append(f"RG Data Emissão '{rg_dt}' inválida. Use DD/MM/AAAA.")

    # CTPS UF
    ctps_uf = dados.get("ctps_uf")
    if ctps_uf and _tem_valor(ctps_uf):
        if str(ctps_uf).strip().upper() not in UF_SIGLAS:
            erros.append(f"CTPS UF '{ctps_uf}' inválida.")

    return erros


# ============================================================================
# CRIAÇÃO DE REGISTROS
# ============================================================================


def _criar_funcionario(dados_func, dados_vinculo, filial):
    """Cria instância de Funcionário com os dados validados."""
    # Buscar vínculo
    departamento = Departamento.objects.model.objects.filter(
        nome__iexact=str(dados_vinculo["departamento"]).strip(),
        filial=filial,
        ativo=True,
    ).first()

    cargo = Cargo.objects.model.objects.filter(
        nome__iexact=str(dados_vinculo["cargo"]).strip(),
        filial=filial,
        ativo=True,
    ).first()

    cliente = None
    if dados_vinculo.get("cliente") and _tem_valor(dados_vinculo["cliente"]):
        cliente = Cliente.objects.model.objects.filter(
            nome__iexact=str(dados_vinculo["cliente"]).strip(),
            filial=filial,
        ).first()

    # Dados do funcionário
    nome = str(dados_func["nome_completo"]).strip()
    matricula = str(dados_func["matricula"]).strip()
    data_admissao = _parse_data(dados_func["data_admissao"])

    data_demissao = None
    if dados_func.get("data_demissao") and _tem_valor(dados_func["data_demissao"]):
        data_demissao = _parse_data(dados_func["data_demissao"])

    data_nascimento = None
    if dados_func.get("data_nascimento") and _tem_valor(dados_func["data_nascimento"]):
        data_nascimento = _parse_data(dados_func["data_nascimento"])

    sexo = None
    if dados_func.get("sexo") and _tem_valor(dados_func["sexo"]):
        sexo = str(dados_func["sexo"]).strip().upper()

    email_pessoal = None
    if dados_func.get("email_pessoal") and _tem_valor(dados_func["email_pessoal"]):
        email_pessoal = str(dados_func["email_pessoal"]).strip()

    telefone = ""
    if dados_func.get("telefone") and _tem_valor(dados_func["telefone"]):
        telefone = str(dados_func["telefone"]).strip()

    salario = Decimal(
        str(dados_func["salario"]).strip().replace(",", ".")
    )

    status = str(dados_func["status"]).strip().upper()

    funcionario = Funcionario(
        nome_completo=nome,
        matricula=matricula,
        data_admissao=data_admissao,
        data_demissao=data_demissao,
        data_nascimento=data_nascimento,
        sexo=sexo,
        email_pessoal=email_pessoal,
        telefone=telefone,
        salario=salario,
        status=status,
        departamento=departamento,
        cargo=cargo,
        cliente=cliente,
        filial=filial,
    )
    funcionario.full_clean()
    funcionario.save()
    return funcionario


def _criar_documentos(dados_docs, funcionario, filial):
    """Cria documentos vinculados ao funcionário (CPF, RG, CTPS, PIS)."""

    # ── CPF ──
    cpf = dados_docs.get("cpf")
    if cpf and _tem_valor(cpf):
        cpf_digitos = re.sub(r"\D", "", str(cpf).strip())
        cpf_formatado = (
            f"{cpf_digitos[:3]}.{cpf_digitos[3:6]}.{cpf_digitos[6:9]}-{cpf_digitos[9:]}"
        )
        Documento.objects.create(
            funcionario=funcionario,
            tipo_documento="CPF",
            numero=cpf_formatado,
            filial=filial,
        )

    # ── RG ──
    rg_numero = dados_docs.get("rg_numero")
    if rg_numero and _tem_valor(rg_numero):
        rg_orgao = (
            str(dados_docs["rg_orgao"]).strip()
            if dados_docs.get("rg_orgao") and _tem_valor(dados_docs["rg_orgao"])
            else None
        )
        rg_uf = (
            str(dados_docs["rg_uf"]).strip().upper()
            if dados_docs.get("rg_uf") and _tem_valor(dados_docs["rg_uf"])
            else None
        )
        rg_data_emissao = None
        if dados_docs.get("rg_data_emissao") and _tem_valor(dados_docs["rg_data_emissao"]):
            rg_data_emissao = _parse_data(dados_docs["rg_data_emissao"])

        rg_nome_mae = (
            str(dados_docs["rg_nome_mae"]).strip()
            if dados_docs.get("rg_nome_mae") and _tem_valor(dados_docs["rg_nome_mae"])
            else None
        )
        rg_nome_pai = (
            str(dados_docs["rg_nome_pai"]).strip()
            if dados_docs.get("rg_nome_pai") and _tem_valor(dados_docs["rg_nome_pai"])
            else None
        )
        rg_naturalidade = (
            str(dados_docs["rg_naturalidade"]).strip()
            if dados_docs.get("rg_naturalidade") and _tem_valor(dados_docs["rg_naturalidade"])
            else None
        )

        Documento.objects.create(
            funcionario=funcionario,
            tipo_documento="RG",
            numero=str(rg_numero).strip(),
            orgao_expedidor=rg_orgao,
            uf_expedidor=rg_uf,
            data_emissao=rg_data_emissao,
            rg_nome_mae=rg_nome_mae,
            rg_nome_pai=rg_nome_pai,
            rg_naturalidade=rg_naturalidade,
            filial=filial,
        )

    # ── CTPS ──
    ctps_numero = dados_docs.get("ctps_numero")
    if ctps_numero and _tem_valor(ctps_numero):
        ctps_serie = (
            str(dados_docs["ctps_serie"]).strip()
            if dados_docs.get("ctps_serie") and _tem_valor(dados_docs["ctps_serie"])
            else None
        )
        ctps_uf = (
            str(dados_docs["ctps_uf"]).strip().upper()
            if dados_docs.get("ctps_uf") and _tem_valor(dados_docs["ctps_uf"])
            else None
        )

        Documento.objects.create(
            funcionario=funcionario,
            tipo_documento="CTPS",
            numero=str(ctps_numero).strip(),
            ctps_serie=ctps_serie,
            ctps_uf=ctps_uf,
            filial=filial,
        )

    # ── PIS/PASEP ──
    pis = dados_docs.get("pis")
    if pis and _tem_valor(pis):
        Documento.objects.create(
            funcionario=funcionario,
            tipo_documento="PIS",
            numero=str(pis).strip(),
            filial=filial,
        )


# ============================================================================
# UTILITÁRIOS
# ============================================================================


def _tem_valor(valor):
    """Verifica se o valor não é vazio/nulo."""
    return valor is not None and str(valor).strip() != ""


def _parse_data(valor):
    """Converte valor para date. Aceita datetime, date, ou string DD/MM/AAAA."""
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    valor_str = str(valor).strip()
    formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"]
    for fmt in formatos:
        try:
            return datetime.strptime(valor_str, fmt).date()
        except ValueError:
            continue

    return None

