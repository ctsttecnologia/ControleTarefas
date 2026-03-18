
import json
from urllib.request import urlopen
from urllib.error import URLError, HTTPError
from django.http import HttpResponse, Http404, JsonResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View
from django.http import HttpResponse, Http404
from django.utils.translation import gettext_lazy as _
from django.db import transaction

import io
import base64
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import Logradouro, Filial
from .forms import LogradouroForm, UploadFileForm
from .constant import ESTADOS_BRASIL
from core.mixins import SSTPermissionMixin, ViewFilialScopedMixin, FilialCreateMixin
import requests as http_requests


# =============================================================================
# CRUD — Logradouro
# =============================================================================

class LogradouroListView(
    LoginRequiredMixin,
    SSTPermissionMixin,
    ViewFilialScopedMixin,
    ListView,
):
    """Lista logradouros com escopo de filial, busca e paginação."""

    model = Logradouro
    template_name = 'logradouro/listar_logradouros.html'
    context_object_name = 'logradouros'
    paginate_by = 15
    permission_required = 'logradouro.view_logradouro'

    def get_queryset(self):
        # ViewFilialScopedMixin já filtra pela filial
        queryset = super().get_queryset()

        # Total ANTES da busca (para o cabeçalho)
        self.total_logradouros_na_filial = queryset.count()

        # Filtros de busca
        q_endereco = self.request.GET.get('q_endereco', '').strip()
        q_cep = self.request.GET.get('q_cep', '').strip()

        if q_endereco:
            queryset = queryset.filter(endereco__icontains=q_endereco)
        if q_cep:
            queryset = queryset.filter(cep__icontains=q_cep)

        return queryset.order_by('endereco')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_logradouros'] = getattr(self, 'total_logradouros_na_filial', 0)

        # ── Contagem dos resultados filtrados (sem query extra) ──
        context['total_filtrados'] = context['paginator'].count if context.get('paginator') else 0

        # ── Preservar filtros na paginação ──
        query_params = self.request.GET.copy()
        query_params.pop('page', None)
        context['query_string'] = query_params.urlencode()

        return context


class LogradouroCreateView(
    LoginRequiredMixin,
    SSTPermissionMixin,
    FilialCreateMixin,
    CreateView,
):
    """Cria logradouro associando à filial do usuário automaticamente."""

    model = Logradouro
    form_class = LogradouroForm
    template_name = 'logradouro/form_logradouro.html'
    success_url = reverse_lazy('logradouro:listar_logradouros')
    permission_required = 'logradouro.add_logradouro'

    def get_initial(self):
        initial = super().get_initial()
        if hasattr(self.request.user, 'filial_ativa'):
            initial['filial'] = self.request.user.filial_ativa
        return initial

    def form_valid(self, form):
        messages.success(self.request, _('Endereço cadastrado com sucesso!'))
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, _('Por favor, corrija os erros abaixo.'))
        return super().form_invalid(form)


class LogradouroUpdateView(
    LoginRequiredMixin,
    SSTPermissionMixin,
    ViewFilialScopedMixin,
    UpdateView,
):
    """Edita logradouro (somente da filial do usuário)."""

    model = Logradouro
    form_class = LogradouroForm
    template_name = 'logradouro/form_logradouro.html'
    success_url = reverse_lazy('logradouro:listar_logradouros')
    permission_required = 'logradouro.change_logradouro'

    def form_valid(self, form):
        messages.success(self.request, _('Endereço atualizado com sucesso!'))
        return super().form_valid(form)


class LogradouroDeleteView(
    LoginRequiredMixin,
    SSTPermissionMixin,
    ViewFilialScopedMixin,
    DeleteView,
):
    """Exclui logradouro (somente da filial do usuário)."""

    model = Logradouro
    template_name = 'logradouro/confirmar_exclusao.html'
    success_url = reverse_lazy('logradouro:listar_logradouros')
    context_object_name = 'logradouro'
    permission_required = 'logradouro.delete_logradouro'

    def form_valid(self, form):
        messages.success(self.request, _('Endereço excluído com sucesso!'))
        return super().form_valid(form)


# =============================================================================
# EXPORTAÇÃO — Excel
# =============================================================================

class LogradouroExportExcelView(LoginRequiredMixin, SSTPermissionMixin, View):
    """Exporta logradouros da filial do usuário para Excel."""

    permission_required = 'logradouro.view_logradouro'

    def get(self, request, *args, **kwargs):
        logradouros = Logradouro.objects.for_request(request).order_by('endereco')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logradouros"

        headers = [
            "Endereço", "Número", "CEP", "Complemento", "Bairro", "Cidade",
            "Estado", "País", "Ponto Referência", "Latitude", "Longitude",
            "Data Cadastro", "Data Atualização",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        center_aligned = Alignment(horizontal='center')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_aligned
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        for row_num, log in enumerate(logradouros, 2):
            row_data = [
                log.endereco, log.numero, log.cep_formatado,
                log.complemento, log.bairro, log.cidade,
                log.get_estado_display(), log.pais, log.ponto_referencia,
                log.latitude, log.longitude,
                log.data_cadastro.strftime('%d/%m/%Y %H:%M') if log.data_cadastro else "",
                log.data_atualizacao.strftime('%d/%m/%Y %H:%M') if log.data_atualizacao else "",
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="logradouros.xlsx"'},
        )
        wb.save(response)
        return response


# =============================================================================
# IMPORTAÇÃO — Upload em massa
# =============================================================================

class UploadLogradourosView(LoginRequiredMixin, SSTPermissionMixin, View):
    """Upload de planilha para importação em massa de logradouros."""

    permission_required = 'logradouro.add_logradouro'
    form_class = UploadFileForm
    template_name = 'logradouro/upload_logradouros.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        request.session.pop('relatorio_erros', None)
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        context = {'form': form}

        if not form.is_valid():
            return render(request, self.template_name, context)

        file = request.FILES['file']

        # ── Etapa 1: Leitura do arquivo ──
        try:
            df = pd.read_excel(file)
            df.dropna(how='all', inplace=True)
        except Exception as e:
            messages.error(
                request,
                f"Não foi possível ler o arquivo. Pode estar corrompido ou "
                f"em formato inválido. Erro: {e}",
            )
            return render(request, self.template_name, context)

        # ── Etapa 2: Validação estrutural ──
        required_cols = ['filial_id', 'endereco', 'numero', 'bairro', 'cep', 'cidade', 'estado']
        cols_faltando = [col for col in required_cols if col not in df.columns]
        if cols_faltando:
            messages.error(
                request,
                f"Colunas obrigatórias ausentes: {', '.join(cols_faltando)}. "
                f"Verifique o modelo e tente novamente.",
            )
            return render(request, self.template_name, context)

        # ── Etapa 3: Validação dos dados (linha a linha) ──
        linhas_com_erro = []
        enderecos_para_criar = []

        try:
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        if row[required_cols].isnull().any():
                            raise ValueError("Contém valores vazios em colunas obrigatórias.")

                        filial = Filial.objects.get(pk=int(row['filial_id']))
                        cep_limpo = str(row['cep']).split('.')[0]
                        lat = row.get('latitude')
                        lon = row.get('longitude')

                        logradouro_obj = Logradouro(
                            filial=filial,
                            endereco=row['endereco'],
                            numero=int(row['numero']),
                            complemento=row.get('complemento'),
                            bairro=row['bairro'],
                            cep=cep_limpo.zfill(8),
                            cidade=row['cidade'],
                            estado=row['estado'],
                            pais=row.get('pais', 'Brasil') or 'Brasil',
                            ponto_referencia=row.get('ponto_referencia'),
                            latitude=None if pd.isna(lat) else lat,
                            longitude=None if pd.isna(lon) else lon,
                        )
                        logradouro_obj.full_clean()
                        enderecos_para_criar.append(logradouro_obj)

                    except Exception as e:
                        linha_original = row.to_dict()
                        linha_original['Erro_Detectado'] = f"Linha {index + 2}: {e}"
                        linhas_com_erro.append(linha_original)

                if linhas_com_erro:
                    raise ValueError("Importação cancelada devido a erros.")

            # Sem erros — salvar
            Logradouro.objects.bulk_create(enderecos_para_criar)
            messages.success(request, f"{len(enderecos_para_criar)} endereços importados com sucesso!")
            return redirect('logradouro:listar_logradouros')

        except ValueError:
            messages.error(
                request,
                "A importação falhou. Verifique os erros no relatório abaixo.",
            )
            df_erros = pd.DataFrame(linhas_com_erro)
            buffer = io.BytesIO()
            df_erros.to_excel(buffer, index=False)
            buffer.seek(0)

            file_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            request.session['relatorio_erros'] = {
                'filename': 'relatorio_de_erros_logradouros.xlsx',
                'content': file_base64,
            }
            context['relatorio_disponivel'] = True
            context['total_erros'] = len(linhas_com_erro)

        return render(request, self.template_name, context)


class DownloadErroRelatorioView(LoginRequiredMixin, View):
    """Download do relatório de erros da importação."""

    def get(self, request, *args, **kwargs):
        relatorio_data = request.session.get('relatorio_erros')
        if not relatorio_data:
            raise Http404("Nenhum relatório de erros encontrado.")

        file_content = base64.b64decode(relatorio_data['content'])
        filename = relatorio_data['filename']
        del request.session['relatorio_erros']

        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class DownloadTemplateView(LoginRequiredMixin, View):
    """Gera modelo .xlsx formatado para download."""

    def get(self, request, *args, **kwargs):
        headers = [
            'filial_id', 'endereco', 'numero', 'complemento', 'bairro',
            'cep', 'cidade', 'estado', 'pais', 'ponto_referencia',
            'latitude', 'longitude',
        ]
        example_data = [[
            1, 'Avenida Paulista', 1578, 'Andar 10', 'Bela Vista',
            '01310200', 'São Paulo', 'SP', 'Brasil', 'Em frente ao MASP',
            -23.561350, -46.656530,
        ]]

        df = pd.DataFrame(example_data, columns=headers)
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, sheet_name='Enderecos')
        buffer.seek(0)

        workbook = load_workbook(buffer)
        ws_enderecos = workbook['Enderecos']

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')

        for col_num, column_cell in enumerate(ws_enderecos[1], 1):
            column_cell.font = header_font
            column_cell.fill = header_fill
            column_cell.alignment = center_align
            col_letter = ws_enderecos.cell(row=1, column=col_num).column_letter
            max_length = max(
                (len(str(cell.value or '')) for cell in ws_enderecos[col_letter]),
                default=10,
            )
            ws_enderecos.column_dimensions[col_letter].width = max_length + 4

        # Validação de estado (coluna H)
        estados_list = [item[0] for item in ESTADOS_BRASIL]
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(estados_list)}"',
            allow_blank=True,
        )
        dv.error = 'O valor deve ser uma sigla de estado válida.'
        dv.errorTitle = 'Entrada Inválida'
        dv.prompt = 'Selecione um estado da lista'
        dv.promptTitle = 'Seleção de Estado'
        ws_enderecos.add_data_validation(dv)
        dv.add('H2:H1048576')
        ws_enderecos.freeze_panes = 'A2'

        # Aba de instruções
        ws_instrucoes = workbook.create_sheet(title="Instruções", index=0)
        ws_instrucoes.append(['Coluna', 'Descrição', 'Obrigatório?', 'Exemplo'])
        for cell in ws_instrucoes[1]:
            cell.font = header_font
            cell.fill = header_fill

        instructions = [
            ['filial_id', 'ID numérico da Filial (deve existir no sistema)', 'Sim', '1'],
            ['endereco', 'Nome da rua, avenida (Máx: 150)', 'Sim', 'Avenida Paulista'],
            ['numero', 'Número do imóvel (inteiro positivo)', 'Sim', '1578'],
            ['complemento', 'Andar, sala, bloco (Máx: 50)', 'Não', 'Andar 10'],
            ['bairro', 'Nome do bairro (Máx: 60)', 'Sim', 'Bela Vista'],
            ['cep', 'CEP com 8 dígitos, sem pontos/traços', 'Sim', '01310200'],
            ['cidade', 'Nome da cidade (Máx: 60)', 'Sim', 'São Paulo'],
            ['estado', 'Sigla do estado (selecione da lista)', 'Sim', 'SP'],
            ['pais', 'Nome do país (padrão: Brasil)', 'Não', 'Brasil'],
            ['ponto_referencia', 'Referência para localização (Máx: 100)', 'Não', 'Em frente ao MASP'],
            ['latitude', 'Coordenada decimal com ponto', 'Não', '-23.561350'],
            ['longitude', 'Coordenada decimal com ponto', 'Não', '-46.656530'],
        ]
        for row in instructions:
            ws_instrucoes.append(row)

        for col_letter in ['A', 'B', 'C', 'D']:
            ws_instrucoes.column_dimensions[col_letter].auto_size = True

        workbook.active = ws_enderecos

        final_buffer = io.BytesIO()
        workbook.save(final_buffer)

        response = HttpResponse(
            final_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="enderecos.xlsx"'
        return response


def consulta_cep(request):
    """
    Consulta o CEP na API ViaCEP e retorna JSON.
    GET /logradouro/consulta-cep/?cep=01001000
    """
    cep = request.GET.get("cep", "").replace("-", "").replace(".", "").strip()

    if not cep or len(cep) != 8 or not cep.isdigit():
        return JsonResponse(
            {"erro": "CEP inválido. Informe 8 dígitos numéricos."}, status=400
        )

    url = f"https://viacep.com.br/ws/{cep}/json/"

    try:
        with urlopen(url, timeout=5) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("erro"):
            return JsonResponse({"erro": "CEP não encontrado."}, status=404)

        return JsonResponse({
            "cep": data.get("cep", ""),
            "endereco": data.get("logradouro", ""),
            "complemento": data.get("complemento", ""),
            "bairro": data.get("bairro", ""),
            "cidade": data.get("localidade", ""),
            "estado": data.get("uf", ""),
        })

    except HTTPError:
        return JsonResponse({"erro": "CEP não encontrado."}, status=404)
    except URLError:
        return JsonResponse({"erro": "Não foi possível conectar ao serviço de CEP."}, status=504)
    except Exception as e:
        return JsonResponse({"erro": f"Erro inesperado: {e}"}, status=500)
