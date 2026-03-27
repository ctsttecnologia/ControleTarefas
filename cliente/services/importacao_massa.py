
"""
Service para importação em massa de Clientes via planilha Excel.

Responsabilidades:
  - Gerar planilha modelo (.xlsx) com instruções e validações
  - Processar upload: validar, criar Logradouro + Cliente
  - Retornar relatório detalhado (sucessos, erros por linha)
"""

import re
from datetime import datetime
from io import BytesIO

from django.db import transaction
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cliente.models import Cliente
from logradouro.constant import ESTADOS_BRASIL, TIPOS_LOGRADOURO
from logradouro.models import Logradouro


# ============================================================================
# CONSTANTES
# ============================================================================

COLUNAS_CLIENTE = [
    # (header, campo, largura, obrigatório, dica)
    ("Razão Social*", "razao_social", 35, True, "Nome completo da empresa"),
    ("Nome Fantasia*", "nome", 30, True, "Nome comercial da empresa"),
    ("CNPJ*", "cnpj", 22, True, "Formato: 00.000.000/0000-00"),
    ("Contrato (CM)", "contrato", 15, False, "Até 4 caracteres. Padrão: 0"),
    ("Unidade/Filial", "unidade", 15, False, "Número inteiro positivo"),
    ("Inscrição Estadual", "inscricao_estadual", 20, False, "Até 20 caracteres"),
    ("Inscrição Municipal", "inscricao_municipal", 20, False, "Até 20 caracteres"),
    ("Telefone", "telefone", 20, False, "(00) 00000-0000 ou (00) 0000-0000"),
    ("E-mail", "email", 30, False, "email@exemplo.com.br"),
    ("Data de Início*", "data_de_inicio", 18, True, "Formato: DD/MM/AAAA"),
    ("Data Encerramento", "data_encerramento", 18, False, "Formato: DD/MM/AAAA"),
    ("Ativo?*", "estatus", 10, True, "SIM ou NÃO"),
    ("Observações", "observacoes", 40, False, "Texto livre"),
]

COLUNAS_ENDERECO = [
    ("Tipo Logradouro*", "tipo_logradouro", 18, True, "Rua, Avenida, Praça, etc."),
    ("Endereço*", "endereco", 35, True, "Nome do logradouro (sem Rua/Av). Ex: das Flores"),
    ("Número*", "numero", 10, True, "Número inteiro (mín. 1)"),
    ("CEP*", "cep", 12, True, "Apenas 8 dígitos (sem traço)"),
    ("Complemento", "complemento", 20, False, "Sala, Andar, Bloco..."),
    ("Bairro*", "bairro", 20, True, "Nome do bairro"),
    ("Cidade*", "cidade", 20, True, "Nome da cidade"),
    ("Estado*", "estado", 10, True, "Sigla UF (ex: SP, RJ, MG)"),
    ("País", "pais", 15, False, "Padrão: Brasil"),
    ("Ponto de Referência", "ponto_referencia", 25, False, "Próximo a..."),
    ("Latitude", "latitude", 14, False, "Ex: -23.550520"),
    ("Longitude", "longitude", 14, False, "Ex: -46.633308"),
]

TODAS_COLUNAS = COLUNAS_CLIENTE + COLUNAS_ENDERECO

SIGLAS_ESTADOS = [sigla for sigla, _ in ESTADOS_BRASIL]
SIGLAS_TIPOS = [sigla for sigla, _ in TIPOS_LOGRADOURO]
NOMES_TIPOS = {nome.upper(): sigla for sigla, nome in TIPOS_LOGRADOURO}
SIGLAS_PARA_NOME = {sigla: nome for sigla, nome in TIPOS_LOGRADOURO}


# ============================================================================
# ESTILOS
# ============================================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
REQUIRED_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
INSTRUCAO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INSTRUCAO_FONT = Font(name="Calibri", size=10, italic=True, color="7F6000")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2E75B6")
TEXT_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="center")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ============================================================================
# GERAÇÃO DA PLANILHA MODELO
# ============================================================================


def gerar_planilha_modelo():
    """
    Gera planilha Excel modelo para importação em massa de clientes.

    Returns:
        BytesIO: Buffer com o arquivo .xlsx pronto para download.
    """
    wb = Workbook()

    # ----- ABA DE INSTRUÇÕES -----
    ws_instrucoes = wb.active
    ws_instrucoes.title = "Instruções"
    ws_instrucoes.sheet_properties.tabColor = "FFC000"
    _criar_aba_instrucoes(ws_instrucoes)

    # ----- ABA DE DADOS -----
    ws_dados = wb.create_sheet(title="Dados Clientes")
    ws_dados.sheet_properties.tabColor = "1F4E79"
    _criar_aba_dados(ws_dados)

    # ----- ABA DE REFERÊNCIA -----
    ws_ref = wb.create_sheet(title="Referência")
    ws_ref.sheet_properties.tabColor = "70AD47"
    _criar_aba_referencia(ws_ref)

    # Proteger abas de instruções e referência
    ws_instrucoes.protection.sheet = True
    ws_instrucoes.protection.password = "modelo"
    ws_ref.protection.sheet = True
    ws_ref.protection.password = "modelo"

    # Aba de dados como ativa ao abrir
    wb.active = ws_dados

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _criar_aba_instrucoes(ws):
    """Cria a aba com instruções detalhadas de preenchimento."""
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 80

    # Título
    cell = ws["B2"]
    cell.value = "INSTRUÇÕES PARA IMPORTAÇÃO EM MASSA DE CLIENTES"
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")

    instrucoes = [
        ("", ""),
        ("REGRAS GERAIS", None),
        ("", "1. Preencha os dados na aba 'Dados Clientes'. NÃO altere os cabeçalhos."),
        ("", "2. Campos marcados com * (asterisco) são OBRIGATÓRIOS."),
        ("", "3. Cada linha representa UM cliente com seu respectivo endereço."),
        ("", "4. O CNPJ deve ser único — duplicatas serão rejeitadas."),
        ("", "5. Não deixe linhas em branco entre os registros."),
        ("", ""),
        ("FORMATO DOS CAMPOS", None),
        ("", "• CNPJ: 00.000.000/0000-00 (com pontuação)"),
        ("", "• Telefone: (00) 00000-0000 ou (00) 0000-0000"),
        ("", "• CEP: 8 dígitos SEM traço (ex: 01310100)"),
        ("", "• Datas: DD/MM/AAAA (ex: 01/03/2026)"),
        ("", "• Estado: Sigla UF com 2 letras (ex: SP, RJ, MG)"),
        ("", "• Ativo?: SIM ou NÃO"),
        ("", "• Tipo Logradouro: Selecione da lista (Rua, Avenida, Praça...)"),
        ("", "• Endereço: Nome SEM o tipo. Ex: 'das Flores' (não 'Rua das Flores')"),
        ("", "• País: Se não informado, será preenchido como 'Brasil'"),
        ("", "• Contrato: Se não informado, será preenchido como '0'"),
        ("", ""),
        ("TIPOS DE LOGRADOURO ACEITOS", None),
    ]

    row = 3
    for col_a, col_b in instrucoes:
        row += 1
        if col_b is None:
            ws.cell(row=row, column=2, value=col_a).font = SUBTITLE_FONT
        else:
            ws.cell(row=row, column=2, value=col_b).font = TEXT_FONT

    # Lista de tipos de logradouro
    row += 1
    tipos_texto = ", ".join(
        [f"{nome} ({sigla})" for sigla, nome in TIPOS_LOGRADOURO]
    )
    ws.cell(row=row, column=2, value=f"  {tipos_texto}").font = TEXT_FONT

    row += 2
    ws.cell(row=row, column=2, value="OBSERVAÇÕES IMPORTANTES").font = SUBTITLE_FONT
    row += 1
    obs = [
        "• Se o endereço já existir no sistema (mesma combinação de tipo, endereço,",
        "  número, complemento e CEP), ele será reutilizado automaticamente.",
        "• Caso contrário, um novo endereço será criado.",
        "• Erros em uma linha NÃO impedem o processamento das demais.",
        "• Após o upload, você receberá um relatório com sucessos e erros.",
    ]
    for texto in obs:
        ws.cell(row=row, column=2, value=texto).font = TEXT_FONT
        row += 1

    # Tabelas de campos
    row += 2
    ws.cell(row=row, column=2, value="CAMPOS DO CLIENTE (Colunas A–M)").font = SUBTITLE_FONT
    row += 1
    _inserir_tabela_campos(ws, row, COLUNAS_CLIENTE, "CLIENTE")

    row += len(COLUNAS_CLIENTE) + 3
    ws.cell(row=row, column=2, value="CAMPOS DO ENDEREÇO (Colunas N–Y)").font = SUBTITLE_FONT
    row += 1
    _inserir_tabela_campos(ws, row, COLUNAS_ENDERECO, "ENDEREÇO")


def _inserir_tabela_campos(ws, start_row, colunas, titulo):
    """Insere tabela descritiva dos campos na aba de instruções."""
    headers = ["Campo", "Obrigatório", "Descrição/Formato"]
    for j, h in enumerate(headers, start=2):
        cell = ws.cell(row=start_row, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGNMENT

    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 50

    for i, (header, _, _, obrig, dica) in enumerate(colunas, start=1):
        r = start_row + i
        ws.cell(row=r, column=2, value=header.replace("*", "")).font = TEXT_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER

        obrig_cell = ws.cell(row=r, column=3, value="SIM" if obrig else "NÃO")
        obrig_cell.font = Font(
            name="Calibri",
            size=10,
            bold=obrig,
            color="C00000" if obrig else "548235",
        )
        obrig_cell.alignment = CENTER_ALIGNMENT
        obrig_cell.border = THIN_BORDER

        ws.cell(row=r, column=4, value=dica).font = TEXT_FONT
        ws.cell(row=r, column=4).border = THIN_BORDER


def _criar_aba_dados(ws):
    """Cria a aba principal de preenchimento de dados."""
    from openpyxl.styles.numbers import FORMAT_TEXT

    for i, (header, campo, _, obrig, dica) in enumerate(TODAS_COLUNAS, start=1):
        col_letter = get_column_letter(i)

        # Dica na linha 1
        hint_cell = ws.cell(row=1, column=i, value=f"💡 {dica}")
        hint_cell.font = INSTRUCAO_FONT
        hint_cell.fill = INSTRUCAO_FILL
        hint_cell.alignment = WRAP_ALIGNMENT
        hint_cell.protection = Protection(locked=True)

        # Header na linha 2
        header_cell = ws.cell(row=2, column=i, value=header)
        header_cell.font = HEADER_FONT
        header_cell.fill = HEADER_FILL
        header_cell.alignment = CENTER_ALIGNMENT
        header_cell.border = THIN_BORDER
        header_cell.protection = Protection(locked=True)

        # Largura da coluna
        ws.column_dimensions[col_letter].width = TODAS_COLUNAS[i - 1][2]

    # =====================================================================
    # IDENTIFICAR COLUNAS QUE DEVEM SER TEXTO (CEP, CNPJ, TELEFONE, etc.)
    # =====================================================================
    campos_texto = {"cep", "cnpj", "telefone", "contrato", "inscricao_estadual", "inscricao_municipal"}

    indices_texto = set()
    for i, (_, campo, _, _, _) in enumerate(TODAS_COLUNAS, start=1):
        if campo in campos_texto:
            indices_texto.add(i)

    # Formatação das células de dados (linhas 3 a 502)
    for row in range(3, 503):
        for i, (_, campo, _, obrig, _) in enumerate(TODAS_COLUNAS, start=1):
            cell = ws.cell(row=row, column=i)
            cell.fill = REQUIRED_FILL if obrig else OPTIONAL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            cell.protection = Protection(locked=False)  # ← DESBLOQUEADA

            # ✅ Forçar formato TEXTO para CEP, CNPJ, Telefone, etc.
            if i in indices_texto:
                cell.number_format = FORMAT_TEXT

    # --- DATA VALIDATIONS ---

    # Tipo Logradouro (dropdown)
    idx_tipo = len(COLUNAS_CLIENTE) + [c[1] for c in COLUNAS_ENDERECO].index("tipo_logradouro") + 1
    col_tipo = get_column_letter(idx_tipo)
    nomes_tipos_lista = [nome for _, nome in TIPOS_LOGRADOURO]
    dv_tipo = DataValidation(
        type="list",
        formula1=f'"{ ",".join(nomes_tipos_lista) }"',
        allow_blank=False,
    )
    dv_tipo.error = "Selecione um tipo de logradouro válido"
    dv_tipo.errorTitle = "Tipo inválido"
    dv_tipo.prompt = "Selecione: Rua, Avenida, Praça..."
    dv_tipo.promptTitle = "Tipo de Logradouro"
    dv_tipo.showInputMessage = True
    dv_tipo.showErrorMessage = True
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f"{col_tipo}3:{col_tipo}502")

    # Estado (dropdown UF)
    idx_estado = len(COLUNAS_CLIENTE) + [c[1] for c in COLUNAS_ENDERECO].index("estado") + 1
    col_estado = get_column_letter(idx_estado)
    dv_estado = DataValidation(
        type="list",
        formula1=f'"{ ",".join(SIGLAS_ESTADOS) }"',
        allow_blank=True,
    )
    dv_estado.error = "Selecione uma UF válida"
    dv_estado.errorTitle = "Estado inválido"
    dv_estado.prompt = "Selecione o estado (UF)"
    dv_estado.promptTitle = "Estado"
    dv_estado.showInputMessage = True
    dv_estado.showErrorMessage = True
    ws.add_data_validation(dv_estado)
    dv_estado.add(f"{col_estado}3:{col_estado}502")

    # Ativo? (SIM/NÃO)
    idx_estatus = [c[1] for c in COLUNAS_CLIENTE].index("estatus") + 1
    col_estatus = get_column_letter(idx_estatus)
    dv_ativo = DataValidation(
        type="list",
        formula1='"SIM,NÃO"',
        allow_blank=False,
    )
    dv_ativo.error = "Selecione SIM ou NÃO"
    dv_ativo.errorTitle = "Valor inválido"
    dv_ativo.prompt = "SIM = Ativo / NÃO = Inativo"
    dv_ativo.promptTitle = "Status"
    dv_ativo.showInputMessage = True
    dv_ativo.showErrorMessage = True
    ws.add_data_validation(dv_ativo)
    dv_ativo.add(f"{col_estatus}3:{col_estatus}502")

    # Congelar painéis
    ws.freeze_panes = "A3"

    # =====================================================================
    # ✅ PROTEÇÃO: Apenas linhas 1 e 2 bloqueadas, dados editáveis
    # =====================================================================
    ws.protection.sheet = True
    ws.protection.password = "modelo"
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.enable()

    # Alturas
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 25


    # Formatação das células de dados (linhas 3 a 502)
    for row in range(3, 503):
        for i, (_, _, _, obrig, _) in enumerate(TODAS_COLUNAS, start=1):
            cell = ws.cell(row=row, column=i)
            cell.fill = REQUIRED_FILL if obrig else OPTIONAL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            cell.protection = Protection(locked=False)

    # --- DATA VALIDATIONS ---

    # Tipo Logradouro (dropdown com nomes legíveis)
    idx_tipo = len(COLUNAS_CLIENTE) + [c[1] for c in COLUNAS_ENDERECO].index("tipo_logradouro") + 1
    col_tipo = get_column_letter(idx_tipo)
    nomes_tipos_lista = [nome for _, nome in TIPOS_LOGRADOURO]
    dv_tipo = DataValidation(
        type="list",
        formula1=f'"{ ",".join(nomes_tipos_lista) }"',
        allow_blank=False,
    )
    dv_tipo.error = "Selecione um tipo de logradouro válido"
    dv_tipo.errorTitle = "Tipo inválido"
    dv_tipo.prompt = "Selecione: Rua, Avenida, Praça..."
    dv_tipo.promptTitle = "Tipo de Logradouro"
    dv_tipo.showInputMessage = True
    dv_tipo.showErrorMessage = True
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f"{col_tipo}3:{col_tipo}502")

    # Estado (dropdown UF)
    idx_estado = len(COLUNAS_CLIENTE) + [c[1] for c in COLUNAS_ENDERECO].index("estado") + 1
    col_estado = get_column_letter(idx_estado)
    dv_estado = DataValidation(
        type="list",
        formula1=f'"{ ",".join(SIGLAS_ESTADOS) }"',
        allow_blank=True,
    )
    dv_estado.error = "Selecione uma UF válida"
    dv_estado.errorTitle = "Estado inválido"
    dv_estado.prompt = "Selecione o estado (UF)"
    dv_estado.promptTitle = "Estado"
    dv_estado.showInputMessage = True
    dv_estado.showErrorMessage = True
    ws.add_data_validation(dv_estado)
    dv_estado.add(f"{col_estado}3:{col_estado}502")

    # Ativo? (SIM/NÃO)
    idx_estatus = [c[1] for c in COLUNAS_CLIENTE].index("estatus") + 1
    col_estatus = get_column_letter(idx_estatus)
    dv_ativo = DataValidation(
        type="list",
        formula1='"SIM,NÃO"',
        allow_blank=False,
    )
    dv_ativo.error = "Selecione SIM ou NÃO"
    dv_ativo.errorTitle = "Valor inválido"
    dv_ativo.prompt = "SIM = Ativo / NÃO = Inativo"
    dv_ativo.promptTitle = "Status"
    dv_ativo.showInputMessage = True
    dv_ativo.showErrorMessage = True
    ws.add_data_validation(dv_ativo)
    dv_ativo.add(f"{col_estatus}3:{col_estatus}502")

    # Congelar painéis
    ws.freeze_panes = "A3"

    # Proteção
    ws.protection.sheet = True
    ws.protection.password = "modelo"
    ws.protection.enable()

    # Alturas
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 25


def _criar_aba_referencia(ws):
    """Cria aba com listas de referência (Estados + Tipos de Logradouro)."""
    # --- TIPOS DE LOGRADOURO ---
    ws.cell(row=1, column=1, value="TIPOS DE LOGRADOURO").font = SUBTITLE_FONT
    ws.merge_cells("A1:B1")

    ws.cell(row=2, column=1, value="Sigla").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.cell(row=2, column=2, value="Nome").font = HEADER_FONT
    ws.cell(row=2, column=2).fill = HEADER_FILL

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20

    for i, (sigla, nome) in enumerate(TIPOS_LOGRADOURO, start=3):
        ws.cell(row=i, column=1, value=sigla).font = TEXT_FONT
        ws.cell(row=i, column=1).alignment = CENTER_ALIGNMENT
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2, value=nome).font = TEXT_FONT
        ws.cell(row=i, column=2).border = THIN_BORDER

    # --- ESTADOS ---
    col_offset = 4  # Coluna D
    ws.cell(row=1, column=col_offset, value="ESTADOS DO BRASIL").font = SUBTITLE_FONT
    ws.merge_cells(
        start_row=1,
        start_column=col_offset,
        end_row=1,
        end_column=col_offset + 1,
    )

    ws.cell(row=2, column=col_offset, value="Sigla").font = HEADER_FONT
    ws.cell(row=2, column=col_offset).fill = HEADER_FILL
    ws.cell(row=2, column=col_offset + 1, value="Estado").font = HEADER_FONT
    ws.cell(row=2, column=col_offset + 1).fill = HEADER_FILL

    ws.column_dimensions[get_column_letter(col_offset)].width = 10
    ws.column_dimensions[get_column_letter(col_offset + 1)].width = 25

    for i, (sigla, nome) in enumerate(ESTADOS_BRASIL, start=3):
        ws.cell(row=i, column=col_offset, value=sigla).font = TEXT_FONT
        ws.cell(row=i, column=col_offset).alignment = CENTER_ALIGNMENT
        ws.cell(row=i, column=col_offset).border = THIN_BORDER
        ws.cell(row=i, column=col_offset + 1, value=nome).font = TEXT_FONT
        ws.cell(row=i, column=col_offset + 1).border = THIN_BORDER


# ============================================================================
# PROCESSAMENTO DA PLANILHA ENVIADA
# ============================================================================


def processar_planilha(arquivo, filial):
    """
    Processa planilha Excel enviada e cria Logradouros + Clientes.

    Args:
        arquivo: InMemoryUploadedFile (arquivo .xlsx)
        filial: Instância de Filial do usuário logado

    Returns:
        dict com total, sucessos, erros, detalhes_sucesso, detalhes_erro
    """
    resultado = {
        "total": 0,
        "sucessos": 0,
        "erros": 0,
        "detalhes_sucesso": [],
        "detalhes_erro": [],
    }

    try:
        wb = load_workbook(arquivo, data_only=True)
    except Exception:
        resultado["detalhes_erro"].append(
            {"linha": 0, "erros": ["Arquivo inválido. Envie um arquivo .xlsx válido."]}
        )
        return resultado

    # Procura aba "Dados Clientes"
    if "Dados Clientes" in wb.sheetnames:
        ws = wb["Dados Clientes"]
    else:
        ws = wb.worksheets[0]

    start_row = 3  # Linha 1=dica, 2=header, 3+=dados
    total_colunas = len(TODAS_COLUNAS)

    for row_idx in range(start_row, ws.max_row + 1):
        valores = []
        for col_idx in range(1, total_colunas + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            valores.append(cell_value)

        # Pular linhas vazias
        if all(v is None or str(v).strip() == "" for v in valores):
            continue

        resultado["total"] += 1
        erros_linha = []

        # Mapear valores
        dados_cliente = {}
        for i, (_, campo, _, _, _) in enumerate(COLUNAS_CLIENTE):
            dados_cliente[campo] = valores[i]

        dados_endereco = {}
        offset = len(COLUNAS_CLIENTE)
        for i, (_, campo, _, _, _) in enumerate(COLUNAS_ENDERECO):
            dados_endereco[campo] = valores[offset + i]

        # Validações
        erros_linha.extend(_validar_cliente(dados_cliente))
        erros_linha.extend(_validar_endereco(dados_endereco))

        if erros_linha:
            resultado["erros"] += 1
            resultado["detalhes_erro"].append(
                {"linha": row_idx, "erros": erros_linha}
            )
            continue

        # Criar registros
        try:
            with transaction.atomic():
                logradouro = _criar_ou_buscar_logradouro(dados_endereco, filial)
                cliente = _criar_cliente(dados_cliente, logradouro, filial)
                resultado["sucessos"] += 1
                resultado["detalhes_sucesso"].append(
                    f"Linha {row_idx}: {cliente.razao_social} "
                    f"(CNPJ: {cliente.cnpj}) — importado com sucesso."
                )
        except Exception as e:
            resultado["erros"] += 1
            resultado["detalhes_erro"].append(
                {"linha": row_idx, "erros": [f"Erro ao salvar: {str(e)}"]}
            )

    return resultado


# ============================================================================
# VALIDAÇÕES
# ============================================================================


def _validar_cliente(dados):
    """Valida campos do cliente e retorna lista de erros."""
    erros = []

    if not dados.get("razao_social") or str(dados["razao_social"]).strip() == "":
        erros.append("Razão Social é obrigatória.")

    if not dados.get("nome") or str(dados["nome"]).strip() == "":
        erros.append("Nome Fantasia é obrigatório.")

    # CNPJ
    cnpj_raw = dados.get("cnpj")
    if not cnpj_raw or str(cnpj_raw).strip() == "":
        erros.append("CNPJ é obrigatório.")
    else:
        cnpj_str = str(cnpj_raw).strip()
        digitos = re.sub(r"\D", "", cnpj_str)
        if len(digitos) != 14:
            erros.append(f"CNPJ '{cnpj_str}' inválido. Deve conter 14 dígitos.")

    # Data de início
    data_inicio = dados.get("data_de_inicio")
    if not data_inicio:
        erros.append("Data de Início é obrigatória.")
    else:
        parsed = _parse_data(data_inicio)
        if parsed is None:
            erros.append(f"Data de Início '{data_inicio}' inválida. Use DD/MM/AAAA.")

    # Data encerramento (opcional)
    data_enc = dados.get("data_encerramento")
    if data_enc and str(data_enc).strip() != "":
        parsed = _parse_data(data_enc)
        if parsed is None:
            erros.append(
                f"Data de Encerramento '{data_enc}' inválida. Use DD/MM/AAAA."
            )

    # Estatus
    estatus = dados.get("estatus")
    if not estatus or str(estatus).strip() == "":
        erros.append("Campo 'Ativo?' é obrigatório (SIM ou NÃO).")
    elif str(estatus).strip().upper() not in ("SIM", "NÃO", "NAO"):
        erros.append(f"Campo 'Ativo?' deve ser SIM ou NÃO. Valor: '{estatus}'.")

    # Contrato
    contrato = dados.get("contrato")
    if contrato and len(str(contrato).strip()) > 4:
        erros.append(f"Contrato deve ter no máximo 4 caracteres. Valor: '{contrato}'.")

    # Unidade
    unidade = dados.get("unidade")
    if unidade is not None and str(unidade).strip() != "":
        try:
            val = int(float(str(unidade)))
            if val < 0:
                erros.append("Unidade deve ser um número positivo.")
        except (ValueError, TypeError):
            erros.append(f"Unidade deve ser um número inteiro. Valor: '{unidade}'.")

    # E-mail
    email = dados.get("email")
    if email and str(email).strip() != "":
        email_str = str(email).strip()
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email_str):
            erros.append(f"E-mail inválido: '{email_str}'.")

    return erros


def _validar_endereco(dados):
    """Valida campos do endereço e retorna lista de erros."""
    erros = []

    # Tipo Logradouro
    tipo = dados.get("tipo_logradouro")
    if not tipo or str(tipo).strip() == "":
        erros.append("Tipo de Logradouro é obrigatório.")
    else:
        tipo_str = str(tipo).strip().upper()
        if tipo_str not in SIGLAS_TIPOS and tipo_str not in NOMES_TIPOS:
            erros.append(
                f"Tipo de Logradouro '{tipo}' inválido. "
                f"Use: {', '.join([n for _, n in TIPOS_LOGRADOURO])}"
            )

    if not dados.get("endereco") or str(dados["endereco"]).strip() == "":
        erros.append("Endereço é obrigatório.")

    # Número
    numero = dados.get("numero")
    if numero is None or str(numero).strip() == "":
        erros.append("Número do endereço é obrigatório.")
    else:
        try:
            val = int(float(str(numero)))
            if val < 1:
                erros.append("Número deve ser no mínimo 1.")
        except (ValueError, TypeError):
            erros.append(f"Número deve ser inteiro. Valor: '{numero}'.")

    # ✅ CEP — Tratar zero à esquerda
    cep = dados.get("cep")
    if not cep or str(cep).strip() == "":
        erros.append("CEP é obrigatório.")
    else:
        cep_digitos = re.sub(r"\D", "", str(cep).strip())
        # Recompor zero à esquerda se Excel removeu
        cep_digitos = cep_digitos.zfill(8)
        if len(cep_digitos) != 8:
            erros.append(f"CEP deve ter 8 dígitos. Valor: '{cep}'.")

    if not dados.get("bairro") or str(dados["bairro"]).strip() == "":
        erros.append("Bairro é obrigatório.")

    if not dados.get("cidade") or str(dados["cidade"]).strip() == "":
        erros.append("Cidade é obrigatória.")

    # Estado
    estado = dados.get("estado")
    if not estado or str(estado).strip() == "":
        erros.append("Estado é obrigatório.")
    elif str(estado).strip().upper() not in SIGLAS_ESTADOS:
        erros.append(f"Estado '{estado}' inválido. Use a sigla UF (ex: SP).")

    return erros


# ============================================================================
# CRIAÇÃO DE REGISTROS
# ============================================================================


def _resolver_tipo_logradouro(valor):
    """
    Converte o valor informado na planilha para a sigla do tipo_logradouro.

    Aceita tanto sigla (RUA, AV) quanto nome (Rua, Avenida).
    """
    val = str(valor).strip().upper()

    # Já é uma sigla válida
    if val in SIGLAS_TIPOS:
        return val

    # É um nome → converte para sigla
    if val in NOMES_TIPOS:
        return NOMES_TIPOS[val]

    # Fallback
    return "RUA"


def _criar_ou_buscar_logradouro(dados, filial):
    """
    Busca logradouro existente ou cria novo.
    """
    tipo_logradouro = _resolver_tipo_logradouro(dados["tipo_logradouro"])
    endereco = str(dados["endereco"]).strip()
    numero = int(float(str(dados["numero"])))

    # ✅ CEP — Recompor zero à esquerda
    cep = re.sub(r"\D", "", str(dados["cep"]).strip()).zfill(8)

    complemento = (
        str(dados["complemento"]).strip()
        if dados.get("complemento") and str(dados["complemento"]).strip() != ""
        else None
    )
    bairro = str(dados["bairro"]).strip()
    cidade = str(dados["cidade"]).strip()
    estado = str(dados["estado"]).strip().upper()
    pais = (
        str(dados["pais"]).strip()
        if dados.get("pais") and str(dados["pais"]).strip() != ""
        else "Brasil"
    )
    ponto_referencia = (
        str(dados["ponto_referencia"]).strip()
        if dados.get("ponto_referencia")
        and str(dados["ponto_referencia"]).strip() != ""
        else None
    )
    latitude = _parse_decimal(dados.get("latitude"))
    longitude = _parse_decimal(dados.get("longitude"))

    # Busca existente
    try:
        logradouro = Logradouro.objects.model.objects.filter(
            tipo_logradouro=tipo_logradouro,
            endereco__iexact=endereco,
            numero=numero,
            complemento=complemento,
            cep=cep,
            filial=filial,
        ).first()
    except Exception:
        logradouro = None

    if logradouro:
        return logradouro

    logradouro = Logradouro(
        tipo_logradouro=tipo_logradouro,
        endereco=endereco,
        numero=numero,
        cep=cep,
        complemento=complemento,
        bairro=bairro,
        cidade=cidade,
        estado=estado,
        pais=pais,
        ponto_referencia=ponto_referencia,
        latitude=latitude,
        longitude=longitude,
        filial=filial,
    )
    logradouro.save()
    return logradouro


def _criar_cliente(dados, logradouro, filial):
    """Cria instância de Cliente com os dados validados."""
    cnpj_str = str(dados["cnpj"]).strip()
    digitos = re.sub(r"\D", "", cnpj_str)
    cnpj_formatado = (
        f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}"
        f"/{digitos[8:12]}-{digitos[12:]}"
    )

    contrato = str(dados.get("contrato") or "0").strip()[:4]

    unidade = None
    if dados.get("unidade") and str(dados["unidade"]).strip() != "":
        unidade = int(float(str(dados["unidade"])))

    telefone = (
        str(dados["telefone"]).strip()
        if dados.get("telefone") and str(dados["telefone"]).strip() != ""
        else None
    )
    email = (
        str(dados["email"]).strip()
        if dados.get("email") and str(dados["email"]).strip() != ""
        else None
    )

    data_inicio = _parse_data(dados["data_de_inicio"])
    data_enc = (
        _parse_data(dados["data_encerramento"])
        if dados.get("data_encerramento")
        and str(dados["data_encerramento"]).strip() != ""
        else None
    )

    estatus_str = str(dados["estatus"]).strip().upper()
    estatus = estatus_str == "SIM"

    inscricao_estadual = (
        str(dados["inscricao_estadual"]).strip()
        if dados.get("inscricao_estadual")
        and str(dados["inscricao_estadual"]).strip() != ""
        else None
    )
    inscricao_municipal = (
        str(dados["inscricao_municipal"]).strip()
        if dados.get("inscricao_municipal")
        and str(dados["inscricao_municipal"]).strip() != ""
        else None
    )
    observacoes = (
        str(dados["observacoes"]).strip()
        if dados.get("observacoes") and str(dados["observacoes"]).strip() != ""
        else None
    )

    cliente = Cliente(
        razao_social=str(dados["razao_social"]).strip(),
        nome=str(dados["nome"]).strip(),
        cnpj=cnpj_formatado,
        contrato=contrato,
        unidade=unidade,
        inscricao_estadual=inscricao_estadual,
        inscricao_municipal=inscricao_municipal,
        telefone=telefone,
        email=email,
        logradouro=logradouro,
        data_de_inicio=data_inicio,
        data_encerramento=data_enc,
        estatus=estatus,
        observacoes=observacoes,
        filial=filial,
    )
    cliente.full_clean()
    cliente.save()
    return cliente


# ============================================================================
# UTILITÁRIOS
# ============================================================================


def _parse_data(valor):
    """Converte valor para date. Aceita datetime, date, ou string DD/MM/AAAA."""
    if valor is None:
        return None

    from datetime import date as date_type

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date_type):
        return valor

    valor_str = str(valor).strip()
    formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"]
    for fmt in formatos:
        try:
            return datetime.strptime(valor_str, fmt).date()
        except ValueError:
            continue

    return None


def _parse_decimal(valor):
    """Converte valor para Decimal ou retorna None."""
    if valor is None or str(valor).strip() == "":
        return None
    try:
        from decimal import Decimal

        return Decimal(str(valor).strip().replace(",", "."))
    except Exception:
        return None

