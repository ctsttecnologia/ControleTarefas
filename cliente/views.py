
# cliente/views.py

"""
Views do app Cliente.

Arquitetura de permissões (em camadas):
┌──────────────────────────────────────────────────────────────────────┐
│ Camada 1: LoginRequiredMixin                                         │
│   → Autenticação obrigatória                                         │
│                                                                      │
│ Camada 2: AppPermissionMixin                                         │
│   → Permissão do app 'cliente' (bloqueia acesso sem módulo liberado) │
│                                                                      │
│ Camada 3: permission_required (Django)                              │
│   → Permissão granular (view/add/change/delete_cliente)              │
│                                                                      │
│ Camada 4: ViewFilialScopedMixin                                      │
│   → Filtra por filial ativa via FilialManager (for_request)          │
│                                                                      │
│ Camada 5: ClienteVisibilityMixin                                     │
│   → Filtra por perfil (superuser / perm global / vínculo funcional) │
└──────────────────────────────────────────────────────────────────────┘
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import ObjectDoesNotExist
from django.db import models as db_models
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView, DeleteView, DetailView, ListView, UpdateView,
)

from core.mixins import AppPermissionMixin, ViewFilialScopedMixin
from core.decorators import app_permission_required
from logradouro.models import Logradouro
from usuario.models import Filial

from .forms import ClienteForm, ImportacaoMassaForm
from .models import Cliente
from .services.importacao_massa import gerar_planilha_modelo, processar_planilha


_APP = 'cliente'


# ═══════════════════════════════════════════════════════════════════════════════
# MIXIN CENTRAL DE FILIAL
# ═══════════════════════════════════════════════════════════════════════════════

class FilialAtivaMixin:
    """
    Obtém a filial ativa com prioridade para o seletor da sessão.
    Fallbacks: user.filial_ativa → funcionario.filial.

    Independente do ViewFilialScopedMixin:
    - ViewFilialScopedMixin → filtra get_queryset() via FilialManager.for_request()
    - FilialAtivaMixin       → fornece utilitários para recuperar a filial ativa
    """

    def get_filial_ativa(self):
        filial_id = self.request.session.get('active_filial_id')

        if filial_id:
            try:
                return Filial.objects.get(pk=filial_id)
            except Filial.DoesNotExist:
                pass

        user = self.request.user
        filial_ativa = getattr(user, 'filial_ativa', None)
        if filial_ativa:
            return filial_ativa

        # Import tardio para evitar ciclos
        from departamento_pessoal.models import Funcionario
        try:
            funcionario = Funcionario.objects.select_related('filial').get(usuario=user)
            return funcionario.filial
        except Funcionario.DoesNotExist:
            pass

        return None

    def get_filial_ativa_id(self):
        filial = self.get_filial_ativa()
        return filial.id if filial else None


# ═══════════════════════════════════════════════════════════════════════════════
# MIXIN DE VISIBILIDADE
# ═══════════════════════════════════════════════════════════════════════════════

class ClienteVisibilityMixin(FilialAtivaMixin):
    """
    Controla a visibilidade dos clientes conforme o perfil do usuário.

    Regras:
    ┌──────────────────────────────────────┬────────────────────────────────────┐
    │ Perfil                               │ Visibilidade                       │
    ├──────────────────────────────────────┼────────────────────────────────────┤
    │ Superuser                            │ Todos os clientes                  │
    │ Permissão cliente.view_all_cliente   │ Todos da filial ativa              │
    │ Funcionario vinculado a Cliente      │ Apenas clientes onde atua          │
    │ Sem vínculo e sem perm global        │ Nenhum cliente                     │
    └──────────────────────────────────────┴────────────────────────────────────┘

    OBS: o vínculo usuário↔cliente é feito via Funcionario.cliente
         (descoberto no model de departamento_pessoal).
    """

    def apply_visibility(self, queryset):
        user = self.request.user

        # Superuser → tudo
        if user.is_superuser:
            return queryset

        # Permissão global → tudo da filial (já filtrada pelo FilialManager)
        if user.has_perm('cliente.view_all_cliente'):
            return queryset

        # Usuário comum → apenas clientes onde o Funcionario dele está alocado
        funcionario = getattr(user, 'funcionario', None)
        if funcionario and funcionario.cliente_id:
            return queryset.filter(pk=funcionario.cliente_id)

        # Sem vínculo → nada
        return queryset.none()


# ═══════════════════════════════════════════════════════════════════════════════
# MIXIN BASE
# ═══════════════════════════════════════════════════════════════════════════════

class ClienteBaseMixin(
    LoginRequiredMixin, AppPermissionMixin,
    ClienteVisibilityMixin, ViewFilialScopedMixin,
):
    """
    Mixin base para as CBVs de Cliente.

    MRO:
      1. LoginRequiredMixin       → autenticação
      2. AppPermissionMixin       → permissão do app 'cliente'
      3. ClienteVisibilityMixin   → apply_visibility() + FilialAtivaMixin
      4. ViewFilialScopedMixin    → get_queryset() filtrado por filial
    """
    model = Cliente
    form_class = ClienteForm
    success_url = reverse_lazy('cliente:lista_clientes')
    app_label_required = _APP


# ═══════════════════════════════════════════════════════════════════════════════
# CRUD
# ═══════════════════════════════════════════════════════════════════════════════

class ClienteListView(ClienteBaseMixin, ListView):
    permission_required = 'cliente.view_cliente'
    template_name = 'cliente/cliente_list.html'
    context_object_name = 'clientes'
    paginate_by = 10

    def dispatch(self, request, *args, **kwargs):
        """Bloqueia usuários sem Funcionario e sem permissão global."""
        if request.user.is_authenticated and not request.user.is_superuser:
            if not request.user.has_perm('cliente.view_all_cliente'):
                try:
                    _ = request.user.funcionario
                except ObjectDoesNotExist:
                    return render(request, 'cliente/acesso_negado.html', {
                        'titulo': 'Acesso Restrito',
                        'mensagem': (
                            'Sua conta não está vinculada a um registro de '
                            'funcionário, por isso não pode acessar o módulo '
                            'de Clientes.'
                        ),
                    }, status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        # Camada 1: filial (ViewFilialScopedMixin → for_request)
        queryset = super().get_queryset().select_related('logradouro')

        # Camada 2: visibilidade por perfil
        queryset = self.apply_visibility(queryset)

        # Camada 3: busca do usuário
        termo = self.request.GET.get('q', '').strip()
        if termo:
            queryset = queryset.filter(
                Q(nome__icontains=termo)
                | Q(razao_social__icontains=termo)
                | Q(cnpj__icontains=termo)
                | Q(contrato__icontains=termo)
            )

        return queryset.order_by('nome')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['termo_pesquisa'] = self.request.GET.get('q', '')
        context['filial_ativa'] = self.get_filial_ativa()
        context['can_view_all'] = (
            self.request.user.is_superuser
            or self.request.user.has_perm('cliente.view_all_cliente')
        )

        if context.get('paginator'):
            context['total_clientes'] = context['paginator'].count
        else:
            context['total_clientes'] = self.object_list.count()

        return context


class ClienteCreateView(ClienteBaseMixin, SuccessMessageMixin, CreateView):
    permission_required = 'cliente.add_cliente'
    template_name = 'cliente/cliente_form.html'
    success_message = "✅ Cliente cadastrado com sucesso!"

    def form_valid(self, form):
        # Atribui a filial ativa (prioriza seletor da sessão)
        filial_ativa = self.get_filial_ativa()

        if filial_ativa:
            form.instance.filial = filial_ativa
        elif not self.request.user.is_superuser:
            messages.error(
                self.request,
                "Nenhuma filial selecionada. Escolha uma filial no menu superior."
            )
            return self.form_invalid(form)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filial_ativa'] = self.get_filial_ativa()
        return context


class ClienteUpdateView(ClienteBaseMixin, SuccessMessageMixin, UpdateView):
    permission_required = 'cliente.change_cliente'
    template_name = 'cliente/cliente_form.html'
    success_message = "🔄 Cliente atualizado com sucesso!"

    def get_queryset(self):
        """Só edita clientes que o usuário pode ver."""
        qs = super().get_queryset()
        return self.apply_visibility(qs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filial_ativa'] = self.get_filial_ativa()
        return context


class ClienteDeleteView(ClienteBaseMixin, DeleteView):
    permission_required = 'cliente.delete_cliente'
    template_name = 'cliente/cliente_confirm_delete.html'
    success_url = reverse_lazy('cliente:lista_clientes')

    def get_queryset(self):
        """Só exclui clientes que o usuário pode ver."""
        qs = super().get_queryset()
        return self.apply_visibility(qs)

    def post(self, request, *args, **kwargs):
        """
        Controla exclusão manualmente para tratar ProtectedError
        e evitar problemas do SuccessMessageMixin com DeleteView no Django 5.x.
        """
        self.object = self.get_object()
        nome_cliente = str(self.object)

        try:
            self.object.delete()
            messages.success(
                request,
                f'🗑️ Cliente "{nome_cliente}" excluído com sucesso!'
            )
            return redirect(self.get_success_url())

        except db_models.ProtectedError:
            messages.error(
                request,
                "❌ Não foi possível excluir este cliente. "
                "Existem registros vinculados (funcionários, contratos, atas, etc.) "
                "que impedem a exclusão."
            )
            return redirect('cliente:lista_clientes')

        except Exception as e:
            messages.error(request, f"❌ Erro inesperado ao excluir: {e}")
            return redirect('cliente:lista_clientes')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filial_ativa'] = self.get_filial_ativa()
        return context


class ClienteDetailView(ClienteBaseMixin, DetailView):
    permission_required = 'cliente.view_cliente'
    template_name = 'cliente/cliente_detail.html'
    context_object_name = 'cliente'

    def get_queryset(self):
        """Só mostra detalhes de clientes que o usuário pode ver."""
        qs = super().get_queryset().prefetch_related('documentos_cliente')
        return self.apply_visibility(qs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['arquivos_cliente'] = self.object.documentos_cliente.all()
        context['filial_ativa'] = self.get_filial_ativa()
        return context


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAÇÃO
# ═══════════════════════════════════════════════════════════════════════════════

class ExportarClientesExcelView(
    LoginRequiredMixin, AppPermissionMixin,
    ClienteVisibilityMixin, ViewFilialScopedMixin, ListView,
):
    """Exporta clientes para Excel respeitando filial + visibilidade."""
    model = Cliente
    app_label_required = _APP
    permission_required = 'cliente.view_cliente'

    def get_queryset(self):
        qs = super().get_queryset().select_related('logradouro')
        qs = self.apply_visibility(qs)
        return qs.order_by('nome')

    def get(self, request, *args, **kwargs):
        clientes = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="0d6efd", end_color="0d6efd", fill_type="solid"
        )

        headers = [
            "ID", "Nome Fantasia", "Razão Social", "CNPJ", "Contrato",
            "Data de Início", "Endereço", "Telefone", "Email", "Status",
        ]

        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_num)].width = 25

        for row_num, cliente in enumerate(clientes, 2):
            ws.cell(row=row_num, column=1, value=cliente.pk)
            ws.cell(row=row_num, column=2, value=cliente.nome)
            ws.cell(row=row_num, column=3, value=cliente.razao_social)
            ws.cell(row=row_num, column=4, value=cliente.cnpj_formatado)
            ws.cell(row=row_num, column=5, value=cliente.contrato)
            ws.cell(row=row_num, column=6, value=cliente.data_de_inicio)
            ws.cell(
                row=row_num, column=7,
                value=str(cliente.logradouro) if cliente.logradouro else "-"
            )
            ws.cell(row=row_num, column=8, value=cliente.telefone)
            ws.cell(row=row_num, column=9, value=cliente.email)
            ws.cell(
                row=row_num, column=10,
                value="Ativo" if cliente.estatus else "Inativo"
            )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
        wb.save(response)
        return response


# ═══════════════════════════════════════════════════════════════════════════════
# VIEWS AJAX / AUTOCOMPLETE
# ═══════════════════════════════════════════════════════════════════════════════

def _aplicar_visibilidade_cliente_fbv(request, queryset):
    """Helper para aplicar visibilidade em FBVs (replica ClienteVisibilityMixin)."""
    user = request.user

    if user.is_superuser:
        return queryset

    if user.has_perm('cliente.view_all_cliente'):
        return queryset

    funcionario = getattr(user, 'funcionario', None)
    if funcionario and funcionario.cliente_id:
        return queryset.filter(pk=funcionario.cliente_id)

    return queryset.none()


@login_required
@app_permission_required(_APP)
def cliente_autocomplete_view(request):
    """
    Retorna clientes para o Select2 respeitando filial + visibilidade.
    GET /cliente/autocomplete/?term=termo
    """
    term = request.GET.get('term', '').strip()

    # Camada 1: filial (FilialManager)
    qs = Cliente.objects.for_request(request)

    # Camada 2: visibilidade por perfil
    qs = _aplicar_visibilidade_cliente_fbv(request, qs)

    # Camada 3: busca
    if term:
        qs = qs.filter(
            Q(razao_social__icontains=term) | Q(nome__icontains=term)
        )

    clientes = qs.values('id', 'razao_social')[:10]
    return JsonResponse(list(clientes), safe=False)


@login_required
@app_permission_required(_APP)
def ajax_buscar_logradouros(request):
    """
    Autocomplete TomSelect de logradouros.
    GET /cliente/ajax/logradouros/?q=termo
    """
    q = request.GET.get("q", "").strip()

    if len(q) < 2:
        return JsonResponse([], safe=False)

    qs = Logradouro.objects.filter(
        Q(endereco__icontains=q)
        | Q(bairro__icontains=q)
        | Q(cidade__icontains=q)
        | Q(cep__icontains=q)
    )[:20]

    results = [
        {
            "id": log.pk,
            "text": (
                f"{log.endereco}, {log.numero} - {log.bairro} - "
                f"{log.cidade}/{log.estado} - CEP: {log.cep}"
            ),
        }
        for log in qs
    ]
    return JsonResponse(results, safe=False)


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTAÇÃO EM MASSA
# ═══════════════════════════════════════════════════════════════════════════════

def _resolver_filial_ativa(request):
    """Helper para FBVs: prioriza sessão, depois user.filial_ativa."""
    filial_id = request.session.get('active_filial_id')
    if filial_id:
        try:
            return Filial.objects.get(pk=filial_id)
        except Filial.DoesNotExist:
            pass
    return getattr(request.user, 'filial_ativa', None)


@login_required
@app_permission_required(_APP)
def download_modelo_view(request):
    """Gera e retorna a planilha modelo para download."""
    buffer = gerar_planilha_modelo()
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="modelo_importacao_clientes.xlsx"'
    )
    return response


@login_required
@app_permission_required(_APP)
def importacao_massa_view(request):
    """Upload e processamento da planilha de clientes."""
    # Exige permissão granular de criação
    if not request.user.has_perm('cliente.add_cliente'):
        messages.error(
            request,
            "Você não tem permissão para importar clientes em massa."
        )
        return redirect('cliente:lista_clientes')

    filial = _resolver_filial_ativa(request)

    if request.method == "POST":
        form = ImportacaoMassaForm(request.POST, request.FILES)
        if form.is_valid():
            if not filial and not request.user.is_superuser:
                messages.error(
                    request,
                    "Nenhuma filial selecionada. Escolha uma filial no menu superior."
                )
                return redirect(request.path)

            arquivo = form.cleaned_data["arquivo"]
            resultado = processar_planilha(arquivo, filial)

            return render(
                request,
                "cliente/importacao_massa_resultado.html",
                {"resultado": resultado, "form": ImportacaoMassaForm()},
            )
    else:
        form = ImportacaoMassaForm()

    return render(request, "cliente/importacao_massa.html", {
        "form": form,
        "filial_ativa": filial,
    })

