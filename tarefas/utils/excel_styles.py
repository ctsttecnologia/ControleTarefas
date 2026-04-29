
"""
Estilos padronizados para exportação Excel — mesma identidade visual dos PDFs.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===== CORES (mesmas do PDF) =====
COR_PRIMARIA = "0D6EFD"      # Azul corporativo
COR_HEADER_BG = "0D6EFD"
COR_HEADER_FG = "FFFFFF"
COR_ZEBRA = "F8F9FA"
COR_BORDA = "DEE2E6"

# Status
COR_PENDENTE  = "FFF3CD"
COR_ANDAMENTO = "CFE2FF"
COR_CONCLUIDA = "D1E7DD"
COR_ATRASADA  = "F8D7DA"

# ===== ESTILOS =====
FONTE_TITULO = Font(name="Calibri", size=16, bold=True, color=COR_PRIMARIA)
FONTE_SUBTITULO = Font(name="Calibri", size=10, color="6C757D")
FONTE_HEADER = Font(name="Calibri", size=10, bold=True, color=COR_HEADER_FG)
FONTE_BODY = Font(name="Calibri", size=10, color="212529")
FONTE_BOLD = Font(name="Calibri", size=10, bold=True, color="212529")

FILL_HEADER = PatternFill("solid", fgColor=COR_HEADER_BG)
FILL_ZEBRA = PatternFill("solid", fgColor=COR_ZEBRA)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

_thin = Side(style="thin", color=COR_BORDA)
BORDA_PADRAO = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def aplicar_cabecalho_relatorio(ws, titulo, subtitulo, data_emissao, num_colunas):
    """Aplica cabeçalho institucional padrão na planilha."""
    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_colunas)
    cell = ws.cell(row=1, column=1, value=titulo)
    cell.font = FONTE_TITULO
    cell.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 28

    # Subtítulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_colunas)
    cell = ws.cell(row=2, column=1, value=subtitulo)
    cell.font = FONTE_SUBTITULO
    cell.alignment = ALIGN_LEFT

    # Data emissão
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_colunas)
    cell = ws.cell(row=3, column=1, value=f"Emitido em: {data_emissao}")
    cell.font = FONTE_SUBTITULO
    cell.alignment = ALIGN_RIGHT

    return 5  # Linha onde a tabela começa


def aplicar_estilo_tabela(ws, header_row, total_rows, total_cols):
    """Aplica zebra, bordas e formatação na tabela."""
    # Cabeçalho
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = FONTE_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDA_PADRAO
    ws.row_dimensions[header_row].height = 26

    # Corpo
    for row in range(header_row + 1, header_row + total_rows + 1):
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = FONTE_BODY
            cell.border = BORDA_PADRAO
            if (row - header_row) % 2 == 0:
                cell.fill = FILL_ZEBRA

    # Auto-width simples
    for col in range(1, total_cols + 1):
        letra = get_column_letter(col)
        max_len = 12
        for row in ws[letra]:
            if row.value:
                max_len = max(max_len, min(len(str(row.value)) + 2, 50))
        ws.column_dimensions[letra].width = max_len

    # Congelar cabeçalho
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

