
# tarefas/views.py

import json
import logging
from datetime import timedelta
from multiprocessing import context
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Q, Count
from django.db.models.functions import TruncWeek
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
)
from core.mixins import FuncionarioRequiredMixin, ViewFilialScopedMixin, TarefaAccessMixin, AppPermissionMixin
from .forms import TarefaForm, ComentarioForm
from .models import HistoricoTarefa, Tarefas
from .services import (
    preparar_contexto_relatorio,
    gerar_pdf_relatorio,
    gerar_csv_relatorio,
    gerar_docx_relatorio,
    registrar_alteracao_status,
)
from notifications.services import notificar_tarefa_criada, notificar_tarefa_comentario
from openpyxl import Workbook
from django.http import HttpResponse
from .utils.excel_styles import (
    aplicar_cabecalho_relatorio, aplicar_estilo_tabela
)

User = get_user_model()
logger = logging.getLogger(__name__)
_APP = 'tarefas'


# =============================================================================
# HELPER — Filtro de visibilidade reutilizável
# =============================================================================

def aplicar_filtro_visibilidade(queryset, user):
    """
    Aplica o filtro de visibilidade ao queryset:
    - Superusuário ou quem tem 'tarefas.view_all_tarefas': vê tudo
    - Demais: só tarefas onde é responsável, participante ou criador
    """
    if user.is_superuser or user.has_perm('tarefas.view_all_tarefas'):
        return queryset
    return queryset.filter(
        Q(responsavel=user) | Q(participantes=user) | Q(usuario=user)
    ).distinct()



class TarefasBaseMixin(
    FuncionarioRequiredMixin,
    AppPermissionMixin,
    ViewFilialScopedMixin,
    TarefaAccessMixin,
):
    """Mixin base para todas as views de Tarefas."""
    app_name = 'tarefas'
    modulo_nome = 'Tarefas'

# =============================================================================
# CRUD
# =============================================================================

class TarefaListView(TarefasBaseMixin, ListView):

    model = Tarefas
    template_name = 'tarefas/listar_tarefas.html'
    context_object_name = 'object_list'
    paginate_by = 20

    def _get_base_queryset(self):
        """
        Queryset base com escopo de filial + filtro de visibilidade.
        Reutilizado no get_queryset e no get_context_data.
        """
        qs = super().get_queryset()
        return aplicar_filtro_visibilidade(qs, self.request.user)

    def get_queryset(self):
        qs = self._get_base_queryset()

        # --- Filtros da URL ---
        status     = self.request.GET.get('status', '')
        projeto    = self.request.GET.get('projeto', '')
        query      = self.request.GET.get('q', '')
        responsavel = self.request.GET.get('responsavel')

        if status:
            qs = qs.filter(status=status)

        if projeto:
            qs = qs.filter(projeto=projeto)

        if responsavel:
            qs = qs.filter(responsavel_id=responsavel)

        if query:
            qs = qs.filter(
                Q(titulo__icontains=query)                         |
                Q(descricao__icontains=query)                      |
                Q(projeto__icontains=query)                        |
                # Responsável
                Q(responsavel__first_name__icontains=query)        |
                Q(responsavel__last_name__icontains=query)         |
                Q(responsavel__username__icontains=query)          |
                # Participantes (M2M)
                Q(participantes__first_name__icontains=query)      |
                Q(participantes__last_name__icontains=query)       |
                Q(participantes__username__icontains=query)
            ).distinct()

        return (
            qs
            .select_related('usuario', 'responsavel', 'filial')
            .prefetch_related('participantes')
            .order_by('-prazo', 'prioridade')
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        base_qs = self._get_base_queryset()

        # Estatísticas
        context['total_tarefas']      = base_qs.count()
        context['tarefas_concluidas'] = base_qs.filter(status='concluida').count()
        context['tarefas_pendentes']  = base_qs.exclude(
            status__in=['concluida', 'cancelada']
        ).count()

        # Opções dos filtros (choices fixos do model)
        context['status_options']     = Tarefas.STATUS_CHOICES
        context['prioridade_options'] = Tarefas.PRIORIDADE_CHOICES

        # 🆕 Opções de projeto — gerado dinamicamente dos valores existentes
        projetos_distintos = (
            base_qs
            .exclude(projeto__isnull=True)
            .exclude(projeto__exact='')
            .values_list('projeto', flat=True)
            .distinct()
            .order_by('projeto')
        )
        context['projeto_options'] = [(p, p) for p in projetos_distintos]

        # Responsáveis ativos
        context['responsaveis'] = User.objects.filter(
            is_active=True
        ).order_by('first_name', 'last_name')

        # Valores atuais dos filtros (para manter selecionados após submit)
        context['responsavel_atual'] = self.request.GET.get('responsavel', '')
        context['status_atual']      = self.request.GET.get('status', '')
        context['projeto_atual']     = self.request.GET.get('projeto', '')
        context['query_atual']       = self.request.GET.get('q', '')

        return context


class TarefaDetailView(TarefasBaseMixin, DetailView):
   
    model = Tarefas
    template_name = 'tarefas/tarefa_detail.html'
    context_object_name = 'object'

    def get_queryset(self):
        """Restringe o queryset para respeitar a visibilidade."""
        qs = super().get_queryset()
        return aplicar_filtro_visibilidade(qs, self.request.user)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        tarefa = self.object
        user = self.request.user

        ctx['comentarios'] = tarefa.comentarios.select_related('autor').all()

        # ★ Usar o novo HistoricoTarefa (historicos_v2)
        ctx['historicos'] = tarefa.historicos_v2.select_related(
            'alterado_por'
        ).order_by('-data_alteracao')[:50]

        ctx['form'] = ComentarioForm()
        ctx['status_choices'] = Tarefas.STATUS_CHOICES

        # Permissões para o template
        ctx['can_edit'] = TarefaAccessMixin.user_can_edit(user, tarefa)
        ctx['can_change_status'] = TarefaAccessMixin.user_can_access(user, tarefa)

        return ctx

    def post(self, request, *args, **kwargs):
        """Salvar comentário — qualquer pessoa com acesso pode comentar."""
        self.object = self.get_object()
        form = ComentarioForm(request.POST, request.FILES)
        if form.is_valid():
            comentario = form.save(commit=False)
            comentario.tarefa = self.object
            comentario.autor = request.user
            comentario.filial = self.object.filial
            comentario.save()

            # Notificar (sino + e-mail)
            notificar_tarefa_comentario(
                tarefa=self.object,
                autor=request.user,
                texto_comentario=comentario.texto,
            )

            return redirect('tarefas:tarefa_detail', pk=self.object.pk)

        ctx = self.get_context_data()
        ctx['form'] = form
        return self.render_to_response(ctx)


class TarefaCreateView(TarefasBaseMixin, CreateView):
    model = Tarefas
    form_class = TarefaForm
    template_name = 'tarefas/tarefa_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.usuario = self.request.user
        form.instance.filial = self.request.user.filial_ativa

        response = super().form_valid(form)  # Agora self.object existe
        self.object.participantes.add(self.request.user)
        # ✅ Notificar DEPOIS do save
        notificar_tarefa_criada(
            tarefa=self.object,
            criador=self.request.user,
        )

        return response


class TarefaUpdateView(TarefasBaseMixin, UpdateView):

    model = Tarefas
    form_class = TarefaForm
    template_name = 'tarefas/tarefa_form.html'

    def get_queryset(self):
        """Restringe o queryset para respeitar a visibilidade."""
        qs = super().get_queryset()
        return aplicar_filtro_visibilidade(qs, self.request.user)

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not TarefaAccessMixin.user_can_edit(self.request.user, obj):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied('Apenas o criador, responsável ou admin pode editar.')
        return obj

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance._user = self.request.user
        return super().form_valid(form)


class TarefaDeleteView(TarefasBaseMixin, DeleteView):
   
    model = Tarefas
    template_name = 'tarefas/confirmar_exclusao.html'
    success_url = reverse_lazy('tarefas:listar_tarefas')

    def get_queryset(self):
        """Restringe o queryset para respeitar a visibilidade."""
        qs = super().get_queryset()
        return aplicar_filtro_visibilidade(qs, self.request.user)

    def form_valid(self, form):
        messages.success(self.request, "Tarefa excluída com sucesso!")
        return super().form_valid(form)


class ConcluirTarefaView(TarefasBaseMixin, View):
    """Marca uma tarefa como concluída via POST."""

    def post(self, request, pk):
        filial_id = request.session.get('active_filial_id')

        # Base: filial + visibilidade
        qs = Tarefas.objects.all()
        if filial_id:
            qs = qs.filter(filial_id=filial_id)
        qs = aplicar_filtro_visibilidade(qs, request.user)

        # Só responsável ou criador pode concluir
        tarefa = get_object_or_404(
            qs.filter(Q(usuario=request.user) | Q(responsavel=request.user)),
            pk=pk,
        )

        if tarefa.status != 'concluida':
            tarefa._user = request.user
            tarefa.status = 'concluida'
            tarefa.save()
            messages.success(request, f'Tarefa "{tarefa.titulo}" concluída!')
        else:
            messages.info(request, 'Esta tarefa já estava concluída.')

        return redirect('tarefas:listar_tarefas')


# =============================================================================
# KANBAN
# =============================================================================

class KanbanView(TarefasBaseMixin, TemplateView):

    """View do Kanban Board."""
    template_name = 'tarefas/kanban_board.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        filial_id = self.request.session.get('active_filial_id')

        # Queryset base respeitando filial
        qs = Tarefas.objects.select_related('responsavel', 'usuario')
        if filial_id:
            qs = qs.filter(filial_id=filial_id)

        # ★ Aplica filtro de visibilidade centralizado
        qs = aplicar_filtro_visibilidade(qs, user)

        # ═══ FILTRO POR RESPONSÁVEL ═══
        responsavel_id = self.request.GET.get('responsavel')
        if responsavel_id:
            qs = qs.filter(responsavel_id=responsavel_id)

        # Lista de responsáveis disponíveis (só quem tem tarefa)
        responsaveis = User.objects.filter(
            pk__in=qs.values_list('responsavel', flat=True).distinct()
        ).exclude(pk__isnull=True).order_by('first_name', 'username')

        # Agrupar por status
        ctx['colunas'] = []
        for key, label in Tarefas.STATUS_CHOICES:
            ctx['colunas'].append({
                'key': key,
                'label': label,
                'tarefas': qs.filter(status=key),
            })

        ctx['status_choices'] = Tarefas.STATUS_CHOICES
        ctx['responsaveis'] = responsaveis
        ctx['responsavel_atual'] = responsavel_id or ''
        ctx['now'] = timezone.now()
        return ctx


@login_required
@require_POST
def update_task_status(request):
    """
    Endpoint AJAX para o Kanban — mover tarefa entre colunas.
    Espera: task_id e new_status via POST form-encoded.
    """
    task_id = request.POST.get('task_id')
    new_status = request.POST.get('new_status', '').strip()

    if not task_id or not new_status:
        return JsonResponse({
            'success': False,
            'message': 'Dados incompletos.'
        })

    # ★ Busca com filtro de visibilidade
    filial_id = request.session.get('active_filial_id')
    qs = Tarefas.objects.all()
    if filial_id:
        qs = qs.filter(filial_id=filial_id)
    qs = aplicar_filtro_visibilidade(qs, request.user)

    try:
        tarefa = qs.get(pk=task_id)
    except Tarefas.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Tarefa não encontrada ou sem permissão.'
        })

    # Verificar permissão de alteração
    if not TarefaAccessMixin.user_can_access(request.user, tarefa):
        return JsonResponse({
            'success': False,
            'message': 'Sem permissão.'
        })

    # Validar status
    status_validos = dict(Tarefas.STATUS_CHOICES)
    if new_status not in status_validos:
        return JsonResponse({
            'success': False,
            'message': f'Status "{new_status}" inválido.'
        })

    # Se não mudou, retorna sucesso sem alterar
    if tarefa.status == new_status:
        return JsonResponse({'success': True, 'changed': False})

    old_status_key = tarefa.status

    # Registrar no novo histórico
    registrar_alteracao_status(
        tarefa=tarefa,
        status_anterior_key=old_status_key,
        novo_status_key=new_status,
        alterado_por=request.user,
    )

    # Atualizar via .update() para não disparar save() e duplicar histórico
    update_fields = {'status': new_status}
    if new_status == 'concluida':
        update_fields['concluida_em'] = timezone.now()
    elif old_status_key == 'concluida':
        update_fields['concluida_em'] = None

    Tarefas.objects.filter(pk=task_id).update(**update_fields)

    # Manter compatibilidade com HistoricoStatus antigo
    try:
        from .models import HistoricoStatus
        HistoricoStatus.objects.create(
            tarefa=tarefa,
            status_anterior=old_status_key,
            novo_status=new_status,
            alterado_por=request.user,
            filial=tarefa.filial,
        )
    except Exception:
        pass

    return JsonResponse({
        'success': True,
        'changed': True,
        'task_id': task_id,
        'new_status': new_status,
        'new_status_display': status_validos[new_status],
    })


# =============================================================================
# CALENDÁRIO
# =============================================================================

class CalendarioTarefasView(TarefasBaseMixin, ListView):
   
    model = Tarefas
    template_name = 'tarefas/calendario.html'

    def get_queryset(self):
        qs = super().get_queryset()
        # ★ Aplica filtro de visibilidade
        qs = aplicar_filtro_visibilidade(qs, self.request.user)
        return qs.filter(
            prazo__isnull=False
        ).exclude(
            status='cancelada'
        ).select_related('responsavel')

    def _classe_css_evento(self, tarefa):
        """Determina classe CSS com base na data."""
        hoje = timezone.now().date()
        prazo = tarefa.prazo.date() if hasattr(tarefa.prazo, 'date') else tarefa.prazo
        data_criacao = tarefa.data_criacao.date()

        if prazo < hoje:
            return 'fc-event-status-atrasada'
        if data_criacao >= hoje - timedelta(days=3):
            return 'fc-event-status-recente'
        if prazo <= hoje + timedelta(days=7):
            return 'fc-event-status-proximo-prazo'
        return 'fc-event-status-em-andamento'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        eventos = []

        for t in self.get_queryset():
            try:
                prazo_fmt = t.prazo.strftime('%d/%m/%Y')
                classe = self._classe_css_evento(t)
                eventos.append({
                    'start': t.data_criacao.strftime('%Y-%m-%d'),
                    'title': f'{t.titulo} (Prazo: {prazo_fmt})',
                    'url': reverse('tarefas:tarefa_detail', kwargs={'pk': t.pk}),
                    'className': f'fc-event-prioridade-{t.prioridade} {classe}',
                    'allDay': True,
                    'extendedProps': {
                        'status': t.status,
                        'data_criacao': t.data_criacao.strftime('%Y-%m-%d %H:%M'),
                        'prazo': prazo_fmt,
                    },
                })
            except Exception as e:
                logger.error(f"Erro ao processar tarefa {t.pk}: {e}")

        context['eventos_json'] = json.dumps(eventos)
        return context


# =============================================================================
# API — UpdateTaskStatusView (class-based, legado)
# =============================================================================

class UpdateTaskStatusView(TarefasBaseMixin, View):
    """API AJAX para atualizar status de tarefa (Kanban drag & drop) — versão legada."""

    def post(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Requisição inválida.'}, status=400)

        task_id = request.POST.get('task_id')
        new_status = request.POST.get('new_status')
        filial_id = request.session.get('active_filial_id')

        # Valida status
        valid_statuses = [s[0] for s in Tarefas.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'message': 'Status inválido.'}, status=400)

        try:
            # ★ Aplica filtro de visibilidade + só responsável/criador pode alterar
            qs = Tarefas.objects.all()
            if filial_id:
                qs = qs.filter(filial_id=filial_id)
            qs = aplicar_filtro_visibilidade(qs, request.user)

            tarefa = qs.filter(
                Q(usuario=request.user) | Q(responsavel=request.user)
            ).get(pk=task_id)

            tarefa._user = request.user
            tarefa.status = new_status
            tarefa.save()

            return JsonResponse({'success': True, 'message': 'Status atualizado.'})
        except Tarefas.DoesNotExist:
            return JsonResponse(
                {'success': False, 'message': 'Tarefa não encontrada ou sem permissão.'},
                status=404,
            )
        except Exception as e:
            logger.error(f"Erro ao atualizar tarefa {task_id}: {e}")
            return JsonResponse(
                {'success': False, 'message': 'Erro interno.'},
                status=500,
            )


# =============================================================================
# RELATÓRIOS
# =============================================================================
@login_required
def _build_queryset_relatorio(request):
    """
    Helper centralizado: aplica filial + visibilidade + filtros GET/POST.
    Usado por todas as views de relatório.
    """
    filial_id = request.session.get('active_filial_id')
    qs = Tarefas.objects.all()

    if filial_id:
        qs = qs.filter(filial_id=filial_id)

    qs = aplicar_filtro_visibilidade(qs, request.user)

    # Aceita filtros tanto via GET quanto POST
    params = request.POST if request.method == 'POST' else request.GET

    if status := params.get('status'):
        qs = qs.filter(status=status)
    if prioridade := params.get('prioridade'):
        qs = qs.filter(prioridade=prioridade)
    if responsavel := params.get('responsavel'):
        qs = qs.filter(responsavel_id=responsavel)
    if projeto := params.get('projeto'):
        qs = qs.filter(projeto_id=projeto)

    return qs.select_related('responsavel', 'filial').order_by('-data_criacao')

@login_required
def _exportar_relatorio(request, formato):
    """
    Roteador de exportação. Recebe o formato e devolve o HttpResponse correto.
    """
    qs = _build_queryset_relatorio(request)
    context = preparar_contexto_relatorio(qs)
    context['request'] = request
    context['now'] = timezone.now()
    context['data_emissao'] = timezone.now().strftime("%d/%m/%Y %H:%M")

    exporters = {
        'pdf':  gerar_pdf_relatorio,
        'csv':  gerar_csv_relatorio,
        'docx': gerar_docx_relatorio,
        'xlsx': gerar_xlsx_relatorio,
    }
    exporter = exporters.get(formato)
    if not exporter:
        messages.error(request, 'Formato de exportação inválido.')
        return redirect('tarefas:relatorio_tarefas')

    return exporter(context)


# =============================================================================
# VIEW PRINCIPAL — Tela de relatório com filtros + tabela + exportação
# =============================================================================
class RelatorioTarefasView(TarefasBaseMixin, ListView):
    """
    Página única de relatórios:
    - GET: exibe tabela + gráficos + filtros
    - POST: exporta no formato escolhido
    """
    model = Tarefas
    template_name = 'tarefas/relatorio_tarefas.html'

    def get_queryset(self):
        return _build_queryset_relatorio(self.request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        context.update(preparar_contexto_relatorio(qs))

        # JSON para gráficos
        context['status_data_json'] = json.dumps(
            context.get('status_data', []), cls=DjangoJSONEncoder
        )
        context['priority_data_json'] = json.dumps(
            context.get('prioridade_data', []), cls=DjangoJSONEncoder
        )

        # Choices para os selects
        context['status_choices'] = Tarefas.STATUS_CHOICES
        context['prioridade_choices'] = Tarefas.PRIORIDADE_CHOICES

        # Responsáveis disponíveis (respeitando filial + visibilidade)
        base_qs = aplicar_filtro_visibilidade(
            Tarefas.objects.filter(
                filial_id=self.request.session.get('active_filial_id')
            ) if self.request.session.get('active_filial_id') else Tarefas.objects.all(),
            self.request.user
        )
        context['responsaveis'] = User.objects.filter(
            pk__in=base_qs.values_list('responsavel', flat=True).distinct()
        ).exclude(pk__isnull=True).order_by('first_name', 'username')

        # Filtros ativos (para manter selecionados no form)
        context['current_filters'] = {
            'status':      self.request.GET.get('status', ''),
            'prioridade':  self.request.GET.get('prioridade', ''),
            'responsavel': self.request.GET.get('responsavel', ''),
            'projeto':     self.request.GET.get('projeto', ''),
        }
        return context

    def post(self, request, *args, **kwargs):
        """Exportação via formulário (offcanvas/modal)."""
        formato = request.POST.get('export_format')
        if not formato:
            return self.get(request, *args, **kwargs)
        return _exportar_relatorio(request, formato)


# =============================================================================
# EXPORTAÇÃO DIRETA — Endpoint para botões de download (links GET)
# =============================================================================
class ExportarRelatorioView(TarefasBaseMixin, View):
    """
    Exportação rápida via GET. Útil para botões/links diretos:
    /tarefas/relatorio/exportar/?formato=pdf&status=pendente
    """
    def get(self, request, *args, **kwargs):
        formato = request.GET.get('formato', '')
        return _exportar_relatorio(request, formato)


# =============================================================================
# EXPORTADOR EXCEL — agora integrado e seguro
# =============================================================================
def gerar_xlsx_relatorio(context):
    """
    Gera arquivo Excel do relatório seguindo a identidade visual padrão.
    Recebe o mesmo `context` que os outros exportadores (pdf/csv/docx).
    """
    from openpyxl import Workbook
    from .utils.excel_styles import (
        aplicar_cabecalho_relatorio,
        aplicar_estilo_tabela,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Tarefas"

    headers = ["Qtda.", "Título", "Responsável", "Status", "Prioridade",
               "Projeto", "Criação", "Prazo"]

    # Cabeçalho institucional padronizado
    start_row = aplicar_cabecalho_relatorio(
        ws,
        titulo="Relatório de Tarefas",  
        subtitulo="Acompanhamento e análise de tarefas do sistema",
        data_emissao=context.get('data_emissao', timezone.now().strftime("%d/%m/%Y %H:%M")),
        num_colunas=len(headers),
    )

    # Cabeçalho da tabela
    for col, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=h)

    # Dados (vêm do context, já filtrados!)
    tarefas = context.get('tarefas', [])
    for idx, t in enumerate(tarefas, start=1):
        row = start_row + idx
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=t.titulo)
        ws.cell(row=row, column=3,
                value=t.responsavel.get_full_name() if t.responsavel else "—")
        ws.cell(row=row, column=4, value=t.get_status_display())
        ws.cell(row=row, column=5, value=t.get_prioridade_display())
        ws.cell(row=row, column=6,
                value=str(t.projeto) if t.projeto else "—")
        ws.cell(row=row, column=7,
                value=t.data_criacao.strftime("%d/%m/%Y") if t.data_criacao else "—")
        ws.cell(row=row, column=8,
                value=t.prazo.strftime("%d/%m/%Y") if t.prazo else "—")

    aplicar_estilo_tabela(
        ws,
        header_row=start_row,
        total_rows=len(tarefas),
        total_cols=len(headers),
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="relatorio_tarefas_{timezone.now():%Y%m%d_%H%M}.xlsx"'
    )
    wb.save(response)
    return response


# =============================================================================
# DASHBOARD ANALÍTICO
# =============================================================================

class DashboardAnaliticoView(TarefasBaseMixin, TemplateView):
    
    template_name = 'tarefas/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        filial_id = self.request.session.get('active_filial_id')
        agora = timezone.now()

        # ★ Queryset base com filial + visibilidade
        base_qs = Tarefas.objects.all()
        if filial_id:
            base_qs = base_qs.filter(filial_id=filial_id)
        base_qs = aplicar_filtro_visibilidade(base_qs, user)

        # KPIs
        total = base_qs.count()
        concluidas = base_qs.filter(status='concluida').count()
        pendentes = base_qs.filter(status='pendente').count()
        em_andamento = base_qs.filter(status='andamento').count()
        atrasadas = base_qs.filter(prazo__lt=agora).exclude(
            status__in=['concluida', 'cancelada']
        ).count()
        pausadas = base_qs.filter(status='pausada').count()
        taxa_conclusao = round((concluidas / total * 100), 1) if total > 0 else 0

        context.update({
            'total_tarefas': total,
            'tarefas_concluidas': concluidas,
            'tarefas_pendentes': pendentes,
            'tarefas_andamento': em_andamento,
            'tarefas_atrasadas': atrasadas,
            'tarefas_pausadas': pausadas,
            'taxa_conclusao': taxa_conclusao,
        })

        # Dados por status e prioridade (para gráficos)
        status_counts = (
            base_qs.values('status')
            .annotate(total=Count('id'))
            .order_by('status')
        )
        prioridade_counts = (
            base_qs.values('prioridade')
            .annotate(total=Count('id'))
            .order_by('prioridade')
        )

        status_labels = dict(Tarefas.STATUS_CHOICES)
        prioridade_labels = dict(Tarefas.PRIORIDADE_CHOICES)

        context['status_data'] = [
            {'status': status_labels.get(s['status'], s['status']), 'total': s['total']}
            for s in status_counts
        ]
        context['prioridade_data'] = [
            {'prioridade': prioridade_labels.get(p['prioridade'], p['prioridade']), 'total': p['total']}
            for p in prioridade_counts
        ]

        # Tarefas recentes
        context['tarefas_recentes'] = base_qs.select_related(
            'responsavel'
        ).order_by('-data_criacao')[:5]

        # Performance dos usuários
        usuarios = User.objects.filter(is_active=True)
        if filial_id:
            usuarios = usuarios.filter(filiais_permitidas__id=filial_id)

        thirty_days_ago = agora - timedelta(days=30)
        usuarios_performance = []
        for usuario in usuarios:
            ativas = base_qs.filter(
                responsavel=usuario,
                status__in=['pendente', 'andamento', 'atrasada']
            ).count()
            concluidas_user = base_qs.filter(
                responsavel=usuario,
                status='concluida',
                concluida_em__gte=thirty_days_ago
            ).count()
            if ativas > 0 or concluidas_user > 0:
                usuarios_performance.append({
                    'username': usuario.get_full_name() or usuario.username,
                    'tarefas_ativas': ativas,
                    'tarefas_concluidas_30d': concluidas_user,
                })
        context['usuarios_performance'] = usuarios_performance

        # Gráfico de tendência (últimas 6 semanas)
        six_weeks_ago = agora - timedelta(weeks=6)

        criadas_qs = (
            base_qs.filter(data_criacao__gte=six_weeks_ago)
            .annotate(semana=TruncWeek('data_criacao'))
            .values('semana')
            .annotate(total=Count('id'))
            .order_by('semana')
        )
        concluidas_qs = (
            base_qs.filter(concluida_em__gte=six_weeks_ago, status='concluida')
            .annotate(semana=TruncWeek('concluida_em'))
            .values('semana')
            .annotate(total=Count('id'))
            .order_by('semana')
        )

        dados_criadas = {item['semana'].date(): item['total'] for item in criadas_qs}
        dados_concluidas = {item['semana'].date(): item['total'] for item in concluidas_qs}

        hoje = agora.date()
        semana_inicio = hoje - timedelta(weeks=5)
        current_week = semana_inicio - timedelta(days=semana_inicio.weekday())

        labels, criadas_list, concluidas_list = [], [], []
        while current_week <= hoje:
            labels.append(current_week)
            criadas_list.append(dados_criadas.get(current_week, 0))
            concluidas_list.append(dados_concluidas.get(current_week, 0))
            current_week += timedelta(weeks=1)

        charts_data = {
            'tendencia_labels': labels,
            'tendencia_criadas': criadas_list,
            'tendencia_concluidas': concluidas_list,
            'performance_equipe': usuarios_performance,
            'status_data': context['status_data'],
            'prioridade_data': context['prioridade_data'],
        }
        context['charts_data_json'] = json.dumps(charts_data, cls=DjangoJSONEncoder)

        return context


# =============================================================================
# ADMIN (SUPERUSER)
# =============================================================================

class TarefaAdminListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Lista TODAS as tarefas (sem filtro de filial). Apenas superuser."""
    model = Tarefas
    template_name = 'tarefas/tarefas_admin.html'
    context_object_name = 'tarefas'
    ordering = ['-data_criacao']
    paginate_by = 50

    def test_func(self):
        return self.request.user.is_superuser


# =============================================================================
# ALTERAR STATUS — AJAX (Detail page)
# =============================================================================

@login_required
@require_POST
def alterar_status_tarefa(request, pk):
    """Altera o status de uma tarefa via AJAX (usado no detalhe)."""
    if not request.user.is_superuser:
        user_perms = request.user.get_all_permissions()
        if not any(p.startswith('tarefas.') for p in user_perms):
            return JsonResponse({'error': 'Sem permissão.'}, status=403)

    # ★ Busca com filtro de visibilidade
    filial_id = request.session.get('active_filial_id')
    qs = Tarefas.objects.all()
    if filial_id:
        qs = qs.filter(filial_id=filial_id)
    qs = aplicar_filtro_visibilidade(qs, request.user)

    tarefa = get_object_or_404(qs, pk=pk)

    # Verificar acesso
    if not TarefaAccessMixin.user_can_access(request.user, tarefa):
        return JsonResponse({'error': 'Sem permissão.'}, status=403)

    novo_status = request.POST.get('status', '').strip()

    status_validos = dict(Tarefas.STATUS_CHOICES)
    if novo_status not in status_validos:
        return JsonResponse({'error': 'Status inválido.'}, status=400)

    if tarefa.status == novo_status:
        return JsonResponse({'error': 'A tarefa já possui este status.'}, status=400)

    status_anterior_key = tarefa.status
    status_anterior_display = tarefa.get_status_display()

    # Registrar no novo histórico
    registrar_alteracao_status(
        tarefa=tarefa,
        status_anterior_key=status_anterior_key,
        novo_status_key=novo_status,
        alterado_por=request.user,
    )

    # Atualizar sem disparar save() (evita duplicação de histórico)
    update_fields = {'status': novo_status}
    if novo_status == 'concluida':
        update_fields['concluida_em'] = timezone.now()
    elif status_anterior_key == 'concluida':
        update_fields['concluida_em'] = None

    Tarefas.objects.filter(pk=pk).update(**update_fields)
    tarefa.refresh_from_db()

    # Manter compatibilidade com HistoricoStatus antigo
    try:
        from .models import HistoricoStatus
        HistoricoStatus.objects.create(
            tarefa=tarefa,
            status_anterior=status_anterior_key,
            novo_status=novo_status,
            alterado_por=request.user,
            filial=tarefa.filial,
        )
    except Exception:
        pass

    return JsonResponse({
        'ok': True,
        'novo_status': novo_status,
        'novo_status_display': tarefa.get_status_display(),
        'status_anterior_display': status_anterior_display,
        'progresso': tarefa.progresso,
        'data_atualizacao': tarefa.data_atualizacao.strftime('%d/%m/%Y %H:%M'),
    })

