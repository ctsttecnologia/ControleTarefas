
"""
Views completas para o módulo PGR
Refatorado com:
  - Filtro por filial (ViewFilialScopedMixin / for_request)
  - Permissões por grupo (SSTPermissionMixin)
  - Escopo de técnico (TecnicoScopeMixin)
  - Sem duplicatas
"""
import json
import openpyxl
from io import BytesIO
from datetime import date, datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Q, Count, Sum, Avg
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views.decorators.http import require_POST
from django.forms import inlineformset_factory

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dal import autocomplete

from core.mixins import (
    SSTPermissionMixin,
    ViewFilialScopedMixin,
    TecnicoScopeMixin,
    FilialCreateMixin,
)

from .models import (
    PGRDocumento, PGRRevisao, PGRDocumentoResponsavel,
    PGRSecaoTexto, PGRSecaoTextoPadrao,
    GESGrupoExposicao, RiscoIdentificado, AvaliacaoQuantitativa,
    PlanoAcaoPGR, AcompanhamentoPlanoAcao, AnexoPlanoAcao,
    CronogramaAcaoPGR, MedidaControle, RiscoMedidaControle,
    RiscoEPIRecomendado, RiscoTreinamentoNecessario,
    Empresa, LocalPrestacaoServico, ProfissionalResponsavel,
    AmbienteTrabalho, TipoRisco,
    STATUS_CHOICES, STATUS_PGR_CHOICES, PRIORIDADE_CHOICES,
    CLASSIFICACAO_RISCO_CHOICES, TIPO_ACAO_CHOICES,
    PERIODICIDADE_CHOICES, TIPO_AVALIACAO_CHOICES,
    UNIDADE_MEDIDA_CHOICES, TIPO_RESPONSABILIDADE_CHOICES,
)
from .forms import (
    AvaliacaoQuantitativaForm, CronogramaAcaoPGRForm, EmpresaForm,
    LocalPrestacaoServicoForm, PGRDocumentoForm, PGRRevisaoForm,
    GESGrupoExposicaoForm, ProfissionalResponsavelForm,
    RiscoIdentificadoForm, PlanoAcaoPGRForm,
    AcompanhamentoPlanoAcaoForm, ResponsavelFormSet,
)
from .utils.cont_seguranca import validar_acesso_documento
from cliente.models import Cliente



# =============================================================================
# HELPERS INTERNOS (para FBVs)
# =============================================================================

def _safe_for_request(model_class, request):
    """
    Chama .for_request(request) se disponível, senão .all().
    Após a refatoração dos models, TODOS devem ter for_request.
    Este helper existe como rede de segurança.
    """
    qs = model_class.objects.all()
    if hasattr(qs, 'for_request'):
        return qs.for_request(request)
    return qs


def _aplicar_filtro_filial_e_tecnico(request, model_class):
    """
    Helper para FBVs: aplica filtro de filial + detecta técnico.
    Retorna (queryset_filtrado, is_tecnico).
    """
    qs = _safe_for_request(model_class, request)

    user = request.user
    if not hasattr(user, '_is_tecnico'):
        user._is_tecnico = user.groups.filter(name='TÉCNICO').exists()

    return qs, user._is_tecnico


def _filtrar_tecnico_documentos(qs, user, is_tecnico):
    """Filtra documentos PGR para técnicos (só vê os que criou)."""
    if is_tecnico:
        return qs.filter(criado_por=user)
    return qs


def _filtrar_tecnico_riscos(qs, user, is_tecnico):
    """Filtra riscos para técnicos."""
    if is_tecnico:
        return qs.filter(pgr_documento__criado_por=user)
    return qs


def _filtrar_tecnico_planos(qs, user, is_tecnico):
    """Filtra planos para técnicos."""
    if is_tecnico:
        return qs.filter(risco_identificado__pgr_documento__criado_por=user)
    return qs


def _get_filial_info(request):
    """Retorna info da filial ativa para o contexto do template."""
    filial_ativa_id = request.session.get('active_filial_id')
    filial_info = None
    acesso_global = request.user.is_superuser

    if filial_ativa_id:
        from usuario.models import Filial
        try:
            filial_info = Filial.objects.get(pk=filial_ativa_id).nome
        except Filial.DoesNotExist:
            pass

    return filial_info, acesso_global


# =============================================================================
# AUTOCOMPLETE
# =============================================================================

class ClienteAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return Cliente.objects.none()

        qs = Cliente.objects.for_request(self.request)
        if self.q:
            qs = qs.filter(
                Q(razao_social__icontains=self.q) |
                Q(nome__icontains=self.q)
            )
        return qs.order_by('razao_social')


# =============================================================================
# DASHBOARDS (FBVs)
# =============================================================================

# =============================================================================
# DASHBOARD UNIFICADO (FBV)
# =============================================================================

@login_required
@permission_required('pgr_gestao.view_pgrdocumento', raise_exception=True)
def dashboard_gerencial_view(request):
    """Dashboard gerencial unificado — PGR completo, filtrado por filial e técnico."""
    user = request.user

    # ══════════════════════════════════════════════════════════
    # QUERYSETS FILTRADOS POR FILIAL
    # ══════════════════════════════════════════════════════════
    documentos_qs, is_tecnico = _aplicar_filtro_filial_e_tecnico(request, PGRDocumento)
    riscos_qs = _safe_for_request(RiscoIdentificado, request)
    planos_qs = _safe_for_request(PlanoAcaoPGR, request)
    ges_qs = _safe_for_request(GESGrupoExposicao, request)
    revisoes_qs = _safe_for_request(PGRRevisao, request)

    # ══════════════════════════════════════════════════════════
    # ESCOPO TÉCNICO
    # ══════════════════════════════════════════════════════════
    documentos_qs = _filtrar_tecnico_documentos(documentos_qs, user, is_tecnico)
    riscos_qs = _filtrar_tecnico_riscos(riscos_qs, user, is_tecnico)
    planos_qs = _filtrar_tecnico_planos(planos_qs, user, is_tecnico)
    if is_tecnico:
        ges_qs = ges_qs.filter(pgr_documento__criado_por=user)
        revisoes_qs = revisoes_qs.filter(pgr_documento__criado_por=user)

    # ══════════════════════════════════════════════════════════
    # CARD 1: DOCUMENTOS
    # ══════════════════════════════════════════════════════════
    total_documentos = documentos_qs.count()
    docs_por_status = documentos_qs.values('status').annotate(total=Count('status'))
    status_docs_data = {item['status']: item['total'] for item in docs_por_status}

    documentos_vigentes = status_docs_data.get('vigente', 0)
    documentos_vencidos = status_docs_data.get('vencido', 0)
    documentos_em_revisao = status_docs_data.get('em_revisao', 0)

    # Documentos a vencer (próximos 60 dias)
    data_alerta = date.today() + timedelta(days=60)
    documentos_a_vencer = documentos_qs.filter(
        data_vencimento__lte=data_alerta,
        data_vencimento__gte=date.today(),
    ).select_related('empresa').order_by('data_vencimento')

    # ══════════════════════════════════════════════════════════
    # CARD 2: RISCOS
    # ══════════════════════════════════════════════════════════
    total_riscos = riscos_qs.count()

    # Mapa de classificação → display legível
    CLASSIFICACAO_DISPLAY = dict(CLASSIFICACAO_RISCO_CHOICES)

    riscos_por_classificacao_qs = riscos_qs.values('classificacao_risco').annotate(
        total=Count('classificacao_risco')
    ).order_by('classificacao_risco')

    # Lista com label legível (template usa {{ item.label }})
    riscos_por_classificacao = [
        {
            'classificacao_risco': item['classificacao_risco'],
            'label': CLASSIFICACAO_DISPLAY.get(
                item['classificacao_risco'], item['classificacao_risco']
            ),
            'total': item['total'],
        }
        for item in riscos_por_classificacao_qs
    ]

    riscos_por_status = riscos_qs.values(
        'status_controle'
    ).annotate(total=Count('id'))

    # Riscos críticos não controlados
    riscos_criticos_pendentes = riscos_qs.filter(
        classificacao_risco__in=['critico', 'muito_grave'],
        status_controle__in=['identificado', 'em_controle']
    ).select_related(
        'tipo_risco', 'ges', 'pgr_documento'
    ).order_by('-classificacao_risco')[:10]

    # ══════════════════════════════════════════════════════════
    # CARD 3: PLANOS DE AÇÃO
    # ══════════════════════════════════════════════════════════
    total_planos = planos_qs.count()
    planos_por_status = planos_qs.values('status').annotate(total=Count('status'))
    status_planos_data = {item['status']: item['total'] for item in planos_por_status}

    planos_pendentes = status_planos_data.get('pendente', 0)
    planos_em_andamento = status_planos_data.get('em_andamento', 0)
    planos_concluidos = status_planos_data.get('concluido', 0)

    planos_atrasados = planos_qs.filter(
        status__in=['pendente', 'em_andamento'],
        data_prevista__lt=date.today()
    ).count()
    status_planos_data['atrasado'] = planos_atrasados

    # ══════════════════════════════════════════════════════════
    # COMPLEMENTOS
    # ══════════════════════════════════════════════════════════
    total_ges = ges_qs.filter(ativo=True).count()

    ultimas_revisoes = revisoes_qs.select_related(
        'pgr_documento', 'pgr_documento__empresa'
    ).order_by('-data_realizacao')[:5]

    ultimos_planos = planos_qs.select_related(
        'risco_identificado', 'risco_identificado__pgr_documento'
    ).order_by('-criado_em')[:5]

    # Riscos por mês (últimos 6 meses)
    seis_meses_atras = date.today() - timedelta(days=180)
    riscos_por_mes = riscos_qs.filter(
        data_identificacao__gte=seis_meses_atras
    ).annotate(
        mes=TruncMonth('data_identificacao')
    ).values('mes').annotate(total=Count('id')).order_by('mes')

    filial_info, acesso_global = _get_filial_info(request)

    # ══════════════════════════════════════════════════════════
    # CONTEXTO COMPLETO
    # ══════════════════════════════════════════════════════════
    context = {
        # Documentos
        'total_documentos': total_documentos,
        'documentos_vigentes': documentos_vigentes,
        'documentos_vencidos': documentos_vencidos,
        'documentos_em_revisao': documentos_em_revisao,
        'status_docs_data': status_docs_data,
        'documentos_a_vencer': documentos_a_vencer,
        'documentos_proximo_vencimento': documentos_a_vencer,

        # Riscos
        'total_riscos': total_riscos,
        'riscos_por_classificacao': riscos_por_classificacao,
        'riscos_por_status': riscos_por_status,
        'riscos_criticos': riscos_criticos_pendentes,
        'riscos_criticos_pendentes': riscos_criticos_pendentes,

        # Planos
        'total_planos': total_planos,
        'planos_pendentes': planos_pendentes,
        'planos_em_andamento': planos_em_andamento,
        'planos_concluidos': planos_concluidos,
        'planos_atrasados': planos_atrasados,
        'status_planos_data': status_planos_data,

        # Complementos
        'total_ges': total_ges,
        'ultimas_revisoes': ultimas_revisoes,
        'ultimos_planos': ultimos_planos,
        'riscos_por_mes': riscos_por_mes,

        # Filial
        'filial_info': filial_info,
        'acesso_global': acesso_global,
    }
    return render(request, 'pgr_gestao/dashboard_gerencial.html', context)



# =============================================================================
# EMPRESAS
# =============================================================================

class EmpresaListView(SSTPermissionMixin, ViewFilialScopedMixin, ListView):
    model = Empresa
    template_name = 'pgr_gestao/empresa_list.html'
    context_object_name = 'empresas'
    paginate_by = 20
    permission_required = 'pgr_gestao.view_empresa'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('cliente', 'filial')

        tipo = self.request.GET.get('tipo')
        grau_risco = self.request.GET.get('grau_risco')
        ativo = self.request.GET.get('ativo')
        search = self.request.GET.get('search')

        if tipo:
            queryset = queryset.filter(tipo_empresa=tipo)
        
        if grau_risco:
            queryset = queryset.filter(grau_risco=grau_risco)
        
        if ativo:
            queryset = queryset.filter(ativo=(ativo == 'true'))
        
        if search:
            queryset = queryset.filter(
                Q(cliente__razao_social__icontains=search) |
                Q(cliente__nome__icontains=search) |
                Q(cnpj__icontains=search) |
                Q(descricao_cnae__icontains=search)
            )
        
        return queryset.order_by('cliente__razao_social')

class EmpresaDetailView(SSTPermissionMixin, ViewFilialScopedMixin, DetailView):
    model = Empresa
    template_name = 'pgr_gestao/empresa_detail.html'
    context_object_name = 'empresa'
    permission_required = 'pgr_gestao.view_empresa'


class EmpresaCreateView(SSTPermissionMixin, FilialCreateMixin, CreateView):
    model = Empresa
    form_class = EmpresaForm
    template_name = 'pgr_gestao/empresa_form.html'
    permission_required = 'pgr_gestao.add_empresa'
    success_url = reverse_lazy('pgr_gestao:empresa_list')


class EmpresaUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, UpdateView):
    model = Empresa
    form_class = EmpresaForm
    template_name = 'pgr_gestao/empresa_form.html'
    permission_required = 'pgr_gestao.change_empresa'

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:empresa_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Empresa atualizada com sucesso!')
        return super().form_valid(form)


class EmpresaDeleteView(SSTPermissionMixin, ViewFilialScopedMixin, DeleteView):
    model = Empresa
    template_name = 'pgr_gestao/empresa_confirm_delete.html'
    permission_required = 'pgr_gestao.delete_empresa'
    success_url = reverse_lazy('pgr_gestao:empresa_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Empresa excluída com sucesso!')
        return super().delete(request, *args, **kwargs)


# =============================================================================
# DOCUMENTOS PGR
# =============================================================================

class PGRDocumentoListView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, ListView):
    model = PGRDocumento
    template_name = 'pgr_gestao/documento_list.html'
    context_object_name = 'documentos'
    paginate_by = 20
    permission_required = 'pgr_gestao.view_pgrdocumento'
    tecnico_scope_lookup = 'criado_por'

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'empresa', 'local_prestacao'
        ).prefetch_related('responsaveis')

        status = self.request.GET.get('status')
        empresa_id = self.request.GET.get('empresa')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)
        if empresa_id:
            queryset = queryset.filter(empresa_id=empresa_id)
        if search:
            queryset = queryset.filter(
                Q(codigo_documento__icontains=search) |
                Q(empresa__razao_social__icontains=search)
            )
        return queryset.order_by('-data_elaboracao')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Empresas também filtradas por filial
        context['empresas'] = Empresa.objects.for_request(self.request).filter(ativo=True)
        context['status_choices'] = STATUS_PGR_CHOICES
        return context


class PGRDocumentoDetailView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, DetailView):
    model = PGRDocumento
    template_name = 'pgr_gestao/documento_detail.html'
    context_object_name = 'documento'
    permission_required = 'pgr_gestao.view_pgrdocumento'
    tecnico_scope_lookup = 'criado_por'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        documento = self.object

        if self.request.method == 'POST':
            context['responsaveis_formset'] = ResponsavelFormSet(
                self.request.POST, instance=documento, prefix='responsaveis'
            )
        else:
            context['responsaveis_formset'] = ResponsavelFormSet(
                instance=documento, prefix='responsaveis'
            )

        context['revisoes'] = documento.revisoes.all().order_by('-numero_revisao')
        context['grupos_exposicao'] = documento.grupos_exposicao.filter(
            ativo=True
        ).select_related('cargo', 'funcao')
        context['riscos'] = RiscoIdentificado.objects.filter(
            pgr_documento=documento
        ).select_related('tipo_risco', 'ges').order_by('-prioridade_acao')
        context['cronograma'] = documento.cronograma_acoes.all().order_by('numero_item')
        context['planos_acao'] = PlanoAcaoPGR.objects.filter(
            risco_identificado__pgr_documento=documento
        ).select_related('risco_identificado').order_by('-prioridade', 'data_prevista')

        riscos_qs = context['riscos']
        context['total_riscos'] = riscos_qs.count()
        context['riscos_criticos'] = riscos_qs.filter(
            classificacao_risco__in=['critico', 'muito_grave']
        ).count()
        context['riscos_controlados'] = riscos_qs.filter(status_controle='controlado').count()
        context['riscos_stats'] = riscos_qs.values('classificacao_risco').annotate(
            total=Count('id')
        ).order_by('classificacao_risco')

        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        formset = ResponsavelFormSet(request.POST, instance=self.object, prefix='responsaveis')
        if formset.is_valid():
            formset.save()
            return redirect(reverse('pgr_gestao:documento_detail', kwargs={'pk': self.object.pk}))
        else:
            return self.render_to_response(self.get_context_data(form=formset))


class PGRDocumentoCreateView(SSTPermissionMixin, CreateView):
    model = PGRDocumento
    form_class = PGRDocumentoForm
    template_name = 'pgr_gestao/documento_form.html'
    permission_required = 'pgr_gestao.add_pgrdocumento'
    success_url = reverse_lazy('pgr_gestao:documento_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        filial_do_usuario = getattr(self.request.user, 'filial_ativa', None)
        if not filial_do_usuario:
            form.add_error(None, "Seu usuário não tem uma filial ativa configurada.")
            return self.form_invalid(form)

        form.instance.filial = filial_do_usuario
        form.instance.criado_por = self.request.user
        messages.success(self.request, "Documento PGR criado com sucesso!")
        return super().form_valid(form)


class PGRDocumentoUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, UpdateView):
    model = PGRDocumento
    form_class = PGRDocumentoForm
    template_name = 'pgr_gestao/documento_form.html'
    permission_required = 'pgr_gestao.change_pgrdocumento'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:documento_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Documento PGR atualizado com sucesso!')
        return super().form_valid(form)


class PGRDocumentoDeleteView(SSTPermissionMixin, ViewFilialScopedMixin, DeleteView):
    model = PGRDocumento
    template_name = 'pgr_gestao/documento_confirm_delete.html'
    permission_required = 'pgr_gestao.delete_pgrdocumento'
    success_url = reverse_lazy('pgr_gestao:documento_list')


# =============================================================================
# REVISÕES
# =============================================================================

class PGRRevisaoCreateView(SSTPermissionMixin, CreateView):
    model = PGRRevisao
    form_class = PGRRevisaoForm
    template_name = 'pgr_gestao/revisao_form.html'
    permission_required = 'pgr_gestao.add_pgrrevisao'

    def get_initial(self):
        initial = super().get_initial()
        pgr_id = self.kwargs.get('pgr_id')
        if pgr_id:
            pgr = get_object_or_404(PGRDocumento, pk=pgr_id)
            initial['pgr_documento'] = pgr
            ultima_revisao = pgr.revisoes.order_by('-numero_revisao').first()
            initial['numero_revisao'] = (ultima_revisao.numero_revisao + 1) if ultima_revisao else 0
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pgr_id = self.kwargs.get('pgr_id')
        if pgr_id:
            context['pgr_documento'] = get_object_or_404(PGRDocumento, pk=pgr_id)
        return context

    def form_valid(self, form):
        form.instance.criado_por = self.request.user
        form.instance.filial = self.request.user.filial_ativa

        pgr = form.instance.pgr_documento
        pgr.data_ultima_revisao = form.instance.data_realizacao
        pgr.versao_atual = form.instance.numero_revisao + 1
        pgr.save()

        messages.success(self.request, 'Revisão registrada com sucesso!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:documento_detail', kwargs={'pk': self.object.pgr_documento.pk})


class PGRRevisaoDetailView(SSTPermissionMixin, ViewFilialScopedMixin, DetailView):
    model = PGRRevisao
    template_name = 'pgr_gestao/revisao_detail.html'
    context_object_name = 'revisao'
    permission_required = 'pgr_gestao.view_pgrrevisao'


# =============================================================================
# GES - Grupos de Exposição Similar
# =============================================================================

class GESListView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, ListView):
    model = GESGrupoExposicao
    template_name = 'pgr_gestao/ges_list.html'
    context_object_name = 'grupos_exposicao'
    paginate_by = 20
    permission_required = 'pgr_gestao.view_gesgrupoexposicao'
    tecnico_scope_lookup = 'pgr_documento__criado_por'

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'pgr_documento', 'ambiente_trabalho', 'cargo', 'funcao'
        ).prefetch_related('riscos')

        pgr_id = self.request.GET.get('pgr')
        ambiente_id = self.request.GET.get('ambiente')
        ativo = self.request.GET.get('ativo')
        search = self.request.GET.get('search')

        if pgr_id:
            queryset = queryset.filter(pgr_documento_id=pgr_id)
        if ambiente_id:
            queryset = queryset.filter(ambiente_trabalho_id=ambiente_id)
        if ativo:
            queryset = queryset.filter(ativo=ativo == 'true')
        if search:
            queryset = queryset.filter(
                Q(codigo__icontains=search) |
                Q(nome__icontains=search) |
                Q(descricao_atividades__icontains=search)
            )
        return queryset.order_by('codigo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Documentos e ambientes filtrados por filial
        context['documentos_pgr'] = PGRDocumento.objects.for_request(
            self.request
        ).filter(status='vigente').select_related('empresa')
        context['ambientes'] = AmbienteTrabalho.objects.for_request(self.request)

        grupos = self.get_queryset()
        context['total_ges'] = grupos.count()
        context['total_trabalhadores'] = grupos.aggregate(
            total=Sum('numero_trabalhadores')
        )['total'] or 0
        context['total_riscos'] = RiscoIdentificado.objects.for_request(
            self.request
        ).filter(ges__in=grupos).count()
        context['ges_ativos'] = grupos.filter(ativo=True).count()
        return context


class GESDetailView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, DetailView):
    model = GESGrupoExposicao
    template_name = 'pgr_gestao/ges_detail.html'
    context_object_name = 'ges'
    permission_required = 'pgr_gestao.view_gesgrupoexposicao'
    tecnico_scope_lookup = 'pgr_documento__criado_por'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ges = self.object

        context['riscos'] = ges.riscos.select_related(
            'tipo_risco'
        ).prefetch_related('avaliacoes_quantitativas', 'planos_acao')

        context['total_riscos'] = context['riscos'].count()
        context['riscos_controlados'] = context['riscos'].filter(
            status_controle='controlado'
        ).count()
        context['riscos_pendentes'] = context['riscos'].filter(
            status_controle='identificado'
        ).count()

        context['epis_recomendados'] = RiscoEPIRecomendado.objects.filter(
            risco_identificado__ges=ges
        ).select_related('equipamento').distinct()

        context['treinamentos_necessarios'] = RiscoTreinamentoNecessario.objects.filter(
            risco_identificado__ges=ges
        ).select_related('tipo_curso').distinct()

        return context


class GESCreateView(SSTPermissionMixin, FilialCreateMixin, CreateView):
    model = GESGrupoExposicao
    form_class = GESGrupoExposicaoForm
    template_name = 'pgr_gestao/ges_form.html'
    permission_required = 'pgr_gestao.add_gesgrupoexposicao'
    success_url = reverse_lazy('pgr_gestao:ges_list')

    def form_valid(self, form):
        form.instance.criado_por = self.request.user
        return super().form_valid(form)


class GESUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, UpdateView):
    model = GESGrupoExposicao
    form_class = GESGrupoExposicaoForm
    template_name = 'pgr_gestao/ges_form.html'
    permission_required = 'pgr_gestao.change_gesgrupoexposicao'

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:ges_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'GES atualizado com sucesso!')
        return super().form_valid(form)


class GESDeleteView(SSTPermissionMixin, ViewFilialScopedMixin, DeleteView):
    model = GESGrupoExposicao
    template_name = 'pgr_gestao/ges_confirm_delete.html'
    permission_required = 'pgr_gestao.delete_gesgrupoexposicao'
    success_url = reverse_lazy('pgr_gestao:ges_list')


@login_required
@require_POST
def ges_toggle_ativo(request, pk):
    """Toggle rápido para ativar/inativar um GES via AJAX."""
    try:
        ges = GESGrupoExposicao.objects.for_request(request).get(pk=pk)

        if not request.user.has_perm('pgr_gestao.change_gesgrupoexposicao'):
            return JsonResponse({
                'success': False,
                'error': 'Você não tem permissão para alterar este GES.'
            }, status=403)

        ges.ativo = not ges.ativo
        ges.save(update_fields=['ativo'])

        return JsonResponse({
            'success': True,
            'ativo': ges.ativo,
            'message': f'GES "{ges.nome}" {"ativado" if ges.ativo else "inativado"} com sucesso!',
        })
    except GESGrupoExposicao.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'GES não encontrado.'
        }, status=404)


# =============================================================================
# RISCOS IDENTIFICADOS
# =============================================================================

class RiscoIdentificadoListView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, ListView):
    model = RiscoIdentificado
    template_name = 'pgr_gestao/risco_list.html'
    context_object_name = 'riscos'
    paginate_by = 20
    permission_required = 'pgr_gestao.view_riscoidentificado'
    tecnico_scope_lookup = 'pgr_documento__criado_por'

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'pgr_documento', 'ges', 'tipo_risco', 'ambiente_trabalho', 'cargo'
        ).prefetch_related('avaliacoes_quantitativas', 'planos_acao')

        pgr_id = self.request.GET.get('pgr')
        classificacao = self.request.GET.get('classificacao')
        status = self.request.GET.get('status')
        prioridade = self.request.GET.get('prioridade')
        tipo_risco = self.request.GET.get('tipo_risco')
        search = self.request.GET.get('search')

        if pgr_id:
            queryset = queryset.filter(pgr_documento_id=pgr_id)
        if classificacao:
            queryset = queryset.filter(classificacao_risco=classificacao)
        if status:
            queryset = queryset.filter(status_controle=status)
        if prioridade:
            queryset = queryset.filter(prioridade_acao=prioridade)
        if tipo_risco:
            queryset = queryset.filter(tipo_risco_id=tipo_risco)
        if search:
            queryset = queryset.filter(
                Q(codigo_risco__icontains=search) |
                Q(agente__icontains=search) |
                Q(fonte_geradora__icontains=search)
            )
        return queryset.order_by('-prioridade_acao', '-data_identificacao')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['documentos_pgr'] = PGRDocumento.objects.for_request(
            self.request
        ).filter(status='vigente')
        context['tipos_risco'] = TipoRisco.objects.filter(ativo=True)

        queryset = self.get_queryset()
        context['total_riscos'] = queryset.count()
        context['riscos_criticos'] = queryset.filter(
            classificacao_risco__in=['critico', 'muito_grave']
        ).count()
        context['riscos_controlados'] = queryset.filter(
            status_controle='controlado'
        ).count()
        return context


class RiscoIdentificadoDetailView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, DetailView):
    model = RiscoIdentificado
    template_name = 'pgr_gestao/risco_detail.html'
    context_object_name = 'risco'
    permission_required = 'pgr_gestao.view_riscoidentificado'
    tecnico_scope_lookup = 'pgr_documento__criado_por'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        risco = self.object

        context['avaliacoes'] = risco.avaliacoes_quantitativas.all().order_by('-data_avaliacao')
        context['medidas_controle'] = risco.medidas_controle.select_related(
            'medida_controle'
        ).order_by('medida_controle__prioridade')
        context['planos_acao'] = risco.planos_acao.all().order_by('-prioridade', 'data_prevista')
        context['epis_recomendados'] = risco.epis_recomendados.select_related('equipamento').all()
        context['treinamentos_necessarios'] = risco.treinamentos_necessarios.select_related(
            'tipo_curso'
        ).all()
        return context


class RiscoIdentificadoCreateView(SSTPermissionMixin, FilialCreateMixin, CreateView):
    model = RiscoIdentificado
    form_class = RiscoIdentificadoForm
    template_name = 'pgr_gestao/risco_form.html'
    permission_required = 'pgr_gestao.add_riscoidentificado'

    def get_initial(self):
        initial = super().get_initial()
        ges_id = self.request.GET.get('ges')
        if ges_id:
            ges = get_object_or_404(GESGrupoExposicao, pk=ges_id)
            initial['ges'] = ges
            initial['pgr_documento'] = ges.pgr_documento
            initial['ambiente_trabalho'] = ges.ambiente_trabalho
            initial['cargo'] = ges.cargo
        initial['data_identificacao'] = timezone.now().date()
        return initial

    def form_valid(self, form):
        form.instance.criado_por = self.request.user
        if not form.instance.classificacao_risco:
            form.instance.classificacao_risco = self._calcular_classificacao_risco(
                form.instance.gravidade_g,
                form.instance.exposicao_e,
                form.instance.severidade_s
            )
        return super().form_valid(form)

    def _calcular_classificacao_risco(self, gravidade, exposicao, severidade):
        matriz = {
            'A': 'negligenciavel', 'B': 'marginal', 'C': 'moderado',
            'D': 'muito_grave', 'E': 'critico'
        }
        return matriz.get(severidade, 'moderado')

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:risco_detail', kwargs={'pk': self.object.pk})


class RiscoIdentificadoUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, UpdateView):
    model = RiscoIdentificado
    form_class = RiscoIdentificadoForm
    template_name = 'pgr_gestao/risco_form.html'
    permission_required = 'pgr_gestao.change_riscoidentificado'

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:risco_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Risco atualizado com sucesso!')
        return super().form_valid(form)


class RiscoIdentificadoDeleteView(SSTPermissionMixin, ViewFilialScopedMixin, DeleteView):
    model = RiscoIdentificado
    template_name = 'pgr_gestao/risco_confirm_delete.html'
    permission_required = 'pgr_gestao.delete_riscoidentificado'
    success_url = reverse_lazy('pgr_gestao:risco_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Risco excluído com sucesso!')
        return super().delete(request, *args, **kwargs)


# =============================================================================
# AVALIAÇÕES QUANTITATIVAS
# =============================================================================

class AvaliacaoQuantitativaListView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, ListView):
    model = AvaliacaoQuantitativa
    template_name = 'pgr_gestao/avaliacao_list.html'
    context_object_name = 'avaliacoes'
    paginate_by = 20
    permission_required = 'pgr_gestao.view_avaliacaoquantitativa'
    tecnico_scope_lookup = 'risco_identificado__pgr_documento__criado_por'

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'risco_identificado', 'risco_identificado__pgr_documento',
            'risco_identificado__ges'
        )

        tipo = self.request.GET.get('tipo')
        conforme = self.request.GET.get('conforme')
        risco_id = self.request.GET.get('risco')
        search_query = self.request.GET.get('search')

        if tipo:
            queryset = queryset.filter(tipo_avaliacao=tipo)
        if conforme in ['true', 'false']:
            queryset = queryset.filter(conforme=(conforme == 'true'))
        if risco_id:
            queryset = queryset.filter(risco_identificado_id=risco_id)
        if search_query:
            queryset = queryset.filter(
                Q(equipamento_utilizado__icontains=search_query) |
                Q(risco_identificado__agente__icontains=search_query)
            )
        return queryset.order_by('-data_avaliacao')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tipos_avaliacao'] = TIPO_AVALIACAO_CHOICES
        context['riscos_para_filtro'] = RiscoIdentificado.objects.for_request(
            self.request
        ).order_by('agente')
        context['get_params'] = self.request.GET.urlencode()
        return context


class AvaliacaoQuantitativaDetailView(SSTPermissionMixin, ViewFilialScopedMixin, DetailView):
    model = AvaliacaoQuantitativa
    template_name = 'pgr_gestao/avaliacao_detail.html'
    context_object_name = 'avaliacao'
    permission_required = 'pgr_gestao.view_avaliacaoquantitativa'


class AvaliacaoQuantitativaCreateView(SSTPermissionMixin, FilialCreateMixin, CreateView):
    model = AvaliacaoQuantitativa
    form_class = AvaliacaoQuantitativaForm
    template_name = 'pgr_gestao/avaliacao_form.html'
    permission_required = 'pgr_gestao.add_avaliacaoquantitativa'

    def get_initial(self):
        initial = super().get_initial()
        risco_pk = self.kwargs.get('risco_pk') or self.request.GET.get('risco')
        if risco_pk:
            initial['risco_identificado'] = get_object_or_404(RiscoIdentificado, pk=risco_pk)
        initial['data_avaliacao'] = timezone.now().date()
        return initial

    def form_valid(self, form):
        form.instance.criado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:risco_detail', kwargs={'pk': self.object.risco_identificado.pk})


class AvaliacaoQuantitativaUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, UpdateView):
    model = AvaliacaoQuantitativa
    form_class = AvaliacaoQuantitativaForm
    template_name = 'pgr_gestao/avaliacao_form.html'
    permission_required = 'pgr_gestao.change_avaliacaoquantitativa'

    def form_valid(self, form):
        messages.success(self.request, 'Avaliação quantitativa atualizada com sucesso!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:risco_detail', kwargs={'pk': self.object.risco_identificado.pk})


class AvaliacaoQuantitativaDeleteView(SSTPermissionMixin, ViewFilialScopedMixin, DeleteView):
    model = AvaliacaoQuantitativa
    template_name = 'pgr_gestao/confirm_delete.html'
    permission_required = 'pgr_gestao.delete_avaliacaoquantitativa'

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:risco_detail', kwargs={'pk': self.object.risco_identificado.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Avaliação quantitativa excluída com sucesso.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_name'] = 'Avaliação Quantitativa'
        context['object_info'] = f"'{self.object.get_tipo_avaliacao_display()}' de {self.object.data_avaliacao}"
        return context


# =============================================================================
# PLANOS DE AÇÃO
# =============================================================================

# Formset para acompanhamentos
AcompanhamentoFormSet = inlineformset_factory(
    PlanoAcaoPGR,
    AcompanhamentoPlanoAcao,
    fields=[
        'data_acompanhamento', 'status_atual', 'percentual_conclusao',
        'descricao', 'evidencias', 'arquivo_evidencia', 'dificuldades',
        'proximos_passos', 'responsavel_acompanhamento', 'observacoes',
    ],
    extra=1,
    can_delete=True,
)


class PlanoAcaoPGRListView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, ListView):
    model = PlanoAcaoPGR
    template_name = 'pgr_gestao/plano_acao_list.html'
    context_object_name = 'planos'
    paginate_by = 20
    permission_required = 'pgr_gestao.view_planoacaopgr'
    tecnico_scope_lookup = 'risco_identificado__pgr_documento__criado_por'

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'risco_identificado', 'risco_identificado__pgr_documento',
            'risco_identificado__tipo_risco'
        )

        status = self.request.GET.get('status')
        prioridade = self.request.GET.get('prioridade')
        tipo_acao = self.request.GET.get('tipo_acao')
        atrasado = self.request.GET.get('atrasado')
        pgr_id = self.request.GET.get('pgr')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)
        if prioridade:
            queryset = queryset.filter(prioridade=prioridade)
        if tipo_acao:
            queryset = queryset.filter(tipo_acao=tipo_acao)
        if atrasado == 'true':
            queryset = queryset.filter(
                status__in=['pendente', 'em_andamento'],
                data_prevista__lt=timezone.now().date()
            )
        if pgr_id:
            queryset = queryset.filter(risco_identificado__pgr_documento_id=pgr_id)
        if search:
            queryset = queryset.filter(
                Q(descricao_acao__icontains=search) |
                Q(responsavel__icontains=search) |
                Q(risco_identificado__agente__icontains=search)
            )
        return queryset.order_by('-prioridade', 'data_prevista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context['total_planos'] = queryset.count()
        context['planos_pendentes'] = queryset.filter(status='pendente').count()
        context['planos_em_andamento'] = queryset.filter(status='em_andamento').count()
        context['planos_concluidos'] = queryset.filter(status='concluido').count()
        context['planos_atrasados'] = queryset.filter(
            status__in=['pendente', 'em_andamento'],
            data_prevista__lt=timezone.now().date()
        ).count()
        context['documentos_pgr'] = PGRDocumento.objects.for_request(
            self.request
        ).filter(status='vigente')
        return context


class PlanoAcaoPGRDetailView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, DetailView):
    model = PlanoAcaoPGR
    template_name = 'pgr_gestao/plano_acao_detail.html'
    context_object_name = 'plano'
    permission_required = 'pgr_gestao.view_planoacaopgr'
    tecnico_scope_lookup = 'risco_identificado__pgr_documento__criado_por'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plano = self.object
        context['risco'] = plano.risco_identificado
        context['esta_atrasado'] = plano.esta_atrasado

        if plano.data_prevista:
            delta = date.today() - plano.data_prevista
            if delta.days > 0:
                context['dias_atraso'] = delta.days
            else:
                context['dias_restantes'] = abs(delta.days)
        return context


class PlanoAcaoPGRCreateView(SSTPermissionMixin, FilialCreateMixin, CreateView):
    model = PlanoAcaoPGR
    form_class = PlanoAcaoPGRForm
    template_name = 'pgr_gestao/plano_acao_form.html'
    permission_required = 'pgr_gestao.add_planoacaopgr'

    def get_initial(self):
        initial = super().get_initial()
        risco_id = self.request.GET.get('risco')
        if risco_id:
            risco = get_object_or_404(RiscoIdentificado, pk=risco_id)
            initial['risco_identificado'] = risco
            initial['prioridade'] = risco.prioridade_acao
            if risco.classificacao_risco in ['critico', 'muito_grave']:
                initial['data_prevista'] = timezone.now().date() + timedelta(days=30)
            elif risco.classificacao_risco == 'moderado':
                initial['data_prevista'] = timezone.now().date() + timedelta(days=90)
            else:
                initial['data_prevista'] = timezone.now().date() + timedelta(days=180)
        return initial

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['formset'] = AcompanhamentoFormSet(self.request.POST)
        else:
            data['formset'] = AcompanhamentoFormSet()
        return data

    def form_valid(self, form):
        form.instance.criado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:plano_acao_detail', kwargs={'pk': self.object.pk})


class PlanoAcaoPGRUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, UpdateView):
    model = PlanoAcaoPGR
    form_class = PlanoAcaoPGRForm
    template_name = 'pgr_gestao/plano_acao_form.html'
    permission_required = 'pgr_gestao.change_planoacaopgr'

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:plano_acao_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['formset'] = AcompanhamentoFormSet(self.request.POST, instance=self.object)
        else:
            data['formset'] = AcompanhamentoFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        messages.success(self.request, 'Plano de ação atualizado com sucesso!')
        return super().form_valid(form)


@login_required
@permission_required('pgr_gestao.change_planoacaopgr', raise_exception=True)
def concluir_plano_acao(request, pk):
    """Marca um plano de ação como concluído."""
    plano = get_object_or_404(
        PlanoAcaoPGR.objects.for_request(request), pk=pk
    )

    if request.method == 'POST':
        plano.status = 'concluido'
        plano.data_conclusao = timezone.now().date()
        plano.evidencia_conclusao = request.POST.get('evidencia_conclusao', '')
        plano.save()
        messages.success(request, 'Plano de ação concluído com sucesso!')
        return redirect('pgr_gestao:plano_acao_detail', pk=pk)

    return render(request, 'pgr_gestao/plano_acao_concluir.html', {'plano': plano})


# =============================================================================
# CRONOGRAMA DE AÇÕES
# =============================================================================

class CronogramaAcaoListView(SSTPermissionMixin, TecnicoScopeMixin, ViewFilialScopedMixin, ListView):
    model = CronogramaAcaoPGR
    template_name = 'pgr_gestao/cronograma_list.html'
    context_object_name = 'acoes'
    paginate_by = 20
    permission_required = 'pgr_gestao.view_cronogramaacaopgr'
    tecnico_scope_lookup = 'pgr_documento__criado_por'

    def get_queryset(self):
        qs = super().get_queryset().select_related('pgr_documento').order_by('-data_proxima_avaliacao')

        pgr = self.request.GET.get('pgr') or self.request.GET.get('pgr_documento')
        status = self.request.GET.get('status')
        search = self.request.GET.get('search') or self.request.GET.get('responsavel')

        if pgr:
            qs = qs.filter(pgr_documento_id=pgr)
        if status:
            qs = qs.filter(status=status)
        if search:
            qs = qs.filter(
                Q(acao_necessaria__icontains=search) |
                Q(responsavel__icontains=search)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = STATUS_CHOICES
        context['documentos_pgr'] = PGRDocumento.objects.for_request(self.request)

        todas = self.get_queryset()
        context['total_acoes'] = todas.count()
        context['acoes_pendentes'] = todas.filter(status='pendente').count()
        context['acoes_em_andamento'] = todas.filter(status='em_andamento').count()
        context['acoes_concluidas'] = todas.filter(status='concluido').count()
        context['acoes_atrasadas'] = todas.filter(
            status__in=['pendente', 'em_andamento'],
            data_proxima_avaliacao__lt=timezone.now().date()
        ).count()
        return context


class CronogramaAcaoDetailView(SSTPermissionMixin, ViewFilialScopedMixin, DetailView):
    model = CronogramaAcaoPGR
    template_name = 'pgr_gestao/cronograma_detail.html'
    context_object_name = 'acao'
    permission_required = 'pgr_gestao.view_cronogramaacaopgr'

    def get_queryset(self):
        return super().get_queryset().select_related('pgr_documento')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = STATUS_CHOICES
        return context


class CronogramaAcaoCreateView(SSTPermissionMixin, SuccessMessageMixin, FilialCreateMixin, CreateView):
    model = CronogramaAcaoPGR
    form_class = CronogramaAcaoPGRForm
    template_name = 'pgr_gestao/cronograma_form.html'
    success_url = reverse_lazy('pgr_gestao:cronograma_list')
    success_message = "Ação do cronograma criada com sucesso!"
    permission_required = 'pgr_gestao.add_cronogramaacaopgr'

    def get_initial(self):
        initial = super().get_initial()
        pgr_pk = self.kwargs.get('pgr_pk')
        if pgr_pk:
            initial['pgr_documento'] = pgr_pk
            ultimo = CronogramaAcaoPGR.objects.filter(
                pgr_documento_id=pgr_pk
            ).order_by('-numero_item').first()
            initial['numero_item'] = (ultimo.numero_item + 1) if ultimo else 1
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = 'Nova Ação no Cronograma'
        context['botao'] = 'Cadastrar'
        pgr_pk = self.kwargs.get('pgr_pk')
        if pgr_pk:
            context['documento_pgr'] = get_object_or_404(PGRDocumento, pk=pgr_pk)
        return context

    def get_success_url(self):
        pgr_pk = self.kwargs.get('pgr_pk')
        if pgr_pk:
            return reverse_lazy('pgr_gestao:documento_detail', kwargs={'pk': pgr_pk})
        return reverse_lazy('pgr_gestao:cronograma_list')


class CronogramaAcaoUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, SuccessMessageMixin, UpdateView):
    model = CronogramaAcaoPGR
    form_class = CronogramaAcaoPGRForm
    template_name = 'pgr_gestao/cronograma_form.html'
    success_message = "Ação do cronograma atualizada com sucesso!"
    permission_required = 'pgr_gestao.change_cronogramaacaopgr'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = f'Editar Ação #{self.object.numero_item:02d}'
        context['botao'] = 'Salvar Alterações'
        context['editando'] = True
        return context

    def get_success_url(self):
        pgr_pk = self.kwargs.get('pgr_pk')
        if pgr_pk:
            return reverse_lazy('pgr_gestao:documento_detail', kwargs={'pk': pgr_pk})
        return reverse_lazy('pgr_gestao:cronograma_detail', kwargs={'pk': self.object.pk})


class CronogramaAcaoDeleteView(SSTPermissionMixin, ViewFilialScopedMixin, DeleteView):
    model = CronogramaAcaoPGR
    template_name = 'pgr_gestao/cronograma_confirm_delete.html'
    permission_required = 'pgr_gestao.delete_cronogramaacaopgr'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_pagina'] = 'Excluir Ação'
        context['mensagem'] = f"Confirma exclusão da ação '{self.object.acao_necessaria}'?"
        return context

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Ação do cronograma excluída com sucesso!")
        return super().delete(request, *args, **kwargs)

    def get_success_url(self):
        pgr_pk = self.kwargs.get('pgr_pk')
        if pgr_pk:
            return reverse_lazy('pgr_gestao:documento_detail', kwargs={'pk': pgr_pk})
        return reverse_lazy('pgr_gestao:cronograma_list')


@login_required
@permission_required('pgr_gestao.change_cronogramaacaopgr', raise_exception=True)
def atualizar_status_acao(request, pk):
    """Atualiza o status de uma ação do cronograma."""
    if request.method == 'POST':
        acao = get_object_or_404(CronogramaAcaoPGR.objects.for_request(request), pk=pk)
        novo_status = request.POST.get('status')

        if novo_status in dict(STATUS_CHOICES):
            acao.status = novo_status
            if novo_status == 'concluido' and not acao.data_realizacao:
                acao.data_realizacao = timezone.now().date()
            acao.save()
            messages.success(request, f'Status atualizado para "{acao.get_status_display()}".')
        else:
            messages.error(request, 'Status inválido!')

    return redirect('pgr_gestao:cronograma_detail', pk=pk)


@login_required
@permission_required('pgr_gestao.change_cronogramaacaopgr', raise_exception=True)
def adicionar_anexo_plano(request, pk):
    """Adiciona anexo a um plano de ação."""
    plano = get_object_or_404(PlanoAcaoPGR.objects.for_request(request), pk=pk)

    if request.method == 'POST':
        nome = request.POST.get('nome_arquivo')
        arquivo = request.FILES.get('arquivo')

        if nome and arquivo:
            AnexoPlanoAcao.objects.create(
                plano_acao=plano,
                nome_arquivo=nome,
                arquivo=arquivo,
                criado_por=request.user
            )
            messages.success(request, 'Anexo adicionado com sucesso!')
        else:
            messages.error(request, 'Preencha todos os campos.')

    return redirect('pgr_gestao:plano_acao_detail', pk=pk)


# =============================================================================
# PROFISSIONAIS RESPONSÁVEIS
# =============================================================================

class ProfissionalResponsavelListView(SSTPermissionMixin, ViewFilialScopedMixin, ListView):
    model = ProfissionalResponsavel
    template_name = 'pgr_gestao/profissional_list.html'
    context_object_name = 'profissionais'
    paginate_by = 20
    permission_required = 'pgr_gestao.view_profissionalresponsavel'

    def get_queryset(self):
        queryset = super().get_queryset()

        ativo = self.request.GET.get('ativo')
        funcao = self.request.GET.get('funcao')
        search = self.request.GET.get('search')

        if ativo:
            queryset = queryset.filter(ativo=ativo == 'true')
        if funcao:
            queryset = queryset.filter(funcao__icontains=funcao)
        if search:
            queryset = queryset.filter(
                Q(nome_completo__icontains=search) |
                Q(registro_classe__icontains=search) |
                Q(email__icontains=search)
            )
        return queryset.order_by('nome_completo')


class ProfissionalResponsavelCreateView(SSTPermissionMixin, FilialCreateMixin, CreateView):
    model = ProfissionalResponsavel
    form_class = ProfissionalResponsavelForm
    template_name = 'pgr_gestao/profissional_form.html'
    permission_required = 'pgr_gestao.add_profissionalresponsavel'
    success_url = reverse_lazy('pgr_gestao:profissional_list')


class ProfissionalResponsavelUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, UpdateView):
    model = ProfissionalResponsavel
    form_class = ProfissionalResponsavelForm
    template_name = 'pgr_gestao/profissional_form.html'
    permission_required = 'pgr_gestao.change_profissionalresponsavel'
    success_url = reverse_lazy('pgr_gestao:profissional_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Editar Profissional: {self.object.nome_completo}'
        return context


class ProfissionalResponsavelDeleteView(SSTPermissionMixin, ViewFilialScopedMixin, DeleteView):
    model = ProfissionalResponsavel
    template_name = 'pgr_gestao/profissional_confirm_delete.html'
    permission_required = 'pgr_gestao.delete_profissionalresponsavel'
    success_url = reverse_lazy('pgr_gestao:profissional_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Confirmar Exclusão: {self.object.nome_completo}'
        context['cancel_url'] = reverse_lazy('pgr_gestao:profissional_list')
        return context


# =============================================================================
# RELATÓRIOS E ESTATÍSTICAS (UMA ÚNICA VIEW — SEM DUPLICATAS)
# =============================================================================

@login_required
@permission_required('pgr_gestao.view_pgrdocumento', raise_exception=True)
def relatorios_pgr(request):
    """Página principal de relatórios — filtrada por filial e técnico."""
    user = request.user

    # Querysets filtrados
    riscos_qs = RiscoIdentificado.objects.for_request(request)
    planos_qs = PlanoAcaoPGR.objects.for_request(request)
    documentos_pgr = PGRDocumento.objects.for_request(request).select_related('empresa')

    # Escopo técnico
    is_tecnico = user.groups.filter(name='TÉCNICO').exists()
    if is_tecnico:
        documentos_pgr = documentos_pgr.filter(criado_por=user)
        riscos_qs = riscos_qs.filter(pgr_documento__criado_por=user)
        planos_qs = planos_qs.filter(risco_identificado__pgr_documento__criado_por=user)

    # Filtros do formulário
    pgr_id = request.GET.get('pgr')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if pgr_id and documentos_pgr.filter(pk=pgr_id).exists():
        riscos_qs = riscos_qs.filter(pgr_documento_id=pgr_id)
        planos_qs = planos_qs.filter(risco_identificado__pgr_documento_id=pgr_id)

    if data_inicio:
        riscos_qs = riscos_qs.filter(criado_em__date__gte=data_inicio)
        planos_qs = planos_qs.filter(criado_em__date__gte=data_inicio)
    if data_fim:
        riscos_qs = riscos_qs.filter(criado_em__date__lte=data_fim)
        planos_qs = planos_qs.filter(criado_em__date__lte=data_fim)

    # Cards
    total_riscos = riscos_qs.count()
    riscos_criticos = riscos_qs.filter(
        classificacao_risco__in=['critico', 'muito_grave', 'intoleravel', 'substancial']
    ).count()
    riscos_controlados = riscos_qs.filter(status_controle='controlado').count()
    planos_pendentes = planos_qs.filter(status__in=['pendente', 'em_andamento']).count()

    # Gráficos
    CLASSIFICACAO_LABELS = {
        'trivial': 'Trivial', 'toleravel': 'Tolerável', 'moderado': 'Moderado',
        'substancial': 'Substancial', 'intoleravel': 'Intolerável',
        'critico': 'Crítico', 'muito_grave': 'Muito Grave',
    }
    CATEGORIA_LABELS = {
        'fisico': 'Físico', 'quimico': 'Químico', 'biologico': 'Biológico',
        'ergonomico': 'Ergonômico', 'acidente': 'Acidente', 'mecanico': 'Mecânico',
    }
    STATUS_LABELS = {
        'nao_controlado': 'Não Controlado', 'parcialmente_controlado': 'Parcialmente',
        'controlado': 'Controlado', 'nao_aplicavel': 'Não Aplicável',
    }
    PLANO_LABELS = {
        'pendente': 'Pendente', 'em_andamento': 'Em Andamento',
        'concluido': 'Concluído', 'cancelado': 'Cancelado',
    }

    def _build_chart_data(qs, field, labels_map):
        data = qs.values(field).annotate(total=Count('id')).order_by(field)
        return {
            'labels': [labels_map.get(i[field], i[field] or 'N/A') for i in data],
            'valores': [i['total'] for i in data],
        }

    riscos_por_classificacao = _build_chart_data(riscos_qs, 'classificacao_risco', CLASSIFICACAO_LABELS)
    riscos_por_categoria = _build_chart_data(riscos_qs, 'tipo_risco__categoria', CATEGORIA_LABELS)
    riscos_por_status_controle = _build_chart_data(riscos_qs, 'status_controle', STATUS_LABELS)
    planos_por_status = _build_chart_data(planos_qs, 'status', PLANO_LABELS)

    # Top 10 Riscos
    top_riscos = riscos_qs.select_related(
        'tipo_risco', 'ges'
    ).order_by('-classificacao_risco', '-gravidade_g')[:10]

    # Permissões
    pode_exportar_pdf = user.has_perm('pgr_gestao.view_pgrdocumento')
    pode_exportar_excel = user.has_perm('pgr_gestao.view_riscoidentificado')
    pode_ver_estatistico = user.has_perm('pgr_gestao.view_pgrdocumento')

    filial_info, acesso_global = _get_filial_info(request)

    context = {
        'total_riscos': total_riscos,
        'riscos_criticos': riscos_criticos,
        'riscos_controlados': riscos_controlados,
        'planos_pendentes': planos_pendentes,
        'riscos_por_classificacao': json.dumps(riscos_por_classificacao),
        'riscos_por_categoria': json.dumps(riscos_por_categoria),
        'riscos_por_status_controle': json.dumps(riscos_por_status_controle),
        'planos_por_status': json.dumps(planos_por_status),
        'top_riscos': top_riscos,
        'documentos_pgr': documentos_pgr,
        'pode_exportar_pdf': pode_exportar_pdf,
        'pode_exportar_excel': pode_exportar_excel,
        'pode_ver_estatistico': pode_ver_estatistico,
        'filial_info': filial_info,
        'acesso_global': acesso_global,
    }
    return render(request, 'pgr_gestao/relatorios.html', context)


@login_required
@permission_required('pgr_gestao.view_riscoidentificado', raise_exception=True)
def relatorio_riscos_por_classificacao(request):
    """Relatório de riscos por classificação — filtrado por filial."""
    riscos_qs = RiscoIdentificado.objects.for_request(request)
    if request.user.groups.filter(name='TÉCNICO').exists():
        riscos_qs = riscos_qs.filter(pgr_documento__criado_por=request.user)

    riscos_stats = riscos_qs.values(
        'classificacao_risco', 'status_controle'
    ).annotate(total=Count('id')).order_by('classificacao_risco', 'status_controle')

    dados_grafico = {}
    for stat in riscos_stats:
        classificacao = stat['classificacao_risco']
        status = stat['status_controle']
        if classificacao not in dados_grafico:
            dados_grafico[classificacao] = {}
        dados_grafico[classificacao][status] = stat['total']

    riscos_por_tipo = riscos_qs.values(
        'tipo_risco__nome', 'tipo_risco__categoria'
    ).annotate(total=Count('id')).order_by('-total')[:10]

    context = {
        'dados_grafico': dados_grafico,
        'riscos_stats': riscos_stats,
        'riscos_por_tipo': riscos_por_tipo,
    }
    return render(request, 'pgr_gestao/relatorio_riscos_classificacao.html', context)


@login_required
@permission_required('pgr_gestao.view_planoacaopgr', raise_exception=True)
def relatorio_planos_acao(request):
    """Relatório de planos de ação — filtrado por filial."""
    planos_qs = PlanoAcaoPGR.objects.for_request(request)
    if request.user.groups.filter(name='TÉCNICO').exists():
        planos_qs = planos_qs.filter(risco_identificado__pgr_documento__criado_por=request.user)

    planos_por_status = planos_qs.values('status').annotate(
        total=Count('id'), custo_total=Sum('custo_estimado')
    )
    planos_por_prioridade = planos_qs.values('prioridade').annotate(
        total=Count('id')
    ).order_by('-prioridade')
    planos_atrasados = planos_qs.filter(
        status__in=['pendente', 'em_andamento'],
        data_prevista__lt=timezone.now().date()
    ).select_related('risco_identificado')[:20]

    total_planos = planos_qs.count()
    planos_concluidos = planos_qs.filter(status='concluido').count()
    taxa_conclusao = (planos_concluidos / total_planos * 100) if total_planos > 0 else 0

    context = {
        'planos_por_status': planos_por_status,
        'planos_por_prioridade': planos_por_prioridade,
        'planos_atrasados': planos_atrasados,
        'taxa_conclusao': taxa_conclusao,
        'total_planos': total_planos,
        'planos_concluidos': planos_concluidos,
    }
    return render(request, 'pgr_gestao/relatorio_planos_acao.html', context)


# =============================================================================
# CONFORMIDADE
# =============================================================================

@login_required
@permission_required('pgr_gestao.view_pgrdocumento', raise_exception=True)
def verificar_conformidade_pgr(request, pk):
    """Verifica a conformidade do PGR com requisitos legais."""
    documento = get_object_or_404(
        PGRDocumento.objects.for_request(request), pk=pk
    )

    verificacoes = []

    verificacoes.append({
        'item': 'Código do documento definido',
        'conforme': bool(documento.codigo_documento),
        'observacao': documento.codigo_documento or 'Não definido',
        'peso': 1
    })

    dias_vencimento = documento.dias_para_vencimento
    verificacoes.append({
        'item': 'Data de vencimento definida e válida',
        'conforme': dias_vencimento > 0,
        'observacao': f'{dias_vencimento} dias até vencimento' if dias_vencimento > 0 else 'Documento vencido',
        'peso': 2
    })

    total_responsaveis = documento.responsaveis.count()
    verificacoes.append({
        'item': 'Responsáveis técnicos cadastrados',
        'conforme': total_responsaveis >= 1,
        'observacao': f'{total_responsaveis} responsável(is)',
        'peso': 2
    })

    total_revisoes = documento.revisoes.count()
    verificacoes.append({
        'item': 'Histórico de revisões',
        'conforme': total_revisoes >= 1,
        'observacao': f'{total_revisoes} revisão(ões)',
        'peso': 1
    })

    total_ges = documento.grupos_exposicao.filter(ativo=True).count()
    verificacoes.append({
        'item': 'GES cadastrados',
        'conforme': total_ges > 0,
        'observacao': f'{total_ges} GES ativo(s)',
        'peso': 2
    })

    total_riscos = RiscoIdentificado.objects.filter(pgr_documento=documento).count()
    verificacoes.append({
        'item': 'Inventário de riscos preenchido',
        'conforme': total_riscos > 0,
        'observacao': f'{total_riscos} risco(s)',
        'peso': 3
    })

    riscos_criticos = RiscoIdentificado.objects.filter(
        pgr_documento=documento,
        classificacao_risco__in=['critico', 'muito_grave']
    )
    total_criticos = riscos_criticos.count()
    criticos_com_plano = riscos_criticos.filter(planos_acao__isnull=False).distinct().count()
    verificacoes.append({
        'item': 'Riscos críticos com plano de ação',
        'conforme': total_criticos == 0 or criticos_com_plano == total_criticos,
        'observacao': f'{criticos_com_plano}/{total_criticos} com plano',
        'peso': 3
    })

    total_acoes = documento.cronograma_acoes.count()
    verificacoes.append({
        'item': 'Cronograma de ações (mín. 10)',
        'conforme': total_acoes >= 10,
        'observacao': f'{total_acoes} ação(ões)',
        'peso': 2
    })

    riscos_com_epi = RiscoEPIRecomendado.objects.filter(
        risco_identificado__pgr_documento=documento
    ).values('risco_identificado').distinct().count()
    verificacoes.append({
        'item': 'EPIs recomendados',
        'conforme': riscos_com_epi > 0,
        'observacao': f'{riscos_com_epi} risco(s) com EPIs',
        'peso': 1
    })

    riscos_com_treino = RiscoTreinamentoNecessario.objects.filter(
        risco_identificado__pgr_documento=documento
    ).values('risco_identificado').distinct().count()
    verificacoes.append({
        'item': 'Treinamentos necessários',
        'conforme': riscos_com_treino > 0,
        'observacao': f'{riscos_com_treino} risco(s) com treinamentos',
        'peso': 1
    })

    peso_total = sum(v['peso'] for v in verificacoes)
    peso_conforme = sum(v['peso'] for v in verificacoes if v['conforme'])
    percentual = (peso_conforme / peso_total * 100) if peso_total > 0 else 100
    nao_conformidades_criticas = [v for v in verificacoes if not v['conforme'] and v['peso'] >= 2]

    context = {
        'documento': documento,
        'verificacoes': verificacoes,
        'percentual_conformidade': percentual,
        'conformidade_total': percentual == 100,
        'nao_conformidades_criticas': nao_conformidades_criticas,
    }
    return render(request, 'pgr_gestao/verificacao_conformidade.html', context)


# =============================================================================
# AJAX / API ENDPOINTS (filtrados por filial)
# =============================================================================

@login_required
def get_locais_prestacao_ajax(request, empresa_id):
    """Retorna locais de prestação de uma empresa via AJAX."""
    locais = LocalPrestacaoServico.objects.for_request(request).filter(
        empresa_id=empresa_id, ativo=True
    ).values('id', 'razao_social')
    return JsonResponse(list(locais), safe=False)


@login_required
def get_ges_ajax(request, pgr_id):
    """Retorna GES de um documento PGR via AJAX."""
    ges = GESGrupoExposicao.objects.for_request(request).filter(
        pgr_documento_id=pgr_id, ativo=True
    ).values('id', 'codigo', 'nome')
    return JsonResponse(list(ges), safe=False)


@login_required
def dashboard_stats_ajax(request):
    """Retorna estatísticas para o dashboard via AJAX — filtrado por filial."""
    docs_qs = PGRDocumento.objects.for_request(request)
    riscos_qs = RiscoIdentificado.objects.for_request(request)
    planos_qs = PlanoAcaoPGR.objects.for_request(request)

    stats = {
        'total_documentos': docs_qs.count(),
        'documentos_vigentes': docs_qs.filter(status='vigente').count(),
        'total_riscos': riscos_qs.count(),
        'riscos_criticos': riscos_qs.filter(
            classificacao_risco__in=['critico', 'muito_grave']
        ).count(),
        'planos_pendentes': planos_qs.filter(status='pendente').count(),
        'planos_atrasados': planos_qs.filter(
            status__in=['pendente', 'em_andamento'],
            data_prevista__lt=timezone.now().date()
        ).count(),
    }
    return JsonResponse(stats)


@login_required
def ajax_get_ges(request, pgr_id):
    """Retorna GES para formulários dinâmicos."""
    ges_list = GESGrupoExposicao.objects.for_request(request).filter(
        pgr_documento_id=pgr_id
    ).values('id', 'codigo', 'nome')
    return JsonResponse(list(ges_list), safe=False)


@login_required
def ajax_get_riscos(request, pgr_id):
    """Retorna riscos para formulários dinâmicos."""
    riscos_list = RiscoIdentificado.objects.for_request(request).filter(
        pgr_documento_id=pgr_id
    ).values('id', 'codigo_risco', 'agente')
    return JsonResponse(list(riscos_list), safe=False)


@login_required
def load_locais_prestacao(request, empresa_id):
    """Carrega locais de prestação filtrados por filial."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Não autenticado'}, status=403)

    locais = LocalPrestacaoServico.objects.for_request(request).filter(empresa_id=empresa_id)
    data = list(locais.values('id', 'razao_social'))
    return JsonResponse(data, safe=False)


# =============================================================================
# EXPORTAÇÕES (filtradas por filial com validação de acesso)
# =============================================================================

@login_required
@permission_required('pgr_gestao.view_pgrdocumento', raise_exception=True)
def gerar_relatorio_completo_pdf(request, pk):
    """Gera PDF — validado por filial."""
    documento = validar_acesso_documento(request, pk)

    from .utils.pdf_generator import gerar_pdf_pgr

    try:
        pdf_buffer = gerar_pdf_pgr(documento)
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        filename = f"PGR_{documento.codigo_documento}_{documento.empresa.razao_social.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        messages.error(request, f'Erro ao gerar PDF: {str(e)}')
        return redirect('pgr_gestao:documento_detail', pk=pk)


@login_required
@permission_required('pgr_gestao.view_riscoidentificado', raise_exception=True)
def exportar_inventario_riscos_excel(request, pk):
    """Exporta inventário de riscos para Excel — validado por filial."""
    documento = validar_acesso_documento(request, pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventário de Riscos"

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'GES', 'Código Risco', 'Tipo', 'Agente', 'Fonte Geradora',
        'Perfil Exposição', 'Gravidade', 'Exposição', 'Severidade',
        'Probabilidade', 'Classificação', 'Status', 'Prioridade',
        'Medidas Controle', 'Data Identificação'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    riscos = RiscoIdentificado.objects.filter(
        pgr_documento=documento
    ).select_related('ges', 'tipo_risco').order_by('ges__codigo', 'codigo_risco')

    for row, risco in enumerate(riscos, 2):
        data = [
            risco.ges.nome if risco.ges else 'N/A',
            risco.codigo_risco or 'N/A',
            risco.tipo_risco.nome,
            risco.agente,
            risco.fonte_geradora or '',
            risco.get_perfil_exposicao_display(),
            risco.gravidade_g,
            risco.exposicao_e,
            risco.severidade_s,
            risco.probabilidade_p,
            risco.get_classificacao_risco_display(),
            risco.get_status_controle_display(),
            risco.get_prioridade_acao_display(),
            risco.medidas_controle_existentes or '',
            risco.data_identificacao.strftime('%d/%m/%Y')
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['N'].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Inventario_Riscos_{documento.codigo_documento}.xlsx"'
    return response


@login_required
@permission_required('pgr_gestao.view_planoacaopgr', raise_exception=True)
def exportar_planos_acao_excel(request):
    """Exporta planos de ação para Excel — filtrado por filial."""
    queryset = PlanoAcaoPGR.objects.for_request(request).select_related(
        'risco_identificado'
    ).order_by('-data_prevista')

    # Escopo técnico
    if request.user.groups.filter(name='TÉCNICO').exists():
        queryset = queryset.filter(risco_identificado__pgr_documento__criado_por=request.user)

    # Filtros
    status = request.GET.get('status')
    prioridade = request.GET.get('prioridade')
    tipo_acao = request.GET.get('tipo_acao')
    search_query = request.GET.get('search')

    if status:
        queryset = queryset.filter(status=status)
    if prioridade:
        queryset = queryset.filter(prioridade=prioridade)
    if tipo_acao:
        queryset = queryset.filter(tipo_acao=tipo_acao)
    if request.GET.get('atrasado') == 'true':
        queryset = queryset.filter(
            data_prevista__lt=timezone.now().date(),
            status__in=['pendente', 'em_andamento']
        )
    if search_query:
        queryset = queryset.filter(
            Q(descricao_acao__icontains=search_query) |
            Q(responsavel__icontains=search_query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planos de Ação"

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'ID', 'Descrição da Ação', 'Risco Associado', 'Tipo de Ação', 'Prioridade',
        'Status', 'Responsável', 'Data Prevista', 'Data de Conclusão',
        'Custo Estimado (R$)', 'Custo Real (R$)', 'Evidência'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for row, plano in enumerate(queryset, 2):
        data = [
            plano.pk,
            plano.descricao_acao,
            plano.risco_identificado.agente if plano.risco_identificado else 'N/A',
            plano.get_tipo_acao_display(),
            plano.get_prioridade_display(),
            plano.get_status_display(),
            plano.responsavel,
            plano.data_prevista,
            plano.data_conclusao,
            plano.custo_estimado,
            plano.custo_real,
            plano.evidencia.url if plano.evidencia else 'N/A'
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = left_align
            if isinstance(value, (date, datetime)):
                cell.number_format = 'DD/MM/YYYY'

    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['G'].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"planos_de_acao_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@permission_required('pgr_gestao.view_cronogramaacaopgr', raise_exception=True)
def exportar_cronograma_acoes(request, pk):
    """Exporta cronograma de ações para Excel — validado por filial."""
    documento = validar_acesso_documento(request, pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cronograma de Ações"

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'Item', 'Ação Necessária', 'Público Alvo', 'Periodicidade',
        'Responsável', 'Status', 'Próxima Avaliação'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    acoes = documento.cronograma_acoes.all().order_by('numero_item')

    for row_num, acao in enumerate(acoes, 2):
        data = [
            acao.numero_item,
            acao.acao_necessaria,
            acao.publico_alvo,
            acao.get_periodicidade_display(),
            acao.responsavel,
            acao.get_status_display(),
            acao.data_proxima_avaliacao.strftime('%d/%m/%Y') if acao.data_proxima_avaliacao else 'N/A'
        ]
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = left_align

    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['G'].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Cronograma_PGR_{documento.codigo_documento}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# =============================================================================
# LOCAL DE PRESTAÇÃO DE SERVIÇOS
# =============================================================================

class LocalPrestacaoCreateView(SSTPermissionMixin, CreateView):
    model = LocalPrestacaoServico
    form_class = LocalPrestacaoServicoForm
    template_name = 'pgr_gestao/local_prestacao_form.html'
    permission_required = 'pgr_gestao.add_localprestacaoservico'

    def form_valid(self, form):
        form.instance.filial = self.request.user.filial_ativa
        messages.success(self.request, 'Local de Prestação cadastrado com sucesso!')
        self.object = form.save()

        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'id': self.object.id,
                'razao_social': self.object.razao_social,
            })
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
        return super().form_invalid(form)

    def get_initial(self):
        initial = super().get_initial()
        empresa_id = self.request.GET.get('empresa')
        if empresa_id:
            initial['empresa'] = get_object_or_404(Empresa, pk=empresa_id)
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Novo Local de Prestação'
        return context

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:empresa_detail', kwargs={'pk': self.object.empresa.pk})


class LocalPrestacaoUpdateView(SSTPermissionMixin, ViewFilialScopedMixin, UpdateView):
    model = LocalPrestacaoServico
    form_class = LocalPrestacaoServicoForm
    template_name = 'pgr_gestao/local_prestacao_form.html'
    permission_required = 'pgr_gestao.change_localprestacaoservico'

    def get_success_url(self):
        return reverse_lazy('pgr_gestao:empresa_detail', kwargs={'pk': self.object.empresa.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Editar Local de Prestação'
        return context


